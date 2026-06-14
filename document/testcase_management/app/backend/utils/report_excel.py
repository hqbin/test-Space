"""
测试报告Excel生成工具

与审核报告页面(ReportReview.vue)格式完全一致
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
from datetime import datetime
import os
from utils.report_conclusion import get_conclusion_and_criteria, get_selected_fields, filter_cover_rows, is_field_visible, get_zmind_conclusion


def generate_report_excel(report_data, test_results, zmind_stats, test_cases, output_path):
    """
    生成测试报告Excel文件，与审核报告页面格式完全一致
    
    Args:
        report_data: 报告基本信息（含 include_pr_closed, has_zmind_csv）
        test_results: 测试结果列表（test_cases已减去NA）
        zmind_stats: Zmind统计信息字典
        test_cases: 用例详细列表
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # ===== 样式定义 =====
    # 字体
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    small_font = Font(name='Arial', size=9)
    bold_font = Font(name='Arial', size=10, bold=True)
    white_bold_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    
    # 对齐
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    left_top_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 填充色 - 与审核页面一致
    orange_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')  # 封面标签列
    gray_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')    # 表头+Module列
    light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # section标题
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')    # Zmind表头
    
    # 黑色边框
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 列宽
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    
    include_pr_closed = report_data.get('include_pr_closed', 0)
    has_zmind_csv = report_data.get('has_zmind_csv', False)
    
    current_row = 1
    
    # ==================== Logo + 版权信息 ====================
    # 添加Logo到左上角
    logo_path = report_data.get('logo_path', '')
    if logo_path and os.path.exists(logo_path):
        try:
            img = XlImage(logo_path)
            img.width = 80
            img.height = 24
            ws.add_image(img, 'A1')
        except Exception:
            pass
    
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    year = datetime.now().year
    cell.value = f'Confidential & Proprietary\nCopyright © {year} Whale TV Information Technology'
    cell.font = Font(name='Arial', size=9, color='666666')
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # ==================== 报告标题 ====================
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    project_name = report_data.get('project_name', '')
    title_text = f'{project_name} Report'
    cell.value = title_text
    # 根据标题长度动态调整字体大小
    if len(title_text) > 60:
        title_font_size = 14
    elif len(title_text) > 45:
        title_font_size = 16
    else:
        title_font_size = 18
    cell.font = Font(name='Arial', size=title_font_size, bold=True)
    cell.alignment = center_wrap
    # 根据标题长度动态调整行高
    title_row_height = max(35, (len(title_text) // 50 + 1) * 20)
    ws.row_dimensions[current_row].height = title_row_height
    current_row += 1

    # ==================== 封面表格 ====================
    # 构建测试结论和Release Criteria（使用共享结论生成工具）
    total_cases = sum(r.get('test_cases', 0) for r in test_results)
    total_pass = sum(r.get('pass', 0) for r in test_results)
    total_passing_rate = (total_pass / total_cases * 100) if total_cases > 0 else 0
    
    template_config = report_data.get('report_template_config')
    selected_fields = get_selected_fields(template_config)
    conclusion, release_criteria = get_conclusion_and_criteria(
        template_config, total_passing_rate, zmind_stats,
        has_zmind_csv, include_pr_closed, html_escape=False
    )
    
    cover_rows_with_keys = [
        (None, '项目名称\nProject name', report_data.get('project_name', '')),
        ('test_cycle', '测试周期\nTest Cycle', report_data.get('test_cycle', '')),
        ('testers', '测试人员\nTesters', report_data.get('testers', '')),
        ('reviewer_name', '审核人员\nReviewer', report_data.get('reviewer_name', '')),
        (None, '验证环境\nVerified environment', report_data.get('verify_env', '')),
        (None, '提测内容\nRelease Note', report_data.get('release_note', '')),
        ('test_conclusion', '测试结论\nTest Conclusion', conclusion),
        ('release_criteria', '测试通过标准\nRelease Criteria', release_criteria),
        (None, '风险评估\nRisk Assessment', report_data.get('risk_assessment', '')),
        ('remark', '备注\nRemark', report_data.get('report_remark', '')),
    ]
    cover_rows = filter_cover_rows(cover_rows_with_keys, selected_fields)
    
    for label, value in cover_rows:
        ws.merge_cells(f'B{current_row}:H{current_row}')
        # 标签列
        cell_a = ws[f'A{current_row}']
        cell_a.value = label
        cell_a.font = header_font
        cell_a.alignment = center_wrap
        cell_a.fill = orange_fill
        cell_a.border = black_border
        # 内容列
        cell_b = ws[f'B{current_row}']
        cell_b.value = value
        cell_b.font = normal_font
        cell_b.alignment = left_top_wrap
        cell_b.border = black_border
        # 合并区域的边框
        for col in range(3, 9):
            ws.cell(row=current_row, column=col).border = black_border
        
        # 行高自适应：根据内容长度和换行数计算
        content_str = str(value) if value else ''
        # 计算换行数
        line_count = content_str.count('\n') + 1
        # 估算每行字符数（合并后B-H列总宽度约79字符宽）
        merged_char_width = 75
        for line in content_str.split('\n'):
            line_count += max(0, len(line) // merged_char_width)
        # 标签列也可能有换行
        label_lines = label.count('\n') + 1
        line_count = max(line_count, label_lines)
        # 每行约15pt，最小30
        row_height = max(30, line_count * 15)
        ws.row_dimensions[current_row].height = row_height
        
        current_row += 1
    
    # 空行
    current_row += 1

    # ==================== Test Result Detail（按模板 selected_fields 控制） ====================
    if is_field_visible(selected_fields, 'test_results'):
        # 检测是否有协测列
        has_assist = report_data.get('has_assist', False)
        if not has_assist:
            has_assist = any(r.get('assist', 0) > 0 for r in test_results)
        
        num_cols = 9 if has_assist else 8
        col_letter = chr(ord('A') + num_cols - 1)  # H or I
        
        ws.merge_cells(f'A{current_row}:{col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'Test Result Detail'
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = center_wrap
        cell.fill = light_gray_fill
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        if has_assist:
            headers = ['Module', 'TestCases', 'PASS', 'FAIL', 'BLOCK', 'NT', 'NA', '协测', 'Passing rate']
            ws.column_dimensions['I'].width = 15
        else:
            headers = ['Module', 'TestCases', 'PASS', 'FAIL', 'BLOCK', 'NT', 'NA', 'Passing rate']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_wrap
            cell.fill = gray_fill
            cell.border = black_border
        current_row += 1
        
        t_cases = 0
        t_pass = 0
        t_fail = 0
        t_block = 0
        t_nt = 0
        t_na = 0
        t_assist = 0
        
        for result in test_results:
            tc = result.get('test_cases', 0)
            ps = result.get('pass', 0)
            fl = result.get('fail', 0)
            bk = result.get('block', 0)
            nt = result.get('nt', 0)
            na = result.get('na', 0)
            ast = result.get('assist', 0)
            pr = (ps / tc * 100) if tc > 0 else 0
            
            if has_assist:
                row_data = [result.get('module', ''), tc, ps, fl, bk, nt, na, ast, f'{pr:.2f}%']
            else:
                row_data = [result.get('module', ''), tc, ps, fl, bk, nt, na, f'{pr:.2f}%']
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = center_wrap
                cell.border = black_border
            ws.cell(row=current_row, column=1).fill = gray_fill
            
            t_cases += tc
            t_pass += ps
            t_fail += fl
            t_block += bk
            t_nt += nt
            t_na += na
            t_assist += ast
            current_row += 1
        
        total_pr = (t_pass / t_cases * 100) if t_cases > 0 else 0
        if has_assist:
            total_row = ['Total', t_cases, t_pass, t_fail, t_block, t_nt, t_na, t_assist, f'{total_pr:.2f}%']
        else:
            total_row = ['Total', t_cases, t_pass, t_fail, t_block, t_nt, t_na, f'{total_pr:.2f}%']
        for col_idx, val in enumerate(total_row, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = bold_font
            cell.alignment = center_wrap
            cell.border = black_border
        ws.cell(row=current_row, column=1).fill = gray_fill
        current_row += 1
        
        total_executed = t_cases
        pass_pct = (t_pass / total_executed * 100) if total_executed > 0 else 0
        fail_pct = (t_fail / total_executed * 100) if total_executed > 0 else 0
        block_pct = (t_block / total_executed * 100) if total_executed > 0 else 0
        nt_pct = (t_nt / total_executed * 100) if total_executed > 0 else 0
        ok_ng = 'OK' if total_pr >= 95 else 'NG'

        if has_assist:
            passed_row = ['Test Case Passed', f'{pass_pct + fail_pct + block_pct + nt_pct:.2f}%', f'{pass_pct:.2f}%', f'{fail_pct:.2f}%',
                          f'{block_pct:.2f}%', f'{nt_pct:.2f}%', '', '', ok_ng]
        else:
            passed_row = ['Test Case Passed', f'{pass_pct + fail_pct + block_pct + nt_pct:.2f}%', f'{pass_pct:.2f}%', f'{fail_pct:.2f}%',
                          f'{block_pct:.2f}%', f'{nt_pct:.2f}%', '', ok_ng]
        for col_idx, val in enumerate(passed_row, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = bold_font
            cell.alignment = center_wrap
            cell.border = black_border
        ws.cell(row=current_row, column=1).fill = gray_fill
        # OK/NG单元格背景色和字体色
        ok_ng_col = num_cols  # 最后一列
        ok_ng_cell = ws.cell(row=current_row, column=ok_ng_col)
        if ok_ng == 'OK':
            ok_ng_cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            ok_ng_cell.font = Font(name='Arial', size=10, bold=True, color='2E7D32')
        else:
            ok_ng_cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            ok_ng_cell.font = Font(name='Arial', size=10, bold=True, color='C62828')
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:{col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = '测试结果说明：Pass-测试通过；Fail-测试失败；NT-暂未测试；NA-不适用；'
        cell.font = Font(name='Arial', size=9, color='666666')
        cell.alignment = left_wrap
        cell.fill = gray_fill
        current_row += 1

    # ==================== Zmind-PR 表格（按模板 selected_fields 控制） ====================
    if is_field_visible(selected_fields, 'zmind_stats') and has_zmind_csv and zmind_stats:
        # 空行
        current_row += 1
        
        # Zmind-PR 表头 - 灰色背景（与审核页面一致）
        zmind_headers = ['Zmind-PR', 'PRs', 'Blocker', 'Critical', 'Major', 'Minor', 'Enhancement', 'Conclusion']
        for col_idx, header in enumerate(zmind_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_wrap
            cell.fill = gray_fill
            cell.border = black_border
        current_row += 1
        
        # 计算数据
        total_prs = zmind_stats.get('total_prs', 0)
        blocker = zmind_stats.get('blocker', 0)
        critical = zmind_stats.get('critical', 0)
        major = zmind_stats.get('major', 0)
        minor = zmind_stats.get('minor', 0)
        enhancement = zmind_stats.get('enhancement', 0)
        
        open_blocker = zmind_stats.get('open_blocker', 0)
        open_critical = zmind_stats.get('open_critical', 0)
        open_major = zmind_stats.get('open_major', 0)
        open_minor = zmind_stats.get('open_minor', 0)
        open_enhancement = zmind_stats.get('open_enhancement', 0)
        open_count = open_blocker + open_critical + open_major + open_minor + open_enhancement
        
        # 关闭率
        blocker_cr = ((blocker - open_blocker) / blocker * 100) if blocker > 0 else 100
        critical_cr = ((critical - open_critical) / critical * 100) if critical > 0 else 100
        major_cr = ((major - open_major) / major * 100) if major > 0 else 100
        
        # Conclusion判断（使用模板配置）
        zmind_conclusion = get_zmind_conclusion(template_config, zmind_stats, include_pr_closed)
        
        # 确定Zmind表格行数（用于合并Conclusion列）
        zmind_row_count = 1  # Open行始终有
        if include_pr_closed:
            zmind_row_count = 3  # Open + Total + PR closed
        
        zmind_start_row = current_row
        
        # Open行
        open_row_data = ['Open', open_count, open_blocker, open_critical, open_major, open_minor, open_enhancement, '']
        for col_idx, val in enumerate(open_row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = normal_font
            cell.alignment = center_wrap
            cell.border = black_border
        # Zmind-PR列灰色背景
        ws.cell(row=current_row, column=1).fill = gray_fill
        current_row += 1
        
        if include_pr_closed:
            # Total行
            total_row_data = ['Total', total_prs, blocker, critical, major, minor, enhancement, '']
            for col_idx, val in enumerate(total_row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = center_wrap
                cell.border = black_border
            ws.cell(row=current_row, column=1).fill = gray_fill
            current_row += 1
            
            # PR closed行
            closed_rate = ((total_prs - open_count) / total_prs * 100) if total_prs > 0 else 100
            minor_cr = ((minor - open_minor) / minor * 100) if minor > 0 else 100
            enhancement_cr = ((enhancement - open_enhancement) / enhancement * 100) if enhancement > 0 else 100
            
            closed_row_data = ['PR closed', f'{closed_rate:.2f}%', f'{blocker_cr:.2f}%', f'{critical_cr:.2f}%',
                               f'{major_cr:.2f}%', f'{minor_cr:.2f}%', f'{enhancement_cr:.2f}%', '']
            for col_idx, val in enumerate(closed_row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = center_wrap
                cell.border = black_border
            ws.cell(row=current_row, column=1).fill = gray_fill
            current_row += 1
        
        # 合并Conclusion列并填入值
        if zmind_row_count > 1:
            ws.merge_cells(f'H{zmind_start_row}:H{zmind_start_row + zmind_row_count - 1}')
        conclusion_cell = ws.cell(row=zmind_start_row, column=8)
        conclusion_cell.value = zmind_conclusion
        if zmind_conclusion == 'OK':
            conclusion_cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            conclusion_cell.font = Font(name='Arial', size=10, bold=True, color='2E7D32')
        else:
            conclusion_cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            conclusion_cell.font = Font(name='Arial', size=10, bold=True, color='C62828')
        conclusion_cell.alignment = center_wrap
        conclusion_cell.border = black_border
        
        # Severity说明
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'Severity：Blocker-致命性或阻塞进度的问题；Critical－关键功能及稳定性问题；Major－次要功能问题；Minor-界面或优化问题；Enhancement-增强建议'
        cell.font = Font(name='Arial', size=9, color='666666')
        cell.alignment = left_wrap
        cell.fill = gray_fill
        current_row += 1
    
    # ==================== TestCases Sheet（按模板 selected_fields 控制） ====================
    if is_field_visible(selected_fields, 'test_cases') and test_cases:
        ws2 = wb.create_sheet(title="TestCases")
        
        # 列宽
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 25
        ws2.column_dimensions['D'].width = 20
        ws2.column_dimensions['E'].width = 30
        ws2.column_dimensions['F'].width = 25
        ws2.column_dimensions['G'].width = 10
        ws2.column_dimensions['H'].width = 10
        ws2.column_dimensions['I'].width = 25
        
        # 表头 - 不换行
        header_nowrap = Alignment(horizontal='center', vertical='center', wrap_text=False)
        tc_headers = ['用例编号', '所属模块', '用例标题', '前置条件', '操作步骤', '预期结果', '用例等级', '测试结果', '执行备注']
        for col_idx, header in enumerate(tc_headers, start=1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = header_nowrap
            cell.fill = gray_fill
            cell.border = black_border
        
        result_map = {'PASS': 'PASS', 'FAIL': 'FAIL', 'BLOCK': 'BLOCK', 'NT': 'NT', 'NA': 'NA'}
        
        for row_idx, tc in enumerate(test_cases, start=2):
            row_data = [
                tc.get('case_number', ''),
                tc.get('module', ''),
                tc.get('name', ''),
                tc.get('precondition', ''),
                tc.get('steps', ''),
                tc.get('expected_result', ''),
                tc.get('level', ''),
                result_map.get(tc.get('result', ''), tc.get('result', '')),
                tc.get('remark', ''),
            ]
            max_lines = 1
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = left_top_wrap
                cell.border = black_border
                # 计算行数用于自适应高度
                val_str = str(val) if val else ''
                lines = val_str.count('\n') + 1
                col_letter = chr(64 + col_idx)
                col_width = ws2.column_dimensions[col_letter].width or 15
                for line in val_str.split('\n'):
                    lines += max(0, len(line) // int(col_width))
                max_lines = max(max_lines, lines)
            ws2.row_dimensions[row_idx].height = max(15, max_lines * 15)
    
    # 保存文件
    wb.save(output_path)
    return output_path


def generate_report_excel_stream(report_data, test_results, zmind_stats, test_cases, progress_callback=None):
    """
    生成测试报告Excel到内存流，不保存到文件
    
    Args:
        report_data: 报告基本信息
        test_results: 测试结果列表
        zmind_stats: Zmind统计信息字典
        test_cases: 用例详细列表
    
    Returns:
        BytesIO: Excel文件的内存流
    """
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # ===== 样式定义 =====
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    small_font = Font(name='Arial', size=9)
    bold_font = Font(name='Arial', size=10, bold=True)
    white_bold_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    left_top_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    orange_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
    gray_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    
    include_pr_closed = report_data.get('include_pr_closed', 0)
    has_zmind_csv = report_data.get('has_zmind_csv', False)
    
    current_row = 1
    
    # Logo + 版权信息
    logo_path = report_data.get('logo_path', '')
    if logo_path and os.path.exists(logo_path):
        try:
            img = XlImage(logo_path)
            img.width = 80
            img.height = 24
            ws.add_image(img, 'A1')
        except Exception:
            pass
    
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    year = datetime.now().year
    cell.value = f'Confidential & Proprietary\nCopyright © {year} Whale TV Information Technology'
    cell.font = Font(name='Arial', size=9, color='666666')
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # 报告标题
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    project_name = report_data.get('project_name', '')
    title_text = f'{project_name} Report'
    cell.value = title_text
    # 根据标题长度动态调整字体大小
    if len(title_text) > 60:
        title_font_size = 14
    elif len(title_text) > 45:
        title_font_size = 16
    else:
        title_font_size = 18
    cell.font = Font(name='Arial', size=title_font_size, bold=True)
    cell.alignment = center_wrap
    # 根据标题长度动态调整行高
    title_row_height = max(35, (len(title_text) // 50 + 1) * 20)
    ws.row_dimensions[current_row].height = title_row_height
    current_row += 1

    # 封面表格
    total_cases = sum(r.get('test_cases', 0) for r in test_results)
    total_pass = sum(r.get('pass', 0) for r in test_results)
    total_passing_rate = (total_pass / total_cases * 100) if total_cases > 0 else 0
    
    template_config = report_data.get('report_template_config')
    selected_fields = get_selected_fields(template_config)
    conclusion, release_criteria = get_conclusion_and_criteria(
        template_config, total_passing_rate, zmind_stats,
        has_zmind_csv, include_pr_closed, html_escape=False
    )
    
    cover_rows_with_keys = [
        (None, '项目名称\nProject name', report_data.get('project_name', '')),
        ('test_cycle', '测试周期\nTest Cycle', report_data.get('test_cycle', '')),
        ('testers', '测试人员\nTesters', report_data.get('testers', '')),
        ('reviewer_name', '审核人员\nReviewer', report_data.get('reviewer_name', '')),
        (None, '验证环境\nVerified environment', report_data.get('verify_env', '')),
        (None, '提测内容\nRelease Note', report_data.get('release_note', '')),
        ('test_conclusion', '测试结论\nTest Conclusion', conclusion),
        ('release_criteria', '测试通过标准\nRelease Criteria', release_criteria),
        (None, '风险评估\nRisk Assessment', report_data.get('risk_assessment', '')),
        ('remark', '备注\nRemark', report_data.get('report_remark', '')),
    ]
    cover_rows = filter_cover_rows(cover_rows_with_keys, selected_fields)
    
    for label, value in cover_rows:
        ws.merge_cells(f'B{current_row}:H{current_row}')
        cell_a = ws[f'A{current_row}']
        cell_a.value = label
        cell_a.font = header_font
        cell_a.alignment = center_wrap
        cell_a.fill = orange_fill
        cell_a.border = black_border
        cell_b = ws[f'B{current_row}']
        cell_b.value = value
        cell_b.font = normal_font
        cell_b.alignment = left_top_wrap
        cell_b.border = black_border
        for col in range(3, 9):
            ws.cell(row=current_row, column=col).border = black_border
        
        content_str = str(value) if value else ''
        line_count = content_str.count('\n') + 1
        merged_char_width = 75
        for line in content_str.split('\n'):
            line_count += max(0, len(line) // merged_char_width)
        label_lines = label.count('\n') + 1
        line_count = max(line_count, label_lines)
        row_height = max(30, line_count * 15)
        ws.row_dimensions[current_row].height = row_height
        current_row += 1
    
    current_row += 1

    # Test Result Detail（按模板 selected_fields 控制）
    if is_field_visible(selected_fields, 'test_results'):
        # 检测是否有协测列
        has_assist = report_data.get('has_assist', False)
        if not has_assist:
            has_assist = any(r.get('assist', 0) > 0 for r in test_results)
        
        num_cols = 9 if has_assist else 8
        col_letter = chr(ord('A') + num_cols - 1)
        
        ws.merge_cells(f'A{current_row}:{col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'Test Result Detail'
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = center_wrap
        cell.fill = light_gray_fill
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        if has_assist:
            headers = ['Module', 'TestCases', 'PASS', 'FAIL', 'BLOCK', 'NT', 'NA', '协测', 'Passing rate']
            ws.column_dimensions['I'].width = 15
        else:
            headers = ['Module', 'TestCases', 'PASS', 'FAIL', 'BLOCK', 'NT', 'NA', 'Passing rate']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_wrap
            cell.fill = gray_fill
            cell.border = black_border
        current_row += 1
        
        t_cases = 0
        t_pass = 0
        t_fail = 0
        t_block = 0
        t_nt = 0
        t_na = 0
        t_assist = 0
        
        for result in test_results:
            tc = result.get('test_cases', 0)
            ps = result.get('pass', 0)
            fl = result.get('fail', 0)
            bk = result.get('block', 0)
            nt = result.get('nt', 0)
            na = result.get('na', 0)
            ast = result.get('assist', 0)
            pr = (ps / tc * 100) if tc > 0 else 0
            
            if has_assist:
                row_data = [result.get('module', ''), tc, ps, fl, bk, nt, na, ast, f'{pr:.2f}%']
            else:
                row_data = [result.get('module', ''), tc, ps, fl, bk, nt, na, f'{pr:.2f}%']
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = center_wrap
                cell.border = black_border
            ws.cell(row=current_row, column=1).fill = gray_fill
            
            t_cases += tc
            t_pass += ps
            t_fail += fl
            t_block += bk
            t_nt += nt
            t_na += na
            t_assist += ast
            current_row += 1
        
        total_pr = (t_pass / t_cases * 100) if t_cases > 0 else 0
        if has_assist:
            total_row = ['Total', t_cases, t_pass, t_fail, t_block, t_nt, t_na, t_assist, f'{total_pr:.2f}%']
        else:
            total_row = ['Total', t_cases, t_pass, t_fail, t_block, t_nt, t_na, f'{total_pr:.2f}%']
        for col_idx, val in enumerate(total_row, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = bold_font
            cell.alignment = center_wrap
            cell.border = black_border
        ws.cell(row=current_row, column=1).fill = gray_fill
        current_row += 1
        
        total_executed = t_cases
        pass_pct = (t_pass / total_executed * 100) if total_executed > 0 else 0
        fail_pct = (t_fail / total_executed * 100) if total_executed > 0 else 0
        block_pct = (t_block / total_executed * 100) if total_executed > 0 else 0
        nt_pct = (t_nt / total_executed * 100) if total_executed > 0 else 0
        ok_ng = 'OK' if total_pr >= 95 else 'NG'

        if has_assist:
            passed_row = ['Test Case Passed', f'{pass_pct + fail_pct + block_pct + nt_pct:.2f}%', f'{pass_pct:.2f}%', f'{fail_pct:.2f}%',
                          f'{block_pct:.2f}%', f'{nt_pct:.2f}%', '', '', ok_ng]
        else:
            passed_row = ['Test Case Passed', f'{pass_pct + fail_pct + block_pct + nt_pct:.2f}%', f'{pass_pct:.2f}%', f'{fail_pct:.2f}%',
                          f'{block_pct:.2f}%', f'{nt_pct:.2f}%', '', ok_ng]
        for col_idx, val in enumerate(passed_row, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = bold_font
            cell.alignment = center_wrap
            cell.border = black_border
        ws.cell(row=current_row, column=1).fill = gray_fill
        ok_ng_col = num_cols
        ok_ng_cell = ws.cell(row=current_row, column=ok_ng_col)
        if ok_ng == 'OK':
            ok_ng_cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            ok_ng_cell.font = Font(name='Arial', size=10, bold=True, color='2E7D32')
        else:
            ok_ng_cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            ok_ng_cell.font = Font(name='Arial', size=10, bold=True, color='C62828')
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:{col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = '测试结果说明：Pass-测试通过；Fail-测试失败；NT-暂未测试；NA-不适用；'
        cell.font = Font(name='Arial', size=9, color='666666')
        cell.alignment = left_wrap
        cell.fill = gray_fill
        current_row += 1

    # Zmind-PR 表格（按模板 selected_fields 控制）
    if is_field_visible(selected_fields, 'zmind_stats') and has_zmind_csv and zmind_stats:
        current_row += 1
        
        zmind_headers = ['Zmind-PR', 'PRs', 'Blocker', 'Critical', 'Major', 'Minor', 'Enhancement', 'Conclusion']
        for col_idx, header in enumerate(zmind_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_wrap
            cell.fill = gray_fill
            cell.border = black_border
        current_row += 1
        
        total_prs = zmind_stats.get('total_prs', 0)
        blocker = zmind_stats.get('blocker', 0)
        critical = zmind_stats.get('critical', 0)
        major = zmind_stats.get('major', 0)
        minor = zmind_stats.get('minor', 0)
        enhancement = zmind_stats.get('enhancement', 0)
        
        open_blocker = zmind_stats.get('open_blocker', 0)
        open_critical = zmind_stats.get('open_critical', 0)
        open_major = zmind_stats.get('open_major', 0)
        open_minor = zmind_stats.get('open_minor', 0)
        open_enhancement = zmind_stats.get('open_enhancement', 0)
        open_count = open_blocker + open_critical + open_major + open_minor + open_enhancement
        
        blocker_cr = ((blocker - open_blocker) / blocker * 100) if blocker > 0 else 100
        critical_cr = ((critical - open_critical) / critical * 100) if critical > 0 else 100
        major_cr = ((major - open_major) / major * 100) if major > 0 else 100
        
        zmind_conclusion = get_zmind_conclusion(template_config, zmind_stats, include_pr_closed)
        
        zmind_row_count = 1
        if include_pr_closed:
            zmind_row_count = 3
        
        zmind_start_row = current_row
        
        open_row_data = ['Open', open_count, open_blocker, open_critical, open_major, open_minor, open_enhancement, '']
        for col_idx, val in enumerate(open_row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = normal_font
            cell.alignment = center_wrap
            cell.border = black_border
        ws.cell(row=current_row, column=1).fill = gray_fill
        current_row += 1
        
        if include_pr_closed:
            total_row_data = ['Total', total_prs, blocker, critical, major, minor, enhancement, '']
            for col_idx, val in enumerate(total_row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = center_wrap
                cell.border = black_border
            ws.cell(row=current_row, column=1).fill = gray_fill
            current_row += 1
            
            closed_rate = ((total_prs - open_count) / total_prs * 100) if total_prs > 0 else 100
            minor_cr = ((minor - open_minor) / minor * 100) if minor > 0 else 100
            enhancement_cr = ((enhancement - open_enhancement) / enhancement * 100) if enhancement > 0 else 100
            
            closed_row_data = ['PR closed', f'{closed_rate:.2f}%', f'{blocker_cr:.2f}%', f'{critical_cr:.2f}%',
                               f'{major_cr:.2f}%', f'{minor_cr:.2f}%', f'{enhancement_cr:.2f}%', '']
            for col_idx, val in enumerate(closed_row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = center_wrap
                cell.border = black_border
            ws.cell(row=current_row, column=1).fill = gray_fill
            current_row += 1
        
        if zmind_row_count > 1:
            ws.merge_cells(f'H{zmind_start_row}:H{zmind_start_row + zmind_row_count - 1}')
        conclusion_cell = ws.cell(row=zmind_start_row, column=8)
        conclusion_cell.value = zmind_conclusion
        if zmind_conclusion == 'OK':
            conclusion_cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            conclusion_cell.font = Font(name='Arial', size=10, bold=True, color='2E7D32')
        else:
            conclusion_cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            conclusion_cell.font = Font(name='Arial', size=10, bold=True, color='C62828')
        conclusion_cell.alignment = center_wrap
        conclusion_cell.border = black_border
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'Severity：Blocker-致命性或阻塞进度的问题；Critical－关键功能及稳定性问题；Major－次要功能问题；Minor-界面或优化问题；Enhancement-增强建议'
        cell.font = Font(name='Arial', size=9, color='666666')
        cell.alignment = left_wrap
        cell.fill = gray_fill
        current_row += 1
    
    # TestCases Sheet（按模板 selected_fields 控制）
    if is_field_visible(selected_fields, 'test_cases') and test_cases:
        ws2 = wb.create_sheet(title="TestCases")
        
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 25
        ws2.column_dimensions['D'].width = 20
        ws2.column_dimensions['E'].width = 30
        ws2.column_dimensions['F'].width = 25
        ws2.column_dimensions['G'].width = 10
        ws2.column_dimensions['H'].width = 10
        ws2.column_dimensions['I'].width = 25
        
        header_nowrap = Alignment(horizontal='center', vertical='center', wrap_text=False)
        tc_headers = ['用例编号', '所属模块', '用例标题', '前置条件', '操作步骤', '预期结果', '用例等级', '测试结果', '执行备注']
        for col_idx, header in enumerate(tc_headers, start=1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = header_nowrap
            cell.fill = gray_fill
            cell.border = black_border
        
        result_map = {'PASS': 'PASS', 'FAIL': 'FAIL', 'BLOCK': 'BLOCK', 'NT': 'NT', 'NA': 'NA'}
        
        _total_tc = len(test_cases)
        _progress_start = 15
        _progress_range = 25
        _next_progress = _progress_start + 5
        for row_idx, tc in enumerate(test_cases, start=2):
            row_data = [
                tc.get('case_number', ''),
                tc.get('module', ''),
                tc.get('name', ''),
                tc.get('precondition', ''),
                tc.get('steps', ''),
                tc.get('expected_result', ''),
                tc.get('level', ''),
                result_map.get(tc.get('result', ''), tc.get('result', '')),
                tc.get('remark', ''),
            ]
            max_lines = 1
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = left_top_wrap
                cell.border = black_border
                val_str = str(val) if val else ''
                lines = val_str.count('\n') + 1
                col_letter = chr(64 + col_idx)
                col_width = ws2.column_dimensions[col_letter].width or 15
                for line in val_str.split('\n'):
                    lines += max(0, len(line) // int(col_width))
                max_lines = max(max_lines, lines)
            ws2.row_dimensions[row_idx].height = max(15, max_lines * 15)
            if progress_callback and _total_tc > 0:
                _pct = _progress_start + int((row_idx - 1) / _total_tc * _progress_range)
                if _pct >= _next_progress:
                    _next_progress = _pct + 5
                    progress_callback(_pct, f'正在生成Excel内容 ({row_idx - 1}/{_total_tc})')
    
    # MpList Sheet（在Issues前面，倒二位置）
    mplist_data = report_data.get('mplist_data', {})
    mplist_headers = mplist_data.get('headers', []) if isinstance(mplist_data, dict) else []
    mplist_rows = mplist_data.get('rows', []) if isinstance(mplist_data, dict) else []
    if mplist_headers and mplist_rows:
        from openpyxl.utils import get_column_letter
        ws_mp = wb.create_sheet(title="Mp List")
        
        # 动态列宽
        for col_idx, header in enumerate(mplist_headers):
            col_letter = get_column_letter(col_idx + 1)
            header_len = len(str(header))
            ws_mp.column_dimensions[col_letter].width = max(12, min(50, header_len * 2 + 4))
        
        # 表头
        header_nowrap = Alignment(horizontal='center', vertical='center', wrap_text=False)
        for col_idx, header in enumerate(mplist_headers, start=1):
            cell = ws_mp.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = header_nowrap
            cell.fill = gray_fill
            cell.border = black_border
        
        for row_idx, row_data in enumerate(mplist_rows, start=2):
            max_lines = 1
            for col_idx, val in enumerate(row_data, start=1):
                if col_idx > len(mplist_headers):
                    break
                cell = ws_mp.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = left_top_wrap
                cell.border = black_border
                
                # 计算行数用于自适应高度
                val_str = str(val) if val else ''
                lines = val_str.count('\n') + 1
                col_letter = get_column_letter(col_idx)
                try:
                    col_width = ws_mp.column_dimensions[col_letter].width or 15
                except (KeyError, AttributeError):
                    col_width = 15
                for line in val_str.split('\n'):
                    lines += max(0, len(line) // int(col_width))
                max_lines = max(max_lines, lines)
            
            ws_mp.row_dimensions[row_idx].height = max(15, max_lines * 15)
    
    # Issues Sheet（按模板 selected_fields 控制）
    issue_list = report_data.get('issue_list', [])
    if is_field_visible(selected_fields, 'issue_list') and issue_list:
        ws3 = wb.create_sheet(title="Issues")
        
        # 列宽
        ws3.column_dimensions['A'].width = 12  # PR号
        ws3.column_dimensions['B'].width = 10  # 跟踪
        ws3.column_dimensions['C'].width = 12  # 类别
        ws3.column_dimensions['D'].width = 12  # Severity
        ws3.column_dimensions['E'].width = 15  # 状态
        ws3.column_dimensions['F'].width = 10  # 优先级
        ws3.column_dimensions['G'].width = 50  # 主题
        ws3.column_dimensions['H'].width = 12  # 指派给
        
        # 表头
        issue_headers = ['#(PR号)', '跟踪', '类别', 'Severity', '状态', '优先级', '主题', '指派给']
        header_nowrap = Alignment(horizontal='center', vertical='center', wrap_text=False)
        for col_idx, header in enumerate(issue_headers, start=1):
            cell = ws3.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = header_nowrap
            cell.fill = gray_fill
            cell.border = black_border
        
        # Severity颜色映射
        severity_colors = {
            'blocker': 'FF0000',    # 红色
            'critical': 'FF6600',   # 橙色
            'major': '0066FF',      # 蓝色
            'minor': '00B050',      # 绿色
            'enhancement': '808080'  # 灰色
        }
        
        # Open状态列表（取自共享常量，不含 Pending / Device Issue / App Issue）
        from utils.constants import OPEN_STATUS_ORDER
        open_statuses = set(OPEN_STATUS_ORDER)
        
        for row_idx, issue in enumerate(issue_list, start=2):
            row_data = [
                issue.get('pr_number', ''),
                issue.get('tracker', ''),
                issue.get('category', ''),
                issue.get('severity', ''),
                issue.get('status', ''),
                issue.get('priority', ''),
                issue.get('subject', ''),
                issue.get('assignee', ''),
            ]
            max_lines = 1
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = normal_font
                cell.alignment = center_wrap if col_idx != 7 else left_top_wrap  # 主题列左对齐且自动换行
                cell.border = black_border
                
                # 计算行数用于自适应高度
                val_str = str(val) if val else ''
                lines = val_str.count('\n') + 1
                col_letter = chr(64 + col_idx)
                col_width = ws3.column_dimensions[col_letter].width or 15
                for line in val_str.split('\n'):
                    lines += max(0, len(line) // int(col_width))
                max_lines = max(max_lines, lines)
                
                # Severity列颜色
                if col_idx == 4 and val:
                    severity_lower = val.lower()
                    if severity_lower in severity_colors:
                        cell.font = Font(name='Arial', size=10, bold=True, color=severity_colors[severity_lower])
                
                # 状态列颜色
                if col_idx == 5 and val:
                    if val in open_statuses:
                        cell.font = Font(name='Arial', size=10, bold=True, color='FF0000')  # 红色
                    elif val in ['Closed', 'Confirm Issue', 'Pending', 'Device Issue', 'App Issue']:
                        cell.font = Font(name='Arial', size=10, color='00B050')  # 绿色
            
            # 设置行高自适应
            ws3.row_dimensions[row_idx].height = max(15, max_lines * 15)
    
    # 保存到内存流
    if progress_callback: progress_callback(50, '正在写入Excel文件...')
    stream = io.BytesIO()
    wb.save(stream)
    if progress_callback: progress_callback(85, 'Excel文件生成完成')
    stream.seek(0)
    return stream
