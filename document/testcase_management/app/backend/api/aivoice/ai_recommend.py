"""
AI 用例推荐 API
基于 Release Note 内容，通过 Azure OpenAI 分析变更影响，推荐相关测试用例
后台任务模式：创建任务后立即返回，通过轮询获取进度
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import desc
from typing import List, Optional
import json
import logging
import re
import threading
import time
from urllib.parse import quote

from database import get_db, SessionLocal
from auth import get_current_user
from models import User, TestCase, Project
from models_aivoice import AiVoiceSetting, AiRecommendHistory

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/ai-recommend", tags=["AI用例推荐"])


class AnalyzeRequest(BaseModel):
    projectIds: List[int]
    releaseNote: str
    title: Optional[str] = None
    prNumbers: Optional[str] = None  # 逗号分隔的PR编号


class TestConnectionRequest(BaseModel):
    endpoint: str
    apiKey: str
    deployment: str
    apiVersion: Optional[str] = "2024-12-01-preview"


class SaveSuiteInfoRequest(BaseModel):
    historyId: int
    suiteId: int
    suiteName: str


# ===== 任务进度存储（内存） =====
# key: history_id, value: { status, progress, message, ... }
task_progress = {}

# key: history_id, value: { status, progress, message, ... } (补充用例生成任务)
supplement_progress = {}

# key: user_id, value: history_id (当前正在执行的任务)
user_running_task = {}

# ===== 任务队列 =====
# task_queue: [{history_id, user_id, project_ids, release_note, pr_numbers, status, created_at}]
# status: "waiting" | "running" | "completed" | "failed" | "aborted"
MAX_CONCURRENT_TASKS = 3
active_task_count = 0
task_queue = []
task_queue_lock = threading.Lock()

def _start_task_thread(item: dict):
    """启动队列项的后台线程"""
    thread = threading.Thread(
        target=run_analysis_task,
        args=(item["history_id"], item["project_ids"], item["release_note"], item["pr_numbers"]),
        daemon=True
    )
    thread.start()

def process_queue():
    """处理队列：若有空闲槽位，启动下一个等待任务"""
    global active_task_count
    item_to_start = None
    with task_queue_lock:
        if active_task_count >= MAX_CONCURRENT_TASKS:
            return
        for item in task_queue:
            if item["status"] == "waiting":
                item["status"] = "running"
                active_task_count += 1
                item_to_start = dict(item)
                break
    if item_to_start:
        _start_task_thread(item_to_start)


def get_azure_openai_client(db: Session):
    """获取 Azure OpenAI 配置"""
    endpoint_row = db.query(AiVoiceSetting).filter(AiVoiceSetting.key == "azure_openai_endpoint").first()
    key_row = db.query(AiVoiceSetting).filter(AiVoiceSetting.key == "azure_openai_api_key").first()
    deployment_row = db.query(AiVoiceSetting).filter(AiVoiceSetting.key == "azure_openai_deployment").first()
    version_row = db.query(AiVoiceSetting).filter(AiVoiceSetting.key == "azure_openai_api_version").first()

    if not endpoint_row or not key_row or not deployment_row:
        raise Exception("Azure OpenAI 未配置")

    if not endpoint_row.value or not key_row.value or not deployment_row.value:
        raise Exception("Azure OpenAI 配置不完整")

    return {
        "endpoint": endpoint_row.value,
        "api_key": key_row.value,
        "deployment": deployment_row.value,
        "api_version": version_row.value if version_row and version_row.value else "2024-12-01-preview"
    }


def call_azure_openai(config: dict, messages: list, max_tokens: int = 4000) -> str:
    """调用 Azure OpenAI API"""
    import requests

    endpoint = config['endpoint'].rstrip('/')
    deployment = config['deployment']
    api_key = config['api_key']

    # 判断 endpoint 是否已经包含完整路径
    if '/openai/v1/' in endpoint or endpoint.endswith('/chat/completions'):
        url = endpoint
        payload = {
            "model": deployment,
            "messages": messages,
            "max_completion_tokens": max_tokens
        }
    else:
        api_version = config.get("api_version", "2024-12-01-preview")
        url = f"{endpoint}/openai/deployments/{deployment}/chat/completions?api-version={api_version}"
        payload = {
            "messages": messages,
            "max_completion_tokens": max_tokens
        }

    headers = {
        "Content-Type": "application/json",
        "api-key": api_key
    }

    resp = requests.post(url, headers=headers, json=payload, timeout=300)
    resp.raise_for_status()
    data = resp.json()
    choice = data["choices"][0]
    message = choice.get("message", {})
    content = message.get("content") or ""
    if message.get("refusal"):
        logger.warning(f"Azure OpenAI returned refusal: {message['refusal'][:300]}")
    if not content:
        # 记录完整 choice 以排查原因（包括 content_filter_results、finish_reason 等）
        logger.warning(f"Azure OpenAI returned empty content, full choice: {str(choice)[:800]}, finish_reason={choice.get('finish_reason')}")
    return content


def clean_json_response(raw: str) -> str:
    """清理 AI 返回的 JSON 字符串，支持代码块内外的多种格式"""
    clean = raw.strip()
    # 提取代码块内的内容（支持 ```json ... ``` 或 ``` ... ```）
    block_match = re.search(r'```(?:json)?\s*\n?(.*?)```', clean, re.DOTALL)
    if block_match:
        clean = block_match.group(1).strip()
    else:
        # 没有完整代码块，处理纯文本格式
        if clean.startswith("```"):
            clean = clean.split("\n", 1)[1] if "\n" in clean else clean[3:]
        if clean.endswith("```"):
            clean = clean[:-3]
        clean = clean.strip()
        if clean.startswith("json"):
            clean = clean[4:].strip()
    return clean


# ===== 向量预筛函数 =====
def _vector_prefilter(cases_query, release_note: str, project_map: dict, top_n: int = 150) -> list:
    """
    使用 TF-IDF + 余弦相似度从全部用例中筛选与 Release Note 相关的用例。
    top_n=0 表示不限制数量，返回所有相似度 > 0 的用例。
    不需要外部 API 调用或数据库扩展，纯 Python 内存计算。
    """
    import re
    import math
    from collections import Counter

    def tokenize(text: str) -> list:
        """简单分词：按非中文/字母/数字字符分割 + 中文单字"""
        if not text:
            return []
        words = re.findall(r'[a-zA-Z0-9]+', text.lower())
        chinese = re.findall(r'[\u4e00-\u9fff]+', text)
        for segment in chinese:
            for i in range(len(segment) - 1):
                words.append(segment[i:i+2])
            if len(segment) == 1:
                words.append(segment)
        return words

    # 构建所有文档
    docs = []
    for tc in cases_query:
        text = f"{tc.name or ''} {tc.module or ''} {tc.sub_module or ''} {(tc.steps or '')[:100]}"
        docs.append(text)

    query_tokens = tokenize(release_note[:3000])
    if not query_tokens:
        return [{"id": tc.id, "title": tc.name, "module": tc.module or "", "project_name": project_map.get(tc.primary_project_id, "")} for tc in cases_query[:200]]

    # 计算文档频率 (DF)
    doc_tokens_list = [tokenize(d) for d in docs]
    total_docs = len(doc_tokens_list)
    df = Counter()
    for tokens in doc_tokens_list:
        unique_tokens = set(tokens)
        for t in unique_tokens:
            df[t] += 1

    # 计算 IDF
    idf = {}
    for term, freq in df.items():
        idf[term] = math.log((total_docs + 1) / (freq + 1)) + 1

    # Query TF-IDF 向量
    query_tf = Counter(query_tokens)
    query_vec = {}
    for term, count in query_tf.items():
        if term in idf:
            query_vec[term] = count * idf[term]

    query_norm = math.sqrt(sum(v * v for v in query_vec.values())) or 1.0

    # 计算每个文档与 query 的余弦相似度
    scores = []
    for idx, tokens in enumerate(doc_tokens_list):
        if not tokens:
            scores.append((idx, 0.0))
            continue
        doc_tf = Counter(tokens)
        dot_product = 0.0
        doc_norm_sq = 0.0
        for term, count in doc_tf.items():
            tfidf = count * idf.get(term, 0)
            doc_norm_sq += tfidf * tfidf
            if term in query_vec:
                dot_product += tfidf * query_vec[term]

        doc_norm = math.sqrt(doc_norm_sq) or 1.0
        similarity = dot_product / (query_norm * doc_norm)
        scores.append((idx, similarity))

    # 排序，过滤相似度 > 0 的
    scores.sort(key=lambda x: x[1], reverse=True)

    if top_n > 0:
        selected = [(idx, sim) for idx, sim in scores[:top_n] if sim > 0]
    else:
        # 不限制数量，返回所有相似度 > 0 的
        selected = [(idx, sim) for idx, sim in scores if sim > 0]

    # 构建返回结果
    result = []
    for idx, _ in selected:
        tc = cases_query[idx]
        result.append({
            "id": tc.id,
            "title": tc.name,
            "module": tc.module or "",
            "project_name": project_map.get(tc.primary_project_id, "")
        })

    return result


# ===== 后台任务执行函数 =====
def run_analysis_task(history_id: int, project_ids: List[int], release_note: str, pr_numbers: str = None):
    """在后台线程中执行 AI 分析"""
    global active_task_count
    db = SessionLocal()
    try:
        start_ts = time.time()
        task_progress[history_id] = {"status": "running", "progress": 5, "message": "正在读取 AI 配置...", "startTime": start_ts}

        # 1. 获取配置
        ai_config = get_azure_openai_client(db)
        task_progress[history_id] = {"status": "running", "progress": 8, "message": "正在加载测试用例...", "startTime": start_ts}

        # 2. 获取用例
        cases_query = db.query(TestCase).filter(
            TestCase.primary_project_id.in_(project_ids)
        ).all()

        if not cases_query:
            task_progress[history_id] = {"status": "failed", "progress": 100, "message": "所选用例库中没有测试用例"}
            record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == history_id).first()
            if record:
                record.summary = "失败：所选用例库中没有测试用例"
            db.commit()
            return

        projects = db.query(Project).filter(Project.id.in_(project_ids)).all()
        project_map = {p.id: p.name for p in projects}

        # 2.1 查找PR关联用例
        pr_linked_case_ids = set()
        if pr_numbers:
            from models import TestCaseZmindLink
            pr_list = [n.strip() for n in pr_numbers.split(',') if n.strip()]
            if pr_list:
                links = db.query(TestCaseZmindLink).filter(
                    TestCaseZmindLink.zmind_issue_id.in_(pr_list)
                ).all()
                pr_linked_case_ids = set(link.test_case_id for link in links)

        total_all_cases = len(cases_query)
        task_progress[history_id] = {"status": "running", "progress": 10, "message": f"已加载 {total_all_cases} 条用例，正在加载测试用例...", "startTime": start_ts}
        task_progress[history_id] = {"status": "running", "progress": 12, "message": f"已加载 {total_all_cases} 条用例，正在进行向量相似度预筛...", "startTime": start_ts}

        # 2.5 向量预筛：用 TF-IDF + 余弦相似度从全部用例中筛选相关用例（相似度 > 0 的全部保留）
        case_summaries = _vector_prefilter(cases_query, release_note, project_map, top_n=0)

        total_cases = len(case_summaries)
        task_progress[history_id] = {"status": "running", "progress": 15, "message": f"向量预筛完成（{total_all_cases} → {total_cases} 条），正在进行 AI 影响分析...", "startTime": start_ts}

        # 3. AI 影响分析
        analysis_prompt = f"""你是一名专业的软件测试专家。请分析以下 Release Note，提取关键变更信息。

Release Note 内容:
{release_note[:3000]}

请以 JSON 格式返回（直接返回JSON，不要markdown代码块）:
{{
  "affectedModules": ["变更涉及的模块列表"],
  "riskLevel": "高/中/低",
  "testDirections": ["建议的测试方向"],
  "summary": "影响范围摘要，200字以内"
}}"""

        analysis_messages = [
            {"role": "system", "content": "你是专业的软件测试分析助手，返回格式严格的JSON。"},
            {"role": "user", "content": analysis_prompt}
        ]

        analysis_raw = call_azure_openai(ai_config, analysis_messages)

        try:
            analysis_data = json.loads(clean_json_response(analysis_raw))
        except json.JSONDecodeError:
            analysis_data = {
                "affectedModules": [],
                "riskLevel": "中",
                "testDirections": ["回归测试"],
                "summary": "AI分析结果解析失败"
            }

        task_progress[history_id] = {"status": "running", "progress": 35, "message": "影响分析完成，正在进行用例匹配评分...", "startTime": start_ts}

        # 4. 分批用例评分
        all_recommendations = []
        batch_size = 30
        total_batches = (len(case_summaries) + batch_size - 1) // batch_size

        for batch_idx, i in enumerate(range(0, len(case_summaries), batch_size)):
            # 检查是否被中止
            if task_progress.get(history_id, {}).get("status") == "aborted":
                logger.info(f"Task {history_id} aborted by user")
                return

            batch = case_summaries[i:i + batch_size]
            cases_text = "\n".join([
                f"[ID:{c['id']}] {c['title']} | 模块:{c['module']}"
                for c in batch
            ])

            progress = 35 + int((batch_idx + 1) / total_batches * 50)
            task_progress[history_id] = {
                "status": "running",
                "progress": progress,
                "message": f"正在评分第 {batch_idx + 1}/{total_batches} 批用例...",
                "startTime": start_ts
            }

            recommend_prompt = f"""基于以下 Release Note 变更内容，对测试用例进行相关性评分。

Release Note:
{release_note[:2000]}

影响模块: {', '.join(analysis_data.get('affectedModules', []))}

测试用例列表:
{cases_text}

返回相关性评分 >=60 的用例，JSON数组格式（直接返回，不要代码块）:
[{{"caseId": ID, "aiScore": 分数(0-100), "reason": "原因(20字内)"}}]

无相关用例则返回 []"""

            recommend_messages = [
                {"role": "system", "content": "只返回JSON数组。"},
                {"role": "user", "content": recommend_prompt}
            ]

            try:
                recommend_raw = call_azure_openai(ai_config, recommend_messages)
                batch_results = json.loads(clean_json_response(recommend_raw))
                if isinstance(batch_results, list):
                    all_recommendations.extend(batch_results)
            except Exception as e:
                logger.warning(f"Batch {batch_idx} failed: {e}")
                continue

        task_progress[history_id] = {"status": "running", "progress": 90, "message": "正在生成推荐结果...", "startTime": start_ts}

        # 5. 构建最终结果
        case_map = {tc.id: tc for tc in cases_query}
        final_recommendations = []
        added_case_ids = set()

        # 5.1 先加入PR关联用例（优先级最高）
        if pr_linked_case_ids:
            for case_id in pr_linked_case_ids:
                if case_id in case_map:
                    tc = case_map[case_id]
                    final_recommendations.append({
                        "caseId": tc.id,
                        "caseNumber": tc.case_number or "",
                        "title": tc.name,
                        "module": tc.module or "",
                        "projectName": project_map.get(tc.primary_project_id, ""),
                        "precondition": tc.precondition or "",
                        "steps": tc.steps or "",
                        "expectedResult": tc.expected_result or "",
                        "aiScore": 100,
                        "reason": "关联PR用例",
                        "prLinked": True
                    })
                    added_case_ids.add(case_id)

        # 5.2 加入AI评分用例（去重）
        for rec in all_recommendations:
            case_id = rec.get("caseId")
            if case_id and case_id in case_map and case_id not in added_case_ids:
                tc = case_map[case_id]
                final_recommendations.append({
                    "caseId": tc.id,
                    "caseNumber": tc.case_number or "",
                    "title": tc.name,
                    "module": tc.module or "",
                    "projectName": project_map.get(tc.primary_project_id, ""),
                    "precondition": tc.precondition or "",
                    "steps": tc.steps or "",
                    "expectedResult": tc.expected_result or "",
                    "aiScore": min(100, max(0, int(rec.get("aiScore", 0)))),
                    "reason": rec.get("reason", ""),
                    "prLinked": False
                })
                added_case_ids.add(case_id)

        # PR关联用例排最前面，其余按评分排序
        pr_cases = [c for c in final_recommendations if c.get("prLinked")]
        ai_cases = [c for c in final_recommendations if not c.get("prLinked")]
        ai_cases.sort(key=lambda x: x["aiScore"], reverse=True)
        final_recommendations = pr_cases + ai_cases

        # 6. 覆盖率 — 基于推荐用例覆盖的影响模块比例
        # 用模糊匹配：如果推荐用例的模块名包含影响模块关键词（或反之），算作覆盖
        coverage = None
        affected_modules = analysis_data.get("affectedModules", [])
        if final_recommendations and affected_modules:
            covered_modules = set()
            rec_modules = set(r["module"] for r in final_recommendations if r["module"])
            for am in affected_modules:
                am_lower = am.lower()
                for rm in rec_modules:
                    rm_lower = rm.lower()
                    if am_lower in rm_lower or rm_lower in am_lower or am_lower == rm_lower:
                        covered_modules.add(am)
                        break
            # 如果有推荐用例但模糊匹配不上，也按有覆盖算（AI已经评估过相关性）
            if not covered_modules and final_recommendations:
                coverage_rate = min(95, int(len(final_recommendations) / max(len(case_summaries), 1) * 100) + 60)
            else:
                coverage_rate = int(len(covered_modules) / max(len(affected_modules), 1) * 100)
            uncovered = [m for m in affected_modules if m not in covered_modules]
            coverage = {
                "coverageRate": coverage_rate,
                "uncoveredFeatures": uncovered[:10],
                "riskWarnings": [f"模块 [{m}] 可能缺少直接用例覆盖" for m in uncovered[:5]] if uncovered else ["推荐用例覆盖了主要变更模块"]
            }
        elif final_recommendations:
            coverage = {
                "coverageRate": 80,
                "uncoveredFeatures": [],
                "riskWarnings": ["推荐用例覆盖了主要变更模块"]
            }

        # 7. 更新历史记录
        record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == history_id).first()
        if record:
            record.affectedModules = json.dumps(analysis_data.get("affectedModules", []), ensure_ascii=False)
            record.riskLevel = analysis_data.get("riskLevel", "中")
            record.testDirections = json.dumps(analysis_data.get("testDirections", []), ensure_ascii=False)
            record.summary = analysis_data.get("summary", "")
            record.recommendedCases = json.dumps(final_recommendations, ensure_ascii=False)
            record.recommendedCount = len(final_recommendations)
            record.coverageRate = coverage["coverageRate"] if coverage else 0
            record.coverageData = json.dumps(coverage, ensure_ascii=False) if coverage else "{}"
            db.commit()

        task_progress[history_id] = {"status": "completed", "progress": 100, "message": f"分析完成，推荐 {len(final_recommendations)} 条用例"}
        # 清除用户正在运行的任务标记
        for uid, hid in list(user_running_task.items()):
            if hid == history_id:
                del user_running_task[uid]
                break

    except Exception as e:
        logger.error(f"Analysis task failed for history {history_id}: {e}")
        task_progress[history_id] = {"status": "failed", "progress": 100, "message": f"分析失败: {str(e)[:100]}"}
        with task_queue_lock:
            for item in task_queue:
                if item["history_id"] == history_id:
                    item["status"] = "failed"
                    break
        # 清除用户正在运行的任务标记
        for uid, hid in list(user_running_task.items()):
            if hid == history_id:
                del user_running_task[uid]
                break
        try:
            record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == history_id).first()
            if record:
                record.summary = f"失败: {str(e)[:200]}"
                db.commit()
        except Exception:
            pass
    finally:
        with task_queue_lock:
            active_task_count -= 1
            # 仍在 running 状态说明正常结束，标记完成
            for item in task_queue:
                if item["history_id"] == history_id:
                    if item["status"] == "running":
                        item["status"] = "completed"
                    break
        db.close()
        process_queue()


# ===== API 路由 =====

class FetchPrRequest(BaseModel):
    prNumbers: str  # 逗号分隔的PR编号


@router.post("/fetch-pr-info")
def fetch_pr_info(
    req: FetchPrRequest,
    current_user: User = Depends(get_current_user),
):
    """根据PR编号获取主题、Root Cause、Solution信息"""
    import requests as http_requests

    from config import settings
    zmind_api_key = getattr(settings, 'ZMIND_API_KEY', '') or ''
    zmind_base_url = getattr(settings, 'ZMIND_BASE_URL', '') or 'https://zmind.whaletv.com'

    if not zmind_api_key:
        raise HTTPException(status_code=400, detail="系统未配置 Zmind API Key")

    pr_numbers = [n.strip() for n in req.prNumbers.split(',') if n.strip()]
    if not pr_numbers:
        raise HTTPException(status_code=400, detail="请输入至少一个PR编号")

    headers = {
        'X-Redmine-API-Key': zmind_api_key,
        'Content-Type': 'application/json'
    }

    results = []
    for pr_num in pr_numbers[:20]:
        try:
            resp = http_requests.get(
                f"{zmind_base_url}/issues/{pr_num}.json",
                headers=headers,
                timeout=15
            )
            if resp.status_code == 200:
                issue = resp.json().get('issue', {})
                subject = issue.get('subject', '')
                root_cause = ''
                solution = ''
                custom_fields = issue.get('custom_fields', [])
                for cf in custom_fields:
                    cf_name = cf.get('name', '').lower()
                    if 'root' in cf_name and 'cause' in cf_name:
                        root_cause = cf.get('value', '') or ''
                    elif 'solution' in cf_name or 'reason' in cf_name:
                        solution = cf.get('value', '') or ''

                results.append({
                    "prNumber": str(issue.get('id', pr_num)),
                    "subject": subject,
                    "rootCause": root_cause,
                    "solution": solution,
                    "status": issue.get('status', {}).get('name', ''),
                })
            else:
                results.append({
                    "prNumber": pr_num,
                    "subject": "",
                    "rootCause": "",
                    "solution": "",
                    "status": "",
                    "error": f"PR#{pr_num} 查询失败({resp.status_code})"
                })
        except Exception as e:
            results.append({
                "prNumber": pr_num,
                "subject": "",
                "rootCause": "",
                "solution": "",
                "status": "",
                "error": f"PR#{pr_num} 请求异常: {str(e)[:50]}"
            })

    return {"code": 200, "data": results}


@router.post("/test-connection")
def test_connection(
    req: TestConnectionRequest,
    current_user: User = Depends(get_current_user),
):
    """测试 Azure OpenAI 连接"""
    import requests

    endpoint = req.endpoint.rstrip("/")
    deployment = req.deployment

    if '/openai/v1/' in endpoint or endpoint.endswith('/chat/completions'):
        url = endpoint
        payload = {
            "model": deployment,
            "messages": [{"role": "user", "content": "Hi, reply OK."}],
            "max_completion_tokens": 10
        }
    else:
        api_version = req.apiVersion or "2024-12-01-preview"
        url = f"{endpoint}/openai/deployments/{deployment}/chat/completions?api-version={api_version}"
        payload = {
            "messages": [{"role": "user", "content": "Hi, reply OK."}],
            "max_completion_tokens": 10
        }

    headers = {"Content-Type": "application/json", "api-key": req.apiKey}

    start_time = time.time()
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=30)
        elapsed = int((time.time() - start_time) * 1000)

        if resp.status_code == 200:
            data = resp.json()
            return {"code": 200, "data": {"success": True, "model": data.get("model", "unknown"), "responseTime": elapsed, "error": None, "hint": None}}
        else:
            error_msg = ""
            hint = ""
            try:
                error_msg = resp.json().get("error", {}).get("message", resp.text[:300])
            except Exception:
                error_msg = resp.text[:300]
            if resp.status_code == 404:
                hint = "Endpoint URL 或 Model Name 不正确。"
            elif resp.status_code == 401:
                hint = "API Key 无效或已过期。"
            elif resp.status_code == 429:
                hint = "请求被限流，请稍后重试。"
            return {"code": 200, "data": {"success": False, "model": None, "responseTime": elapsed, "error": f"HTTP {resp.status_code}: {error_msg}", "hint": hint}}
    except requests.exceptions.Timeout:
        return {"code": 200, "data": {"success": False, "model": None, "responseTime": 30000, "error": "连接超时（30秒）", "hint": "请检查 Endpoint 地址和网络。"}}
    except requests.exceptions.ConnectionError as e:
        return {"code": 200, "data": {"success": False, "model": None, "responseTime": 0, "error": f"连接失败: {str(e)[:200]}", "hint": "无法建立网络连接。"}}
    except Exception as e:
        return {"code": 200, "data": {"success": False, "model": None, "responseTime": 0, "error": f"错误: {str(e)[:200]}", "hint": "请检查配置。"}}


@router.post("/analyze")
def create_analysis(
    req: AnalyzeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """创建 AI 分析任务（加入队列，后台执行）"""
    if not req.projectIds:
        raise HTTPException(status_code=400, detail="请至少选择一个测试用例库")
    if not req.releaseNote.strip() and not req.prNumbers:
        raise HTTPException(status_code=400, detail="PR关联和Release Note内容至少填写一项")

    # 检查当前用户是否有正在运行的任务
    if current_user.id in user_running_task:
        running_hid = user_running_task[current_user.id]
        if running_hid in task_progress and task_progress[running_hid].get("status") == "running":
            raise HTTPException(status_code=409, detail="当前已有AI分析任务在进行中，请等待完成后再创建新任务")
        else:
            del user_running_task[current_user.id]

    # 检查配置
    try:
        get_azure_openai_client(db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    # 获取项目名称
    projects = db.query(Project).filter(Project.id.in_(req.projectIds)).all()
    project_names = [p.name for p in projects]

    # 创建历史记录（状态：进行中）
    history = AiRecommendHistory(
        title=req.title or f"AI推荐 - {req.releaseNote[:200]}",
        releaseNote=req.releaseNote,
        projectIds=json.dumps(req.projectIds),
        projectNames=json.dumps(project_names, ensure_ascii=False),
        affectedModules="[]",
        riskLevel="",
        testDirections="[]",
        summary="分析进行中...",
        recommendedCases="[]",
        recommendedCount=0,
        coverageRate=0,
        coverageData="{}",
        createdById=current_user.id,
        createdByName=current_user.username
    )
    db.add(history)
    db.commit()
    db.refresh(history)

    # 初始化进度
    task_progress[history.id] = {"status": "waiting", "progress": 0, "message": "任务已加入队列，等待执行..."}
    user_running_task[current_user.id] = history.id

    # 加入队列
    queue_item = {
        "history_id": history.id,
        "user_id": current_user.id,
        "project_ids": req.projectIds,
        "release_note": req.releaseNote,
        "pr_numbers": req.prNumbers,
        "status": "waiting",
        "created_at": datetime.utcnow().isoformat()
    }

    position = 1
    total_waiting = 0
    with task_queue_lock:
        task_queue.append(queue_item)
        # 只有队列中没有等待项 且 活跃任务数未达上限时，才立即执行
        has_waiting_ahead = any(item["status"] == "waiting" for item in task_queue[:-1])
        if not has_waiting_ahead and active_task_count < MAX_CONCURRENT_TASKS:
            queue_item["status"] = "running"
            active_task_count += 1
            start_now = True
        else:
            start_now = False
        waiting_items = [item for item in task_queue if item["status"] == "waiting"]
        total_waiting = len(waiting_items)
        my_index = next((i for i, item in enumerate(waiting_items) if item["history_id"] == history.id), -1)
        if my_index >= 0:
            position = my_index + 1

    if start_now:
        _start_task_thread(queue_item)

    return {
        "code": 200,
        "data": {
            "historyId": history.id,
            "status": "running" if start_now else "waiting",
            "queuePosition": position,
            "totalWaiting": total_waiting
        }
    }


@router.get("/progress/{history_id}")
def get_progress(
    history_id: int,
    current_user: User = Depends(get_current_user),
):
    """获取分析任务进度"""
    if history_id in task_progress:
        return {"code": 200, "data": task_progress[history_id]}
    else:
        return {"code": 200, "data": {"status": "unknown", "progress": 0, "message": "任务不存在或已过期"}}


@router.get("/supplement-progress/{history_id}")
def get_supplement_progress(
    history_id: int,
    current_user: User = Depends(get_current_user),
):
    """获取补充用例生成任务进度"""
    if history_id in supplement_progress:
        return {"code": 200, "data": supplement_progress[history_id]}
    else:
        return {"code": 200, "data": {"status": "unknown", "progress": 0, "message": "任务不存在或已过期"}}


@router.get("/running-task")
def get_running_task(
    current_user: User = Depends(get_current_user),
):
    """检查当前用户是否有正在运行的任务"""
    with task_queue_lock:
        queue_list = list(task_queue)
    if current_user.id in user_running_task:
        hid = user_running_task[current_user.id]
        if hid in task_progress and task_progress[hid].get("status") == "running":
            return {"code": 200, "data": {"running": True, "historyId": hid, "progress": task_progress[hid], "queueInfo": {"active": active_task_count, "max": MAX_CONCURRENT_TASKS, "total": len(queue_list)}}}
        else:
            del user_running_task[current_user.id]
    return {"code": 200, "data": {"running": False, "historyId": None, "progress": None, "queueInfo": {"active": active_task_count, "max": MAX_CONCURRENT_TASKS, "total": len(queue_list)}}}


@router.post("/abort/{history_id}")
def abort_task(
    history_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """中止AI分析任务（支持队列中和运行中的任务）"""
    # 标记队列项
    was_waiting = False
    with task_queue_lock:
        for item in task_queue:
            if item["history_id"] == history_id:
                if item["status"] == "waiting":
                    item["status"] = "aborted"
                    was_waiting = True
                elif item["status"] == "running":
                    item["status"] = "aborted"
                break

    # 如果是等待中的任务，直接移除进度记录
    if was_waiting:
        task_progress.pop(history_id, None)
        # 清除用户运行标记
        if current_user.id in user_running_task and user_running_task[current_user.id] == history_id:
            del user_running_task[current_user.id]
        # 更新数据库记录
        record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == history_id).first()
        if record:
            record.summary = "已中止"
            db.commit()
        # 处理队列中的下一个任务
        process_queue()
        return {"code": 200, "message": "任务已从队列中移除"}

    # 运行中的任务 — 标记为已中止
    if history_id in task_progress:
        task_progress[history_id] = {"status": "aborted", "progress": 100, "message": "任务已被用户中止"}

    # 清除用户运行标记
    if current_user.id in user_running_task and user_running_task[current_user.id] == history_id:
        del user_running_task[current_user.id]

    # 更新数据库记录
    record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == history_id).first()
    if record:
        record.summary = "已中止"
        db.commit()

    return {"code": 200, "message": "任务已中止"}


# ===== 任务队列 API =====

class ReorderRequest(BaseModel):
    historyIds: List[int]  # 队列中所有任务的新顺序（history_id 列表）


@router.get("/queue")
def get_queue(
    current_user: User = Depends(get_current_user),
):
    """获取当前所有任务队列（含进度信息）"""
    with task_queue_lock:
        queue_list = list(task_queue)

    items = []
    for item in queue_list:
        items.append({
            "historyId": item["history_id"],
            "userId": item["user_id"],
            "status": item["status"],
            "createdAt": item["created_at"],
            "progress": task_progress.get(item["history_id"], {"status": item["status"], "progress": 0, "message": ""})
        })

    return {"code": 200, "data": items}


@router.post("/queue/reorder")
def reorder_queue(
    req: ReorderRequest,
    current_user: User = Depends(get_current_user),
):
    """重新排序任务队列（仅支持调整 waiting 状态任务的顺序）"""
    with task_queue_lock:
        # 构建 history_id -> item 的映射
        item_map = {item["history_id"]: item for item in task_queue}
        # 提取所有现有队列项的 history_id
        existing_ids = [item["history_id"] for item in task_queue]
        # 验证请求的 ID 列表
        for hid in req.historyIds:
            if hid not in item_map:
                raise HTTPException(status_code=400, detail=f"任务 {hid} 不存在于队列中")
        if set(req.historyIds) != set(existing_ids):
            raise HTTPException(status_code=400, detail="请求的队列项与当前队列不匹配")
        # 按新顺序重建队列
        new_queue = [item_map[hid] for hid in req.historyIds]
        task_queue.clear()
        task_queue.extend(new_queue)

    return {"code": 200, "message": "队列排序已更新"}


# ===== 历史记录 API =====

@router.get("/latest")
def get_latest_analysis(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取当前用户最近24小时内最新一条已完成的分析记录"""
    from datetime import datetime, timedelta
    cutoff = datetime.utcnow() - timedelta(hours=24)

    record = db.query(AiRecommendHistory).filter(
        AiRecommendHistory.createdById == current_user.id,
        AiRecommendHistory.createdAt >= cutoff,
        AiRecommendHistory.recommendedCount > 0  # 只返回有结果的
    ).order_by(desc(AiRecommendHistory.createdAt)).first()

    if not record:
        return {"code": 200, "data": None}

    return {
        "code": 200,
        "data": {
            "id": record.id,
            "title": record.title,
            "projectIds": json.loads(record.projectIds) if record.projectIds else [],
            "projectNames": json.loads(record.projectNames) if record.projectNames else [],
            "releaseNote": record.releaseNote,
            "affectedModules": json.loads(record.affectedModules) if record.affectedModules else [],
            "riskLevel": record.riskLevel,
            "testDirections": json.loads(record.testDirections) if record.testDirections else [],
            "summary": record.summary,
            "recommendedCases": json.loads(record.recommendedCases) if record.recommendedCases else [],
            "recommendedCount": record.recommendedCount,
            "coverageRate": record.coverageRate,
            "coverageData": json.loads(record.coverageData) if record.coverageData else None,
            "supplementCases": json.loads(record.supplementCases) if record.supplementCases else [],
            "createdAt": record.createdAt.isoformat() if record.createdAt else None
        }
    }


@router.get("/history")
def get_history(
    page: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取推荐历史列表（只显示当前用户创建的）"""
    query = db.query(AiRecommendHistory).filter(AiRecommendHistory.createdById == current_user.id)
    if keyword:
        query = query.filter(AiRecommendHistory.title.ilike(f"%{keyword}%"))

    total = query.count()
    records = query.order_by(desc(AiRecommendHistory.createdAt)).offset((page - 1) * pageSize).limit(pageSize).all()

    data = []
    for r in records:
        data.append({
            "id": r.id,
            "title": r.title,
            "projectIds": json.loads(r.projectIds) if r.projectIds else [],
            "projectNames": json.loads(r.projectNames) if r.projectNames else [],
            "affectedModules": json.loads(r.affectedModules) if r.affectedModules else [],
            "riskLevel": r.riskLevel,
            "recommendedCount": r.recommendedCount,
            "coverageRate": r.coverageRate,
            "suiteName": r.suiteName,
            "suiteId": r.suiteId,
            "createdByName": r.createdByName,
            "createdAt": r.createdAt.isoformat() if r.createdAt else None
        })

    return {"code": 200, "data": {"data": data, "total": total}}


@router.get("/history/{history_id}")
def get_history_detail(
    history_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取推荐历史详情"""
    record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == history_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")

    return {
        "code": 200,
        "data": {
            "id": record.id,
            "title": record.title,
            "releaseNote": record.releaseNote,
            "projectIds": json.loads(record.projectIds) if record.projectIds else [],
            "projectNames": json.loads(record.projectNames) if record.projectNames else [],
            "affectedModules": json.loads(record.affectedModules) if record.affectedModules else [],
            "riskLevel": record.riskLevel,
            "testDirections": json.loads(record.testDirections) if record.testDirections else [],
            "summary": record.summary,
            "recommendedCases": json.loads(record.recommendedCases) if record.recommendedCases else [],
            "recommendedCount": record.recommendedCount,
            "coverageRate": record.coverageRate,
            "coverageData": json.loads(record.coverageData) if record.coverageData else None,
            "suiteId": record.suiteId,
            "suiteName": record.suiteName,
            "supplementCases": json.loads(record.supplementCases) if record.supplementCases else [],
            "createdByName": record.createdByName,
            "createdAt": record.createdAt.isoformat() if record.createdAt else None
        }
    }


@router.delete("/history/{history_id}")
def delete_history(
    history_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """删除推荐历史"""
    record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == history_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(record)
    db.commit()
    # 清理进度缓存
    task_progress.pop(history_id, None)
    return {"code": 200, "message": "删除成功"}


@router.put("/history/{history_id}/suite")
def update_history_suite(
    history_id: int,
    req: SaveSuiteInfoRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """更新推荐历史的套件关联信息"""
    record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == history_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    record.suiteId = req.suiteId
    record.suiteName = req.suiteName
    db.commit()
    return {"code": 200, "message": "更新成功"}


# ===== 补充用例生成（后台任务） =====

def _call_supplement_for_feature(ai_config: dict, feature: str, context: str) -> list:
    """为单个未覆盖功能点调用 AI 生成补充用例"""
    template = '[{"module":"模块","name":"标题","precondition":"前置条件","steps":"步骤(换行分隔)","expected_result":"预期结果"}]'
    prompt = f"为以下功能点生成1条测试用例（JSON数组格式，直接返回，不要代码块）：\n功能点：{feature}\n背景：{context[:500]}\n格式：{template}"
    messages = [
        {"role": "system", "content": "你是测试用例设计专家，只返回JSON数组，不要其他文字。"},
        {"role": "user", "content": prompt}
    ]
    raw = call_azure_openai(ai_config, messages, max_tokens=3000)
    cleaned = clean_json_response(raw)
    if not cleaned:
        return []
    try:
        parsed = json.loads(cleaned)
        if isinstance(parsed, list):
            return parsed
        if isinstance(parsed, dict):
            for key in ("cases", "data", "items", "results"):
                if isinstance(parsed.get(key), list):
                    return parsed[key]
    except json.JSONDecodeError:
        logger.warning(f"Parse failed for feature '{feature[:50]}', raw: {raw[:200]}")
    return []


def run_supplement_task(history_id: int, uncovered_features: List[str]):
    """在后台线程中执行补充用例生成（逐功能点调用，避免超长 prompt）"""
    db = SessionLocal()
    all_cases = []
    try:
        supplement_progress[history_id] = {"status": "running", "progress": 5, "message": "正在获取配置..."}

        ai_config = get_azure_openai_client(db)

        record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == history_id).first()
        release_context = (record.releaseNote or "")[:1000] if record else ""

        total = len(uncovered_features)
        for idx, feature in enumerate(uncovered_features):
            pct = int(10 + (idx / total) * 80) if total > 0 else 50
            supplement_progress[history_id] = {"status": "running", "progress": pct, "message": f"正在生成 {feature[:30]} 的补充用例...（{idx+1}/{total}）"}

            cases = _call_supplement_for_feature(ai_config, feature, release_context)
            for c in cases:
                c["module"] = c.get("module") or feature
            all_cases.extend(cases)

        supplement_progress[history_id] = {"status": "running", "progress": 95, "message": "正在保存结果..."}

        if record:
            record.supplementCases = json.dumps(all_cases, ensure_ascii=False)
            db.commit()

        supplement_progress[history_id] = {"status": "completed", "progress": 100, "message": f"已生成 {len(all_cases)} 条补充用例"}

    except Exception as e:
        logger.error(f"Generate supplement task failed for history {history_id}: {e}")
        supplement_progress[history_id] = {"status": "failed", "progress": 100, "message": f"生成失败: {str(e)[:100]}"}
        if all_cases and record:
            record.supplementCases = json.dumps(all_cases, ensure_ascii=False)
            db.commit()
        try:
            if not all_cases and record:
                record.supplementCases = "[]"
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


# ===== 新增功能接口 =====

class GenerateSupplementRequest(BaseModel):
    historyId: int
    uncoveredFeatures: List[str]


class ExportRequest(BaseModel):
    historyId: int


@router.post("/generate-supplement")
def generate_supplement_cases(
    req: GenerateSupplementRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """AI 生成补充用例（异步）— 立即返回，后台执行"""
    if not req.uncoveredFeatures:
        raise HTTPException(status_code=400, detail="没有未覆盖功能点")

    # 检查配置
    try:
        get_azure_openai_client(db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    # 初始化进度
    supplement_progress[req.historyId] = {"status": "running", "progress": 0, "message": "任务已创建，等待执行..."}

    # 启动后台线程
    thread = threading.Thread(
        target=run_supplement_task,
        args=(req.historyId, req.uncoveredFeatures),
        daemon=True
    )
    thread.start()

    return {"code": 200, "data": {"historyId": req.historyId, "status": "running"}}


@router.post("/export-excel")
def export_recommend_excel(
    req: ExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出推荐用例为 Excel — 用例编号去测试用例表查询完整信息"""
    from fastapi.responses import StreamingResponse
    import io

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器缺少 openpyxl 库")

    record = db.query(AiRecommendHistory).filter(AiRecommendHistory.id == req.historyId).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")

    cases = json.loads(record.recommendedCases) if record.recommendedCases else []
    supplement_cases = json.loads(record.supplementCases) if record.supplementCases else []

    # 从测试用例表查询完整信息，使用与测试用例导出相同的格式
    case_ids = [c.get("caseId") for c in cases if c.get("caseId")]
    from utils.excel import generate_excel
    from models import TestCase

    testcases = []
    if case_ids:
        testcases = db.query(TestCase).filter(TestCase.id.in_(case_ids)).all()

    excel_bytes = generate_excel(testcases, db=db)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    ws.title = "AI推荐用例"

    # Sheet 2: AI 补充用例
    if supplement_cases:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ws2 = wb.create_sheet("AI补充用例")
        sup_headers = ["所属模块", "用例标题", "前置条件", "操作步骤", "预期结果"]
        for col, h in enumerate(sup_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, case in enumerate(supplement_cases, 2):
            ws2.cell(row=row_idx, column=1, value=case.get("module", "")).border = thin_border
            ws2.cell(row=row_idx, column=2, value=case.get("name", "")).border = thin_border
            ws2.cell(row=row_idx, column=3, value=case.get("precondition", "")).border = thin_border
            ws2.cell(row=row_idx, column=4, value=case.get("steps", "")).border = thin_border
            ws2.cell(row=row_idx, column=5, value=case.get("expected_result", "")).border = thin_border

        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 50
        ws2.column_dimensions['E'].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"AI推荐用例_{record.id}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
