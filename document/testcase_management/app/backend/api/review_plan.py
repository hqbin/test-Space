from fastapi import APIRouter, Depends, HTTPException, Request
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, text, cast, Integer, case
from typing import Optional
from datetime import datetime
import json
from database import get_db
from models import ReviewPlan, ReviewPlanTestCase, TestCase, User, Project, UserProject, UserTeam, TeamProject
from schemas import (
    ReviewPlanCreate, ReviewPlanUpdate, ReviewPlanResponse,
    AddTestCasesToPlanRequest, ReviewTestCaseRequest, BatchReviewRequest
)
from utils.notification_helper import trigger_assignment_notification
from auth import get_current_user
from utils.logger import log_operation, LogAction, LogModule
from utils.permissions import is_admin

router = APIRouter()


def update_review_plan_status(db: Session, plan_id: int):
    """根据评审进度自动更新评审计划状态（只更新到进行中，不自动完成）"""
    # 获取评审计划
    plan = db.query(ReviewPlan).filter(ReviewPlan.id == plan_id).first()
    if not plan:
        return

    # 如果已经是已完成状态，不再修改
    if plan.status == 'COMPLETED':
        return

    # 统计用例数
    total_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        ReviewPlanTestCase.review_plan_id == plan_id
    ).scalar() or 0

    if total_testcases == 0:
        return

    # 进度判定：已正式评审（review_status != PENDING）或已有评审草稿（pending_review_result 非空）
    # 都视为"已开始评审"
    active_count = db.query(func.count(ReviewPlanTestCase.id)).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            or_(
                ReviewPlanTestCase.review_status != 'PENDING',
                ReviewPlanTestCase.pending_review_result.isnot(None)
            )
        )
    ).scalar() or 0

    # 根据进度更新状态（只更新到进行中）
    if active_count == 0:
        # 进度为0，保持未开始状态
        if plan.status == 'PENDING':
            pass  # 不需要更新
    else:
        # 进度不为0，变为进行中（不自动完成）
        if plan.status == 'PENDING':
            plan.status = 'IN_PROGRESS'
            db.commit()


@router.get("/projects/{project_id}/members")
def get_project_members(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取项目成员列表"""
    # 验证项目是否存在
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 获取直接授权的用户
    user_projects = db.query(UserProject).filter(UserProject.project_id == project_id).all()
    direct_user_ids = set([up.user_id for up in user_projects])
    
    # 获取通过项目组授权的用户
    team_projects = db.query(TeamProject).filter(TeamProject.project_id == project_id).all()
    team_ids = [tp.team_id for tp in team_projects]
    
    team_user_ids = set()
    if team_ids:
        user_teams = db.query(UserTeam).filter(UserTeam.team_id.in_(team_ids)).all()
        team_user_ids = set([ut.user_id for ut in user_teams])
    
    # 合并所有用户ID
    all_user_ids = direct_user_ids | team_user_ids
    
    if not all_user_ids:
        return {"code": 200, "message": "success", "data": []}
    
    # 获取用户信息
    users = db.query(User).filter(
        User.id.in_(all_user_ids),
        User.status == 1
    ).all()
    
    members = []
    for user in users:
        members.append({
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "email": user.email
        })
    
    return {"code": 200, "message": "success", "data": members}


@router.post("")
def create_review_plan(
    plan: ReviewPlanCreate,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建评审计划"""
    try:
        # 验证项目是否存在
        project = db.query(Project).filter(Project.id == plan.project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="项目不存在")
        
        # 处理评审人ID列表
        reviewer_ids_json = None
        if plan.reviewer_ids:
            # 如果是普通用户且没有指定评审人，默认为自己
            if not is_admin(current_user) and not plan.reviewer_ids:
                reviewer_ids_json = json.dumps([current_user.id])
            else:
                reviewer_ids_json = json.dumps(plan.reviewer_ids)
        else:
            # 默认为创建人自己
            reviewer_ids_json = json.dumps([current_user.id])
        
        # 创建评审计划
        db_plan = ReviewPlan(
            name=plan.name,
            description=plan.description,
            project_id=plan.project_id,
            team_id=plan.team_id,
            reviewer_ids=reviewer_ids_json,
            start_time=plan.start_time,
            end_time=plan.end_time,
            status='PENDING',
            created_by=current_user.id
        )
        db.add(db_plan)
        db.commit()
        db.refresh(db_plan)
        
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.REVIEW_PLAN,
            action=LogAction.CREATE,
            description=f"创建评审计划：{db_plan.name}（ID: {db_plan.id}）",
            request=req
        )
        
        # 触发评审人分配通知
        try:
            reviewer_ids = json.loads(reviewer_ids_json) if reviewer_ids_json else []
            if reviewer_ids:
                end_time_str = plan.end_time.strftime("%Y-%m-%d %H:%M:%S") if plan.end_time else "未设置"
                
                trigger_assignment_notification(
                    db=db,
                    notification_type='testcase',
                    event_type='review_invitation',
                    title=f'您被邀请参与评审计划「{db_plan.name}」',
                    content=f'评审计划「{db_plan.name}」已创建，您被指定为评审人。\n\n'
                            f'创建人：{current_user.username}\n'
                            f'截止时间：{end_time_str}',
                    related_id=db_plan.id,
                    related_type='review_plan',
                    sender_id=current_user.id,
                    recipient_user_ids=reviewer_ids,
                    project_id=plan.project_id,
                    team_id=plan.team_id,
                    extra_context={
                        'plan_name': db_plan.name,
                        'operator': current_user.username,
                        'end_time': end_time_str,
                        'action': '创建',
                    }
                )
        except Exception as e:
            print(f"触发评审分配通知失败: {e}")
        
        return {
            "code": 200,
            "message": "创建成功",
            "data": {
                "id": db_plan.id,
                "name": db_plan.name,
                "description": db_plan.description,
                "project_id": db_plan.project_id,
                "status": db_plan.status,
                "created_by": db_plan.created_by,
                "created_at": db_plan.created_at.strftime('%Y-%m-%d %H:%M:%S') if db_plan.created_at else None,
                "updated_at": db_plan.updated_at.strftime('%Y-%m-%d %H:%M:%S') if db_plan.updated_at else None
            }
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"创建评审计划失败: {str(e)}")


@router.get("")
def list_review_plans(
    page: int = 1,
    size: int = 10,
    project_id: Optional[int] = None,
    project_ids: Optional[str] = None,
    team_id: Optional[int] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取评审计划列表 - 用户可以看到自己创建的和自己作为评审人的计划"""
    query = db.query(ReviewPlan)
    
    # 按项目组过滤（优先级最高，确保项目组隔离）
    if team_id:
        query = query.filter(ReviewPlan.team_id == team_id)
    
    # 筛选条件 - 支持多个用例库ID
    if project_ids:
        try:
            pid_list = [int(pid.strip()) for pid in project_ids.split(',') if pid.strip()]
            if pid_list:
                query = query.filter(ReviewPlan.project_id.in_(pid_list))
        except ValueError:
            pass
    elif project_id:
        query = query.filter(ReviewPlan.project_id == project_id)
    if status:
        # 支持多个状态筛选，用逗号分隔
        if ',' in status:
            status_list = [s.strip() for s in status.split(',')]
            query = query.filter(ReviewPlan.status.in_(status_list))
        else:
            query = query.filter(ReviewPlan.status == status)
    if keyword:
        query = query.filter(ReviewPlan.name.ilike(f"%{keyword}%"))
    
    # 权限筛选：超级管理员 > 组织负责人 > content_permissions
    if not is_admin(current_user):
        from utils.data_permission import (
            get_user_content_permission, get_user_organization_ids,
            get_organization_project_ids, get_user_team_ids, get_team_project_ids,
            get_managed_department_ids
        )
        
        user_id_str = str(current_user.id)
        
        # 组织负责人：看管理的所有组织下的评审计划
        managed_dept_ids = get_managed_department_ids(current_user, db)
        if managed_dept_ids:
            org_project_ids = list(get_organization_project_ids(managed_dept_ids, db))
            if org_project_ids:
                query = query.filter(ReviewPlan.project_id.in_(org_project_ids))
            else:
                query = query.filter(ReviewPlan.id == -1)
        else:
            # 非组织负责人：按 content_permissions 配置
            permission_level = get_user_content_permission(current_user, 'testplan', db)
            
            if permission_level == 'personal':
                query = query.filter(
                    or_(
                        ReviewPlan.created_by == current_user.id,
                        ReviewPlan.reviewer_ids.like(f'%{user_id_str}%')
                    )
                )
            elif permission_level == 'project':
                team_ids = get_user_team_ids(current_user, db)
                team_project_ids = list(get_team_project_ids(team_ids, db)) if team_ids else []
                if team_project_ids:
                    query = query.filter(
                        or_(
                            ReviewPlan.created_by == current_user.id,
                            ReviewPlan.reviewer_ids.like(f'%{user_id_str}%'),
                            ReviewPlan.project_id.in_(team_project_ids)
                        )
                    )
                else:
                    query = query.filter(
                        or_(
                            ReviewPlan.created_by == current_user.id,
                            ReviewPlan.reviewer_ids.like(f'%{user_id_str}%')
                        )
                    )
            elif permission_level == 'all':
                org_ids = get_user_organization_ids(current_user, db)
                if org_ids:
                    org_project_ids = list(get_organization_project_ids(org_ids, db))
                    if org_project_ids:
                        query = query.filter(
                            or_(
                                ReviewPlan.created_by == current_user.id,
                                ReviewPlan.reviewer_ids.like(f'%{user_id_str}%'),
                                ReviewPlan.project_id.in_(org_project_ids)
                            )
                        )
                    else:
                        query = query.filter(
                            or_(
                                ReviewPlan.created_by == current_user.id,
                                ReviewPlan.reviewer_ids.like(f'%{user_id_str}%')
                            )
                        )
                else:
                    team_ids = get_user_team_ids(current_user, db)
                    team_project_ids = list(get_team_project_ids(team_ids, db)) if team_ids else []
                    if team_project_ids:
                        query = query.filter(
                            or_(
                                ReviewPlan.created_by == current_user.id,
                                ReviewPlan.reviewer_ids.like(f'%{user_id_str}%'),
                                ReviewPlan.project_id.in_(team_project_ids)
                            )
                        )
                    else:
                        query = query.filter(
                            or_(
                                ReviewPlan.created_by == current_user.id,
                                ReviewPlan.reviewer_ids.like(f'%{user_id_str}%')
                            )
                        )
    
    total = query.count()
    
    # 排序和分页
    plans = query.order_by(ReviewPlan.created_at.desc(), ReviewPlan.id.desc()).offset((page - 1) * size).limit(size).all()
    
    if not plans:
        return {
            "code": 200,
            "message": "success",
            "data": {
                "records": [],
                "total": total,
                "page": page,
                "size": size
            }
        }

    plan_ids = [p.id for p in plans]

    # ---- 批量查询，消除 N+1 ----

    # 1. 批量拉取创建人和项目（IN 查询各 1 次）
    creator_ids = list({p.created_by for p in plans if p.created_by})
    project_ids_batch = list({p.project_id for p in plans if p.project_id})

    creators_map = {u.id: u for u in db.query(User).filter(User.id.in_(creator_ids)).all()} if creator_ids else {}
    projects_map = {p.id: p for p in db.query(Project).filter(Project.id.in_(project_ids_batch)).all()} if project_ids_batch else {}

    # 2. 批量拉取所有评审人 ID（先解析 reviewer_ids JSON，再一次 IN 查询）
    all_reviewer_ids = set()
    plan_reviewer_ids_map = {}  # plan.id -> List[int]
    for plan in plans:
        ids = []
        if plan.reviewer_ids:
            try:
                ids = json.loads(plan.reviewer_ids)
            except Exception:
                ids = []
        plan_reviewer_ids_map[plan.id] = ids
        all_reviewer_ids.update(ids)

    reviewers_map = {u.id: u for u in db.query(User).filter(User.id.in_(list(all_reviewer_ids))).all()} if all_reviewer_ids else {}

    # 3. 一条 SQL 用 GROUP BY + CASE WHEN 统计所有计划的各类用例数
    stats_rows = db.query(
        ReviewPlanTestCase.review_plan_id,
        func.count(ReviewPlanTestCase.id).label("total"),
        func.sum(case((ReviewPlanTestCase.review_status != 'PENDING', 1), else_=0)).label("reviewed"),
        func.sum(case((ReviewPlanTestCase.review_result == 'APPROVED', 1), else_=0)).label("passed"),
        func.sum(case((ReviewPlanTestCase.review_result == 'DEPRECATED', 1), else_=0)).label("deprecated"),
        func.sum(case(
            (or_(
                ReviewPlanTestCase.pending_review_result.isnot(None),
                ReviewPlanTestCase.review_status != 'PENDING'
            ), 1),
            else_=0
        )).label("executed"),
        func.sum(case(
            (or_(
                ReviewPlanTestCase.pending_review_result == 'APPROVED',
                and_(
                    ReviewPlanTestCase.pending_review_result.is_(None),
                    ReviewPlanTestCase.review_result == 'APPROVED'
                )
            ), 1),
            else_=0
        )).label("executed_passed"),
        func.sum(case(
            (or_(
                ReviewPlanTestCase.pending_review_result == 'REJECTED',
                and_(
                    ReviewPlanTestCase.pending_review_result.is_(None),
                    ReviewPlanTestCase.review_result == 'REJECTED'
                )
            ), 1),
            else_=0
        )).label("executed_rejected"),
        func.sum(case(
            (or_(
                ReviewPlanTestCase.pending_review_result == 'DEPRECATED',
                and_(
                    ReviewPlanTestCase.pending_review_result.is_(None),
                    ReviewPlanTestCase.review_result == 'DEPRECATED'
                )
            ), 1),
            else_=0
        )).label("executed_deprecated"),
    ).filter(
        ReviewPlanTestCase.review_plan_id.in_(plan_ids)
    ).group_by(ReviewPlanTestCase.review_plan_id).all()

    # 转为 plan_id -> stats dict
    stats_map = {row.review_plan_id: row for row in stats_rows}

    # ---- 组装响应 ----
    result_plans = []
    for plan in plans:
        creator = creators_map.get(plan.created_by)
        creator_name = creator.username if creator else "-"

        project = projects_map.get(plan.project_id)
        project_name = project.name if project else "-"

        reviewer_ids = plan_reviewer_ids_map.get(plan.id, [])
        reviewer_names = [reviewers_map[rid].username for rid in reviewer_ids if rid in reviewers_map]

        s = stats_map.get(plan.id)
        total_testcases = int(s.total) if s and s.total else 0
        reviewed_testcases = int(s.reviewed) if s and s.reviewed else 0
        passed_testcases = int(s.passed) if s and s.passed else 0
        deprecated_testcases = int(s.deprecated) if s and s.deprecated else 0
        executed_testcases = int(s.executed) if s and s.executed else 0
        executed_passed_testcases = int(s.executed_passed) if s and s.executed_passed else 0
        executed_rejected_testcases = int(s.executed_rejected) if s and s.executed_rejected else 0
        executed_deprecated_testcases = int(s.executed_deprecated) if s and s.executed_deprecated else 0

        result_plans.append({
            "id": plan.id,
            "name": plan.name,
            "description": plan.description,
            "project_id": plan.project_id,
            "project_name": project_name,
            "reviewer_ids": reviewer_ids,
            "reviewer_names": reviewer_names,
            "start_time": plan.start_time,
            "end_time": plan.end_time,
            "status": plan.status,
            "created_by": plan.created_by,
            "creator_name": creator_name,
            "created_at": plan.created_at,
            "updated_at": plan.updated_at,
            "total_testcases": total_testcases,
            "reviewed_testcases": reviewed_testcases,
            "executed_testcases": executed_testcases,
            "executed_passed_testcases": executed_passed_testcases,
            "executed_rejected_testcases": executed_rejected_testcases,
            "executed_deprecated_testcases": executed_deprecated_testcases,
            "passed_testcases": passed_testcases,
            "deprecated_testcases": deprecated_testcases
        })
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "records": result_plans,
            "total": total,
            "page": page,
            "size": size
        }
    }


@router.get("/{plan_id}")
def get_review_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取评审计划详情"""
    plan = db.query(ReviewPlan).filter(ReviewPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="评审计划不存在")
    
    # 获取创建人信息
    creator = db.query(User).filter(User.id == plan.created_by).first()
    creator_name = creator.username if creator else "-"
    
    # 获取项目信息
    project = db.query(Project).filter(Project.id == plan.project_id).first()
    project_name = project.name if project else "-"
    
    # 解析评审人ID列表
    reviewer_ids = []
    reviewer_names = []
    if plan.reviewer_ids:
        try:
            reviewer_ids = json.loads(plan.reviewer_ids)
            if reviewer_ids:
                reviewers = db.query(User).filter(User.id.in_(reviewer_ids)).all()
                reviewer_names = [r.username for r in reviewers]
        except:
            pass
    
    # 统计信息
    total_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        ReviewPlanTestCase.review_plan_id == plan_id
    ).scalar() or 0
    
    reviewed_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            ReviewPlanTestCase.review_status != 'PENDING'
        )
    ).scalar() or 0
    
    approved_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            ReviewPlanTestCase.review_result == 'APPROVED'
        )
    ).scalar() or 0
    
    rejected_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            ReviewPlanTestCase.review_result == 'REJECTED'
        )
    ).scalar() or 0
    
    deprecated_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            ReviewPlanTestCase.review_result == 'DEPRECATED'
        )
    ).scalar() or 0

    executed_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            or_(
                ReviewPlanTestCase.pending_review_result.isnot(None),
                ReviewPlanTestCase.review_status != 'PENDING'
            )
        )
    ).scalar() or 0

    # 以"执行"为口径的结果分布（暂存优先，否则官方）
    executed_passed_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            or_(
                ReviewPlanTestCase.pending_review_result == 'APPROVED',
                and_(
                    ReviewPlanTestCase.pending_review_result.is_(None),
                    ReviewPlanTestCase.review_result == 'APPROVED'
                )
            )
        )
    ).scalar() or 0

    executed_rejected_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            or_(
                ReviewPlanTestCase.pending_review_result == 'REJECTED',
                and_(
                    ReviewPlanTestCase.pending_review_result.is_(None),
                    ReviewPlanTestCase.review_result == 'REJECTED'
                )
            )
        )
    ).scalar() or 0

    executed_deprecated_testcases = db.query(func.count(ReviewPlanTestCase.id)).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            or_(
                ReviewPlanTestCase.pending_review_result == 'DEPRECATED',
                and_(
                    ReviewPlanTestCase.pending_review_result.is_(None),
                    ReviewPlanTestCase.review_result == 'DEPRECATED'
                )
            )
        )
    ).scalar() or 0

    pending_testcases = total_testcases - reviewed_testcases

    return {
        "code": 200,
        "message": "success",
        "data": {
            "id": plan.id,
            "name": plan.name,
            "description": plan.description,
            "project_id": plan.project_id,
            "project_name": project_name,
            "reviewer_ids": reviewer_ids,
            "reviewer_names": reviewer_names,
            "start_time": plan.start_time,
            "end_time": plan.end_time,
            "status": plan.status,
            "created_by": plan.created_by,
            "creator_name": creator_name,
            "created_at": plan.created_at,
            "updated_at": plan.updated_at,
            "total_testcases": total_testcases,
            "reviewed_testcases": reviewed_testcases,
            "executed_testcases": executed_testcases,
            "executed_passed_testcases": executed_passed_testcases,
            "executed_rejected_testcases": executed_rejected_testcases,
            "executed_deprecated_testcases": executed_deprecated_testcases,
            "passed_testcases": approved_testcases,
            "deprecated_testcases": deprecated_testcases,
            "statistics": {
                "total": total_testcases,
                "pending": pending_testcases,
                "approved": approved_testcases,
                "rejected": rejected_testcases,
                "deprecated": deprecated_testcases
            }
        }
    }


@router.put("/{plan_id}")
def update_review_plan(
    plan_id: int,
    plan_update: ReviewPlanUpdate,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新评审计划"""
    db_plan = db.query(ReviewPlan).filter(ReviewPlan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="评审计划不存在")
    
    # 更新字段
    if plan_update.name is not None:
        db_plan.name = plan_update.name
    if plan_update.description is not None:
        db_plan.description = plan_update.description
    if plan_update.reviewer_ids is not None:
        db_plan.reviewer_ids = json.dumps(plan_update.reviewer_ids)
    if plan_update.start_time is not None:
        db_plan.start_time = plan_update.start_time
    if plan_update.end_time is not None:
        db_plan.end_time = plan_update.end_time
    if plan_update.status is not None:
        db_plan.status = plan_update.status
    
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.REVIEW_PLAN,
        action=LogAction.UPDATE,
        description=f"更新评审计划：{db_plan.name}（ID: {plan_id}）",
        request=req
    )
    
    return {"code": 200, "message": "success", "data": None}


@router.delete("/{plan_id}")
def delete_review_plan(
    plan_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除评审计划"""
    from utils.permissions import is_admin
    
    db_plan = db.query(ReviewPlan).filter(ReviewPlan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="评审计划不存在")
    
    # 权限检查：超管可删除任意状态的计划，其他用户只能删除PENDING状态的自己的计划
    is_super_admin = is_admin(current_user)
    
    if not is_super_admin:
        # 非超管：检查是否是创建人 且 状态为PENDING
        if db_plan.created_by != current_user.id:
            raise HTTPException(status_code=403, detail="只有创建者可以删除评审计划")
        if db_plan.status != 'PENDING':
            raise HTTPException(status_code=403, detail="只能删除未开始状态的评审计划")
    
    plan_name = db_plan.name
    
    try:
        db.query(ReviewPlanTestCase).filter(ReviewPlanTestCase.review_plan_id == plan_id).delete()
        db.delete(db_plan)
        db.commit()
        
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.REVIEW_PLAN,
            action=LogAction.DELETE,
            description=f"删除评审计划：{plan_name}（ID: {plan_id}）",
            request=req
        )
        
        return {"code": 200, "message": "success", "data": None}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"删除评审计划失败: {str(e)}")


@router.post("/{plan_id}/testcases")
def add_testcases_to_plan(
    plan_id: int,
    request: AddTestCasesToPlanRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """添加用例到评审计划"""
    # 验证评审计划是否存在
    plan = db.query(ReviewPlan).filter(ReviewPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="评审计划不存在")
    
    added_count = 0
    skipped_count = 0
    errors = []
    
    for testcase_id in request.testcase_ids:
        # 验证用例是否存在
        testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
        if not testcase:
            errors.append(f"用例 ID {testcase_id} 不存在")
            continue
        
        # 检查是否已添加
        existing = db.query(ReviewPlanTestCase).filter(
            and_(
                ReviewPlanTestCase.review_plan_id == plan_id,
                ReviewPlanTestCase.testcase_id == testcase_id
            )
        ).first()
        
        if existing:
            skipped_count += 1
            errors.append(f"用例 {testcase.case_number} 已在评审计划中")
            continue
        
        # 添加用例到计划
        plan_testcase = ReviewPlanTestCase(
            review_plan_id=plan_id,
            testcase_id=testcase_id,
            review_status='PENDING'
        )
        db.add(plan_testcase)
        added_count += 1
    
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.REVIEW_PLAN,
        action=LogAction.CREATE,
        description=f"向评审计划 {plan.name} 添加 {added_count} 个用例",
        request=req
    )
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "added": added_count,
            "skipped": skipped_count,
            "errors": errors
        }
    }


@router.get("/{plan_id}/testcases")
def get_plan_testcases(
    plan_id: int,
    page: int = 1,
    size: int = 20,
    review_status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取评审计划的用例列表"""
    # 验证评审计划是否存在
    plan = db.query(ReviewPlan).filter(ReviewPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="评审计划不存在")
    
    # 构建查询 - 添加排序：按照添加到评审计划的顺序（ReviewPlanTestCase.id）
    query = db.query(ReviewPlanTestCase, TestCase).join(
        TestCase, ReviewPlanTestCase.testcase_id == TestCase.id
    ).filter(ReviewPlanTestCase.review_plan_id == plan_id)
    
    # 筛选条件 - 以"执行"为口径（暂存优先，否则官方）
    if review_status:
        if review_status == 'REVIEWED':
            # 已执行：有暂存或已提交
            query = query.filter(
                or_(
                    ReviewPlanTestCase.pending_review_result.isnot(None),
                    ReviewPlanTestCase.review_status != 'PENDING'
                )
            )
        elif review_status == 'PENDING':
            # 未执行：既无暂存也未提交
            query = query.filter(
                and_(
                    ReviewPlanTestCase.pending_review_result.is_(None),
                    ReviewPlanTestCase.review_status == 'PENDING'
                )
            )
        elif review_status == 'APPROVED':
            query = query.filter(
                or_(
                    ReviewPlanTestCase.pending_review_result == 'APPROVED',
                    and_(
                        ReviewPlanTestCase.pending_review_result.is_(None),
                        ReviewPlanTestCase.review_result == 'APPROVED'
                    )
                )
            )
        elif review_status == 'REJECTED':
            query = query.filter(
                or_(
                    ReviewPlanTestCase.pending_review_result == 'REJECTED',
                    and_(
                        ReviewPlanTestCase.pending_review_result.is_(None),
                        ReviewPlanTestCase.review_result == 'REJECTED'
                    )
                )
            )
        elif review_status == 'DEPRECATED':
            query = query.filter(
                or_(
                    ReviewPlanTestCase.pending_review_result == 'DEPRECATED',
                    and_(
                        ReviewPlanTestCase.pending_review_result.is_(None),
                        ReviewPlanTestCase.review_result == 'DEPRECATED'
                    )
                )
            )
        else:
            query = query.filter(ReviewPlanTestCase.review_status == review_status)
    if keyword:
        query = query.filter(
            (TestCase.case_number.ilike(f"%{keyword}%")) |
            (TestCase.name.ilike(f"%{keyword}%"))
        )
    
    # 按模块sort_order排序（替代按添加顺序排序）
    from utils.module_sort import build_module_sort_key_expr
    from sqlalchemy import text, cast, Integer
    
    # 收集该评审计划所有用例的实际 project_id（用例可能来自多个用例库）
    all_project_ids = db.query(TestCase.primary_project_id).join(
        ReviewPlanTestCase, ReviewPlanTestCase.testcase_id == TestCase.id
    ).filter(
        ReviewPlanTestCase.review_plan_id == plan_id,
        TestCase.primary_project_id.isnot(None)
    ).distinct().all()
    project_ids_for_sort = list({row[0] for row in all_project_ids})
    if plan.project_id and plan.project_id not in project_ids_for_sort:
        project_ids_for_sort.append(plan.project_id)
    
    module_sort = build_module_sort_key_expr(db, project_ids_for_sort)
    case_num_sort = cast(text("SUBSTRING(test_cases.case_number FROM '([0-9]+)$')"), Integer).asc()
    query = query.order_by(module_sort, TestCase.sort_order.asc(), case_num_sort, TestCase.id.asc())

    total = query.count()
    
    # 分页
    results = query.offset((page - 1) * size).limit(size).all()
    
    # 构建响应数据
    testcases = []
    for plan_tc, testcase in results:
        # 获取评审人信息
        reviewer_name = None
        if plan_tc.reviewer_id:
            reviewer = db.query(User).filter(User.id == plan_tc.reviewer_id).first()
            reviewer_name = reviewer.username if reviewer else None
        
        # 根据评审计划状态决定使用实时状态还是快照状态
        if plan.status == 'COMPLETED' and plan_tc.testcase_status_snapshot:
            # 已完成的评审计划使用快照状态
            testcase_status = plan_tc.testcase_status_snapshot
        else:
            # 未完成的评审计划使用实时状态
            testcase_status = testcase.status
        
        testcases.append({
            "id": testcase.id,
            "case_number": testcase.case_number,
            "name": testcase.name,
            "module": testcase.module,
            "sub_module": testcase.sub_module,
            "precondition": testcase.precondition,
            "steps": testcase.steps,
            "expected_result": testcase.expected_result,
            "level": testcase.level,
            "status": testcase_status,  # 使用计算后的状态
            "review_status": plan_tc.review_status,
            "review_result": plan_tc.review_result,
            "reviewer_id": plan_tc.reviewer_id,
            "reviewer_name": reviewer_name,
            "review_comment": plan_tc.review_comment,
            "reviewed_at": plan_tc.reviewed_at,
            "plan_testcase_id": plan_tc.id,
            # 草稿字段：评审页填了但还没点"提交评审"时的状态
            "pending_review_result": plan_tc.pending_review_result,
            "pending_review_comment": plan_tc.pending_review_comment,
            "pending_reviewer_id": plan_tc.pending_reviewer_id,
            "pending_reviewed_at": plan_tc.pending_reviewed_at
        })
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "records": testcases,
            "total": total,
            "page": page,
            "size": size
        }
    }


@router.delete("/{plan_id}/testcases/{testcase_id}")
def remove_testcase_from_plan(
    plan_id: int,
    testcase_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """从评审计划中移除用例"""
    plan_testcase = db.query(ReviewPlanTestCase).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            ReviewPlanTestCase.testcase_id == testcase_id
        )
    ).first()
    
    if not plan_testcase:
        raise HTTPException(status_code=404, detail="用例不在评审计划中")
    
    db.delete(plan_testcase)
    db.commit()
    
    return {"code": 200, "message": "success", "data": None}


@router.post("/{plan_id}/testcases/{testcase_id}/review")
def review_testcase(
    plan_id: int,
    testcase_id: int,
    review_request: ReviewTestCaseRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """评审单个用例（写入草稿，不直接更新用例状态）"""
    # 验证评审结果
    if review_request.result not in ['APPROVED', 'REJECTED', 'DEPRECATED']:
        raise HTTPException(status_code=400, detail="评审结果必须是 APPROVED、REJECTED 或 DEPRECATED")

    # 查找评审计划用例
    plan_testcase = db.query(ReviewPlanTestCase).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            ReviewPlanTestCase.testcase_id == testcase_id
        )
    ).first()

    if not plan_testcase:
        raise HTTPException(status_code=404, detail="用例不在评审计划中")

    # 查找测试用例
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")

    try:
        # 仅写入草稿字段（pending_*），不动正式评审字段与用例状态
        plan_testcase.pending_review_result = review_request.result
        plan_testcase.pending_review_comment = review_request.comment
        plan_testcase.pending_reviewer_id = current_user.id
        plan_testcase.pending_reviewed_at = datetime.now()

        db.commit()

        # 检查并更新评审计划状态（开始评审后从 PENDING 切到 IN_PROGRESS）
        update_review_plan_status(db, plan_id)

        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.REVIEW_PLAN,
            action=LogAction.UPDATE,
            description=f"填写评审草稿 {testcase.case_number}：{review_request.result}",
            request=req
        )

        return {"code": 200, "message": "已记为草稿", "data": None}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"评审失败: {str(e)}")


@router.post("/{plan_id}/testcases/batch-review")
def batch_review_testcases(
    plan_id: int,
    review_request: BatchReviewRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """批量评审用例（写入草稿，不直接更新用例状态）"""
    # 验证评审结果
    if review_request.result not in ['APPROVED', 'REJECTED', 'DEPRECATED']:
        raise HTTPException(status_code=400, detail="评审结果必须是 APPROVED、REJECTED 或 DEPRECATED")

    reviewed_count = 0
    errors = []
    now = datetime.now()

    for testcase_id in review_request.testcase_ids:
        try:
            # 查找评审计划用例
            plan_testcase = db.query(ReviewPlanTestCase).filter(
                and_(
                    ReviewPlanTestCase.review_plan_id == plan_id,
                    ReviewPlanTestCase.testcase_id == testcase_id
                )
            ).first()

            if not plan_testcase:
                errors.append(f"用例 ID {testcase_id} 不在评审计划中")
                continue

            # 查找测试用例
            testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
            if not testcase:
                errors.append(f"用例 ID {testcase_id} 不存在")
                continue

            # 仅写入草稿字段（pending_*）
            plan_testcase.pending_review_result = review_request.result
            plan_testcase.pending_review_comment = review_request.comment
            plan_testcase.pending_reviewer_id = current_user.id
            plan_testcase.pending_reviewed_at = now

            reviewed_count += 1
        except Exception as e:
            errors.append(f"用例 ID {testcase_id} 评审失败: {str(e)}")

    db.commit()

    # 检查并更新评审计划状态
    update_review_plan_status(db, plan_id)

    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.REVIEW_PLAN,
        action=LogAction.UPDATE,
        description=f"批量填写评审草稿 {reviewed_count} 个用例：{review_request.result}",
        request=req
    )

    return {
        "code": 200,
        "message": "已记为草稿",
        "data": {
            "reviewed": reviewed_count,
            "errors": errors
        }
    }


@router.get("/{plan_id}/export")
def export_review_report(
    plan_id: int,
    format: str = 'excel',  # excel 或 pdf
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出评审报告"""
    from fastapi.responses import StreamingResponse
    import io
    from utils.testcase_utils import parse_steps_and_expected, format_steps_for_excel, format_steps_for_pdf
    
    # 获取评审计划
    plan = db.query(ReviewPlan).filter(ReviewPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="评审计划不存在")
    
    # 获取项目信息
    project = db.query(Project).filter(Project.id == plan.project_id).first()
    project_name = project.name if project else "-"
    
    # 获取创建人信息
    creator = db.query(User).filter(User.id == plan.created_by).first()
    creator_name = creator.username if creator else "-"
    
    # 获取评审人信息
    reviewer_names = []
    if plan.reviewer_ids:
        try:
            reviewer_ids = json.loads(plan.reviewer_ids)
            if reviewer_ids:
                reviewers = db.query(User).filter(User.id.in_(reviewer_ids)).all()
                reviewer_names = [r.username for r in reviewers]
        except:
            pass
    
    # 获取用例列表（按用例标准排序：模块顺序 → sort_order → case_number）
    testcases_query = db.query(ReviewPlanTestCase, TestCase).join(
        TestCase, ReviewPlanTestCase.testcase_id == TestCase.id
    ).filter(ReviewPlanTestCase.review_plan_id == plan_id)

    # 应用标准排序（用 plan.project_id 作为模块排序基准）
    case_num_sort = cast(text("SUBSTRING(test_cases.case_number FROM '([0-9]+)$')"), Integer)
    testcases_query = testcases_query.order_by(
        TestCase.module.asc().nullslast(),
        TestCase.sort_order.asc(),
        case_num_sort.asc(),
        TestCase.id.asc()
    )
    
    testcases_data = []
    for plan_tc, testcase in testcases_query.all():
        reviewer_name = None
        if plan_tc.reviewer_id:
            reviewer = db.query(User).filter(User.id == plan_tc.reviewer_id).first()
            reviewer_name = reviewer.username if reviewer else None
        
        testcases_data.append({
            "case_number": testcase.case_number,
            "name": testcase.name,
            "precondition": testcase.precondition or "",
            "steps": testcase.steps or "",
            "expected_result": testcase.expected_result or "",
            "level": testcase.level,
            "review_status": plan_tc.review_status,
            "review_result": plan_tc.review_result,
            "reviewer_name": reviewer_name or "",
            "review_comment": plan_tc.review_comment or "",
            "reviewed_at": plan_tc.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if plan_tc.reviewed_at else ""
        })
    
    # 统计信息
    total_testcases = len(testcases_data)
    approved_testcases = sum(1 for tc in testcases_data if tc['review_result'] == 'APPROVED')
    rejected_testcases = sum(1 for tc in testcases_data if tc['review_result'] == 'REJECTED')
    pending_testcases = total_testcases - approved_testcases - rejected_testcases
    
    if format == 'excel':
        # 导出Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "评审报告"
        
        # 标题
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"测试用例评审报告 - {plan.name}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 基本信息
        row = 3
        info_data = [
            ["评审计划", plan.name],
            ["创建人", creator_name, "评审人", ", ".join(reviewer_names) if reviewer_names else "-"],
            ["开始时间", plan.start_time.strftime('%Y-%m-%d %H:%M') if plan.start_time else "-", 
             "结束时间", plan.end_time.strftime('%Y-%m-%d %H:%M') if plan.end_time else "-"],
            ["总用例数", str(total_testcases), "已通过", str(approved_testcases)],
            ["未通过", str(rejected_testcases), "待评审", str(pending_testcases)]
        ]
        
        for info_row in info_data:
            ws.cell(row, 1, info_row[0]).font = Font(bold=True)
            ws.cell(row, 2, info_row[1])
            if len(info_row) > 2:
                ws.cell(row, 3, info_row[2]).font = Font(bold=True)
                ws.cell(row, 4, info_row[3])
            row += 1
        
        # 表头
        row += 1
        headers = ["用例编号", "用例标题", "前置条件", "操作步骤", "预期结果", "用例等级", "评审状态", "评审意见", "评审人"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据
        row += 1
        for tc in testcases_data:
            # 使用统一的工具函数解析步骤和预期结果
            steps_list, expected_list = parse_steps_and_expected(tc['steps'], tc['expected_result'])
            steps_text = format_steps_for_excel(steps_list)
            expected_text = format_steps_for_excel(expected_list)
            
            ws.cell(row, 1, tc['case_number'])
            ws.cell(row, 2, tc['name'])
            
            precondition_cell = ws.cell(row, 3, tc['precondition'] or "")
            precondition_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # 设置操作步骤单元格，启用自动换行
            steps_cell = ws.cell(row, 4, steps_text)
            steps_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # 设置预期结果单元格，启用自动换行
            expected_cell = ws.cell(row, 5, expected_text)
            expected_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.cell(row, 6, tc['level'])
            
            status_text = {"PENDING": "待评审", "APPROVED": "已通过", "REJECTED": "未通过"}.get(tc['review_status'], tc['review_status'])
            ws.cell(row, 7, status_text)
            
            comment_cell = ws.cell(row, 8, tc['review_comment'] or "")
            comment_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.cell(row, 9, tc['reviewer_name'] or "")
            
            # 改进的行高计算：考虑文本长度和列宽
            def calculate_text_lines(text, column_width):
                """计算文本在指定列宽下需要的行数"""
                if not text:
                    return 1
                lines = text.split('\n')
                total_lines = 0
                # 每个字符大约占0.8个单位宽度（中文字符）
                # 英文字符大约占0.5个单位宽度
                for line in lines:
                    if not line:
                        total_lines += 1
                        continue
                    # 估算字符宽度（中文按1计算，英文按0.5计算）
                    char_count = 0
                    for char in line:
                        if ord(char) > 127:  # 中文字符
                            char_count += 1.2
                        else:  # 英文字符
                            char_count += 0.6
                    # 计算需要的行数（向上取整）
                    line_count = max(1, int((char_count / column_width) + 0.99))
                    total_lines += line_count
                return total_lines
            
            # 计算各列需要的行数
            precondition_lines = calculate_text_lines(tc['precondition'] or "", 30)
            steps_lines = calculate_text_lines(steps_text, 50)
            expected_lines = calculate_text_lines(expected_text, 50)
            comment_lines = calculate_text_lines(tc['review_comment'] or "", 35)
            
            # 取最大值，并增加一些额外空间
            max_lines = max(precondition_lines, steps_lines, expected_lines, comment_lines)
            ws.row_dimensions[row].height = max(25, max_lines * 18 + 5)
            
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 25  # 用例编号
        ws.column_dimensions['B'].width = 35  # 用例标题
        ws.column_dimensions['C'].width = 30  # 前置条件
        ws.column_dimensions['D'].width = 50  # 操作步骤
        ws.column_dimensions['E'].width = 50  # 预期结果
        ws.column_dimensions['F'].width = 12  # 用例等级
        ws.column_dimensions['G'].width = 12  # 评审状态
        ws.column_dimensions['H'].width = 35  # 评审意见
        ws.column_dimensions['I'].width = 15  # 评审人
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from urllib.parse import quote
        filename = f"评审报告_{plan.name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        # 使用 RFC 5987 编码，确保兼容性
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    
    elif format == 'pdf':
        # 导出PDF
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.units import cm
        
        # 注册中文字体（需要确保系统有该字体）
        import os
        import platform
        
        font_name = 'Helvetica'
        try:
            # Windows系统字体路径 — ttc 文件必须指定 subfontIndex
            if platform.system() == 'Windows':
                font_candidates = [
                    ('C:/Windows/Fonts/simsun.ttc', 0),     # 宋体
                    ('C:/Windows/Fonts/msyh.ttc', 0),       # 微软雅黑
                    ('C:/Windows/Fonts/simhei.ttf', None),   # 黑体
                ]
            # Linux系统字体路径
            else:
                font_candidates = [
                    ('/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc', 0),
                    ('/usr/share/fonts/truetype/wqy/wqy-microhei.ttc', 0),
                    ('/usr/share/fonts/truetype/arphic/uming.ttc', 0),
                    ('/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc', 0),
                    ('/usr/share/fonts/google-noto-cjk/NotoSansCJK-Regular.ttc', 0),
                    ('/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc', 0),
                ]
            
            # 尝试注册第一个可用的字体
            for font_path, sub_index in font_candidates:
                if os.path.exists(font_path):
                    if sub_index is not None:
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path, subfontIndex=sub_index))
                    else:
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                    font_name = 'ChineseFont'
                    break
        except Exception as e:
            print(f"TTF字体注册失败: {e}")
        
        # 回退到 reportlab 内置 CID 字体
        if font_name == 'Helvetica':
            try:
                from reportlab.pdfbase.cidfonts import UnicodeCIDFont
                pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
                font_name = 'STSong-Light'
            except Exception as e:
                print(f"CID字体注册失败: {e}")
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        
        # 标题
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=18,
            alignment=1
        )
        elements.append(Paragraph(f"测试用例评审报告 - {plan.name}", title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # 基本信息表格
        info_data = [
            ["评审计划", plan.name, "", ""],
            ["创建人", creator_name, "评审人", ", ".join(reviewer_names) if reviewer_names else "-"],
            ["开始时间", plan.start_time.strftime('%Y-%m-%d %H:%M') if plan.start_time else "-", 
             "结束时间", plan.end_time.strftime('%Y-%m-%d %H:%M') if plan.end_time else "-"],
            ["总用例数", str(total_testcases), "已通过", str(approved_testcases)],
            ["未通过", str(rejected_testcases), "待评审", str(pending_testcases)]
        ]
        
        info_table = Table(info_data, colWidths=[3*cm, 5*cm, 3*cm, 5*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('BACKGROUND', (2, 1), (2, -1), colors.lightgrey),
            ('SPAN', (1, 0), (3, 0)),  # 第一行：评审计划值跨3列
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # 用例详细信息 - 使用Paragraph处理长文本
        content_style = ParagraphStyle(
            'ContentStyle',
            fontName=font_name,
            fontSize=8,
            leading=10,
            wordWrap='CJK'
        )
        
        # 为每个测试用例创建详细表格
        for idx, tc in enumerate(testcases_data, 1):
            # 用例标题
            case_title_style = ParagraphStyle(
                'CaseTitleStyle',
                fontName=font_name,
                fontSize=10,
                leading=12,
                textColor=colors.HexColor('#333333'),
                spaceAfter=6
            )
            elements.append(Paragraph(f"<b>{idx}. {tc['case_number']} - {tc['name']}</b>", case_title_style))
            
            # 使用统一的工具函数解析步骤和预期结果
            steps_list, expected_list = parse_steps_and_expected(tc['steps'], tc['expected_result'])
            steps_text = format_steps_for_pdf(steps_list)
            expected_text = format_steps_for_pdf(expected_list)
            
            # 用例详细信息表格
            status_text = {"PENDING": "待评审", "APPROVED": "已通过", "REJECTED": "未通过"}.get(tc['review_status'], tc['review_status'])
            
            detail_data = [
                ["用例等级", tc['level'], "评审状态", status_text],
                ["前置条件", Paragraph(tc['precondition'] or "-", content_style), "", ""],
                ["操作步骤", Paragraph(steps_text, content_style), "", ""],
                ["预期结果", Paragraph(expected_text, content_style), "", ""],
                ["评审意见", Paragraph(tc['review_comment'] or "-", content_style), "评审人", tc['reviewer_name'] or "-"],
            ]
            
            detail_table = Table(detail_data, colWidths=[2.5*cm, 10*cm, 2.5*cm, 10*cm])
            detail_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('SPAN', (1, 1), (3, 1)),  # 前置条件跨列
                ('SPAN', (1, 2), (3, 2)),  # 操作步骤跨列
                ('SPAN', (1, 3), (3, 3)),  # 预期结果跨列
            ]))
            elements.append(detail_table)
            elements.append(Spacer(1, 0.3*cm))
        
        doc.build(elements)
        output.seek(0)
        
        from urllib.parse import quote
        filename = f"评审报告_{plan.name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        # 使用 RFC 5987 编码，确保兼容性
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    
    else:
        raise HTTPException(status_code=400, detail="不支持的导出格式")



@router.post("/{plan_id}/submit")
def submit_review_plan(
    plan_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """提交评审计划：把所有 pending_* 草稿升级为正式评审，并标记计划为已完成"""
    # 获取评审计划
    plan = db.query(ReviewPlan).filter(ReviewPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="评审计划不存在")

    # 防止重复提交
    if plan.status == 'COMPLETED':
        raise HTTPException(status_code=400, detail="评审计划已提交，无需重复操作")

    # 加载所有 plan_testcases
    plan_testcases = db.query(ReviewPlanTestCase).filter(
        ReviewPlanTestCase.review_plan_id == plan_id
    ).all()

    total_testcases = len(plan_testcases)

    # 检查每个用例是否都有草稿结论（pending_review_result 不为空）
    missing = [pt for pt in plan_testcases if not pt.pending_review_result]
    if missing:
        # 仅取前 5 个用例的 case_number 作为提示，避免返回过长
        case_numbers = []
        for pt in missing[:5]:
            tc = db.query(TestCase).filter(TestCase.id == pt.testcase_id).first()
            if tc:
                case_numbers.append(tc.case_number)
        suffix = f" 等{len(missing)}个" if len(missing) > 5 else ""
        cn_text = "、".join(case_numbers) + suffix if case_numbers else f"{len(missing)}个"
        raise HTTPException(
            status_code=400,
            detail=f"还有 {len(missing)} 个用例未填写评审草稿（{cn_text}），无法提交"
        )

    try:
        # 把草稿升级为正式评审，并写用例状态
        for plan_tc in plan_testcases:
            testcase = db.query(TestCase).filter(TestCase.id == plan_tc.testcase_id).first()
            if not testcase:
                continue

            # pending_review_result 是校验过的非空值
            result = plan_tc.pending_review_result

            # 升级正式评审字段
            plan_tc.review_status = result
            plan_tc.review_result = result
            plan_tc.reviewer_id = plan_tc.pending_reviewer_id
            plan_tc.review_comment = plan_tc.pending_review_comment
            plan_tc.reviewed_at = plan_tc.pending_reviewed_at

            # 更新用例状态
            if result == 'APPROVED':
                testcase.status = 'REVIEWED'
            elif result == 'REJECTED':
                testcase.status = 'REJECTED'
            else:  # DEPRECATED
                testcase.status = 'DEPRECATED'

            # 保存状态快照（在更新 testcase.status 之后保存，反映最终状态）
            plan_tc.testcase_status_snapshot = testcase.status

            # 清空草稿字段（升级为正式后无需保留）
            plan_tc.pending_review_result = None
            plan_tc.pending_review_comment = None
            plan_tc.pending_reviewer_id = None
            plan_tc.pending_reviewed_at = None

        # 标记计划为已完成
        plan.status = 'COMPLETED'
        db.commit()

        # 记录日志
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.REVIEW_PLAN,
            action=LogAction.UPDATE,
            description=f"提交评审计划: {plan.name}（共 {total_testcases} 个用例）",
            request=req
        )

        return {"code": 200, "message": "提交成功", "data": None}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"提交失败: {str(e)}")


@router.get("/{plan_id}/testcases/{testcase_id}/detail")
def get_testcase_review_detail(
    plan_id: int,
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取用例评审详情（用于详情页评审）"""
    # 获取评审计划
    plan = db.query(ReviewPlan).filter(ReviewPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="评审计划不存在")
    
    # 获取用例和评审信息
    result = db.query(ReviewPlanTestCase, TestCase).join(
        TestCase, ReviewPlanTestCase.testcase_id == TestCase.id
    ).filter(
        and_(
            ReviewPlanTestCase.review_plan_id == plan_id,
            ReviewPlanTestCase.testcase_id == testcase_id
        )
    ).first()
    
    if not result:
        raise HTTPException(status_code=404, detail="用例不在评审计划中")
    
    plan_tc, testcase = result
    
    # 获取评审人信息
    reviewer_name = None
    if plan_tc.reviewer_id:
        reviewer = db.query(User).filter(User.id == plan_tc.reviewer_id).first()
        reviewer_name = reviewer.username if reviewer else None
    
    # 根据评审计划状态决定使用实时状态还是快照状态
    if plan.status == 'COMPLETED' and plan_tc.testcase_status_snapshot:
        # 已完成的评审计划使用快照状态
        testcase_status = plan_tc.testcase_status_snapshot
    else:
        # 未完成的评审计划使用实时状态
        testcase_status = testcase.status
    
    # 获取所有用例ID列表（用于上一个/下一个导航）
    all_testcase_ids = db.query(ReviewPlanTestCase.testcase_id).filter(
        ReviewPlanTestCase.review_plan_id == plan_id
    ).all()
    all_testcase_ids = [tc_id[0] for tc_id in all_testcase_ids]
    # 按模块sort_order排序
    from utils.module_sort import get_module_sort_map as _nav_sort
    _nav_map = _nav_sort(db, plan.project_id)
    from models import TestCase as _TC
    _tc_modules = {tc.id: tc.module for tc in db.query(_TC).filter(_TC.id.in_(all_testcase_ids)).all()}
    all_testcase_ids.sort(key=lambda tid: _nav_map.get(_tc_modules.get(tid, ""), "9999999999"))
    
    # 找到当前用例的索引
    current_index = all_testcase_ids.index(testcase_id) if testcase_id in all_testcase_ids else -1
    
    # 计算上一个和下一个用例ID
    prev_testcase_id = all_testcase_ids[current_index - 1] if current_index > 0 else None
    next_testcase_id = all_testcase_ids[current_index + 1] if current_index < len(all_testcase_ids) - 1 else None
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "testcase": {
                "id": testcase.id,
                "case_number": testcase.case_number,
                "name": testcase.name,
                "module": testcase.module,
                "sub_module": testcase.sub_module,
                "precondition": testcase.precondition,
                "steps": testcase.steps,
                "expected_result": testcase.expected_result,
                "level": testcase.level,
                "status": testcase_status,
                "feedback": testcase.feedback,
                "primary_project_id": testcase.primary_project_id,
                "project_id": testcase.primary_project_id,
                "case_type": testcase.case_type,
                "remarks": testcase.remarks,
                "automation": testcase.automation
            },
            "review": {
                "review_status": plan_tc.review_status,
                "review_result": plan_tc.review_result,
                "reviewer_id": plan_tc.reviewer_id,
                "reviewer_name": reviewer_name,
                "review_comment": plan_tc.review_comment,
                "reviewed_at": plan_tc.reviewed_at,
                "plan_testcase_id": plan_tc.id,
                # 草稿字段
                "pending_review_result": plan_tc.pending_review_result,
                "pending_review_comment": plan_tc.pending_review_comment,
                "pending_reviewer_id": plan_tc.pending_reviewer_id,
                "pending_reviewed_at": plan_tc.pending_reviewed_at
            },
            "navigation": {
                "current_index": current_index + 1,
                "total_count": len(all_testcase_ids),
                "prev_testcase_id": prev_testcase_id,
                "next_testcase_id": next_testcase_id
            }
        }
    }
