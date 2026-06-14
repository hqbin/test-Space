from openpyxl import Workbook, load_workbook
from io import BytesIO
from sqlalchemy import func
import json
import re

def split_steps_by_number(text):
    """
    解析步骤/预期结果文本：
    - 有序号（至少2行以 1. 2. 1、等开头）→ 按序号分隔成多条，续行合并到上一条
    - 无序号 → 整体作为1条，保留内部换行
    注意：小数点不被视为序号分隔符（如 177.5MHz 不会被拆分）
    """
    if not text:
        return []
    lines = text.split('\n')
    # 序号行：数字+标点开头，. 后不能紧跟数字（排除小数如 177.5）
    numbered_lines = [l for l in lines if re.match(r'^\s*\d+\s*([。、]|\.)(?!\d)', l.strip())]
    if len(numbered_lines) >= 2:
        items = []
        current = None
        for line in lines:
            stripped = line.strip()
            if re.match(r'^\s*\d+\s*([。、]|\.)(?!\d)\s*', stripped):
                if current is not None:
                    items.append(current.strip())
                # 去掉序号前缀
                current = re.sub(r'^\s*\d+\s*([。、]|\.)(?!\d)\s*', '', stripped)
            elif current is not None:
                # 续行合并，保留换行
                current += '\n' + line
        if current is not None:
            items.append(current.strip())
        return [i for i in items if i]
    else:
        # 无序号：整体1条，保留内部换行
        stripped = text.strip()
        return [stripped] if stripped else []


def translate_automation(value):
    if not value:
        return ""
    if value == "Y":
        return "是"
    if value.startswith("N-"):
        labels = {
            "N-HW_PHYSICAL": "否-硬件物理交互类",
            "N-VISUAL_JUDGE": "否-视觉主观判断类",
            "N-BOOT_PROCESS": "否-开机启动流程类",
            "N-MEDIA_PLAY": "否-媒体播放验证类",
            "N-OTA_UPGRADE": "否-OTA升级专项类",
            "N-DATA_CONFIG": "否-数据配置核对类",
            "N-LOG_CHECK": "否-日志数据核对类",
            "N-BACKEND_CONFIG": "否-后台配置交互类",
            "N-DATA_DYNAMIC": "否-数据动态变化类",
            "N-OTHER_SPECIAL": "否-其他特殊场景类"
        }
        return labels.get(value, f"否-{value}")
    if value == "N":
        return "否"
    return value

def parse_excel_v2(file_content: bytes, user_id: int, db=None, task_id=None, task_mgr=None):
    """
    解析Excel文件，返回测试用例列表
    必须包含"所属用例库"列，模块列支持"模块/子模块"层级格式
    表头位置灵活，不限定列顺序
    task_id/task_mgr: 可选，用于在解析过程中更新进度
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # 检查文件内容是否有效
    if not file_content:
        logger.error("Excel文件内容为空")
        raise ValueError("无法读取Excel文件: 文件内容为空")
    
    file_size = len(file_content)
    logger.info(f"开始解析Excel文件, 大小: {file_size} bytes")
    
    if file_size < 100:
        logger.error(f"Excel文件内容不完整, 大小: {file_size} bytes")
        raise ValueError(f"无法读取Excel文件: 文件内容不完整 (大小: {file_size} bytes)")
    
    # 检查文件头是否是有效的zip文件（xlsx本质是zip）
    if file_content[0:2] != b'PK':
        logger.error(f"Excel文件头无效: {file_content[0:10]}")
        raise ValueError("无法读取Excel文件: 文件格式无效，不是有效的Excel文件")
    
    try:
        wb = load_workbook(BytesIO(file_content), data_only=True)
    except Exception as e:
        error_msg = str(e)
        logger.error(f"load_workbook失败: {error_msg}")
        # 提供更友好的错误提示
        if "worksheets from None" in error_msg or "invalid XML" in error_msg.lower():
            raise ValueError("无法读取Excel文件: 文件可能损坏或不是有效的Excel文件。请尝试：1)重新下载模板；2)使用WPS重新保存为Excel格式；3)检查文件是否在传输过程中损坏")
        raise ValueError(f"无法读取Excel文件: {error_msg}")
    
    if not wb:
        logger.error("workbook对象为None")
        raise ValueError("无法读取Excel文件: 工作簿为空")
    
    if not wb.sheetnames:
        logger.error("workbook没有工作表")
        raise ValueError("无法读取Excel文件: 工作簿没有工作表")
    
    logger.info(f"Excel文件解析成功, 工作表数量: {len(wb.sheetnames)}, 工作表名称: {wb.sheetnames}")
    
    testcases = []
    valid_sheet_count = 0
    
    # 遍历所有sheet
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            
            # 检查sheet是否为空
            if ws.max_row < 2:
                logger.info(f"跳过空sheet: {sheet_name}")
                continue
            
            # 读取表头
            first_row_list = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not first_row_list or not first_row_list[0]:
                logger.info(f"跳过空sheet（无表头）: {sheet_name}")
                continue
            
            headers = [str(cell).strip() if cell else "" for cell in first_row_list[0]]
            
            # 必须包含"所属用例库"列，否则跳过（非数据sheet）
            project_col_keywords = {"所属用例库", "所属用例库*", "project_name"}
            has_project_col = any(h in project_col_keywords for h in headers)
            if not has_project_col:
                logger.info(f"跳过非数据sheet（无'所属用例库'列）: {sheet_name}")
                continue
            
            # 表头映射（灵活匹配，不限定列顺序）
            header_mapping = {
                '所属用例库': 'project_name',
                '所属用例库*': 'project_name',
                'project_name': 'project_name',
                '用例编号': 'case_number',
                'case_number': 'case_number',
                '用例编号*': 'case_number',
                '用例类型': 'case_type',
                'case_type': 'case_type',
                '模块': 'module_path',
                'module': 'module_path',
                '模块*': 'module_path',
                '用例标题': 'name',
                'name': 'name',
                '用例标题*': 'name',
                '标题': 'name',
                '前置条件': 'precondition',
                'precondition': 'precondition',
                '操作步骤': 'steps',
                'steps': 'steps',
                '操作步骤*': 'steps',
                '测试步骤': 'steps',
                '预期结果': 'expected_result',
                'expected_result': 'expected_result',
                '预期结果*': 'expected_result',
                '期望结果': 'expected_result',
                '用例等级': 'level',
                'level': 'level',
                '用例等级*': 'level',
                '等级': 'level',
                '优先级': 'level',
                '备注': 'remarks',
                'remarks': 'remarks',
                '说明': 'remarks',
                '自动化': 'automation',
                'automation': 'automation'
            }
            
            # 建立列索引映射
            column_map = {}
            for idx, header in enumerate(headers):
                if header in header_mapping:
                    column_map[header_mapping[header]] = idx
            
            # 验证必填列（用例编号已改为自动生成，不再是必填字段）
            required_fields = ['project_name', 'module_path', 'name', 'steps', 'expected_result', 'level']
            field_names = {
                'project_name': '所属用例库',
                'module_path': '模块',
                'name': '用例标题',
                'steps': '操作步骤',
                'expected_result': '预期结果',
                'level': '用例等级'
            }
            
            missing_fields = [f for f in required_fields if f not in column_map]
            if missing_fields:
                missing_names = [field_names.get(f, f) for f in missing_fields]
                raise ValueError(f"Sheet '{sheet_name}' 缺少必填列: {', '.join(missing_names)}")
            
            valid_sheet_count += 1
            print(f"开始解析sheet: {sheet_name}")
            print(f"  列映射: {column_map}")
            
            total_rows = ws.max_row - 1  # 减去表头
            
            # 解析每一行数据
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 定期更新进度
                    if task_mgr and task_id and (row_idx - 2) % 1000 == 0:
                        progress = min(90, int((row_idx - 2) / total_rows * 90))
                        task_mgr.update_task(task_id, status="parsing", 
                            message=f"正在解析{sheet_name}：{row_idx-1}/{total_rows}行...", 
                            progress=progress)
                    
                    # 跳过空行（用例标题为空则跳过）
                    def get_cell_value(field, default=None):
                        if field in column_map and column_map[field] < len(row):
                            value = row[column_map[field]]
                            return str(value).strip() if value else default
                        return default
                    
                    project_name = get_cell_value('project_name')
                    # 用例编号已改为自动生成，不再从Excel读取
                    case_number = None
                    case_type = get_cell_value('case_type', 'COMMON')
                    module_path = get_cell_value('module_path', '未分类')
                    name = get_cell_value('name')
                    precondition = get_cell_value('precondition')
                    steps_text = get_cell_value('steps', '')
                    expected_text = get_cell_value('expected_result', '')
                    level = get_cell_value('level')  # 移除默认值，校验不能为空
                    remarks = get_cell_value('remarks')
                    automation = get_cell_value('automation')
                    
                    # 只跳过完全空白的行（所有字段都为空）
                    if not row or (not project_name and not name and not steps_text and not expected_text and not level):
                        continue
                    
                    # 必填校验（用例编号除外）
                    if not project_name:
                        raise ValueError(f"Sheet '{sheet_name}' 第{row_idx}行：所属用例库不能为空")
                    if not module_path:
                        raise ValueError(f"Sheet '{sheet_name}' 第{row_idx}行：模块不能为空")
                    if not name:
                        raise ValueError(f"Sheet '{sheet_name}' 第{row_idx}行：用例标题不能为空")
                    if not steps_text:
                        raise ValueError(f"Sheet '{sheet_name}' 第{row_idx}行：操作步骤不能为空")
                    if not expected_text:
                        raise ValueError(f"Sheet '{sheet_name}' 第{row_idx}行：预期结果不能为空")
                    if not level:
                        raise ValueError(f"Sheet '{sheet_name}' 第{row_idx}行：用例等级不能为空")
                    # 去掉前导斜杠（防止路径格式异常）
                    module_path = module_path.lstrip('/')
                    
                    # 验证用例类型
                    valid_types = ['COMMON', 'PERFORMANCE', 'SECURITY', 'INTERFACE', 'INSTALL', 'CONFIG', 'OTHER']
                    type_map = {
                        '功能测试': 'COMMON',
                        '性能测试': 'PERFORMANCE',
                        '安全测试': 'SECURITY',
                        '接口测试': 'INTERFACE',
                        '安装部署': 'INSTALL',
                        '配置相关': 'CONFIG',
                        '其他': 'OTHER'
                    }
                    if case_type in type_map:
                        case_type = type_map[case_type]
                    elif case_type not in valid_types:
                        case_type = 'COMMON'
                    
                    # 验证等级
                    if not level:
                        raise ValueError(f"Sheet '{sheet_name}' 第{row_idx}行：用例等级不能为空")
                    if level not in ['L1', 'L2', 'L3', 'L4']:
                        raise ValueError(f"Sheet '{sheet_name}' 第{row_idx}行：用例等级必须是L1/L2/L3/L4")
                    
                    # 验证自动化字段
                    valid_automation = ['Y', 'N', '']
                    valid_automation.extend([
                        'N-HW_PHYSICAL', 'N-VISUAL_JUDGE', 'N-BOOT_PROCESS', 'N-MEDIA_PLAY',
                        'N-OTA_UPGRADE', 'N-DATA_CONFIG', 'N-LOG_CHECK', 'N-BACKEND_CONFIG', 'N-DATA_DYNAMIC', 'N-OTHER_SPECIAL'
                    ])
                    if automation and automation not in valid_automation:
                        automation = None
                    
                    # 解析步骤和预期结果
                    # 有序号 → 按序号分隔成多条；无序号 → 整体1条保留换行
                    steps_list = []
                    steps_items = split_steps_by_number(steps_text)
                    expected_items = split_steps_by_number(expected_text)

                    max_len = max(len(steps_items), len(expected_items))
                    for i in range(max_len):
                        step = steps_items[i] if i < len(steps_items) else ""
                        expected = expected_items[i] if i < len(expected_items) else ""
                        if step or expected:
                            steps_list.append({"step": step, "expected": expected})
                    
                    if not steps_list:
                        steps_list.append({"step": "", "expected": ""})
                    
                    steps_json = json.dumps(steps_list, ensure_ascii=False)
                    
                    testcases.append({
                        "project_name": project_name,
                        "case_number": case_number,
                        "case_type": case_type,
                        "module_path": module_path,
                        "name": name,
                        "precondition": precondition,
                        "steps": steps_json,
                        "expected_result": expected_text,
                        "level": level,
                        "remarks": remarks,
                        "automation": automation,
                        "status": "PENDING",
                        "source": "LOCAL",
                        "created_by": user_id,
                    })
                    
                except Exception as e:
                    error_msg = f"Sheet '{sheet_name}' 第{row_idx}行解析失败: {str(e)}"
                    print(error_msg)
                    raise ValueError(error_msg)
                    
        except Exception as e:
            error_msg = f"解析sheet '{sheet_name}' 时出错: {str(e)}"
            print(error_msg)
            raise ValueError(error_msg)
    
    if valid_sheet_count == 0:
        raise ValueError("Excel文件中没有找到有效的测试用例数据。请确保：1)Excel包含'所属用例库'列；2)已删除模板中的说明sheet页；3)使用系统提供的模板。")
    
    if not testcases:
        raise ValueError(f"Excel文件中没有找到任何测试用例数据。请检查是否已删除模板中的说明sheet页。")
    
    print(f"成功解析 {len(testcases)} 个测试用例（来自 {valid_sheet_count} 个sheet）")
    return testcases


def parse_csv(file_content: bytes, user_id: int, db=None):
    """
    解析CSV文件，返回测试用例列表
    与parse_excel_v2使用相同的字段映射和验证逻辑
    支持UTF-8和GBK编码自动检测
    """
    import csv
    from io import StringIO

    # 尝试解码：先UTF-8（去BOM），再GBK
    for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']:
        try:
            text = file_content.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        raise ValueError("无法识别CSV文件编码，请使用UTF-8或GBK编码保存文件")

    text = text.replace('\r\n', '\n').replace('\r', '\n')
    reader = csv.reader(StringIO(text))
    rows = list(reader)

    if len(rows) < 2:
        raise ValueError("CSV文件为空或只有表头，没有数据行")

    headers = [h.strip() for h in rows[0]]

    # 表头映射（与parse_excel_v2一致）
    header_mapping = {
        '所属用例库': 'project_name', '所属用例库*': 'project_name', 'project_name': 'project_name',
        '用例编号': 'case_number', 'case_number': 'case_number', '用例编号*': 'case_number',
        '用例类型': 'case_type', 'case_type': 'case_type',
        '模块': 'module_path', 'module': 'module_path', '模块*': 'module_path',
        '用例标题': 'name', 'name': 'name', '用例标题*': 'name', '标题': 'name',
        '前置条件': 'precondition', 'precondition': 'precondition',
        '操作步骤': 'steps', 'steps': 'steps', '操作步骤*': 'steps', '测试步骤': 'steps',
        '预期结果': 'expected_result', 'expected_result': 'expected_result', '预期结果*': 'expected_result', '期望结果': 'expected_result',
        '用例等级': 'level', 'level': 'level', '用例等级*': 'level', '等级': 'level', '优先级': 'level',
        '备注': 'remarks', 'remarks': 'remarks', '说明': 'remarks',
        '自动化': 'automation', 'automation': 'automation'
    }

    # 建立列索引映射
    column_map = {}
    for idx, header in enumerate(headers):
        if header in header_mapping:
            column_map[header_mapping[header]] = idx

    # 验证必填列
    project_col_keywords = {"所属用例库", "所属用例库*", "project_name"}
    if not any(h in project_col_keywords for h in headers):
        raise ValueError("CSV文件缺少必填列: 所属用例库。请使用系统提供的模板导入。")

    # 验证必填列（用例编号已改为自动生成，不再是必填字段）
    required_fields = ['project_name', 'module_path', 'name', 'steps', 'expected_result', 'level']
    field_names = {
        'project_name': '所属用例库', 'module_path': '模块',
        'name': '用例标题', 'steps': '操作步骤', 'expected_result': '预期结果', 'level': '用例等级'
    }
    missing_fields = [f for f in required_fields if f not in column_map]
    if missing_fields:
        missing_names = [field_names.get(f, f) for f in missing_fields]
        raise ValueError(f"CSV文件缺少必填列: {', '.join(missing_names)}")

    testcases = []
    valid_types = ['COMMON', 'PERFORMANCE', 'SECURITY', 'INTERFACE', 'INSTALL', 'CONFIG', 'OTHER']
    type_map = {
        '功能测试': 'COMMON', '性能测试': 'PERFORMANCE', '安全测试': 'SECURITY',
        '接口测试': 'INTERFACE', '安装部署': 'INSTALL', '配置相关': 'CONFIG', '其他': 'OTHER'
    }

    for row_idx, row in enumerate(rows[1:], start=2):
        # 跳过空行
        if not row or all(not cell.strip() for cell in row):
            continue

        def get_val(field, default=None):
            if field in column_map and column_map[field] < len(row):
                v = row[column_map[field]].strip()
                return v if v else default
            return default

        # 用例编号已改为自动生成，不再从CSV读取
        case_number = None
        project_name = get_val('project_name')
        if not project_name:
            raise ValueError(f"第{row_idx}行：所属用例库不能为空")
        
        name = get_val('name')
        if not name:
            raise ValueError(f"第{row_idx}行：用例标题不能为空")

        module_path = get_val('module_path', '未分类')
        if not module_path:
            raise ValueError(f"第{row_idx}行：模块不能为空")
        
        steps_text = get_val('steps', '')
        if not steps_text:
            raise ValueError(f"第{row_idx}行：操作步骤不能为空")
        
        expected_text = get_val('expected_result', '')
        if not expected_text:
            raise ValueError(f"第{row_idx}行：预期结果不能为空")
        
        level = get_val('level')
        if not level:
            raise ValueError(f"第{row_idx}行：用例等级不能为空")
        if level not in ['L1', 'L2', 'L3', 'L4']:
            raise ValueError(f"第{row_idx}行：用例等级必须是L1/L2/L3/L4")
        # 去掉前导斜杠（CSV可能导出带前导/）
        module_path = module_path.lstrip('/')

        case_type = get_val('case_type', 'COMMON')
        if case_type in type_map:
            case_type = type_map[case_type]
        elif case_type not in valid_types:
            case_type = 'COMMON'

        automation = get_val('automation')
        valid_automation = ['Y', 'N', '']
        valid_automation.extend([
            'N-HW_PHYSICAL', 'N-VISUAL_JUDGE', 'N-BOOT_PROCESS', 'N-MEDIA_PLAY',
            'N-OTA_UPGRADE', 'N-DATA_CONFIG', 'N-LOG_CHECK', 'N-BACKEND_CONFIG', 'N-DATA_DYNAMIC', 'N-OTHER_SPECIAL'
        ])
        if automation and automation not in valid_automation:
            automation = None

        # 解析步骤和预期结果
        # 有序号 → 按序号分隔成多条；无序号 → 整体1条保留换行
        steps_items = split_steps_by_number(steps_text)
        expected_items = split_steps_by_number(expected_text)
        steps_list = []
        for i in range(max(len(steps_items), len(expected_items))):
            step = steps_items[i] if i < len(steps_items) else ""
            expected = expected_items[i] if i < len(expected_items) else ""
            if step or expected:
                steps_list.append({"step": step, "expected": expected})
        if not steps_list:
            steps_list.append({"step": "", "expected": ""})

        testcases.append({
            "project_name": project_name,
            "case_number": case_number,
            "case_type": case_type,
            "module_path": module_path,
            "name": get_val('name', f"测试用例{row_idx}"),
            "precondition": get_val('precondition'),
            "steps": json.dumps(steps_list, ensure_ascii=False),
            "expected_result": expected_text,
            "level": level,
            "remarks": get_val('remarks'),
            "automation": automation,
            "status": "PENDING",
            "source": "LOCAL",
            "created_by": user_id,
        })

    if not testcases:
        raise ValueError("CSV文件中没有找到任何测试用例数据")

    print(f"成功从CSV解析 {len(testcases)} 个测试用例")
    return testcases


def generate_excel(testcases, db=None, default_project_name=None):
    """生成Excel文件
    
    Args:
        testcases: 测试用例列表
        db: 数据库会话，用于查询用例库名称
        default_project_name: 默认用例库名称。当 testcases 为空时，会在表头下方添加一行
            仅填充"所属用例库"列，方便用户直接基于该模板填写用例（避免名称错误）
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    # 表头（包含所属用例库）
    headers = [
        "用例编号", "所属用例库", "用例类型", "模块", "用例标题", "前置条件",
        "操作步骤", "预期结果", "用例等级", "备注", "自动化"
    ]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="8B9AEE", end_color="8B9AEE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15  # 用例编号
    ws.column_dimensions['B'].width = 20  # 所属用例库
    ws.column_dimensions['C'].width = 12  # 用例类型
    ws.column_dimensions['D'].width = 30  # 模块
    ws.column_dimensions['E'].width = 30  # 用例标题
    ws.column_dimensions['F'].width = 30  # 前置条件
    ws.column_dimensions['G'].width = 40  # 操作步骤
    ws.column_dimensions['H'].width = 40  # 预期结果
    ws.column_dimensions['I'].width = 12  # 用例等级
    ws.column_dimensions['J'].width = 30  # 备注
    ws.column_dimensions['K'].width = 10  # 自动化
    
    # 用例类型映射
    case_type_map = {
        'COMMON': '功能测试',
        'PERFORMANCE': '性能测试',
        'SECURITY': '安全测试',
        'INTERFACE': '接口测试',
        'INSTALL': '安装部署',
        'CONFIG': '配置相关',
        'OTHER': '其他'
    }
    
    # 移除文本开头序号的辅助函数
    def remove_leading_numbers(text):
        """移除文本开头的序号（如 1. 2. 3.），排除小数点（如 3.14 不被识别为序号）"""
        if not text:
            return text
        # 持续移除开头的 "数字+标点" 格式，. 后不能紧跟数字（避免误删小数）
        while re.match(r'^\d+([。、]|\.)(?!\d)\s*', text):
            text = re.sub(r'^\d+([。、]|\.)(?!\d)\s*', '', text).strip()
        return text
    
    # 预加载用例库名称缓存
    project_name_cache = {}
    if db:
        from models import Project
        project_ids = set(tc.primary_project_id for tc in testcases if tc.primary_project_id)
        if project_ids:
            projects = db.query(Project).filter(Project.id.in_(project_ids)).all()
            project_name_cache = {p.id: p.name for p in projects}
    
    for tc in testcases:
        # 解析步骤JSON
        steps_text = ""
        expected_text = ""
        
        try:
            steps_data = json.loads(tc.steps)
            if isinstance(steps_data, list):
                steps_lines = []
                expected_lines = []
                for i, item in enumerate(steps_data, 1):
                    if isinstance(item, dict):
                        step = remove_leading_numbers(item.get('step', ''))
                        expected = remove_leading_numbers(item.get('expected', ''))
                        steps_lines.append(f"{i}. {step}")
                        expected_lines.append(f"{i}. {expected}")
                    else:
                        cleaned_item = remove_leading_numbers(str(item))
                        steps_lines.append(f"{i}. {cleaned_item}")
                steps_text = '\n'.join(steps_lines)
                expected_text = '\n'.join(expected_lines)
        except (json.JSONDecodeError, AttributeError):
            # 如果不是JSON格式，使用原始文本并移除序号
            if hasattr(tc, 'steps') and tc.steps:
                lines = tc.steps.split('\n')
                steps_lines = []
                for i, line in enumerate(lines, 1):
                    cleaned_line = remove_leading_numbers(line.strip())
                    if cleaned_line:
                        steps_lines.append(f"{i}. {cleaned_line}")
                steps_text = '\n'.join(steps_lines)
            
            if hasattr(tc, 'expected_result') and tc.expected_result:
                lines = tc.expected_result.split('\n')
                expected_lines = []
                for i, line in enumerate(lines, 1):
                    cleaned_line = remove_leading_numbers(line.strip())
                    if cleaned_line:
                        expected_lines.append(f"{i}. {cleaned_line}")
                expected_text = '\n'.join(expected_lines)
        
        ws.append([
            tc.case_number,
            project_name_cache.get(tc.primary_project_id, "") if tc.primary_project_id else "",
            case_type_map.get(tc.case_type, '功能测试'),
            tc.module or "",
            tc.name,
            tc.precondition or "",
            steps_text,
            expected_text,
            tc.level,
            tc.remarks or "",
            translate_automation(tc.automation)
        ])
    
    # 如果没有测试用例且提供了默认用例库名称，则在表头下方追加一行
    # 仅填充"所属用例库"列，方便用户直接在该模板上补充用例数据（避免手填名称出错）
    if not testcases and default_project_name:
        ws.append([
            "",                    # 用例编号
            default_project_name,  # 所属用例库
            "",                    # 用例类型
            "",                    # 模块
            "",                    # 用例标题
            "",                    # 前置条件
            "",                    # 操作步骤
            "",                    # 预期结果
            "",                    # 用例等级
            "",                    # 备注
            "",                    # 自动化
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
