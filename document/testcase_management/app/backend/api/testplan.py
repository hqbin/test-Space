from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import func as sa_func
from typing import Optional
from datetime import datetime
from database import get_db
from models import TestPlan, User, TestExecution, TestPlanTestCase, TestPlanExecutor, TestPlanViewer, Report, ReportTemplate, TestCase
from schemas import TestPlanCreate
from auth import get_current_user
from utils.logger import log_operation, LogAction, LogModule
from utils.notification_helper import trigger_assignment_notification, _format_datetime_str
from utils.report_conclusion import get_conclusion_and_criteria
import csv
import io
import json
import re

router = APIRouter()


def _get_module_sort_key_with_fallback(module_path, sort_map):
    """模块排序辅助函数：支持子模块fallback - 匹配最长前缀的模块路径
    
    例如：
    - 模块路径 "ISDB/Setting/Network" 会依次尝试匹配:
      1. "ISDB/Setting/Network" (精确匹配)
      2. "ISDB/Setting" 
      3. "ISDB"
    
    返回的排序键会在匹配的父模块后追加 .9999999999 以确保子模块排在父模块的子用例之后
    """
    if not module_path:
        return "9999999999"
    # 尝试精确匹配
    if module_path in sort_map:
        return sort_map[module_path]
    # 尝试逐层匹配：从最长前缀到最短前缀
    parts = module_path.split('/')
    for i in range(len(parts) - 1, 0, -1):
        prefix = '/'.join(parts[:i])
        if prefix in sort_map:
            # 追加一个大的排序值，确保该路径下的所有用例排在已知子模块之后
            return sort_map[prefix] + '.9999999999'
    return "9999999999"


@router.get("")
def list_testplans(
    req: Request,
    page: int = 1,
    size: int = 10,
    project_id: Optional[int] = None,
    project_ids: Optional[str] = None,
    team_id: Optional[int] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from utils.project_helper import get_project_ids_with_children
    from utils.data_permission import apply_testplan_data_permission
    from sqlalchemy import func
    import logging
    import traceback
    logger = logging.getLogger(__name__)
    
    try:
        # 限制 size 最大值，防止内存溢出
        if size > 10000:
            size = 10000
        
        query = db.query(TestPlan)
        
        # 应用数据权限过滤
        logger.info(f"[list_testplans] user={current_user.username}, project_ids={project_ids}, project_id={project_id}, team_id={team_id}, page={page}, size={size}")
        query = apply_testplan_data_permission(query, current_user, db)
        
        # 按项目组过滤（优先级最高，确保项目组隔离）
        if team_id:
            query = query.filter(TestPlan.team_id == team_id)

        # 搜索过滤
        if search:
            query = query.filter(TestPlan.name.ilike(f'%{search}%'))
        
        # 支持多个用例库ID筛选
        if project_ids:
            try:
                pid_list = [int(pid.strip()) for pid in project_ids.split(',') if pid.strip()]
                if pid_list:
                    query = query.filter(TestPlan.project_id.in_(pid_list))
            except ValueError:
                pass
        elif project_id:
            # 获取项目ID列表（包括子项目）
            project_ids_list = get_project_ids_with_children(db, project_id)
            
            if project_ids_list:
                # 查询这些项目的所有测试计划
                query = query.filter(TestPlan.project_id.in_(project_ids_list))
        
        total = query.count()
        testplans = query.order_by(TestPlan.updated_at.desc(), TestPlan.id.desc()).offset((page - 1) * size).limit(size).all()
        logger.info(f"[list_testplans] total={total}, fetched={len(testplans)}")
        
        # 批量获取所有测试计划的统计数据（一次查询）
        plan_ids = [plan.id for plan in testplans]
        
        # 批量查询执行人
        executors_by_plan = {}
        if plan_ids:
            executor_records = db.query(TestPlanExecutor, User).join(
                User, TestPlanExecutor.executor_id == User.id
            ).filter(TestPlanExecutor.test_plan_id.in_(plan_ids)).all()
            
            for executor_record, user in executor_records:
                if executor_record.test_plan_id not in executors_by_plan:
                    executors_by_plan[executor_record.test_plan_id] = []
                executors_by_plan[executor_record.test_plan_id].append(user)
        
        # 批量查询查看人
        viewers_by_plan = {}
        if plan_ids:
            viewer_records = db.query(TestPlanViewer, User).join(
                User, TestPlanViewer.viewer_id == User.id
            ).filter(TestPlanViewer.test_plan_id.in_(plan_ids)).all()
            
            for viewer_record, user in viewer_records:
                if viewer_record.test_plan_id not in viewers_by_plan:
                    viewers_by_plan[viewer_record.test_plan_id] = []
                viewers_by_plan[viewer_record.test_plan_id].append(user)
        
        # 批量查询审核人
        reviewer_ids = list(set([plan.reviewer_id for plan in testplans if plan.reviewer_id]))
        reviewers_dict = {}
        if reviewer_ids:
            reviewers = db.query(User).filter(User.id.in_(reviewer_ids)).all()
            reviewers_dict = {reviewer.id: reviewer for reviewer in reviewers}
        
        # 批量查询所有计划的关联用例（一次查询替代原来的 2N 次循环）
        plan_testcase_ids = {}
        testcase_counts = {}
        if plan_ids:
            all_plan_tcs = db.query(
                TestPlanTestCase.test_plan_id,
                TestPlanTestCase.test_case_id
            ).join(
                TestCase, TestPlanTestCase.test_case_id == TestCase.id
            ).filter(
                TestPlanTestCase.test_plan_id.in_(plan_ids)
            ).all()
            for row in all_plan_tcs:
                plan_testcase_ids.setdefault(row.test_plan_id, set()).add(row.test_case_id)
            for pid in plan_ids:
                testcase_counts[pid] = len(plan_testcase_ids.get(pid, set()))
        
        # 批量查询所有执行记录（只统计关联用例）
        executions = db.query(TestExecution).filter(
            TestExecution.test_plan_id.in_(plan_ids)
        ).all() if plan_ids else []
        
        # 按测试计划分组执行记录（只保留关联用例的执行记录）
        executions_by_plan = {}
        for execution in executions:
            # 只处理在关联列表中的用例
            linked_ids = plan_testcase_ids.get(execution.test_plan_id, set())
            if execution.test_case_id not in linked_ids:
                continue
            if execution.test_plan_id not in executions_by_plan:
                executions_by_plan[execution.test_plan_id] = []
            executions_by_plan[execution.test_plan_id].append(execution)
        
        # 批量查询已审核通过的报告（用于在测试计划列表显示报告名称）
        reports_by_plan = {}
        if plan_ids:
            approved_reports = db.query(Report).filter(
                Report.test_plan_id.in_(plan_ids),
                Report.status == 'APPROVED'
            ).all()
            for report in approved_reports:
                # 每个计划可能有多个报告，取最新的
                if report.test_plan_id not in reports_by_plan:
                    reports_by_plan[report.test_plan_id] = report
                elif report.created_at and reports_by_plan[report.test_plan_id].created_at and report.created_at > reports_by_plan[report.test_plan_id].created_at:
                    reports_by_plan[report.test_plan_id] = report
        
        # 为每个测试计划添加执行人信息和统计数据
        result_list = []
        for plan in testplans:
            # 执行人信息
            executors = executors_by_plan.get(plan.id, [])
            
            # 统计用例数据
            total_testcases = testcase_counts.get(plan.id, 0)
            
            # 统计执行数据 - 只统计每个用例的最新执行结果
            plan_executions = executions_by_plan.get(plan.id, [])
            
            # 按用例ID分组，获取每个用例的最新执行记录
            latest_executions = {}
            for execution in plan_executions:
                if execution.test_case_id not in latest_executions:
                    latest_executions[execution.test_case_id] = execution
                else:
                    # 比较执行时间，保留最新的（防止 None 值）
                    current_at = execution.executed_at
                    existing_at = latest_executions[execution.test_case_id].executed_at
                    if current_at and existing_at and current_at > existing_at:
                        latest_executions[execution.test_case_id] = execution
                    elif current_at and not existing_at:
                        latest_executions[execution.test_case_id] = execution
            
            # 统计各种状态的用例数量（基于最新执行记录，ONGOING不计入已执行）
            executed_testcases = sum(1 for e in latest_executions.values() if e.result != 'ONGOING' and e.result != 'PENDING')
            passed_testcases = sum(1 for e in latest_executions.values() if e.result == 'PASS')
            failed_testcases = sum(1 for e in latest_executions.values() if e.result == 'FAIL')
            blocked_testcases = sum(1 for e in latest_executions.values() if e.result == 'BLOCK')
            na_testcases = sum(1 for e in latest_executions.values() if e.result == 'NA')
            nt_testcases = sum(1 for e in latest_executions.values() if e.result == 'NT')
            
            # 审核人信息
            reviewer = reviewers_dict.get(plan.reviewer_id)
            
            viewers = viewers_by_plan.get(plan.id, [])

            plan_dict = {
                "id": plan.id,
                "project_id": plan.project_id,
                "name": plan.name,
                "description": plan.description,
                "start_time": plan.start_time,
                "end_time": plan.end_time,
                "status": plan.status,
                "reviewer_id": plan.reviewer_id,
                "reviewer_name": reviewer.username if reviewer else None,
                "created_by": plan.created_by,
                "created_at": plan.created_at,
                "updated_at": plan.updated_at,
                "total_testcases": total_testcases,
                "executed_testcases": executed_testcases,
                "passed_testcases": passed_testcases,
                "failed_testcases": failed_testcases,
                "blocked_testcases": blocked_testcases,
                "na_testcases": na_testcases,
                "nt_testcases": nt_testcases,
                "executor_names": [u.username for u in executors],
                "executors": [{"id": u.id, "username": u.username, "full_name": u.full_name} for u in executors],
                "viewer_ids": [u.id for u in viewers],
                "viewer_names": [u.username for u in viewers],
                "viewers": [{"id": u.id, "username": u.username, "full_name": u.full_name} for u in viewers]
            }
            
            # 添加已审核通过的报告信息
            approved_report = reports_by_plan.get(plan.id)
            if approved_report:
                display_name = approved_report.project_name or approved_report.name or ''
                # 去掉 " - 测试报告" 后缀
                if display_name.endswith(' - 测试报告'):
                    display_name = display_name[:-len(' - 测试报告')].rstrip()
                plan_dict["report_id"] = approved_report.id
                plan_dict["report_name"] = display_name
            else:
                plan_dict["report_id"] = None
                plan_dict["report_name"] = None
            
            result_list.append(plan_dict)
        
        return {
            "code": 200,
            "message": "success",
            "data": {
                "records": result_list,
                "total": total
            }
        }
    except Exception as e:
        logger.error(f"[list_testplans] ERROR for user={current_user.username}: {str(e)}")
        logger.error(f"[list_testplans] traceback: {traceback.format_exc()}")
        return {
            "code": 500,
            "message": f"服务器内部错误: {str(e)}",
            "data": {
                "records": [],
                "total": 0
            }
        }

@router.post("")
async def create_testplan(
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 从请求体中获取数据
    body = await req.json()
    
    # 提取测试用例、执行人和审核人信息
    test_case_ids = body.get('test_case_ids', [])
    executor_ids = body.get('executor_ids', [])
    reviewer_id = body.get('reviewer_id')
    suite_id = body.get('suite_id')
    
    # 如果选择了套件且没有手动选用例，从套件加载用例
    if suite_id and not test_case_ids:
        from models import TestSuiteTestCase
        links = db.query(TestSuiteTestCase).filter(TestSuiteTestCase.test_suite_id == suite_id).all()
        test_case_ids = [l.test_case_id for l in links]
    
    # 提取测试用例、执行人、查看人和审核人信息
    test_case_ids = body.get('test_case_ids', [])
    executor_ids = body.get('executor_ids', [])
    viewer_ids = body.get('viewer_ids', [])
    reviewer_id = body.get('reviewer_id')
    suite_id = body.get('suite_id')
    
    # 校验：查看人和执行人不能重叠
    if viewer_ids and executor_ids:
        overlap = set(viewer_ids) & set(executor_ids)
        if overlap:
            overlap_users = db.query(User).filter(User.id.in_(overlap)).all()
            overlap_names = [u.username for u in overlap_users]
            raise HTTPException(
                status_code=400,
                detail=f"用户 {', '.join(overlap_names)} 不能同时设为查看人和执行人"
            )
    
    # 创建测试计划数据
    testplan_data = {
        'name': body.get('name'),
        'description': body.get('description'),
        'start_time': body.get('start_time'),
        'end_time': body.get('end_time'),
        'project_id': body.get('project_id'),
        'team_id': body.get('team_id'),
        'suite_id': suite_id,
        'reviewer_id': reviewer_id
    }
    
    # 转换日期字符串为 datetime 对象
    if testplan_data.get('start_time'):
        try:
            testplan_data['start_time'] = datetime.strptime(testplan_data['start_time'], '%Y-%m-%d')
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="开始时间格式错误，应为 YYYY-MM-DD")
    
    if testplan_data.get('end_time'):
        try:
            testplan_data['end_time'] = datetime.strptime(testplan_data['end_time'], '%Y-%m-%d')
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="结束时间格式错误，应为 YYYY-MM-DD")
    
    db_testplan = TestPlan(**testplan_data, created_by=current_user.id, status="PENDING")
    db.add(db_testplan)
    db.commit()
    db.refresh(db_testplan)
    
    # 存储测试计划与测试用例的关联关系
    if test_case_ids:
        for test_case_id in test_case_ids:
            test_plan_test_case = TestPlanTestCase(
                test_plan_id=db_testplan.id,
                test_case_id=test_case_id
            )
            db.add(test_plan_test_case)
    
    # 存储测试计划与执行人的关联关系
    if executor_ids:
        for executor_id in executor_ids:
            test_plan_executor = TestPlanExecutor(
                test_plan_id=db_testplan.id,
                executor_id=executor_id
            )
            db.add(test_plan_executor)
    
    # 存储测试计划与查看人的关联关系
    if viewer_ids:
        for viewer_id in viewer_ids:
            test_plan_viewer = TestPlanViewer(
                test_plan_id=db_testplan.id,
                viewer_id=viewer_id
            )
            db.add(test_plan_viewer)
    
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTPLANS,
        action=LogAction.CREATE,
        description=f"创建测试计划：{db_testplan.name}（ID: {db_testplan.id}），关联 {len(test_case_ids)} 个用例，{len(executor_ids)} 个执行人，{len(viewer_ids)} 个查看人",
        request=req
    )
    
    # 获取执行人名称
    executor_names = []
    for executor_id in executor_ids:
        user = db.query(User).filter(User.id == executor_id).first()
        if user:
            executor_names.append(user.username)
    
    # 获取查看人名称
    viewer_names = []
    for viewer_id in viewer_ids:
        user = db.query(User).filter(User.id == viewer_id).first()
        if user:
            viewer_names.append(user.username)
    
    # 获取审核人名称
    reviewer_name = ''
    if db_testplan.reviewer_id:
        reviewer_user = db.query(User).filter(User.id == db_testplan.reviewer_id).first()
        if reviewer_user:
            reviewer_name = reviewer_user.username
    
    # 触发分配通知（执行人 + 审核人 + 查看人）
    notify_ids = list(executor_ids) if executor_ids else []
    if db_testplan.reviewer_id and db_testplan.reviewer_id not in notify_ids:
        notify_ids.append(db_testplan.reviewer_id)
    for viewer_id in viewer_ids:
        if viewer_id not in notify_ids:
            notify_ids.append(viewer_id)
    if notify_ids:
        # 使用测试计划已关联的项目组ID
        testplan_team_id = db_testplan.team_id
        
        trigger_assignment_notification(
            db=db,
            notification_type='testplan',
            event_type='created',
            title=f'您被关联到测试计划「{db_testplan.name}」',
            content=f'测试计划「{db_testplan.name}」已创建，您被关联为相关人员。\n\n'
                    f'计划名称：{db_testplan.name}\n'
                    f'分配人：{current_user.username}\n'
                    f'开始时间：{_format_datetime_str(str(testplan_data.get("start_time", ""))) or "未设置"}\n'
                    f'结束时间：{_format_datetime_str(str(testplan_data.get("end_time", ""))) or "未设置"}\n'
                    f'执行人：{"、".join(executor_names) or "未设置"}\n'
                    f'查看人：{"、".join(viewer_names) or "未设置"}\n'
                    f'审核人：{reviewer_name or "未设置"}',
            related_id=db_testplan.id,
            related_type='testplan',
            sender_id=current_user.id,
            recipient_user_ids=notify_ids,
            project_id=db_testplan.project_id,
            team_id=testplan_team_id,
            extra_context={
                'plan_name': db_testplan.name,
                'start_time': _format_datetime_str(str(testplan_data.get("start_time", ""))) or "未设置",
                'end_time': _format_datetime_str(str(testplan_data.get("end_time", ""))) or "未设置",
                'executors': "、".join(executor_names) or "未设置",
                'viewers': "、".join(viewer_names) or "未设置",
                'reviewer': reviewer_name or "未设置",
                'operator': current_user.username,
                'action': '创建',
            }
        )
    
    # 构建包含 test_case_ids、executor_ids 和 viewer_ids 的响应
    result = {
        "id": db_testplan.id,
        "project_id": db_testplan.project_id,
        "name": db_testplan.name,
        "description": db_testplan.description,
        "start_time": db_testplan.start_time,
        "end_time": db_testplan.end_time,
        "status": db_testplan.status,
        "created_at": db_testplan.created_at,
        "test_case_ids": test_case_ids,
        "executor_ids": executor_ids,
        "viewer_ids": viewer_ids
    }
    
    return {"code": 200, "message": "success", "data": result}

@router.get("/{testplan_id}")
def get_testplan(
    req: Request,
    testplan_id: int,
    page: int = 1,
    size: int = 50,
    keyword: str = None,
    status: str = None,
    all_cases: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from typing import Optional
    
    testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    # 获取用例详情 - 使用子查询一次查询获取所有存在的用例
    if testplan_id:
        linked_ids_subq = db.query(TestPlanTestCase.test_case_id).filter(
            TestPlanTestCase.test_plan_id == testplan_id,
            TestPlanTestCase.test_case_id.isnot(None)
        ).subquery()
        
        # 应用用例排序
        from api.testcase import apply_testcase_sort
        query = db.query(TestCase).filter(TestCase.id.in_(linked_ids_subq))
        query = apply_testcase_sort(query, testplan.project_id, db)
        all_test_cases_query = query.all()
        
        test_case_ids = [tc.id for tc in all_test_cases_query]
    else:
        test_case_ids = []
        all_test_cases_query = []
    
    # all_cases=True 时返回所有用例（不分页），用于执行页面
    if all_cases:
        size = len(test_case_ids) if test_case_ids else 50
    
    # 统计总用例数
    total_testcases = len(test_case_ids)
    
    # 从关联表中获取执行人ID
    test_plan_executors = db.query(TestPlanExecutor).filter(TestPlanExecutor.test_plan_id == testplan_id).all()
    executor_ids = [te.executor_id for te in test_plan_executors]
    
    # 如果关联表中没有执行人数据，尝试从关联用例的执行记录中获取
    if not executor_ids and test_case_ids:
        executions = db.query(TestExecution).filter(
            TestExecution.test_plan_id == testplan_id,
            TestExecution.test_case_id.in_(test_case_ids)
        ).all()
        if executions:
            executor_ids = list(set([e.executor_id for e in executions if e.executor_id]))
    
    # 从关联表中获取查看人ID
    test_plan_viewers = db.query(TestPlanViewer).filter(TestPlanViewer.test_plan_id == testplan_id).all()
    viewer_ids = [tv.viewer_id for tv in test_plan_viewers]
    
    # 获取用例详情（应用标准排序）- 分页
    if test_case_ids:
        from api.testcase import apply_testcase_sort
        query = db.query(TestCase).filter(TestCase.id.in_(test_case_ids))
        query = apply_testcase_sort(query, testplan.project_id, db)
        test_cases = query.all()
    else:
        test_cases = []
    
    # 构建用例ID到用例的映射（用于后续处理）
    test_case_map = {tc.id: tc for tc in test_cases}
    
    # 复用 all_test_cases_query 用于统计和项目ID收集
    all_test_cases_map = {tc.id: tc for tc in all_test_cases_query}
    
    # 收集用例的项目ID（使用全部用例）
    test_case_project_ids = set()
    for tc in all_test_cases_query:
        if tc.primary_project_id:
            test_case_project_ids.add(tc.primary_project_id)
    
    # 获取执行人详情
    executors = db.query(User).filter(User.id.in_(executor_ids)).all() if executor_ids else []
    
    # 获取查看人详情
    viewers = db.query(User).filter(User.id.in_(viewer_ids)).all() if viewer_ids else []
    
    # 获取创建人信息
    creator = db.query(User).filter(User.id == testplan.created_by).first() if testplan.created_by else None
    
    # 获取审核人信息
    reviewer = db.query(User).filter(User.id == testplan.reviewer_id).first() if testplan.reviewer_id else None
    
    # 统计执行数据（只统计关联用例的执行记录）
    if test_case_ids:
        executions = db.query(TestExecution).filter(
            TestExecution.test_plan_id == testplan_id,
            TestExecution.test_case_id.in_(test_case_ids)
        ).all()
    else:
        executions = []
    
    # 构建用例ID到最新执行记录的映射（一次遍历）
    latest_execution_map = {}
    for e in executions:
        if e.test_case_id not in latest_execution_map:
            latest_execution_map[e.test_case_id] = e
        else:
            if e.executed_at > latest_execution_map[e.test_case_id].executed_at:
                latest_execution_map[e.test_case_id] = e
    
    # 构建执行人ID到用户名的映射
    executor_user_ids = set(e.executor_id for e in latest_execution_map.values() if e.executor_id)
    executor_user_map = {}
    if executor_user_ids:
        executor_users = db.query(User).filter(User.id.in_(executor_user_ids)).all()
        executor_user_map = {u.id: u.username for u in executor_users}
    
    # 保存原始关联用例ID列表（用于计算模块统计，不受分页/筛选影响）
    all_case_ids = test_case_ids.copy()
    
    # 为每个用例添加执行状态
    case_with_status = []
    test_case_ids_set = set(test_case_ids)  # 用于快速查找
    
    # 分页后的用例ID列表
    paginated_case_ids = set(tc.id for tc in test_cases)
    
    for tc in test_cases:
        execution = latest_execution_map.get(tc.id)
        exec_status = execution.result if execution else 'PENDING'
        exec_remark = execution.remarks if execution else ''
        exec_executor = executor_user_map.get(execution.executor_id, '') if execution and execution.executor_id else ''
        exec_time = execution.executed_at if execution else None
        case_with_status.append({
            "id": tc.id,
            "case_number": tc.case_number,
            "name": tc.name,
            "module": tc.module,
            "sub_module": tc.sub_module,
            "level": tc.level,
            "priority": tc.priority,
            "case_type": tc.case_type,
            "precondition": tc.precondition,
            "steps": tc.steps,
            "expected_result": tc.expected_result,
            "remarks": tc.remarks,
            "automation": tc.automation,
            "execution_status": exec_status,
            "execution_remark": exec_remark,
            "execution_executor": exec_executor,
            "executed_at": exec_time
        })
    
    # 构建原始ID到用例的映射（用于模块统计）- 使用全部用例
    case_with_all_status = []
    for tc_id in all_case_ids:
        case = all_test_cases_map.get(tc_id)
        execution = latest_execution_map.get(tc_id)
        exec_status = execution.result if execution else 'PENDING'
        exec_remark = execution.remarks if execution else ''
        exec_executor = executor_user_map.get(execution.executor_id, '') if execution and execution.executor_id else ''
        exec_time = execution.executed_at if execution else None
        if case:
            case_with_all_status.append({
                "id": case.id,
                "case_number": case.case_number,
                "name": case.name,
                "module": case.module,
                "sub_module": case.sub_module,
                "level": case.level,
                "priority": case.priority,
                "case_type": case.case_type,
                "precondition": case.precondition,
                "steps": case.steps,
                "expected_result": case.expected_result,
                "remarks": case.remarks,
                "automation": case.automation,
                "execution_status": exec_status,
                "execution_remark": exec_remark,
                "execution_executor": exec_executor,
                "executed_at": exec_time
            })
    
    # 计算模块统计（基于全部用例，不受分页/筛选影响）
    module_statsMap = {}
    for tc in case_with_all_status:
        rawModule = tc.get('module') or '未分类'
        mainModule = rawModule.split('/')[0].strip() or '未分类'
        if mainModule not in module_statsMap:
            module_statsMap[mainModule] = {'total': 0, 'executed': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'nt': 0, 'assist': 0}
        m = module_statsMap[mainModule]
        m['total'] += 1
        exec_status = tc.get('execution_status')
        if exec_status and exec_status != 'PENDING':
            m['executed'] += 1
            if exec_status == 'PASS':
                m['passed'] += 1
            elif exec_status == 'FAIL':
                m['failed'] += 1
            elif exec_status == 'BLOCK':
                m['blocked'] += 1
            elif exec_status == 'NA':
                m['na'] += 1
            elif exec_status == 'NT':
                m['nt'] += 1
            elif (exec_status or '').strip() == '协测':
                m['assist'] += 1
    
    # 获取模块的RD Owner - 从测试用例关联的项目中查询
    module_rd_owner_map = {}
    project_ids_for_modules = test_case_project_ids
    if project_ids_for_modules:
        from models import Module
        modules = db.query(Module).filter(
            Module.project_id.in_(project_ids_for_modules),
            Module.parent_id == None
        ).all()
        
        for m in modules:
            if m.rd_owner:
                module_rd_owner_map[m.name] = m.rd_owner
    
    module_stats = []
    for name, m in module_statsMap.items():
        valid = m['total'] - m['na'] - m.get('assist', 0)
        pass_rate = f"{(m['passed'] / valid * 100):.2f}" if valid > 0 else "0.00"
        executed_pct = f"{(m['executed'] / m['total'] * 100):.2f}" if m['total'] > 0 else "0.00"
        module_stats.append({
            "name": name,
            "total": m['total'],
            "executed": m['executed'],
            "passed": m['passed'],
            "failed": m['failed'],
            "blocked": m['blocked'],
            "na": m['na'],
            "nt": m['nt'],
            "assist": m.get('assist', 0),
            "pass_rate": pass_rate,
            "executed_pct": executed_pct,
            "rd_owner": module_rd_owner_map.get(name, "")
        })
    module_stats.sort(key=lambda x: x['total'], reverse=True)
    
    # 填充 RD Owner
    for stat in module_stats:
        stat['rd_owner'] = module_rd_owner_map.get(stat['name'], '')
    
    # 应用筛选条件（基于全部用例，不受分页影响）
    filtered_cases = case_with_all_status
    if keyword:
        keyword_lower = keyword.lower()
        filtered_cases = [tc for tc in filtered_cases if 
            keyword_lower in tc.get('name', '').lower() or 
            keyword_lower in tc.get('case_number', '').lower()]
    if status in ['UNEXECUTED', 'PENDING']:
        filtered_cases = [tc for tc in filtered_cases if tc.get('execution_status') in ['PENDING', 'ONGOING']]
    elif status == 'EXECUTED':
        filtered_cases = [tc for tc in filtered_cases if tc.get('execution_status') not in ['PENDING', 'ONGOING']]
    elif status and status != 'ALL':
        filtered_cases = [tc for tc in filtered_cases if tc.get('execution_status') == status]
    
    filtered_total = len(filtered_cases)
    
    # 分页处理（基于筛选后的用例）
    start = (page - 1) * size
    end = start + size
    paginated_cases = filtered_cases[start:end]
    
    # 统计各种执行结果（基于全部用例，不受筛选影响）
    stat_executed = 0
    stat_passed = 0
    stat_failed = 0
    stat_blocked = 0
    stat_na = 0
    stat_nt = 0
    stat_assist = 0
    for tc in case_with_all_status:
        exec_status = tc.get('execution_status')
        if exec_status and exec_status != 'PENDING' and exec_status != 'ONGOING':
            stat_executed += 1
            if exec_status == 'PASS':
                stat_passed += 1
            elif exec_status == 'FAIL':
                stat_failed += 1
            elif exec_status == 'BLOCK':
                stat_blocked += 1
            elif exec_status == 'NA':
                stat_na += 1
            elif exec_status == 'NT':
                stat_nt += 1
            elif (exec_status or '').strip() == '协测':
                stat_assist += 1
    
    executed_pct = f"{(stat_executed / len(case_with_all_status) * 100):.2f}" if case_with_all_status else "0.00"
    valid_cases = len(case_with_all_status) - stat_na - stat_assist
    pass_rate = f"{(stat_passed / valid_cases * 100):.2f}" if valid_cases > 0 else "0.00"
    
    statistics = {
        "executed": stat_executed,
        "passed": stat_passed,
        "failed": stat_failed,
        "blocked": stat_blocked,
        "na": stat_na,
        "nt": stat_nt,
        "assist": stat_assist
    }
    
    # 使用统计数据赋值（基于全部用例）
    executed_testcases = stat_executed
    passed_count = stat_passed
    failed_count = stat_failed
    blocked_count = stat_blocked
    na_count = stat_na
    nt_count = stat_nt
    
    # 构建返回数据
    result = {
        "id": testplan.id,
        "project_id": testplan.project_id,
        "team_id": testplan.team_id,
        "test_case_project_ids": list(test_case_project_ids),
        "name": testplan.name,
        "description": testplan.description,
        "start_time": testplan.start_time,
        "end_time": testplan.end_time,
        "status": testplan.status,
        "reviewer_id": testplan.reviewer_id,
        "reviewer_name": reviewer.username if reviewer else None,
        "reviewer_info": {"id": reviewer.id, "username": reviewer.username, "full_name": reviewer.full_name} if reviewer else None,
        "created_at": testplan.created_at,
        "updated_at": testplan.updated_at,
        "created_by": testplan.created_by,
        "creator_name": creator.username if creator else None,
        "suite_id": testplan.suite_id,
        "total_testcases": total_testcases,
        "executed_testcases": executed_testcases,
        "passed_testcases": passed_count,
        "failed_testcases": failed_count,
        "blocked_testcases": blocked_count,
        "na_testcases": na_count,
        "nt_testcases": nt_count,
        "module_stats": module_stats,
        "statistics": statistics,
        "executed_pct": executed_pct,
        "pass_rate": pass_rate,
        "test_cases": paginated_cases,
        "all_test_cases": case_with_all_status,
        "all_statistics": statistics,
        "pagination": {
            "page": page,
            "size": size,
            "total": filtered_total if (keyword or (status and status != 'ALL')) else total_testcases
        },
        "executors": [{"id": u.id, "username": u.username, "full_name": u.full_name} for u in executors],
        "executor_names": [u.username for u in executors],
        "executor_ids": executor_ids,
        "viewers": [{"id": u.id, "username": u.username, "full_name": u.full_name} for u in viewers],
        "viewer_names": [u.username for u in viewers],
        "viewer_ids": viewer_ids,
        "test_case_ids": test_case_ids
    }
    
    # 如果测试计划状态为REJECTED，获取最新报告的reject_reason
    if testplan.status == 'REJECTED':
        latest_report = db.query(Report).filter(
            Report.test_plan_id == testplan_id,
            Report.status == 'REJECTED'
        ).order_by(Report.reviewed_at.desc()).first()
        if latest_report:
            result["reject_reason"] = latest_report.reject_reason or ''
            # 获取审核人名称
            reject_reviewer = db.query(User).filter(User.id == latest_report.reviewed_by).first()
            result["rejected_by"] = reject_reviewer.username if reject_reviewer else ''
            result["rejected_at"] = latest_report.reviewed_at.isoformat() if latest_report.reviewed_at else ''
        else:
            result["reject_reason"] = ''
            result["rejected_by"] = ''
            result["rejected_at"] = ''
    
    return {"code": 200, "message": "success", "data": result}

@router.put("/{testplan_id}")
async def update_testplan(
    req: Request,
    testplan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    db_testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not db_testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    # 已完成的计划不允许编辑
    if db_testplan.status == 'COMPLETED':
        raise HTTPException(status_code=403, detail="已完成的测试计划不允许编辑")
    
    # 权限检查：超级管理员、系统管理员、项目组负责人、创建人可以编辑
    from utils.data_permission import is_super_admin
    # 检查是否为系统管理员角色（通过 UserRole + Role 表查询）
    is_role_admin = False
    from models import UserRole, Role
    user_roles = db.query(Role.name).join(UserRole, UserRole.role_id == Role.id).filter(
        UserRole.user_id == current_user.id
    ).all()
    role_names = [r.name for r in user_roles]
    if '系统管理员' in role_names or 'admin' in role_names:
        is_role_admin = True
    
    is_admin_user = is_super_admin(current_user) or is_role_admin
    is_creator = db_testplan.created_by == current_user.id
    
    # 检查是否为项目组负责人
    is_leader = False
    if db_testplan.team_id:
        from models import TeamLeader, Team, DepartmentManager
        # 检查是否是项目组负责人
        team_leader = db.query(TeamLeader).filter(
            TeamLeader.team_id == db_testplan.team_id,
            TeamLeader.user_id == current_user.id
        ).first()
        if team_leader:
            is_leader = True
        else:
            # 检查是否是组织负责人
            team = db.query(Team).filter(Team.id == db_testplan.team_id).first()
            if team and team.department_id:
                org_manager = db.query(DepartmentManager).filter(
                    DepartmentManager.department_id == team.department_id,
                    DepartmentManager.user_id == current_user.id
                ).first()
                if org_manager:
                    is_leader = True
    
    if not (is_admin_user or is_creator or is_leader):
        raise HTTPException(status_code=403, detail="只有项目组负责人或创建人可以编辑测试计划")
    
    # 从请求体中获取数据
    body = await req.json()
    
    # 提取测试用例、执行人、查看人和审核人信息
    test_case_ids = body.get('test_case_ids', [])
    executor_ids = body.get('executor_ids', [])
    viewer_ids = body.get('viewer_ids', [])
    reviewer_id = body.get('reviewer_id')
    suite_id = body.get('suite_id')
    
    # 校验：查看人和执行人不能重叠
    if viewer_ids and executor_ids:
        overlap = set(viewer_ids) & set(executor_ids)
        if overlap:
            overlap_users = db.query(User).filter(User.id.in_(overlap)).all()
            overlap_names = [u.username for u in overlap_users]
            raise HTTPException(
                status_code=400,
                detail=f"用户 {', '.join(overlap_names)} 不能同时设为查看人和执行人"
            )
    
    # 如果选择了套件且没有手动选用例，从套件加载用例
    if suite_id and not test_case_ids:
        from models import TestSuiteTestCase
        links = db.query(TestSuiteTestCase).filter(TestSuiteTestCase.test_suite_id == suite_id).all()
        test_case_ids = [l.test_case_id for l in links]
    
    # ========== 保存更新前的旧值，用于对比变更 ==========
    old_name = db_testplan.name
    old_description = db_testplan.description
    old_start_time = str(db_testplan.start_time) if db_testplan.start_time else ''
    old_end_time = str(db_testplan.end_time) if db_testplan.end_time else ''
    old_reviewer_id = db_testplan.reviewer_id
    
    # 获取旧执行人ID列表
    old_executor_records = db.query(TestPlanExecutor).filter(TestPlanExecutor.test_plan_id == testplan_id).all()
    old_executor_ids = [r.executor_id for r in old_executor_records]
    
    # 获取旧查看人ID列表
    old_viewer_records = db.query(TestPlanViewer).filter(TestPlanViewer.test_plan_id == testplan_id).all()
    old_viewer_ids = [r.viewer_id for r in old_viewer_records]
    
    # 获取旧关联用例数量
    old_testcase_count = db.query(TestPlanTestCase).filter(TestPlanTestCase.test_plan_id == testplan_id).count()
    
    # 更新测试计划数据
    testplan_data = {
        'name': body.get('name'),
        'description': body.get('description'),
        'start_time': body.get('start_time'),
        'end_time': body.get('end_time'),
        'project_id': body.get('project_id'),
        'reviewer_id': reviewer_id,
        'suite_id': suite_id
    }
    
    # 转换日期字符串为 datetime 对象
    if testplan_data.get('start_time'):
        try:
            testplan_data['start_time'] = datetime.strptime(testplan_data['start_time'], '%Y-%m-%d')
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="开始时间格式错误，应为 YYYY-MM-DD")
    
    if testplan_data.get('end_time'):
        try:
            testplan_data['end_time'] = datetime.strptime(testplan_data['end_time'], '%Y-%m-%d')
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="结束时间格式错误，应为 YYYY-MM-DD")
    
    for key, value in testplan_data.items():
        if value is not None:
            setattr(db_testplan, key, value)
    
    # suite_id 允许设为 None（用户清除套件选择）
    if 'suite_id' in body:
        db_testplan.suite_id = suite_id
    
    # 只有在明确提供了 test_case_ids 时才更新用例关联
    if 'test_case_ids' in body:
        # 过滤掉不存在的用例ID（使用子查询一次完成）
        if test_case_ids:
            existing_ids_subq = db.query(TestCase.id).filter(
                TestCase.id.in_(test_case_ids)
            ).subquery()
            test_case_ids = [cid[0] for cid in db.query(TestCase.id).filter(TestCase.id.in_(existing_ids_subq)).all()]
        
        # 删除现有的测试计划与测试用例的关联关系
        db.query(TestPlanTestCase).filter(TestPlanTestCase.test_plan_id == testplan_id).delete()
        
        # 存储新的测试计划与测试用例的关联关系
        if test_case_ids:
            for test_case_id in test_case_ids:
                test_plan_test_case = TestPlanTestCase(
                    test_plan_id=testplan_id,
                    test_case_id=test_case_id
                )
                db.add(test_plan_test_case)
    
    # 只有在明确提供了 executor_ids 时才更新执行人关联
    if 'executor_ids' in body:
        # 删除现有的测试计划与执行人的关联关系
        db.query(TestPlanExecutor).filter(TestPlanExecutor.test_plan_id == testplan_id).delete()
        
        # 存储新的测试计划与执行人的关联关系
        if executor_ids:
            for executor_id in executor_ids:
                test_plan_executor = TestPlanExecutor(
                    test_plan_id=testplan_id,
                    executor_id=executor_id
                )
                db.add(test_plan_executor)
    
    # 只有在明确提供了 viewer_ids 时才更新查看人关联
    if 'viewer_ids' in body:
        # 删除现有的测试计划与查看人的关联关系
        db.query(TestPlanViewer).filter(TestPlanViewer.test_plan_id == testplan_id).delete()
        
        # 存储新的测试计划与查看人的关联关系
        if viewer_ids:
            for viewer_id in viewer_ids:
                test_plan_viewer = TestPlanViewer(
                    test_plan_id=testplan_id,
                    viewer_id=viewer_id
                )
                db.add(test_plan_viewer)
    
    db.commit()
    db.refresh(db_testplan)
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTPLANS,
        action=LogAction.UPDATE,
        description=f"更新测试计划：{db_testplan.name}（ID: {db_testplan.id}），关联 {len(test_case_ids)} 个用例，{len(executor_ids)} 个执行人，{len(viewer_ids)} 个查看人",
        request=req
    )
    
    # ========== 获取更新后的人员名称 ==========
    # 执行人名称
    executor_names = []
    for executor_id in executor_ids:
        user = db.query(User).filter(User.id == executor_id).first()
        if user:
            executor_names.append(user.username)
    
    # 查看人名称
    viewer_names = []
    for viewer_id in viewer_ids:
        user = db.query(User).filter(User.id == viewer_id).first()
        if user:
            viewer_names.append(user.username)
    
    # 审核人名称
    reviewer_name = ''
    if db_testplan.reviewer_id:
        reviewer_user = db.query(User).filter(User.id == db_testplan.reviewer_id).first()
        if reviewer_user:
            reviewer_name = reviewer_user.username
    
    # ========== 获取旧人员名称（用于变更对比） ==========
    old_executor_names = []
    if old_executor_ids:
        old_executors = db.query(User).filter(User.id.in_(old_executor_ids)).all()
        old_executor_names = [u.username for u in old_executors]
    
    old_viewer_names = []
    if old_viewer_ids:
        old_viewers = db.query(User).filter(User.id.in_(old_viewer_ids)).all()
        old_viewer_names = [u.username for u in old_viewers]
    
    old_reviewer_name = ''
    if old_reviewer_id:
        old_reviewer_user = db.query(User).filter(User.id == old_reviewer_id).first()
        if old_reviewer_user:
            old_reviewer_name = old_reviewer_user.username
    
    # ========== 对比变更，生成变更详情 ==========
    changes = []
    
    # 计划名称
    new_name = db_testplan.name
    if old_name != new_name:
        changes.append(f'计划名称：{old_name} → {new_name}')
    
    # 开始时间
    new_start_time = _format_datetime_str(str(db_testplan.start_time)) if db_testplan.start_time else ''
    old_start_fmt = _format_datetime_str(old_start_time)
    if old_start_fmt != new_start_time:
        changes.append(f'开始时间：{old_start_fmt or "未设置"} → {new_start_time or "未设置"}')
    
    # 结束时间
    new_end_time = _format_datetime_str(str(db_testplan.end_time)) if db_testplan.end_time else ''
    old_end_fmt = _format_datetime_str(old_end_time)
    if old_end_fmt != new_end_time:
        changes.append(f'结束时间：{old_end_fmt or "未设置"} → {new_end_time or "未设置"}')
    
    # 审核人
    if old_reviewer_name != reviewer_name:
        changes.append(f'审核人：{old_reviewer_name or "未设置"} → {reviewer_name or "未设置"}')
    
    # 执行人
    old_exec_set = set(old_executor_names)
    new_exec_set = set(executor_names)
    if old_exec_set != new_exec_set:
        added = new_exec_set - old_exec_set
        removed = old_exec_set - new_exec_set
        parts = []
        if added:
            parts.append(f'新增：{"、".join(added)}')
        if removed:
            parts.append(f'移除：{"、".join(removed)}')
        changes.append(f'执行人变更（{"、".join(parts)}）')
    
    # 查看人
    old_viewer_set = set(old_viewer_names)
    new_viewer_set = set(viewer_names)
    if old_viewer_set != new_viewer_set:
        added = new_viewer_set - old_viewer_set
        removed = old_viewer_set - new_viewer_set
        parts = []
        if added:
            parts.append(f'新增：{"、".join(added)}')
        if removed:
            parts.append(f'移除：{"、".join(removed)}')
        changes.append(f'查看人变更（{"、".join(parts)}）')
    
    # 关联用例数量
    new_testcase_count = len(test_case_ids) if 'test_case_ids' in body else old_testcase_count
    if 'test_case_ids' in body and old_testcase_count != new_testcase_count:
        changes.append(f'关联用例数：{old_testcase_count} → {new_testcase_count}')
    
    # 描述
    if 'description' in body and body.get('description') is not None:
        old_desc_brief = (old_description or '').strip()[:50]
        new_desc_brief = (db_testplan.description or '').strip()[:50]
        if old_desc_brief != new_desc_brief:
            changes.append('描述已更新')
    
    # 构建变更详情文本
    changes_text = '\n'.join([f'  - {c}' for c in changes]) if changes else '  无字段变更（仅人员关联调整）'
    
    # 触发分配通知（执行人 + 审核人 + 查看人）
    notify_ids = list(executor_ids) if executor_ids else []
    if db_testplan.reviewer_id and db_testplan.reviewer_id not in notify_ids:
        notify_ids.append(db_testplan.reviewer_id)
    for viewer_id in viewer_ids:
        if viewer_id not in notify_ids:
            notify_ids.append(viewer_id)
    if notify_ids:
        trigger_assignment_notification(
            db=db,
            notification_type='testplan',
            event_type='updated',
            title=f'测试计划「{db_testplan.name}」已更新',
            content=f'测试计划「{db_testplan.name}」已更新，您是相关人员。\n\n'
                    f'变更详情：\n{changes_text}\n\n'
                    f'当前信息：\n'
                    f'  - 计划名称：{db_testplan.name}\n'
                    f'  - 操作人：{current_user.username}\n'
                    f'  - 开始时间：{new_start_time or "未设置"}\n'
                    f'  - 结束时间：{new_end_time or "未设置"}\n'
                    f'  - 执行人：{"、".join(executor_names) or "未设置"}\n'
                    f'  - 查看人：{"、".join(viewer_names) or "未设置"}\n'
                    f'  - 审核人：{reviewer_name or "未设置"}',
            related_id=db_testplan.id,
            related_type='testplan',
            sender_id=current_user.id,
            recipient_user_ids=notify_ids,
            project_id=db_testplan.project_id,
            team_id=db_testplan.team_id,
            extra_context={
                'plan_name': db_testplan.name,
                'start_time': new_start_time or "未设置",
                'end_time': new_end_time or "未设置",
                'executors': "、".join(executor_names) or "未设置",
                'viewers': "、".join(viewer_names) or "未设置",
                'reviewer': reviewer_name or "未设置",
                'operator': current_user.username,
                'action': '更新',
                'changes': changes_text,
                'changes_list': changes,
            }
        )
    
    # 构建包含 test_case_ids、executor_ids 和 viewer_ids 的响应
    result = {
        "id": db_testplan.id,
        "project_id": db_testplan.project_id,
        "name": db_testplan.name,
        "description": db_testplan.description,
        "start_time": db_testplan.start_time,
        "end_time": db_testplan.end_time,
        "status": db_testplan.status,
        "created_at": db_testplan.created_at,
        "test_case_ids": test_case_ids,
        "executor_ids": executor_ids,
        "viewer_ids": viewer_ids
    }
    
    return {"code": 200, "message": "success", "data": result}

@router.post("/{testplan_id}/start")
async def start_testplan(
    req: Request,
    testplan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    db_testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not db_testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    # 从关联表中获取测试用例ID（使用子查询一次完成）
    linked_ids_subq = db.query(TestPlanTestCase.test_case_id).filter(
        TestPlanTestCase.test_plan_id == testplan_id,
        TestPlanTestCase.test_case_id.isnot(None)
    ).subquery()
    
    test_case_ids = [tc.id for tc in db.query(TestCase.id).filter(TestCase.id.in_(linked_ids_subq)).all()]
    
    # 从关联表中获取执行人ID
    test_plan_executors = db.query(TestPlanExecutor).filter(TestPlanExecutor.test_plan_id == testplan_id).all()
    executor_ids = [te.executor_id for te in test_plan_executors]
    
    if not test_case_ids or not executor_ids:
        raise HTTPException(status_code=400, detail="测试计划需要关联测试用例和执行人才可以开始执行")
    
    # Create test executions for each combination of test case and executor
    for test_case_id in test_case_ids:
        for executor_id in executor_ids:
            execution = TestExecution(
                test_plan_id=testplan_id,
                test_case_id=test_case_id,
                executor_id=executor_id,
                result="PENDING"
            )
            db.add(execution)
    
    # Update test plan status
    db_testplan.status = "ACTIVE"
    db.commit()
    db.refresh(db_testplan)
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTPLANS,
        action=LogAction.UPDATE,
        description=f"开始执行测试计划：{db_testplan.name}（ID: {testplan_id}），关联 {len(test_case_ids)} 个用例，{len(executor_ids)} 个执行人",
        request=req
    )
    
    return {"code": 200, "message": "success", "data": db_testplan}

@router.delete("/{testplan_id}")
def delete_testplan(
    req: Request,
    testplan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from utils.data_permission import is_super_admin
    from models import UserRole, Role
    
    db_testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not db_testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    # 检查是否为管理员（超级管理员或系统管理员角色）
    is_role_admin = False
    user_roles = db.query(Role.name).join(UserRole, UserRole.role_id == Role.id).filter(
        UserRole.user_id == current_user.id
    ).all()
    role_names = [r.name for r in user_roles]
    if '系统管理员' in role_names or 'admin' in role_names:
        is_role_admin = True
    
    is_admin_user = is_super_admin(current_user) or is_role_admin
    
    # 非管理员只能删除未开始的计划
    if not is_admin_user and db_testplan.status != 'PENDING':
        raise HTTPException(status_code=403, detail="只有管理员可以删除非未开始状态的计划")
    
    plan_name = db_testplan.name
    
    # Delete related reports
    db.query(Report).filter(Report.test_plan_id == testplan_id).delete()
    
    # Delete related zmind links
    from models import TestCaseZmindLink
    db.query(TestCaseZmindLink).filter(TestCaseZmindLink.test_plan_id == testplan_id).delete()
    
    # Delete related executions
    db.query(TestExecution).filter(TestExecution.test_plan_id == testplan_id).delete()
    
    # Delete related test plan test cases
    db.query(TestPlanTestCase).filter(TestPlanTestCase.test_plan_id == testplan_id).delete()
    
    # Delete related test plan executors
    db.query(TestPlanExecutor).filter(TestPlanExecutor.test_plan_id == testplan_id).delete()
    
    # Delete test plan
    db.delete(db_testplan)
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTPLANS,
        action=LogAction.DELETE,
        description=f"删除测试计划：{plan_name}（ID: {testplan_id}）",
        request=req
    )
    
    return {"code": 200, "message": "success"}


@router.post("/{testplan_id}/testcases")
async def add_testcases_to_testplan(
    req: Request,
    testplan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """添加用例到测试计划"""
    
    # 验证测试计划是否存在
    testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    # 从请求体中获取数据
    body = await req.json()
    testcase_ids = body.get('testcase_ids', [])
    
    if not testcase_ids:
        raise HTTPException(status_code=400, detail="请提供要添加的用例ID列表")
    
    added_count = 0
    skipped_count = 0
    errors = []
    
    for testcase_id in testcase_ids:
        # 验证用例是否存在
        testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
        if not testcase:
            errors.append(f"用例 ID {testcase_id} 不存在")
            continue
        
        # 检查是否已添加
        existing = db.query(TestPlanTestCase).filter(
            TestPlanTestCase.test_plan_id == testplan_id,
            TestPlanTestCase.test_case_id == testcase_id
        ).first()
        
        if existing:
            skipped_count += 1
            errors.append(f"用例 {testcase.case_number} 已在测试计划中")
            continue
        
        # 添加用例到计划
        test_plan_testcase = TestPlanTestCase(
            test_plan_id=testplan_id,
            test_case_id=testcase_id
        )
        db.add(test_plan_testcase)
        added_count += 1
    
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTPLANS,
        action=LogAction.UPDATE,
        description=f"向测试计划 {testplan.name} 添加 {added_count} 个用例",
        request=req
    )
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "added": added_count,
            "skipped": skipped_count,
            "errors": errors
        }
    }


@router.delete("/{testplan_id}/testcases/{testcase_id}")
def remove_testcase_from_testplan(
    req: Request,
    testplan_id: int,
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """从测试计划中移除用例"""
    
    # 验证测试计划是否存在
    testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    # 查找关联记录
    test_plan_testcase = db.query(TestPlanTestCase).filter(
        TestPlanTestCase.test_plan_id == testplan_id,
        TestPlanTestCase.test_case_id == testcase_id
    ).first()
    
    if not test_plan_testcase:
        raise HTTPException(status_code=404, detail="用例不在测试计划中")
    
    # 获取用例信息用于日志
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    testcase_number = testcase.case_number if testcase else str(testcase_id)
    
    # 删除关联记录
    db.delete(test_plan_testcase)
    
    # 同时删除该用例在此测试计划中的执行记录（如果有）
    db.query(TestExecution).filter(
        TestExecution.test_plan_id == testplan_id,
        TestExecution.test_case_id == testcase_id
    ).delete()
    
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTPLANS,
        action=LogAction.UPDATE,
        description=f"从测试计划 {testplan.name} 移除用例 {testcase_number}",
        request=req
    )
    
    return {"code": 200, "message": "success", "data": None}


@router.post("/{testplan_id}/testcases/batch-remove")
async def batch_remove_testcases_from_testplan(
    req: Request,
    testplan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """批量从测试计划中移除用例（单次事务）
    
    请求体: { "testcase_ids": [1, 2, 3] }
    - 只允许移除未执行 / NT / ONGOING 状态的用例（已执行的会被自动跳过）
    - 采用单次 SQL 批量删除，避免逐条请求带来的性能和网络问题
    """
    # 验证测试计划是否存在
    testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")

    body = await req.json()
    testcase_ids = body.get('testcase_ids', [])
    if not isinstance(testcase_ids, list) or not testcase_ids:
        raise HTTPException(status_code=400, detail="请提供要取消关联的用例ID列表")

    # 只保留当前计划中真实存在关联的用例
    linked_rows = db.query(TestPlanTestCase).filter(
        TestPlanTestCase.test_plan_id == testplan_id,
        TestPlanTestCase.test_case_id.in_(testcase_ids)
    ).all()
    linked_ids = {row.test_case_id for row in linked_rows}
    if not linked_ids:
        return {
            "code": 200,
            "message": "success",
            "data": {"removed": 0, "skipped": len(testcase_ids), "skipped_executed": 0}
        }

    # 查询这些用例在本计划中的最新执行记录，过滤掉已执行的
    executed_ids = set()
    executions = db.query(TestExecution).filter(
        TestExecution.test_plan_id == testplan_id,
        TestExecution.test_case_id.in_(linked_ids)
    ).all()
    # 取每个用例的最新执行结果
    latest_map = {}
    for e in executions:
        old = latest_map.get(e.test_case_id)
        if old is None:
            latest_map[e.test_case_id] = e
        elif e.executed_at and (not old.executed_at or e.executed_at > old.executed_at):
            latest_map[e.test_case_id] = e
    for tc_id, e in latest_map.items():
        # PASS / FAIL / BLOCK / NA 视为已执行，不允许移除
        if e.result in ('PASS', 'FAIL', 'BLOCK', 'NA'):
            executed_ids.add(tc_id)

    removable_ids = linked_ids - executed_ids
    if not removable_ids:
        return {
            "code": 200,
            "message": "success",
            "data": {
                "removed": 0,
                "skipped": len(testcase_ids) - len(linked_ids),
                "skipped_executed": len(executed_ids)
            }
        }

    # 批量删除关联记录
    removed = db.query(TestPlanTestCase).filter(
        TestPlanTestCase.test_plan_id == testplan_id,
        TestPlanTestCase.test_case_id.in_(removable_ids)
    ).delete(synchronize_session=False)

    # 同时删除执行记录（保持数据一致）
    db.query(TestExecution).filter(
        TestExecution.test_plan_id == testplan_id,
        TestExecution.test_case_id.in_(removable_ids)
    ).delete(synchronize_session=False)

    db.commit()

    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTPLANS,
        action=LogAction.UPDATE,
        description=f"从测试计划 {testplan.name} 批量移除 {removed} 个用例（跳过 {len(executed_ids)} 个已执行）",
        request=req
    )

    return {
        "code": 200,
        "message": "success",
        "data": {
            "removed": removed,
            "skipped": len(testcase_ids) - len(linked_ids),
            "skipped_executed": len(executed_ids)
        }
    }


def _generate_report_snapshot_on_submit(db, testplan, project_name, verify_env, release_note, risk_assessment, zmind_stats, zmind_issues, include_pr_closed, mplist_data=None, report_remark=None):
    """在提交审核时生成报告快照数据"""
    from models import TestCase, TestPlanExecutor, TestPlanTestCase
    from utils.testcase_utils import parse_steps_and_expected
    
    # 获取关联的用例ID - 优化为子查询一次完成
    linked_ids_subq = db.query(TestPlanTestCase.test_case_id).filter(
        TestPlanTestCase.test_plan_id == testplan.id,
        TestPlanTestCase.test_case_id.isnot(None)
    ).subquery()
    
    all_cases = db.query(TestCase).filter(TestCase.id.in_(linked_ids_subq)).all()
    linked_ids = {tc.id for tc in all_cases}
    
    # 获取执行数据（每个用例只取最新记录，只统计关联用例）
    latest_subq = db.query(
        TestExecution.test_case_id,
        sa_func.max(TestExecution.id).label('max_id')
    ).filter(
        TestExecution.test_plan_id == testplan.id,
        TestExecution.test_case_id.in_(linked_ids)
    ).group_by(
        TestExecution.test_case_id
    ).subquery()

    executions = db.query(TestExecution).join(
        latest_subq,
        TestExecution.id == latest_subq.c.max_id
    ).all()
    
    # 批量加载所有关联的测试用例（避免 N+1 查询）
    test_case_ids = [e.test_case_id for e in executions]
    test_cases_map = {}
    if test_case_ids:
        all_test_cases = db.query(TestCase).filter(TestCase.id.in_(test_case_ids)).all()
        test_cases_map = {tc.id: tc for tc in all_test_cases}
    
    # 从测试计划获取执行周期
    if testplan.start_time and testplan.end_time:
        test_cycle = f"{testplan.start_time.strftime('%Y-%m-%d')} ~ {testplan.end_time.strftime('%Y-%m-%d')}"
    else:
        test_cycle = ""
    
    # 从测试计划执行人表获取测试人员
    plan_executors = db.query(TestPlanExecutor, User).join(
        User, TestPlanExecutor.executor_id == User.id
    ).filter(TestPlanExecutor.test_plan_id == testplan.id).all()
    
    if plan_executors:
        testers = [user.username for _, user in plan_executors]
        testers_str = ", ".join(testers)
    else:
        testers_str = ""
    
    # 获取审核人员
    reviewer = db.query(User).filter(User.id == testplan.reviewer_id).first()
    reviewer_name = reviewer.username if reviewer else ""
    
    # 计算测试结论
    # Passing rate = PASS / (总用例 - NA - 协测)
    # 只统计有效的用例ID（排除NULL）
    plan_total_cases = db.query(TestPlanTestCase).filter(
        TestPlanTestCase.test_plan_id == testplan.id,
        TestPlanTestCase.test_case_id.isnot(None)
    ).join(
        TestCase, TestPlanTestCase.test_case_id == TestCase.id
    ).count()
    na_cases_count = sum(1 for e in executions if e.result == 'NA')
    assist_cases_count = sum(1 for e in executions if (e.result or '').strip() == '协测')
    passed_cases = sum(1 for e in executions if e.result == 'PASS')
    valid_cases = plan_total_cases - na_cases_count - assist_cases_count
    pass_rate = (passed_cases / valid_cases * 100) if valid_cases > 0 else 0
    total_cases = plan_total_cases
    
    if pass_rate >= 95:
        test_conclusion = "测试通过"
    else:
        test_conclusion = "测试未通过"
    
    # 按模块统计测试结果（只统计主模块，子模块归入主模块）
    module_results = {}
    has_assist = False  # 是否有协测结果
    for execution in executions:
        test_case = test_cases_map.get(execution.test_case_id)
        if test_case:
            module = (test_case.module or "未分类").split('/')[0].strip()
            if module not in module_results:
                module_results[module] = {
                    "test_cases": 0,
                    "pass": 0,
                    "fail": 0,
                    "block": 0,
                    "nt": 0,
                    "na": 0,
                    "assist": 0
                }
            module_results[module]["test_cases"] += 1
            result_val = (execution.result or '').strip()
            if result_val == 'PASS':
                module_results[module]["pass"] += 1
            elif result_val == 'FAIL':
                module_results[module]["fail"] += 1
            elif result_val == 'BLOCK':
                module_results[module]["block"] += 1
            elif result_val == 'NT':
                module_results[module]["nt"] += 1
            elif result_val == 'NA':
                module_results[module]["na"] += 1
            elif result_val == '协测':
                module_results[module]["assist"] += 1
                has_assist = True
    
    # 转换为列表格式
    # Passing rate = PASS / (总用例 - NA - 协测)
    test_results = []
    for module, result in module_results.items():
        effective_cases = result["test_cases"] - result["na"] - result["assist"]
        passing_rate = (result["pass"] / effective_cases * 100) if effective_cases > 0 else 0
        row = {
            "module": module,
            "test_cases": effective_cases,
            "pass": result["pass"],
            "fail": result["fail"],
            "block": result["block"],
            "nt": result["nt"],
            "na": result["na"],
            "passing_rate": f"{passing_rate:.2f}%"
        }
        if has_assist:
            row["assist"] = result["assist"]
        test_results.append(row)
    
    # 准备用例详细情况
    # 按模块sort_order排序 test_results
    from utils.module_sort import get_module_sort_map as _get_sort_map
    import re
    
    _sort_map = _get_sort_map(db, testplan.project_id)
    
    def _nat_key(cn):
        if not cn:
            return ('', 0, 0, '', 0, '')
        m = re.match(r'^([^0-9]*)(\d+)([^0-9]?)(\d*)(.*)', cn)
        if m:
            return (m.group(1), int(m.group(2)), 0 if not m.group(3) else 1, m.group(3), int(m.group(4)) if m.group(4) else 0, m.group(5))
        return (cn, 0, 0, '', 0, '')
    
    # test_results 是模块统计，只需按模块排序
    test_results.sort(key=lambda r: _get_module_sort_key_with_fallback(r.get("module", ""), _sort_map))
    
    # 如果有 MpList 数据，解析测试统计并加入 MpList 模块行（在 Total 行上方）
    from utils.test_result_calc import parse_mplist_test_stats
    mplist_stats = None
    if mplist_data and isinstance(mplist_data, dict) and mplist_data.get('headers'):
        mplist_stats = parse_mplist_test_stats(mplist_data)
        if mplist_stats and mplist_stats.get('has_stats'):
            if mplist_stats.get('has_assist'):
                has_assist = True
            mplist_row = {
                "module": "MpList",
                "test_cases": mplist_stats['test_cases'],
                "pass": mplist_stats['pass'],
                "fail": mplist_stats['fail'],
                "block": mplist_stats['block'],
                "nt": mplist_stats['nt'],
                "na": mplist_stats['na'],
                "passing_rate": f"{mplist_stats['passing_rate']:.2f}%"
            }
            if has_assist:
                mplist_row["assist"] = mplist_stats['assist']
                # 回填之前没有 assist 字段的模块行
                for r in test_results:
                    if "assist" not in r:
                        r["assist"] = 0
            test_results.append(mplist_row)

    test_cases = []
    for execution in executions:
        test_case = test_cases_map.get(execution.test_case_id)
        if test_case:
            steps_list, expected_list = parse_steps_and_expected(test_case.steps, test_case.expected_result)
            test_cases.append({
                "case_number": test_case.case_number,
                "module": test_case.module,
                "name": test_case.name,
                "precondition": test_case.precondition or '',
                "steps": "\n".join(steps_list),
                "expected_result": "\n".join(expected_list),
                "level": test_case.level or '',
                "result": execution.result,
                "remark": execution.remarks or '',
                "sort_order": test_case.sort_order or 0  # 保留sort_order用于快照
            })
    
    # 构建快照数据
    # 排序：编号Tag前缀 → 模块sort_order → 用例sort_order → case_number自然排序
    def _nat_key(cn):
        if not cn:
            return ('', 0)
        m = re.search(r'(\d+)$', cn)
        if m:
            return ('', int(m.group(1)))
        return (cn, 0)
    
    # 提取编号前缀Tag（如 TV_AD_, OS10_AD_）用于排序
    def _get_case_prefix(cn):
        if not cn:
            return 'zzz'  # 未知前缀放最后
        # 提取到最后一个数字前的前缀
        m = re.match(r'^([A-Za-z]+(?:_[A-Za-z]+)?)_?\d', cn or '')
        if m:
            return m.group(1).upper()
        return 'zzz'
    
    # 排序辅助函数：支持子模块fallback - 匹配最长前缀的模块路径
    def _get_module_sort_key(module_path):
        if not module_path:
            return "9999999999"
        # 尝试精确匹配
        if module_path in _sort_map:
            return _sort_map[module_path]
        # 尝试逐层匹配：例如 "ISDB/Setting/Network" 尝试匹配 "ISDB/Setting"、"ISDB"
        parts = module_path.split('/')
        for i in range(len(parts) - 1, 0, -1):
            prefix = '/'.join(parts[:i])
            if prefix in _sort_map:
                return _sort_map[prefix] + '.9999999999'
        return "9999999999"
    
    # 排序：模块sort_order → Case前缀 → 用例sort_order → case_number自然排序
    test_cases.sort(key=lambda tc: (
        _get_module_sort_key(tc["module"] or ""),
        _get_case_prefix(tc["case_number"]),
        tc.get("sort_order", 0) or 0,
        _nat_key(tc["case_number"])
    ))
    # 不再删除sort_order，保留用于快照导出

    snapshot = {
        "cover_data": {
            "project_name": project_name or "",
            "verify_env": verify_env or "",
            "release_note": release_note or "",
            "risk_assessment": risk_assessment or "",
            "report_remark": report_remark or "",
            "test_cycle": test_cycle,
            "testers": testers_str,
            "reviewer_name": reviewer_name,
            "test_conclusion": test_conclusion,
            "pass_rate": f"{pass_rate:.2f}%",
            "total_cases": total_cases,
            "passed_cases": passed_cases,
            "pass_rate_value": pass_rate
        },
        "test_results": test_results,
        "test_cases": test_cases,
        "zmind_stats": zmind_stats or {},
        "issue_list": zmind_issues or [],
        "has_zmind_csv": zmind_stats is not None,
        "has_assist": has_assist,
        "include_pr_closed": include_pr_closed or 0,
        "mplist_data": mplist_data or [],
        "snapshot_time": datetime.utcnow().isoformat()
    }
    
    return snapshot


@router.post("/{testplan_id}/submit-review")
async def submit_testplan_for_review(
    req: Request,
    testplan_id: int,
    project_name: str = Form(...),
    verify_env: str = Form(...),
    release_note: str = Form(...),
    risk_assessment: str = Form(""),
    report_remark: str = Form(""),
    include_pr_closed: int = Form(0),
    report_template_id: Optional[int] = Form(None),
    zmind_file: Optional[UploadFile] = File(None),
    zmind_pr_data: Optional[str] = Form(None),
    mplist_file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """提交测试计划审核"""
    try:
        return await _do_submit_review(
            req=req, testplan_id=testplan_id, project_name=project_name,
            verify_env=verify_env, release_note=release_note,
            risk_assessment=risk_assessment, report_remark=report_remark,
            include_pr_closed=include_pr_closed, report_template_id=report_template_id,
            zmind_file=zmind_file, zmind_pr_data=zmind_pr_data,
            mplist_file=mplist_file, db=db, current_user=current_user
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[ERROR] submit-review failed: {tb}")
        return {"code": 500, "message": f"提交审核失败: {str(e)}", "detail": tb}


async def _do_submit_review(
    req, testplan_id, project_name, verify_env, release_note,
    risk_assessment, report_remark, include_pr_closed, report_template_id,
    zmind_file, zmind_pr_data, mplist_file, db, current_user
):
    """提交测试计划审核 - 实际逻辑"""
    # 验证测试计划是否存在
    testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    # 验证测试计划状态
    if testplan.status not in ["IN_PROGRESS", "REJECTED"]:
        raise HTTPException(status_code=400, detail="只有进行中或审核不通过的测试计划才能提交审核")
    
    # 验证是否有审核人
    if not testplan.reviewer_id:
        raise HTTPException(status_code=400, detail="测试计划未设置审核人")
    
    # 验证所有用例是否已执行 - 使用子查询一次完成
    linked_ids_subq = db.query(TestPlanTestCase.test_case_id).filter(
        TestPlanTestCase.test_plan_id == testplan_id,
        TestPlanTestCase.test_case_id.isnot(None)
    ).subquery()
    
    test_case_ids = [tc.id for tc in db.query(TestCase.id).filter(TestCase.id.in_(linked_ids_subq)).all()]
    
    if test_case_ids:
        executed_test_cases = db.query(TestExecution).filter(
            TestExecution.test_plan_id == testplan_id,
            TestExecution.test_case_id.in_(test_case_ids)
        ).distinct(TestExecution.test_case_id).count()
        
        if executed_test_cases < len(test_case_ids):
            raise HTTPException(status_code=400, detail="还有未执行的用例，无法提交审核")
    
    # 解析Zmind-PR CSV附件 或 PR号查询数据
    zmind_stats = None
    zmind_issues = None
    if zmind_file:
        csv_result = await parse_zmind_csv(zmind_file)
        zmind_stats = csv_result.get('stats')
        zmind_issues = csv_result.get('issues')
    elif zmind_pr_data:
        try:
            pr_data = json.loads(zmind_pr_data)
            zmind_stats = pr_data.get('stats')
            zmind_issues = pr_data.get('issues')
        except (json.JSONDecodeError, TypeError):
            raise HTTPException(status_code=400, detail="PR数据格式错误")
    
    # 解析MpList附件（支持xlsx/csv）
    mplist_data = None
    if mplist_file:
        mplist_data = await parse_mplist_file(mplist_file)
    
    # 生成报告快照数据（在提交审核时就生成，后续直接使用）
    snapshot_data = _generate_report_snapshot_on_submit(
        db=db,
        testplan=testplan,
        project_name=project_name,
        verify_env=verify_env,
        release_note=release_note,
        risk_assessment=risk_assessment,
        zmind_stats=zmind_stats,
        zmind_issues=zmind_issues,
        include_pr_closed=include_pr_closed,
        mplist_data=mplist_data,
        report_remark=report_remark
    )
    
    # 创建测试报告
    report_data = {
        "test_plan_id": testplan_id,
        "name": f"{testplan.name} - 测试报告",
        "project_name": project_name,
        "verify_env": verify_env,
        "release_note": release_note,
        "zmind_pr_stats": json.dumps(zmind_stats) if zmind_stats else None,
        "zmind_issue_list": json.dumps(zmind_issues, ensure_ascii=False) if zmind_issues else None,
        "include_pr_closed": include_pr_closed,
        "risk_assessment": risk_assessment or None,
        "mplist_data": json.dumps(mplist_data, ensure_ascii=False) if mplist_data else None,
        "template_id": report_template_id,
        "status": "PENDING_REVIEW",
        "created_by": current_user.id,
        "snapshot_data": json.dumps(snapshot_data, ensure_ascii=False) if snapshot_data else None
    }
    
    report = Report(**report_data)
    db.add(report)
    
    # 更新测试计划状态为审核中
    testplan.status = "IN_REVIEW"
    
    db.commit()
    db.refresh(report)
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTPLANS,
        action=LogAction.UPDATE,
        description=f"提交测试计划 {testplan.name} 审核，生成报告 ID: {report.id}",
        request=req
    )
    
    # 触发通知 - 只通知审核人
    if testplan.reviewer_id:
        # 获取开始时间和结束时间
        start_time_str = _format_datetime_str(str(testplan.start_time)) if testplan.start_time else "未设置"
        end_time_str = _format_datetime_str(str(testplan.end_time)) if testplan.end_time else "未设置"
        
        # 获取执行人列表
        executor_names = []
        test_plan_executors = db.query(TestPlanExecutor).filter(TestPlanExecutor.test_plan_id == testplan_id).all()
        for te in test_plan_executors:
            user = db.query(User).filter(User.id == te.executor_id).first()
            if user:
                executor_names.append(user.username)
        
        trigger_assignment_notification(
            db=db,
            notification_type='testplan',
            event_type='submitted_for_review',
            title=f'测试计划「{testplan.name}」提交审核',
            content=f'测试计划「{testplan.name}」已提交审核，请您审核。\n\n'
                    f'提交人：{current_user.username}',
            related_id=testplan_id,
            related_type='testplan',
            sender_id=current_user.id,
            recipient_user_ids=[testplan.reviewer_id],
            project_id=testplan.project_id,
            team_id=testplan.team_id,
            extra_context={
                'plan_name': testplan.name,
                'operator': current_user.username,
                'report_id': report.id,
                'start_time': start_time_str,
                'end_time': end_time_str,
                'executors': "、".join(executor_names) if executor_names else "未设置",
                'action': '提交审核',
            }
        )
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "testplan_id": testplan_id,
            "report_id": report.id,
            "status": "IN_REVIEW"
        }
    }


@router.post("/{testplan_id}/withdraw-review")
def withdraw_testplan_review(
    req: Request,
    testplan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """撤回测试计划审核，状态回到IN_PROGRESS，删除对应的待审核报告"""
    testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    if testplan.status != "IN_REVIEW":
        raise HTTPException(status_code=400, detail="只有审核中的测试计划才能撤回")
    
    # 删除对应的待审核报告
    pending_report = db.query(Report).filter(
        Report.test_plan_id == testplan_id,
        Report.status == "PENDING_REVIEW"
    ).first()
    if pending_report:
        db.delete(pending_report)
    
    # 状态回到IN_PROGRESS
    testplan.status = "IN_PROGRESS"
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTPLANS,
        action=LogAction.UPDATE,
        description=f"撤回测试计划 {testplan.name} 的审核",
        request=req
    )
    
    # 通知审核人：审核已被撤回（跟提交审核一样，直接指定审核人）
    if testplan.reviewer_id:
        trigger_assignment_notification(
            db=db,
            notification_type='testplan',
            event_type='review_withdrawn',
            title=f'测试计划「{testplan.name}」审核已撤回',
            content=f'测试计划「{testplan.name}」的审核已被撤回。\n\n'
                    f'撤回人：{current_user.username}',
            related_id=testplan_id,
            related_type='testplan',
            sender_id=current_user.id,
            recipient_user_ids=[testplan.reviewer_id],
            project_id=testplan.project_id,
            team_id=testplan.team_id,
            extra_context={
                'plan_name': testplan.name,
                'operator': current_user.username,
                'action': '撤回审核',
            }
        )
    
    return {"code": 200, "message": "success"}


def _get_team_report_template_for_preview(db, team_id):
    """获取项目组的默认报告模板的 criteria_config + selected_fields（用于预览）"""
    if not team_id:
        return None
    template = db.query(ReportTemplate).filter(
        ReportTemplate.team_id == team_id,
        ReportTemplate.is_default == True
    ).first()
    if not template or not template.criteria_config:
        return None
    try:
        result = json.loads(template.criteria_config)
    except (json.JSONDecodeError, TypeError):
        return None
    if template.selected_fields:
        try:
            result['selected_fields'] = json.loads(template.selected_fields)
        except (json.JSONDecodeError, TypeError):
            pass
    return result


@router.get("/{testplan_id}/preview-report")
def preview_testplan_report(
    req: Request,
    testplan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """预览测试计划报告（不创建报告记录，实时生成快照数据）"""
    testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    # 生成实时快照数据（不传project_name等，用空值占位）
    snapshot = _generate_report_snapshot_on_submit(
        db=db,
        testplan=testplan,
        project_name="(预览)",
        verify_env="(预览)",
        release_note="(预览)",
        risk_assessment="",
        zmind_stats=None,
        zmind_issues=None,
        include_pr_closed=0
    )
    
    cover_data = snapshot.get('cover_data', {})
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "report": {
                "id": 0,
                "status": "PREVIEW",
                "project_name": cover_data.get('project_name', ''),
                "verify_env": cover_data.get('verify_env', ''),
                "release_note": cover_data.get('release_note', ''),
                "risk_assessment": cover_data.get('risk_assessment', ''),
                "test_cycle": cover_data.get('test_cycle', ''),
                "testers": cover_data.get('testers', ''),
                "reviewer_name": cover_data.get('reviewer_name', ''),
                "report_remark": cover_data.get('report_remark', ''),
            },
            "test_results": snapshot.get('test_results', []),
            "test_cases": snapshot.get('test_cases', []),
            "zmind_stats": {},
            "issue_list": [],
            "include_pr_closed": 0,
            "has_zmind_csv": False,
            "has_assist": snapshot.get('has_assist', False),
            "is_preview": True,
            "report_template_config": _get_team_report_template_for_preview(db, testplan.team_id)
        }
    }


@router.post("/{testplan_id}/preview-report")
async def preview_testplan_report_with_data(
    req: Request,
    testplan_id: int,
    project_name: str = Form(''),
    verify_env: str = Form(''),
    release_note: str = Form(''),
    risk_assessment: str = Form(''),
    report_remark: str = Form(''),
    include_pr_closed: int = Form(0),
    report_template_id: Optional[int] = Form(None),
    zmind_file: Optional[UploadFile] = File(None),
    zmind_pr_data: Optional[str] = Form(None),
    mplist_file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """带表单数据的报告预览（用于提交前确认）"""
    testplan = db.query(TestPlan).filter(TestPlan.id == testplan_id).first()
    if not testplan:
        raise HTTPException(status_code=404, detail="测试计划不存在")
    
    # 处理Zmind数据
    zmind_stats = None
    zmind_issues = None
    
    if zmind_file and zmind_file.filename:
        csv_result = await parse_zmind_csv(zmind_file)
        if csv_result.get('stats'):
            zmind_stats = csv_result['stats']
        if csv_result.get('issues'):
            zmind_issues = csv_result['issues']
    elif zmind_pr_data:
        try:
            pr_data = json.loads(zmind_pr_data)
            if pr_data.get('stats'):
                zmind_stats = pr_data['stats']
            if pr_data.get('issues'):
                zmind_issues = pr_data['issues']
        except (json.JSONDecodeError, TypeError):
            pass
    
    # 处理MpList数据
    mplist_data = None
    if mplist_file and mplist_file.filename:
        mplist_data = await parse_mplist_file(mplist_file)
    
    # 获取报告模板配置（criteria_config + selected_fields）
    report_template_config = None
    if report_template_id:
        tpl = db.query(ReportTemplate).filter(ReportTemplate.id == report_template_id).first()
        if tpl and tpl.criteria_config:
            try:
                report_template_config = json.loads(tpl.criteria_config)
            except (json.JSONDecodeError, TypeError):
                pass
            if report_template_config and tpl.selected_fields:
                try:
                    report_template_config['selected_fields'] = json.loads(tpl.selected_fields)
                except (json.JSONDecodeError, TypeError):
                    pass
    if not report_template_config:
        report_template_config = _get_team_report_template_for_preview(db, testplan.team_id)
    
    snapshot = _generate_report_snapshot_on_submit(
        db=db,
        testplan=testplan,
        project_name=project_name,
        verify_env=verify_env,
        release_note=release_note,
        risk_assessment=risk_assessment,
        zmind_stats=zmind_stats,
        zmind_issues=zmind_issues,
        include_pr_closed=include_pr_closed,
        mplist_data=mplist_data,
        report_remark=report_remark
    )
    
    cover_data = snapshot.get('cover_data', {})
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "report": {
                "id": 0,
                "status": "SUBMIT_PREVIEW",
                "project_name": cover_data.get('project_name', ''),
                "verify_env": cover_data.get('verify_env', ''),
                "release_note": cover_data.get('release_note', ''),
                "risk_assessment": cover_data.get('risk_assessment', ''),
                "test_cycle": cover_data.get('test_cycle', ''),
                "testers": cover_data.get('testers', ''),
                "reviewer_name": cover_data.get('reviewer_name', ''),
                "report_remark": cover_data.get('report_remark', ''),
            },
            "test_results": snapshot.get('test_results', []),
            "test_cases": snapshot.get('test_cases', []),
            "zmind_stats": snapshot.get('zmind_stats', {}),
            "issue_list": snapshot.get('issue_list', []),
            "include_pr_closed": snapshot.get('include_pr_closed', 0),
            "has_zmind_csv": snapshot.get('has_zmind_csv', False),
            "has_assist": snapshot.get('has_assist', False),
            "mplist_data": snapshot.get('mplist_data', []),
            "is_preview": True,
            "is_submit_preview": True,
            "report_template_config": report_template_config
        }
    }


async def parse_mplist_file(file: UploadFile) -> dict:
    """解析MpList文件（支持xlsx/xls/csv），原样返回表头和数据。
    返回: {'headers': ['col1', 'col2', ...], 'rows': [['val1', 'val2', ...], ...]}
    """
    import logging
    logger = logging.getLogger(__name__)

    content = await file.read()
    filename = (file.filename or '').lower()

    headers = []
    rows = []

    if filename.endswith('.csv'):
        # CSV解析 - 智能编码检测
        best_text = None
        for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']:
            try:
                best_text = content.decode(encoding)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if best_text is None:
            best_text = content.decode('utf-8', errors='replace')

        # 自动检测分隔符
        first_line = best_text.split('\n')[0] if best_text else ''
        delimiter = ','
        if '\t' in first_line and first_line.count('\t') > first_line.count(','):
            delimiter = '\t'
        elif ';' in first_line and first_line.count(';') > first_line.count(','):
            delimiter = ';'

        best_text = best_text.replace('\r\n', '\n').replace('\r', '\n')
        reader = csv.reader(io.StringIO(best_text), delimiter=delimiter)
        all_rows = list(reader)
        if all_rows:
            headers = [str(h).strip() for h in all_rows[0]]
            for row in all_rows[1:]:
                if all(not str(c).strip() for c in row):
                    continue
                # 补齐列数
                padded = [str(c).strip() if c else '' for c in row]
                while len(padded) < len(headers):
                    padded.append('')
                rows.append(padded[:len(headers)])
    else:
        # Excel解析 - 先尝试openpyxl(.xlsx)，失败则用pandas兼容.xls等格式
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                return {'headers': [], 'rows': []}

            all_data = list(ws.iter_rows(values_only=True))
            if all_data:
                headers = [str(h).strip() if h is not None else '' for h in all_data[0]]
                for row in all_data[1:]:
                    if all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    padded = [str(cell).strip() if cell is not None else '' for cell in row]
                    while len(padded) < len(headers):
                        padded.append('')
                    rows.append(padded[:len(headers)])
            wb.close()
        except Exception as openpyxl_err:
            logger.warning(f"MpList: openpyxl解析失败({openpyxl_err})，尝试pandas解析")
            try:
                import pandas as pd
                df = pd.read_excel(io.BytesIO(content), header=0, dtype=str)
                df = df.fillna('')
                headers = [str(h).strip() for h in df.columns.tolist()]
                for _, row in df.iterrows():
                    vals = [str(v).strip() for v in row.tolist()]
                    if all(not v for v in vals):
                        continue
                    rows.append(vals)
            except Exception as pd_err:
                logger.error(f"MpList: pandas解析也失败({pd_err})，返回空数据")
                return {'headers': [], 'rows': []}

    logger.info(f"MpList: 解析到 {len(headers)} 列, {len(rows)} 行数据")

    # 限制最大列数，防止异常文件导致前端渲染卡死
    MAX_COLUMNS = 30
    if len(headers) > MAX_COLUMNS:
        logger.warning(f"MpList: 列数({len(headers)})超过上限{MAX_COLUMNS}，截断处理")
        headers = headers[:MAX_COLUMNS]
        rows = [row[:MAX_COLUMNS] for row in rows]

    return {'headers': headers, 'rows': rows}


async def parse_zmind_csv(file: UploadFile) -> dict:
    """解析Zmind-PR CSV文件，统计各Severity的Open和Total数据，并提取issue列表。
    自动识别Severity列和状态列，不依赖固定列名。
    返回: {'stats': {...统计数据...}, 'issues': [...issue列表...]}
    """
    import logging
    logger = logging.getLogger(__name__)
    
    from utils.constants import CLOSED_STATUSES, OPEN_STATUS_ORDER
    
    SEVERITY_KEYS = ['Blocker', 'Critical', 'Major', 'Minor', 'Enhancement']
    SEVERITY_LOWER = {s.lower(): s for s in SEVERITY_KEYS}
    SEVERITY_ORDER = ['Blocker', 'Critical', 'Major', 'Minor', 'Enhancement']
    
    content = await file.read()
    
    # 智能编码检测
    # 策略：尝试多种编码，选择能正确解析出已知列名关键字最多的那个
    KNOWN_KEYWORDS = ['severity', '跟踪', '类别', '状态', '优先级', '主题', '指派给', 'tracker', 'category', 'status', 'priority', 'subject', 'assigned']
    
    best_text = None
    best_encoding = None
    best_score = -1
    
    for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']:
        try:
            candidate = content.decode(encoding)
            first_line = candidate.split('\n')[0].lower() if candidate else ''
            score = sum(1 for kw in KNOWN_KEYWORDS if kw in first_line)
            if score > best_score:
                best_score = score
                best_text = candidate
                best_encoding = encoding
        except (UnicodeDecodeError, LookupError):
            continue
    
    if best_text is not None:
        text = best_text
    else:
        text = content.decode('utf-8', errors='replace')
        best_encoding = 'utf-8(fallback)'
    
    logger.info(f"CSV 编码检测: {best_encoding} (score={best_score})")
    
    # 自动检测分隔符
    first_line = text.split('\n')[0] if text else ''
    delimiter = ','
    if '\t' in first_line and first_line.count('\t') > first_line.count(','):
        delimiter = '\t'
    elif ';' in first_line and first_line.count(';') > first_line.count(','):
        delimiter = ';'
    
    logger.info(f"CSV 分隔符: {repr(delimiter)}, 首行: {first_line[:200]}")
    
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    
    if not rows:
        return {
            'stats': {'total_prs': 0, 'open': 0, **{s.lower(): 0 for s in SEVERITY_KEYS}, **{f'open_{s.lower()}': 0 for s in SEVERITY_KEYS}},
            'issues': []
        }
    
    # 自动识别各列
    headers = list(rows[0].keys()) if rows else []
    severity_col = None
    status_col = None
    pr_col = None  # #(PR号)
    tracker_col = None  # 跟踪
    category_col = None  # 类别
    priority_col = None  # 优先级
    subject_col = None  # 主题
    assignee_col = None  # 指派给
    author_col = None  # 作者/创建者
    
    for h in headers:
        if not h:
            continue
        hl = h.lower().strip()
        h_stripped = h.strip()
        
        # 识别PR号列（必须是 # 开头或就是 #，避免误匹配 priority 等）
        if h_stripped == '#' or h_stripped.startswith('#(') or h_stripped.startswith('# '):
            if not pr_col:
                pr_col = h
        elif hl == 'pr' or hl == 'pr号':
            if not pr_col:
                pr_col = h
        # 识别跟踪列
        if '跟踪' in h or hl == 'tracker':
            tracker_col = h
        # 识别类别列
        if '类别' in h or hl == 'category':
            category_col = h
        # 识别Severity列
        if 'severity' in hl:
            severity_col = h
        # 识别状态列
        if ('状态' in h or hl == 'status') and '更新' not in h and 'update' not in hl:
            status_col = h
        # 识别优先级列
        if '优先级' in h or hl == 'priority':
            priority_col = h
        # 识别主题列
        if '主题' in h or hl == 'subject' or hl == 'title':
            subject_col = h
        # 识别指派给列
        if '指派给' in h or 'assignee' in hl or 'assigned' in hl:
            assignee_col = h
        # 识别作者/创建者列
        if '作者' in h or '创建者' in h or hl == 'author' or hl == 'creator' or hl == 'created by':
            author_col = h
    
    # 如果通过列名没找到，尝试通过列值内容推断
    logger.info(f"CSV 列名识别结果(列名匹配后): pr={pr_col}, tracker={tracker_col}, category={category_col}, severity={severity_col}, status={status_col}, priority={priority_col}, subject={subject_col}, assignee={assignee_col}, author={author_col}")
    logger.info(f"CSV 原始列名: {headers}")
    if not severity_col or not status_col:
        sample_rows = rows[:min(20, len(rows))]
        for h in headers:
            if not h:
                continue
            vals = [r.get(h, '').strip().lower() for r in sample_rows if r.get(h)]
            if not severity_col:
                severity_matches = sum(1 for v in vals if v in SEVERITY_LOWER)
                if severity_matches >= len(vals) * 0.5 and severity_matches >= 2:
                    severity_col = h
            if not status_col:
                status_keywords = {'new', 'on-going', 'closed', 'confirm', 'info', 'pending', 'verification failed', 're-open', 'partner wfr'}
                status_matches = sum(1 for v in vals if v in status_keywords)
                if status_matches >= len(vals) * 0.3 and status_matches >= 2:
                    status_col = h
    
    logger.info(f"CSV 列名识别最终结果: pr={pr_col}, tracker={tracker_col}, category={category_col}, severity={severity_col}, status={status_col}, priority={priority_col}, subject={subject_col}, assignee={assignee_col}, author={author_col}")
    
    # 初始化统计
    stats = {s: {'open': 0, 'total': 0} for s in SEVERITY_KEYS}
    issues = []
    
    for row in rows:
        severity_val = row.get(severity_col, '').strip() if severity_col else ''
        status_val = row.get(status_col, '').strip() if status_col else ''
        
        # 提取issue数据
        issue = {
            'pr_number': row.get(pr_col, '').strip() if pr_col else '',
            'tracker': row.get(tracker_col, '').strip() if tracker_col else '',
            'category': row.get(category_col, '').strip() if category_col else '',
            'severity': severity_val,
            'status': status_val,
            'priority': row.get(priority_col, '').strip() if priority_col else '',
            'subject': row.get(subject_col, '').strip() if subject_col else '',
            'assignee': row.get(assignee_col, '').strip() if assignee_col else '',
            'author': row.get(author_col, '').strip() if author_col else ''
        }
        
        # 只添加有PR号的issue
        if issue['pr_number']:
            issues.append(issue)
        
        if not severity_val:
            continue
        
        # 匹配Severity（不区分大小写）
        matched_severity = SEVERITY_LOWER.get(severity_val.lower())
        if not matched_severity:
            continue
        
        stats[matched_severity]['total'] += 1
        if status_val not in CLOSED_STATUSES:
            stats[matched_severity]['open'] += 1
    
    # 对issues进行排序
    def get_severity_order(severity):
        """获取Severity排序优先级"""
        severity_upper = severity.capitalize() if severity else ''
        if severity_upper in SEVERITY_ORDER:
            return SEVERITY_ORDER.index(severity_upper)
        return len(SEVERITY_ORDER)  # 未知severity排最后
    
    def is_open_status(status):
        """判断是否为Open状态"""
        return status not in CLOSED_STATUSES
    
    # 排序规则：
    # 1. 先按是否为Open状态分组（Open状态优先）
    # 2. 在每个分组内按Severity排序（Blocker → Critical → Major → Minor → Enhancement）
    issues.sort(key=lambda x: (
        0 if is_open_status(x['status']) else 1,  # Open状态优先
        get_severity_order(x['severity'])          # 然后按Severity排序
    ))
    
    # 构建返回数据
    result_stats = {}
    for s in SEVERITY_KEYS:
        key = s.lower()
        result_stats[key] = stats[s]['total']
        result_stats[f'open_{key}'] = stats[s]['open']
    
    result_stats['total_prs'] = sum(stats[s]['total'] for s in SEVERITY_KEYS)
    result_stats['open'] = sum(stats[s]['open'] for s in SEVERITY_KEYS)
    
    return {
        'stats': result_stats,
        'issues': issues
    }
