from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_
from database import get_db
from models import (
    User, UserProject, UserRole, Project, PositionTag, LoginAttempt,
    TestCase, TestExecution, ReviewPlan, ReviewPlanTestCase, Report, TestPlan,
    Module, TestSuite, Comment, UserTeam, UserDepartment,
    UserProjectGroup, UserNotificationPreference, TestPlanExecutor, TestExecutionProgress,
    DingtalkBot, CaseTemplate, ReportTemplate, Notification, NotificationRecipient,
    DepartmentManager, TeamLeader, Role, Department, Team
)
from schemas import UserCreate
from auth import get_current_user, get_password_hash, verify_password, is_super_admin
from utils.logger import log_operation, LogAction, LogModule
from pydantic import BaseModel, Field
from typing import Optional
import os
import uuid
from datetime import datetime

router = APIRouter()

# 密码修改请求模型
class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str

# 个人资料更新请求模型
class UpdateProfileRequest(BaseModel):
    full_name: Optional[str] = Field(default=None)
    email: Optional[str] = Field(default=None)
    phone: Optional[str] = Field(default=None)

@router.get("")
def list_users(
    page: int = 1,
    size: int = 1000,
    search: str = "",
    role_id: int = None,
    status: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(User)

    # 排除超级管理员账号（admin, super）和待审核用户（status=2）
    query = query.filter(User.username.notin_(['admin', 'super']))
    query = query.filter(User.status != 2)
    
    # 状态筛选
    if status is not None:
        query = query.filter(User.status == status)
    
    # 角色筛选
    if role_id is not None:
        role_user_ids = [ur.user_id for ur in db.query(UserRole.user_id).filter(UserRole.role_id == role_id).all()]
        query = query.filter(User.id.in_(role_user_ids)) if role_user_ids else query.filter(User.id == -1)
    
    # 搜索功能
    if search and search.strip():
        search_pattern = f"%{search.strip()}%"
        query = query.filter(
            or_(
                User.username.ilike(search_pattern),
                User.full_name.ilike(search_pattern),
                User.email.ilike(search_pattern)
            )
        )
    
    total = query.count()
    users = query.order_by(User.id.asc()).offset((page - 1) * size).limit(size).all()
    
    user_list = []
    for user in users:
        # 获取授权用例库
        user_projects = db.query(UserProject.project_id).filter(UserProject.user_id == user.id).all()
        project_ids = [up.project_id for up in user_projects]
        
        # 获取项目名称（保持与project_ids相同的顺序）
        project_names = []
        if project_ids:
            projects = db.query(Project).filter(Project.id.in_(project_ids)).all()
            # 创建ID到名称的映射
            project_map = {p.id: p.name for p in projects}
            # 按照project_ids的顺序构建project_names
            project_names = [project_map.get(pid, f'项目-{pid}') for pid in project_ids]
        
        # 获取角色
        user_role = db.query(UserRole.role_id).filter(UserRole.user_id == user.id).first()
        role_id = user_role.role_id if user_role else None
        
        # 获取项目组
        from models import UserTeam, Team, UserDepartment, Department
        user_teams = db.query(UserTeam).filter(UserTeam.user_id == user.id).all()
        team_ids = [ut.team_id for ut in user_teams]
        team_names = []
        if team_ids:
            teams = db.query(Team).filter(Team.id.in_(team_ids)).all()
            team_names = [t.name for t in teams]
        
        # 获取所属组织
        user_depts = db.query(UserDepartment).filter(UserDepartment.user_id == user.id).all()
        dept_ids = [ud.department_id for ud in user_depts]
        dept_names = []
        if dept_ids:
            depts = db.query(Department).filter(Department.id.in_(dept_ids)).all()
            dept_names = [d.name for d in depts]
        
        user_dict = {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "full_name": user.full_name,
            "status": user.status,
            "projectIds": project_ids,
            "projectNames": project_names,
            "roleId": role_id,
            "teamIds": team_ids,
            "teamNames": team_names,
            "departmentIds": dept_ids,
            "departmentNames": dept_names,
            "position_tag_id": user.position_tag_id,
            "phone": user.phone
        }
        user_list.append(user_dict)
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "records": user_list,
            "total": total
        }
    }

@router.post("")
def create_user(
    user: UserCreate,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    existing_user = db.query(User).filter(User.username == user.username).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="用户名已存在")
    
    # 验证职位Tag是否存在
    if user.position_tag_id:
        position_tag = db.query(PositionTag).filter(PositionTag.id == user.position_tag_id).first()
        if not position_tag:
            raise HTTPException(status_code=404, detail="职位Tag不存在")
    
    # 初始密码为邮箱，首次登录强制修改
    hashed_password = get_password_hash(user.password or user.email)
    db_user = User(
        username=user.username,
        password=hashed_password,
        email=user.email,
        full_name=user.full_name,
        phone=user.phone,
        position_tag_id=user.position_tag_id,
        must_change_password=True,
        status=1
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.CREATE,
        description=f"创建用户：{db_user.username}（ID: {db_user.id}）",
        request=req
    )
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "id": db_user.id,
            "username": db_user.username,
            "email": db_user.email,
            "full_name": db_user.full_name,
            "status": db_user.status,
            "created_at": db_user.created_at
        }
    }

@router.post("/change-password")
def change_password(
    request: ChangePasswordRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """修改当前用户密码"""
    from utils.login_security import validate_password_strength
    
    # 验证旧密码
    if not verify_password(request.old_password, current_user.password):
        raise HTTPException(status_code=400, detail="当前密码不正确")
    
    # 校验新密码强度
    is_valid, error_msg = validate_password_strength(request.new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # 新旧密码不能相同
    if request.old_password == request.new_password:
        raise HTTPException(status_code=400, detail="新密码不能与当前密码相同")
    
    # 更新密码
    current_user.password = get_password_hash(request.new_password)
    current_user.must_change_password = False
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.UPDATE,
        description=f"修改密码",
        request=req
    )
    
    return {"code": 200, "message": "密码修改成功", "data": None}

@router.put("/profile")
def update_profile(
    request: UpdateProfileRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新当前用户个人资料"""
    changes = []
    
    if request.full_name is not None and request.full_name != current_user.full_name:
        changes.append(f"姓名: {current_user.full_name or '(空)'} → {request.full_name}")
        current_user.full_name = request.full_name
    
    if request.email is not None and request.email != current_user.email:
        changes.append(f"邮箱: {current_user.email or '(空)'} → {request.email}")
        current_user.email = request.email
    
    if request.phone is not None and request.phone != current_user.phone:
        changes.append(f"手机: {current_user.phone or '(空)'} → {request.phone}")
        current_user.phone = request.phone
    
    if changes:
        db.commit()
        db.refresh(current_user)
        
        change_detail = "；".join(changes)
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.USERS,
            action=LogAction.UPDATE,
            description=f"更新个人资料：{change_detail}",
            request=req
        )
    
    return {
        "code": 200,
        "message": "个人资料更新成功",
        "data": {
            "id": current_user.id,
            "username": current_user.username,
            "email": current_user.email,
            "full_name": current_user.full_name,
            "phone": current_user.phone,
            "avatar": current_user.avatar
        }
    }

@router.post("/avatar")
async def upload_avatar(
    req: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """上传用户头像（存储为base64到数据库）"""
    import base64
    
    # 验证文件类型
    allowed_types = {"image/jpeg": "jpeg", "image/png": "png", "image/gif": "gif", "image/webp": "webp"}
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="仅支持JPG/PNG/GIF/WEBP格式")
    
    # 读取并验证大小（10MB）
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="头像图片不能超过10MB")
    
    # 转为 base64 data URL
    b64 = base64.b64encode(contents).decode("utf-8")
    data_url = f"data:{file.content_type};base64,{b64}"
    
    # 存入数据库
    current_user.avatar = data_url
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.UPDATE,
        description="上传头像",
        request=req
    )
    
    return {
        "code": 200,
        "message": "头像上传成功",
        "data": {"avatar_url": data_url}
    }

@router.delete("/avatar")
def delete_avatar(
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除用户头像"""
    if not current_user.avatar:
        raise HTTPException(status_code=400, detail="当前没有头像")
    
    current_user.avatar = None
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.DELETE,
        description="删除头像",
        request=req
    )
    
    return {"code": 200, "message": "头像已删除", "data": None}

@router.get("/pending")
def list_pending_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取待审核的注册用户列表"""
    users = db.query(User).filter(User.status == 2).order_by(User.created_at.desc()).all()
    return {
        "code": 200,
        "message": "success",
        "data": [{
            "id": u.id,
            "username": u.username,
            "email": u.email,
            "full_name": u.full_name,
            "phone": u.phone,
            "created_at": u.created_at.isoformat() if u.created_at else None
        } for u in users]
    }


class ApproveUserRequest(BaseModel):
    role_id: Optional[int] = None
    position_tag_id: Optional[int] = None
    department_ids: Optional[list] = None
    team_ids: Optional[list] = None


@router.post("/pending/{user_id}/approve")
def approve_user(
    user_id: int,
    req: Request,
    body: ApproveUserRequest = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """审核通过注册用户，可同时授权角色、数据权限、组织、项目组"""
    user = db.query(User).filter(User.id == user_id, User.status == 2).first()
    if not user:
        raise HTTPException(status_code=404, detail="待审核用户不存在")
    
    user.status = 1
    
    # 可选：分配数据权限标签
    if body and body.position_tag_id:
        user.position_tag_id = body.position_tag_id
    
    db.flush()
    
    # 可选：分配角色
    if body and body.role_id:
        from models import UserRole
        existing_role = db.query(UserRole).filter(UserRole.user_id == user_id).first()
        if existing_role:
            existing_role.role_id = body.role_id
        else:
            db.add(UserRole(user_id=user_id, role_id=body.role_id))
    
    # 可选：分配组织
    if body and body.department_ids is not None:
        from models import UserDepartment
        db.query(UserDepartment).filter(UserDepartment.user_id == user_id).delete()
        for dept_id in body.department_ids:
            db.add(UserDepartment(user_id=user_id, department_id=dept_id))
    
    # 可选：分配项目组
    if body and body.team_ids is not None:
        from models import UserTeam
        db.query(UserTeam).filter(UserTeam.user_id == user_id).delete()
        for team_id in body.team_ids:
            db.add(UserTeam(user_id=user_id, team_id=team_id))
    
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.UPDATE,
        description=f"审核通过注册用户：{user.username}（ID: {user.id}）",
        request=req
    )
    
    return {"code": 200, "message": "审核通过", "data": None}


@router.post("/pending/{user_id}/reject")
def reject_user(
    user_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """拒绝注册用户（直接删除）"""
    user = db.query(User).filter(User.id == user_id, User.status == 2).first()
    if not user:
        raise HTTPException(status_code=404, detail="待审核用户不存在")
    
    username = user.username
    db.delete(user)
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.DELETE,
        description=f"拒绝注册用户：{username}（ID: {user_id}）",
        request=req
    )
    
    return {"code": 200, "message": "已拒绝", "data": None}


@router.put("/{user_id}")
def update_user(
    user_id: int,
    user: UserCreate,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    if user.username != db_user.username:
        existing = db.query(User).filter(User.username == user.username).first()
        if existing:
            raise HTTPException(status_code=400, detail="用户名已存在")
    
    # 验证职位Tag是否存在
    if user.position_tag_id is not None:
        position_tag = db.query(PositionTag).filter(PositionTag.id == user.position_tag_id).first()
        if not position_tag:
            raise HTTPException(status_code=404, detail="职位Tag不存在")
    
    old_username = db_user.username
    changes = []  # 记录变更内容
    
    # 记录变更
    if user.username != db_user.username:
        changes.append(f"用户名: {db_user.username} → {user.username}")
    if user.email != db_user.email:
        changes.append(f"邮箱: {db_user.email or '(空)'} → {user.email or '(空)'}")
    if user.full_name != db_user.full_name:
        changes.append(f"姓名: {db_user.full_name or '(空)'} → {user.full_name or '(空)'}")
    if user.phone is not None and user.phone != db_user.phone:
        changes.append(f"手机: {db_user.phone or '(空)'} → {user.phone or '(空)'}")
    if user.status is not None and user.status != db_user.status:
        status_map = {0: "禁用", 1: "启用"}
        changes.append(f"状态: {status_map.get(db_user.status, db_user.status)} → {status_map.get(user.status, user.status)}")
    if user.position_tag_id is not None and user.position_tag_id != db_user.position_tag_id:
        old_tag = db.query(PositionTag).filter(PositionTag.id == db_user.position_tag_id).first()
        new_tag = db.query(PositionTag).filter(PositionTag.id == user.position_tag_id).first()
        old_tag_name = old_tag.name if old_tag else "(空)"
        new_tag_name = new_tag.name if new_tag else "(空)"
        changes.append(f"职位Tag: {old_tag_name} → {new_tag_name}")
    
    db_user.username = user.username
    if user.email is not None:
        db_user.email = user.email
    if user.full_name is not None:
        db_user.full_name = user.full_name
    if user.phone is not None:
        db_user.phone = user.phone
    if user.status is not None:
        db_user.status = user.status
    if user.position_tag_id is not None:
        db_user.position_tag_id = user.position_tag_id
    
    db.commit()
    db.refresh(db_user)
    
    # 只有在有实际变更时才记录日志
    if changes:
        change_detail = "；".join(changes)
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.USERS,
            action=LogAction.UPDATE,
            description=f"更新用户：{old_username}（ID: {user_id}，{change_detail}）",
            request=req
        )
    
    return {"code": 200, "message": "success", "data": db_user}

@router.put("/{user_id}/toggle-status")
def toggle_user_status(
    user_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    if is_super_admin(user):
        raise HTTPException(status_code=403, detail="超级管理员账号不能被禁用")
    
    old_status = user.status
    user.status = 0 if user.status == 1 else 1
    db.commit()
    db.refresh(user)
    
    status_text = "启用" if user.status == 1 else "禁用"
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.UPDATE,
        description=f"{status_text}用户：{user.username}",
        request=req
    )
    
    return {
        "code": 200,
        "message": "success",
        "data": {"status": user.status}
    }

@router.post("/{user_id}/reset-password")
def reset_user_password(
    user_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """重置用户密码为随机强密码"""
    import secrets
    import string
    
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 生成12位随机强密码（包含大小写+数字+特殊字符）
    alphabet = string.ascii_letters + string.digits
    special = "!@#$%&*"
    # 确保至少包含各类字符
    pwd = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
        secrets.choice(special),
    ]
    pwd += [secrets.choice(alphabet + special) for _ in range(8)]
    secrets.SystemRandom().shuffle(pwd)
    new_password = ''.join(pwd)
    
    user.password = get_password_hash(new_password)
    user.must_change_password = True
    
    # 清除该用户的登录失败记录，避免重置密码后仍被锁定
    db.query(LoginAttempt).filter(
        LoginAttempt.username == user.username,
        LoginAttempt.success == False
    ).delete()
    
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.UPDATE,
        description=f"重置用户密码：{user.username}（ID: {user_id}）",
        request=req
    )
    
    return {
        "code": 200,
        "message": "密码已重置",
        "data": {
            "new_password": new_password
        }
    }


@router.post("/{user_id}/unlock")
def unlock_user_account(
    user_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """管理员解锁被锁定的用户账户（清除登录失败记录）"""
    if not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="仅超级管理员可执行此操作")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    deleted = db.query(LoginAttempt).filter(
        LoginAttempt.username == user.username,
        LoginAttempt.success == False
    ).delete()
    db.commit()

    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.UPDATE,
        description=f"解锁用户账户：{user.username}（ID: {user_id}），清除{deleted}条失败记录",
        request=req
    )

    return {
        "code": 200,
        "message": f"已解锁用户 {user.username}",
        "data": {"cleared": deleted}
    }


@router.get("/template")
def download_user_template(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """下载用户导入模板"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "用户导入模板"
    
    headers = ["用户名*", "邮箱*", "姓名", "手机号", "角色名称", "数据权限名称", "所属组织名称", "项目组名称", "状态(1启用/0禁用)"]
    ws.append(headers)
    
    ws.append(["zhangsan", "zhangsan@example.com", "张三", "13800138000", "研发工程师", "全部数据", "研发中心", "测试组", "1"])
    ws.append(["lisi", "lisi@example.com", "李四", "13900139000", "测试工程师", "本项目组数据", "测试中心", "功能测试组", "1"])
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    
    def iter_file(file: BytesIO):
        yield file.getvalue()
    
    return StreamingResponse(
        iter_file(output),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_import_template.xlsx"}
    )


class BatchImportRequest(BaseModel):
    overwrite: bool = False


import uuid
import threading

BATCH_SIZE = 50

_import_progress = {}
_import_lock = threading.Lock()

def set_progress(task_id: str, total: int, processed: int, status: str, message: str = ""):
    with _import_lock:
        _import_progress[task_id] = {
            "total": total,
            "processed": processed,
            "status": status,
            "message": message
        }

def get_progress(task_id: str):
    with _import_lock:
        return _import_progress.get(task_id, {"status": "not_found"})

@router.get("/batch-import/progress/{task_id}")
def get_import_progress(task_id: str):
    """获取导入进度"""
    progress = get_progress(task_id)
    return progress


@router.post("/batch-import")
async def batch_import_users(
    req: Request,
    file: UploadFile = File(...),
    overwrite: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """批量导入用户"""
    from io import BytesIO
    from openpyxl import load_workbook
    import asyncio
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件(.xlsx, .xls)")
    
    contents = await file.read()
    if len(contents) == 0:
        raise HTTPException(status_code=400, detail="文件为空")
    
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过10MB")
    
    try:
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        username_idx = next((i for i, h in enumerate(headers) if "用户名" in h and "*" in h), -1)
        email_idx = next((i for i, h in enumerate(headers) if "邮箱" in h and "*" in h), -1)
        full_name_idx = next((i for i, h in enumerate(headers) if "姓名" in h), -1)
        phone_idx = next((i for i, h in enumerate(headers) if "手机" in h), -1)
        role_idx = next((i for i, h in enumerate(headers) if "角色" in h), -1)
        position_tag_idx = next((i for i, h in enumerate(headers) if "数据权限" in h), -1)
        department_idx = next((i for i, h in enumerate(headers) if "所属组织" in h), -1)
        team_idx = next((i for i, h in enumerate(headers) if "项目组" in h), -1)
        status_idx = next((i for i, h in enumerate(headers) if "状态" in h), -1)
        
        if username_idx == -1 or email_idx == -1:
            raise HTTPException(status_code=400, detail="模板缺少必需列：用户名*、邮箱*")
        
        rows = list(ws.iter_rows(min_row=2))
        total_rows = len(rows)
        
        if total_rows > 5000:
            raise HTTPException(status_code=400, detail="单次导入最多支持5000条记录")
        
        task_id = str(uuid.uuid4())
        set_progress(task_id, total_rows, 0, "processing", "正在解析文件...")
        
        all_roles = {r.name: r.id for r in db.query(Role).all()}
        all_tags = {t.name: t.id for t in db.query(PositionTag).all()}
        all_departments = {d.name: d.id for d in db.query(Department).all()}
        all_teams = {t.name: t.id for t in db.query(Team).all()}
        all_users = {u.username: u for u in db.query(User).all()}
        
        success_count = 0
        error_rows = []
        
        for batch_start in range(0, total_rows, BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, total_rows)
            batch_rows = rows[batch_start:batch_end]
            
            try:
                for row_idx, row in enumerate(batch_rows, start=batch_start + 2):
                    try:
                        username = str(row[username_idx].value).strip() if row[username_idx].value else None
                        email = str(row[email_idx].value).strip() if row[email_idx].value else None
                        
                        if not username or not email:
                            error_rows.append({"row": row_idx, "error": "用户名或邮箱为空"})
                            continue
                        
                        existing_user = all_users.get(username)
                        if existing_user:
                            if not overwrite:
                                error_rows.append({"row": row_idx, "error": f"用户{username}已存在"})
                                continue
                            
                            existing_user.email = email
                            if full_name_idx != -1 and row[full_name_idx].value:
                                existing_user.full_name = str(row[full_name_idx].value).strip()
                            if phone_idx != -1 and row[phone_idx].value:
                                existing_user.phone = str(row[phone_idx].value).strip()
                            if status_idx != -1 and row[status_idx].value:
                                status_val = str(row[status_idx].value).strip()
                                existing_user.status = 1 if status_val == "1" else 0
                            
                            db_user = existing_user
                            is_update = True
                        else:
                            password = get_password_hash(email)
                            db_user = User(
                                username=username,
                                password=password,
                                email=email,
                                full_name=str(row[full_name_idx].value).strip() if full_name_idx != -1 and row[full_name_idx].value else None,
                                phone=str(row[phone_idx].value).strip() if phone_idx != -1 and row[phone_idx].value else None,
                                status=1,
                                must_change_password=True
                            )
                            db.add(db_user)
                            db.flush()
                            is_update = False
                        
                        if role_idx != -1 and row[role_idx].value:
                            role_name = str(row[role_idx].value).strip()
                            if role_name in all_roles:
                                from models import UserRole
                                if is_update:
                                    db.query(UserRole).filter(UserRole.user_id == db_user.id).delete()
                                db.add(UserRole(user_id=db_user.id, role_id=all_roles[role_name]))
                        
                        if position_tag_idx != -1 and row[position_tag_idx].value:
                            tag_name = str(row[position_tag_idx].value).strip()
                            if tag_name in all_tags:
                                db_user.position_tag_id = all_tags[tag_name]
                        
                        if department_idx != -1 and row[department_idx].value:
                            dept_name = str(row[department_idx].value).strip()
                            if dept_name in all_departments:
                                from models import UserDepartment
                                if is_update:
                                    db.query(UserDepartment).filter(UserDepartment.user_id == db_user.id).delete()
                                db.add(UserDepartment(user_id=db_user.id, department_id=all_departments[dept_name]))
                        
                        if team_idx != -1 and row[team_idx].value:
                            team_name = str(row[team_idx].value).strip()
                            if team_name in all_teams:
                                from models import UserTeam
                                if is_update:
                                    db.query(UserTeam).filter(UserTeam.user_id == db_user.id).delete()
                                db.add(UserTeam(user_id=db_user.id, team_id=all_teams[team_name]))
                        
                        success_count += 1
                    except Exception as e:
                        error_rows.append({"row": row_idx, "error": str(e)})
                
                db.commit()
                
                await asyncio.sleep(0)
                set_progress(task_id, total_rows, batch_end, "processing", f"已处理 {batch_end}/{total_rows} 条")
                
            except Exception as e:
                db.rollback()
                error_rows.append({"row": batch_start + 2, "error": f"批次提交失败: {str(e)}"})
        
        set_progress(task_id, total_rows, total_rows, "completed", f"导入完成，成功{success_count}条")
        
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.USERS,
            action=LogAction.CREATE,
            description=f"批量导入用户：成功{success_count}条，失败{len(error_rows)}条",
            request=req
        )
        
        return {
            "code": 200,
            "message": f"导入完成，成功{success_count}条{'，失败' + str(len(error_rows)) + '条' if error_rows else ''}",
            "data": {
                "task_id": task_id,
                "success": success_count,
                "errors": error_rows[:100]
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.delete("/{user_id}")
def delete_user(
    user_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    if is_super_admin(user):
        raise HTTPException(status_code=403, detail="超级管理员账号不能被删除")
    
    username = user.username
    
    # 删除用户角色关联
    db.query(UserRole).filter(UserRole.user_id == user_id).delete()
    
    # 删除用户项目关联
    db.query(UserProject).filter(UserProject.user_id == user_id).delete()
    
    # 删除用户团队关联
    db.query(UserTeam).filter(UserTeam.user_id == user_id).delete()
    
    # 删除用户部门关联
    db.query(UserDepartment).filter(UserDepartment.user_id == user_id).delete()
    
    # 删除用户项目组关联
    db.query(UserProjectGroup).filter(UserProjectGroup.user_id == user_id).delete()
    
    # 删除用户通知偏好
    db.query(UserNotificationPreference).filter(UserNotificationPreference.user_id == user_id).delete()
    
    # 清理 review_plan_testcases 中的 reviewer_id（设为NULL）
    db.query(ReviewPlanTestCase).filter(ReviewPlanTestCase.reviewer_id == user_id).update({"reviewer_id": None})
    
    # 清理 test_plans 中的 reviewer_id（设为NULL）
    db.query(TestPlan).filter(TestPlan.reviewer_id == user_id).update({"reviewer_id": None})
    
    # 清理 reports 中的 reviewed_by（设为NULL）
    db.query(Report).filter(Report.reviewed_by == user_id).update({"reviewed_by": None})
    
    # 清理 test_plan_executors 中的 executor_id（删除执行人记录）
    db.query(TestPlanExecutor).filter(TestPlanExecutor.executor_id == user_id).delete()
    
    # 清理 test_execution_progress 中的 user_id（设为NULL）
    db.query(TestExecutionProgress).filter(TestExecutionProgress.user_id == user_id).update({"user_id": None})
    
    # 清理 system_logs 中的用户操作记录
    from sqlalchemy import text
    db.execute(text("DELETE FROM system_logs WHERE user_id = :user_id"), {"user_id": user_id})
    
    # 清理测试执行记录中的 executor_id（设为NULL）
    db.query(TestExecution).filter(TestExecution.executor_id == user_id).update({"executor_id": None})
    
    # 删除用户创建的测试用例（将created_by设为NULL）
    db.query(TestCase).filter(TestCase.created_by == user_id).update({"created_by": None})
    
    # 删除用户更新的测试用例（将updated_by设为NULL）
    db.query(TestCase).filter(TestCase.updated_by == user_id).update({"updated_by": None})
    
    # 清理其他 created_by / updated_by 字段
    db.query(Module).filter(Module.created_by == user_id).update({"created_by": None})
    db.query(TestSuite).filter(TestSuite.created_by == user_id).update({"created_by": None})
    db.query(Comment).filter(Comment.author_id == user_id).delete()
    db.query(DingtalkBot).filter(DingtalkBot.created_by == user_id).update({"created_by": None})
    db.query(CaseTemplate).filter(CaseTemplate.created_by == user_id).update({"created_by": None})
    db.query(ReportTemplate).filter(ReportTemplate.created_by == user_id).update({"created_by": None})
    db.query(Notification).filter(Notification.sender_id == user_id).delete()
    db.query(NotificationRecipient).filter(NotificationRecipient.user_id == user_id).delete()
    db.query(DepartmentManager).filter(DepartmentManager.user_id == user_id).delete()
    db.query(TeamLeader).filter(TeamLeader.user_id == user_id).delete()
    
    # 删除用户
    db.delete(user)
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.USERS,
        action=LogAction.DELETE,
        description=f"删除用户：{username}（ID: {user_id}）",
        request=req
    )
    
    return {"code": 200, "message": "success", "data": None}
