"""
测试套件Excel导出工具

生成包含完整报告格式的Excel文件：
- Sheet 1 "Test Report": 封面(空值) + Test Result Detail + Zmind-PR(公式)
- 每个主模块一个Sheet
- 最后一个Sheet "Issues": 空模板供用户填写issue数据
所有统计单元格使用Excel公式，导出后可直接用于执行测试。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io
import re

ISSUES_SHEET_NAME = 'Issues'


def _styles():
    """共享样式"""
    header_font = Font(name='Arial', size=11, bold=True)
    return {
        'bold_font': Font(name='Arial', size=10, bold=True),
        'normal_font': Font(name='Arial', size=10),
        'header_font': header_font,
        'center_wrap': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'left_wrap': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'left_top_wrap': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'orange_fill': PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid'),
        'gray_fill': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
        'light_gray_fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
        'black_border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        ),
    }


def _safe_sheet_name(name):
    """将模块名转为合法的Sheet名称"""
    cleaned = re.sub(r'[\\/*?\[\]:]', '_', name or '未分类')
    return cleaned[:31] if len(cleaned) > 31 else (cleaned or '未分类')


def _build_unique_sheet_names(sorted_modules):
    """预先为所有模块生成唯一的Sheet名称"""
    reserved = {'Report', ISSUES_SHEET_NAME}
    name_map = {}
    for mod_name in sorted_modules:
        base = _safe_sheet_name(mod_name)
        unique = base
        counter = 2
        while unique in reserved:
            suffix = f'_{counter}'
            unique = base[:31 - len(suffix)] + suffix
            counter += 1
        reserved.add(unique)
        name_map[mod_name] = unique
    return name_map


def generate_suite_excel_stream(suite_name, module_cases_map, sorted_modules):
    """
    生成测试套件Excel到内存流。

    Args:
        suite_name: 套件名称
        module_cases_map: {主模块名: [用例列表]}
        sorted_modules: 排序后的主模块名列表

    Returns:
        BytesIO
    """
    s = _styles()
    wb = Workbook()
    sheet_name_map = _build_unique_sheet_names(sorted_modules)

    # ==================== Sheet 1: Report ====================
    ws = wb.active
    ws.title = 'Report'

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15

    row = 1

    # ---- 公司信息行 ----
    c = ws.cell(row=row, column=1)
    c.value = 'whale tv'
    c.font = Font(name='Arial', size=12, bold=True, color='FF0066')
    c.alignment = s['left_wrap']
    ws.merge_cells(f'E{row}:H{row}')
    c = ws.cell(row=row, column=5)
    c.value = 'Confidential & Proprietary\nCopyright © 2026 Whale TV Information Technology'
    c.font = Font(name='Arial', size=8, color='666666')
    c.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 1

    # ---- 报告标题 ----
    ws.merge_cells(f'A{row}:H{row}')
    c = ws[f'A{row}']
    c.value = f'{suite_name} Report'
    c.font = Font(name='Arial', size=18, bold=True)
    c.alignment = s['center_wrap']
    ws.row_dimensions[row].height = 35
    row += 1

    # ---- 封面表格（标签有值，内容留空） ----
    cover_labels = [
        ('项目名称\nProject name', ''),
        ('测试周期\nTest Cycle', ''),
        ('测试人员\nTesters', ''),
        ('审核人员\nReviewer', ''),
        ('验证环境\nVerified environment', ''),
        ('提测内容\nRelease Note', ''),
        ('测试结论\nTest Conclusion', ''),
        ('测试通过标准\nRelease Criteria',
         'Testcase pass rate >=95% & PR closure rate: Blocker=100%, Critical>=95%, Major>=90%'),
        ('风险评估\nRisk Assessment', ''),
        ('备注\nRemark', ''),
    ]
    for label, default_val in cover_labels:
        ws.merge_cells(f'B{row}:H{row}')
        ca = ws[f'A{row}']
        ca.value = label
        ca.font = s['header_font']
        ca.alignment = s['center_wrap']
        ca.fill = s['orange_fill']
        ca.border = s['black_border']
        cb = ws[f'B{row}']
        cb.value = default_val
        cb.font = s['normal_font']
        cb.alignment = s['left_top_wrap']
        cb.border = s['black_border']
        for col in range(3, 9):
            ws.cell(row=row, column=col).border = s['black_border']
        ws.row_dimensions[row].height = 30
        row += 1

    # 空行
    row += 1

    # ---- Test Result Detail ----
    ws.merge_cells(f'A{row}:H{row}')
    c = ws[f'A{row}']
    c.value = 'Test Result Detail'
    c.font = Font(name='Arial', size=14, bold=True)
    c.alignment = s['center_wrap']
    c.fill = s['light_gray_fill']
    ws.row_dimensions[row].height = 30
    row += 1

    # 表头
    headers = ['Module', 'TestCases', 'PASS', 'FAIL', 'BLOCK', 'NT', 'NA', 'Passing rate']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci)
        c.value = h
        c.font = s['bold_font']
        c.alignment = s['center_wrap']
        c.fill = s['gray_fill']
        c.border = s['black_border']
    row += 1

    data_start_row = row

    for mod_name in sorted_modules:
        cases = module_cases_map.get(mod_name, [])
        sn = sheet_name_map[mod_name]
        case_count = len(cases)

        c = ws.cell(row=row, column=1)
        c.value = mod_name
        c.font = s['normal_font']
        c.alignment = s['center_wrap']
        c.fill = s['gray_fill']
        c.border = s['black_border']

        # TestCases = 总数 - NA
        c = ws.cell(row=row, column=2)
        c.value = f'={case_count}-G{row}'
        c.font = s['normal_font']
        c.alignment = s['center_wrap']
        c.border = s['black_border']

        safe_ref = f"'{sn}'!$H:$H"
        for ci, label in [(3, 'PASS'), (4, 'FAIL'), (5, 'BLOCK'), (6, 'NT'), (7, 'NA')]:
            c = ws.cell(row=row, column=ci)
            c.value = f'=COUNTIF({safe_ref},"{label}")'
            c.font = s['normal_font']
            c.alignment = s['center_wrap']
            c.border = s['black_border']

        c = ws.cell(row=row, column=8)
        c.value = f'=IF(B{row}=0,0,C{row}/B{row})'
        c.font = s['normal_font']
        c.alignment = s['center_wrap']
        c.border = s['black_border']
        c.number_format = '0.00%'
        row += 1

    data_end_row = row - 1

    # Total行
    c = ws.cell(row=row, column=1)
    c.value = 'Total'
    c.font = s['bold_font']
    c.alignment = s['center_wrap']
    c.fill = s['gray_fill']
    c.border = s['black_border']
    for ci in range(2, 8):
        cl = get_column_letter(ci)
        c = ws.cell(row=row, column=ci)
        c.value = f'=SUM({cl}{data_start_row}:{cl}{data_end_row})'
        c.font = s['bold_font']
        c.alignment = s['center_wrap']
        c.border = s['black_border']
    c = ws.cell(row=row, column=8)
    c.value = f'=IF(B{row}=0,0,C{row}/B{row})'
    c.font = s['bold_font']
    c.alignment = s['center_wrap']
    c.border = s['black_border']
    c.number_format = '0.00%'
    total_row = row
    row += 1

    # Test Case Passed 行
    c = ws.cell(row=row, column=1)
    c.value = 'Test Case Passed'
    c.font = s['bold_font']
    c.alignment = s['center_wrap']
    c.fill = s['gray_fill']
    c.border = s['black_border']
    # B: TestCases = PASS% + FAIL% + BLOCK% + NT% (即 C+D+E+F 本行的值之和)
    tcp_row = row  # 记录当前行号供公式引用
    c = ws.cell(row=row, column=2)
    c.value = f'=C{tcp_row}+D{tcp_row}+E{tcp_row}+F{tcp_row}'
    c.font = s['bold_font']
    c.alignment = s['center_wrap']
    c.border = s['black_border']
    c.number_format = '0.00%'
    for ci, src in [(3, 'C'), (4, 'D'), (5, 'E'), (6, 'F')]:
        c = ws.cell(row=row, column=ci)
        c.value = f'=IF(B{total_row}=0,0,{src}{total_row}/B{total_row})'
        c.font = s['bold_font']
        c.alignment = s['center_wrap']
        c.border = s['black_border']
        c.number_format = '0.00%'
    c = ws.cell(row=row, column=7)
    c.value = ''
    c.border = s['black_border']
    c = ws.cell(row=row, column=8)
    c.value = f'=IF(H{total_row}>=0.95,"OK","NG")'
    c.font = s['bold_font']
    c.alignment = s['center_wrap']
    c.border = s['black_border']
    row += 1

    # 说明行
    ws.merge_cells(f'A{row}:H{row}')
    c = ws[f'A{row}']
    c.value = '测试结果说明：Pass-测试通过；Fail-测试失败；NT-暂未测试；NA-不适用；'
    c.font = Font(name='Arial', size=9, color='666666')
    c.alignment = s['left_wrap']
    c.fill = s['gray_fill']
    row += 1

    # ---- Zmind-PR 统计（公式引用 Issues Sheet） ----
    # Issues Sheet 结构: A=#(PR号), B=跟踪, C=类别, D=Severity, E=状态, F=优先级, G=主题, H=指派给
    # Severity 值: Blocker/Critical/Major/Minor/Enhancement
    # 状态 值: Open/Closed
    iss = f"'{ISSUES_SHEET_NAME}'"
    row += 1  # 空行

    # 表头
    zmind_headers = ['Zmind-PR', 'PRs', 'Blocker', 'Critical', 'Major', 'Minor', 'Enhancement', 'Conclusion']
    for ci, h in enumerate(zmind_headers, 1):
        c = ws.cell(row=row, column=ci)
        c.value = h
        c.font = s['bold_font']
        c.alignment = s['center_wrap']
        c.fill = s['gray_fill']
        c.border = s['black_border']
    row += 1

    zmind_start_row = row

    # Open行 — 统计状态为Open的各Severity数量 (D=Severity, E=状态)
    open_labels = ['Open',
                   f'=COUNTIF({iss}!$E:$E,"Open")',
                   f'=COUNTIFS({iss}!$D:$D,"Blocker",{iss}!$E:$E,"Open")',
                   f'=COUNTIFS({iss}!$D:$D,"Critical",{iss}!$E:$E,"Open")',
                   f'=COUNTIFS({iss}!$D:$D,"Major",{iss}!$E:$E,"Open")',
                   f'=COUNTIFS({iss}!$D:$D,"Minor",{iss}!$E:$E,"Open")',
                   f'=COUNTIFS({iss}!$D:$D,"Enhancement",{iss}!$E:$E,"Open")',
                   '']
    for ci, val in enumerate(open_labels, 1):
        c = ws.cell(row=row, column=ci)
        c.value = val
        c.font = s['normal_font']
        c.alignment = s['center_wrap']
        c.border = s['black_border']
    ws.cell(row=row, column=1).fill = s['gray_fill']
    open_row = row
    row += 1

    # Total行 — 统计各Severity总数
    total_labels = ['Total',
                    f'=COUNTA({iss}!$A:$A)-1',
                    f'=COUNTIF({iss}!$D:$D,"Blocker")',
                    f'=COUNTIF({iss}!$D:$D,"Critical")',
                    f'=COUNTIF({iss}!$D:$D,"Major")',
                    f'=COUNTIF({iss}!$D:$D,"Minor")',
                    f'=COUNTIF({iss}!$D:$D,"Enhancement")',
                    '']
    for ci, val in enumerate(total_labels, 1):
        c = ws.cell(row=row, column=ci)
        c.value = val
        c.font = s['normal_font']
        c.alignment = s['center_wrap']
        c.border = s['black_border']
    ws.cell(row=row, column=1).fill = s['gray_fill']
    total_zmind_row = row
    row += 1

    # PR closed行 — 关闭率公式（与平台逻辑一致）
    c = ws.cell(row=row, column=1)
    c.value = 'PR closed'
    c.font = s['normal_font']
    c.alignment = s['center_wrap']
    c.fill = s['gray_fill']
    c.border = s['black_border']
    c = ws.cell(row=row, column=2)
    c.value = f'=IF(B{total_zmind_row}=0,1,(B{total_zmind_row}-B{open_row})/B{total_zmind_row})'
    c.font = s['normal_font']
    c.alignment = s['center_wrap']
    c.border = s['black_border']
    c.number_format = '0.00%'
    for ci in range(3, 8):
        cl = get_column_letter(ci)
        c = ws.cell(row=row, column=ci)
        c.value = f'=IF({cl}{total_zmind_row}=0,1,({cl}{total_zmind_row}-{cl}{open_row})/{cl}{total_zmind_row})'
        c.font = s['normal_font']
        c.alignment = s['center_wrap']
        c.border = s['black_border']
        c.number_format = '0.00%'
    c = ws.cell(row=row, column=8)
    c.value = ''
    c.border = s['black_border']
    closed_row = row
    row += 1

    # 合并Conclusion列（3行）并用公式判断OK/NG
    # 逻辑：Blocker关闭率=100% AND Critical关闭率>=95% AND Major关闭率>=90%（与Release Criteria一致）
    ws.merge_cells(f'H{zmind_start_row}:H{zmind_start_row + 2}')
    cc = ws.cell(row=zmind_start_row, column=8)
    cc.value = f'=IF(AND(C{closed_row}>=1,D{closed_row}>=0.95,E{closed_row}>=0.9),"OK","NG")'
    cc.font = s['bold_font']
    cc.alignment = s['center_wrap']
    cc.border = s['black_border']

    # Severity说明
    ws.merge_cells(f'A{row}:H{row}')
    c = ws[f'A{row}']
    c.value = 'Severity：Blocker-致命性或阻塞进度的问题；Critical－关键功能及稳定性问题；Major－次要功能问题；Minor-界面或优化问题；Enhancement-增强建议'
    c.font = Font(name='Arial', size=9, color='666666')
    c.alignment = s['left_wrap']
    c.fill = s['gray_fill']

    # ==================== 每个主模块一个Sheet ====================
    for mod_name in sorted_modules:
        cases = module_cases_map.get(mod_name, [])
        sn = sheet_name_map[mod_name]
        ms = wb.create_sheet(title=sn)

        ms.column_dimensions['A'].width = 15
        ms.column_dimensions['B'].width = 18
        ms.column_dimensions['C'].width = 12
        ms.column_dimensions['D'].width = 30
        ms.column_dimensions['E'].width = 25
        ms.column_dimensions['F'].width = 35
        ms.column_dimensions['G'].width = 30
        ms.column_dimensions['H'].width = 12
        ms.column_dimensions['I'].width = 10
        ms.column_dimensions['J'].width = 25
        ms.column_dimensions['K'].width = 30

        tc_headers = ['用例编号', '所属模块', '子模块', '用例标题', '前置条件',
                      '操作步骤', '预期结果', '测试结果', '用例等级', '执行备注', '注意事项']
        for ci, h in enumerate(tc_headers, 1):
            c = ms.cell(row=1, column=ci)
            c.value = h
            c.font = s['bold_font']
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            c.fill = s['gray_fill']
            c.border = s['black_border']

        for ri, tc in enumerate(cases, 2):
            row_data = [
                tc.get('case_number', ''),
                tc.get('module', ''),
                tc.get('sub_module', ''),
                tc.get('name', ''),
                tc.get('precondition', ''),
                tc.get('steps', ''),
                tc.get('expected_result', ''),
                '',  # 测试结果
                tc.get('level', ''),
                '',  # 执行备注
                tc.get('remarks', ''),  # 注意事项（用例备注）
            ]
            max_lines = 1
            for ci, val in enumerate(row_data, 1):
                c = ms.cell(row=ri, column=ci)
                c.value = val
                c.font = s['normal_font']
                c.alignment = s['left_top_wrap']
                c.border = s['black_border']
                val_str = str(val) if val else ''
                lines = val_str.count('\n') + 1
                col_w = ms.column_dimensions[get_column_letter(ci)].width or 15
                for line in val_str.split('\n'):
                    lines += max(0, len(line) // int(col_w))
                max_lines = max(max_lines, lines)
            ms.row_dimensions[ri].height = max(15, max_lines * 15)

        if cases:
            dv = DataValidation(type='list', formula1='"PASS,FAIL,BLOCK,NT,NA"', allow_blank=True)
            dv.error = '请选择有效的测试结果'
            dv.errorTitle = '无效输入'
            dv.prompt = '选择测试结果'
            dv.promptTitle = '测试结果'
            dv.add(f'H2:H{1 + len(cases)}')
            ms.add_data_validation(dv)

    # ==================== 最后一页: Issues ====================
    iss_ws = wb.create_sheet(title=ISSUES_SHEET_NAME)
    iss_ws.column_dimensions['A'].width = 14
    iss_ws.column_dimensions['B'].width = 14
    iss_ws.column_dimensions['C'].width = 14
    iss_ws.column_dimensions['D'].width = 18
    iss_ws.column_dimensions['E'].width = 12
    iss_ws.column_dimensions['F'].width = 12
    iss_ws.column_dimensions['G'].width = 50
    iss_ws.column_dimensions['H'].width = 16

    iss_headers = ['#', '跟踪', '类别', 'Severity', '状态', '优先级', '主题', '指派给']
    for ci, h in enumerate(iss_headers, 1):
        c = iss_ws.cell(row=1, column=ci)
        c.value = h
        c.font = s['bold_font']
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        c.fill = s['gray_fill']
        c.border = s['black_border']

    # Severity下拉验证 (D列)
    sev_dv = DataValidation(type='list', formula1='"Blocker,Critical,Major,Minor,Enhancement"', allow_blank=True)
    sev_dv.prompt = '选择严重等级'
    sev_dv.promptTitle = 'Severity'
    sev_dv.add('D2:D1000')
    iss_ws.add_data_validation(sev_dv)

    # 状态下拉验证 (E列)
    status_dv = DataValidation(type='list', formula1='"Open,Closed"', allow_blank=True)
    status_dv.prompt = '选择状态'
    status_dv.promptTitle = '状态'
    status_dv.add('E2:E1000')
    iss_ws.add_data_validation(status_dv)

    # 输出
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
