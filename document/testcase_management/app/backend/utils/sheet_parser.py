"""
Sheet解析工具模块
V1.3 - 多Sheet处理功能
"""
import warnings
from typing import List, Dict, Optional, Any
from dataclasses import dataclass
from io import BytesIO
import pandas as pd

# 过滤 openpyxl 的样式警告
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


# 用例字段关键词库
FIELD_KEYWORDS = [
    "用例编号", "用例标题", "测试步骤", "操作步骤", "预期结果",
    "前置条件", "用例等级", "模块", "所属用例库", "优先级",
    "Case ID", "Test Case", "Steps", "Expected Result", "Title",
    "编号", "标题", "步骤", "结果", "等级"
]

# Sheet名称关键词
SHEET_NAME_KEYWORDS = ["用例", "case", "test", "测试", "Case", "Test"]


@dataclass
class SheetInfo:
    """Sheet信息"""
    name: str
    row_count: int
    column_count: int
    preview_rows: List[List[Any]]
    auto_recommend: bool = False
    recommend_reason: Optional[str] = None
    score: int = 0
    detected_header_row: int = 1


def parse_all_sheets(file_content: bytes, file_name: str, preview_rows: int = 10) -> List[SheetInfo]:
    """
    解析Excel文件的所有Sheet
    
    Args:
        file_content: 文件内容
        file_name: 文件名
        preview_rows: 预览行数，默认10行
    
    Returns:
        Sheet信息列表
    """
    try:
        # 根据文件类型选择引擎
        is_xlsx = file_name.lower().endswith('.xlsx')
        
        if is_xlsx:
            # 使用 openpyxl 直接读取，更高效获取行数
            from openpyxl import load_workbook
            
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
            sheets = []
            
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    
                    # 读取预览行并同时计算实际行列数
                    preview_data = []
                    actual_row_count = 0
                    actual_col_count = 0
                    
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                        # 检查是否为空行
                        row_has_data = any(cell is not None and str(cell).strip() for cell in row)
                        
                        if row_has_data:
                            actual_row_count = row_idx + 1
                            # 计算实际列数
                            for col_idx, cell in enumerate(row):
                                if cell is not None and str(cell).strip():
                                    actual_col_count = max(actual_col_count, col_idx + 1)
                        
                        # 只保存预览行
                        if row_idx < preview_rows:
                            cleaned_row = []
                            for cell in row:
                                if cell is None:
                                    cleaned_row.append('')
                                elif isinstance(cell, float) and cell == int(cell):
                                    cleaned_row.append(int(cell))
                                else:
                                    cleaned_row.append(str(cell) if not isinstance(cell, (int, float)) else cell)
                            preview_data.append(cleaned_row)
                    
                    # 检测表头行
                    detected_header = detect_header_row(preview_data) if preview_data else 1
                    
                    sheet_info = SheetInfo(
                        name=sheet_name,
                        row_count=actual_row_count,
                        column_count=actual_col_count,
                        preview_rows=preview_data,
                        detected_header_row=detected_header
                    )
                    sheets.append(sheet_info)
                    
                except Exception as e:
                    sheets.append(SheetInfo(
                        name=sheet_name,
                        row_count=0,
                        column_count=0,
                        preview_rows=[],
                        recommend_reason=f"解析失败: {str(e)}",
                        detected_header_row=1
                    ))
            
            wb.close()
            return sheets
        else:
            # .xls 文件使用 xlrd
            engine = 'xlrd'
            excel_file = pd.ExcelFile(BytesIO(file_content), engine=engine)
            sheet_names = excel_file.sheet_names
            
            sheets = []
            for sheet_name in sheet_names:
                try:
                    # 使用 xlrd 直接获取行数
                    import xlrd
                    xls_book = xlrd.open_workbook(file_contents=file_content)
                    xls_sheet = xls_book.sheet_by_name(sheet_name)
                    
                    # 计算实际有数据的行列数
                    actual_row_count = 0
                    actual_col_count = 0
                    preview_data = []
                    
                    for row_idx in range(xls_sheet.nrows):
                        row = xls_sheet.row_values(row_idx)
                        row_has_data = any(str(cell).strip() for cell in row)
                        
                        if row_has_data:
                            actual_row_count = row_idx + 1
                            for col_idx, cell in enumerate(row):
                                if str(cell).strip():
                                    actual_col_count = max(actual_col_count, col_idx + 1)
                        
                        # 只保存预览行
                        if row_idx < preview_rows:
                            cleaned_row = []
                            for cell in row:
                                if cell == '':
                                    cleaned_row.append('')
                                elif isinstance(cell, float) and cell == int(cell):
                                    cleaned_row.append(int(cell))
                                else:
                                    cleaned_row.append(str(cell) if not isinstance(cell, (int, float)) else cell)
                            preview_data.append(cleaned_row)
                    
                    # 检测表头行
                    detected_header = detect_header_row(preview_data) if preview_data else 1
                    
                    sheet_info = SheetInfo(
                        name=sheet_name,
                        row_count=actual_row_count,
                        column_count=actual_col_count,
                        preview_rows=preview_data,
                        detected_header_row=detected_header
                    )
                    sheets.append(sheet_info)
                    
                except Exception as e:
                    sheets.append(SheetInfo(
                        name=sheet_name,
                        row_count=0,
                        column_count=0,
                        preview_rows=[],
                        recommend_reason=f"解析失败: {str(e)}",
                        detected_header_row=1
                    ))
            
            return sheets
        
    except Exception as e:
        raise ValueError(f"解析Excel文件失败: {str(e)}")


def calculate_sheet_score(sheet: SheetInfo) -> int:
    """
    计算Sheet的推荐分数
    
    Args:
        sheet: Sheet信息
    
    Returns:
        推荐分数
    """
    score = 0
    reasons = []
    
    # 1. 检查Sheet名称关键词
    for keyword in SHEET_NAME_KEYWORDS:
        if keyword.lower() in sheet.name.lower():
            score += 10
            reasons.append(f"名称包含'{keyword}'")
            break
    
    # 2. 检查表头匹配度
    if sheet.preview_rows:
        # 检查前3行是否包含关键词
        for row_idx, row in enumerate(sheet.preview_rows[:3]):
            row_text = ' '.join(str(cell) for cell in row).lower()
            matched_keywords = 0
            for keyword in FIELD_KEYWORDS:
                if keyword.lower() in row_text:
                    matched_keywords += 1
            
            if matched_keywords > 0:
                # 第一行匹配权重更高
                weight = 5 if row_idx == 0 else 3
                score += matched_keywords * weight
                if matched_keywords >= 3:
                    reasons.append(f"第{row_idx + 1}行匹配{matched_keywords}个用例字段")
    
    # 3. 检查数据行数量
    if sheet.row_count > 5:
        score += min(sheet.row_count // 10, 20)  # 最多加20分
        if sheet.row_count > 10:
            reasons.append(f"包含{sheet.row_count}行数据")
    
    # 4. 检查列数量（用例通常有多列）
    if 5 <= sheet.column_count <= 20:
        score += 10
    
    sheet.score = score
    if reasons:
        sheet.recommend_reason = "; ".join(reasons[:3])  # 最多3个原因
    
    return score


def detect_recommended_sheet(sheets: List[SheetInfo]) -> Optional[str]:
    """
    检测推荐的Sheet
    
    Args:
        sheets: Sheet列表
    
    Returns:
        推荐的Sheet名称，如果无法确定则返回None
    """
    if not sheets:
        return None
    
    if len(sheets) == 1:
        sheets[0].auto_recommend = True
        sheets[0].recommend_reason = "唯一Sheet"
        return sheets[0].name
    
    # 计算每个Sheet的分数
    for sheet in sheets:
        calculate_sheet_score(sheet)
    
    # 按分数排序
    sorted_sheets = sorted(sheets, key=lambda s: s.score, reverse=True)
    
    # 如果最高分大于0，则推荐
    if sorted_sheets[0].score > 0:
        sorted_sheets[0].auto_recommend = True
        return sorted_sheets[0].name
    
    return None


def detect_header_row(preview_rows: List[List[Any]], max_check_rows: int = 10) -> int:
    """
    检测表头所在行
    
    Args:
        preview_rows: 预览数据
        max_check_rows: 最大检查行数
    
    Returns:
        表头行号（从1开始）
    """
    if not preview_rows:
        return 1
    
    best_row = 1
    best_score = 0
    
    for row_idx, row in enumerate(preview_rows[:max_check_rows]):
        score = 0
        row_text = ' '.join(str(cell) for cell in row).lower()
        
        # 检查关键词匹配
        for keyword in FIELD_KEYWORDS:
            if keyword.lower() in row_text:
                score += 1
        
        # 检查是否有非空单元格
        non_empty = sum(1 for cell in row if str(cell).strip())
        if non_empty >= 3:
            score += 1
        
        # 检查是否像表头（多为文本，少数字）
        text_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
        if text_count >= len(row) * 0.5:
            score += 1
        
        if score > best_score:
            best_score = score
            best_row = row_idx + 1
    
    return best_row


def get_sheet_data_from_row(
    file_content: bytes, 
    file_name: str, 
    sheet_name: str, 
    header_row: int = 1
) -> tuple:
    """
    从指定行开始获取Sheet数据
    
    Args:
        file_content: 文件内容
        file_name: 文件名
        sheet_name: Sheet名称
        header_row: 表头行号（从1开始）
    
    Returns:
        (headers, data_rows)
    """
    try:
        if file_name.endswith('.xlsx'):
            engine = 'openpyxl'
        else:
            engine = 'xlrd'
        
        # 读取指定Sheet，跳过表头之前的行
        df = pd.read_excel(
            BytesIO(file_content),
            sheet_name=sheet_name,
            engine=engine,
            header=header_row - 1  # pandas的header是0-indexed
        )
        
        # 处理NaN值
        df = df.fillna('')
        
        # 获取表头
        headers = [str(col).strip() for col in df.columns]
        
        # 获取数据行
        data_rows = df.to_dict('records')
        
        # 清理数据
        cleaned_rows = []
        for row in data_rows:
            cleaned_row = {}
            for key, value in row.items():
                clean_key = str(key).strip()
                if pd.isna(value):
                    cleaned_row[clean_key] = ''
                elif isinstance(value, float) and value == int(value):
                    cleaned_row[clean_key] = int(value)
                else:
                    cleaned_row[clean_key] = str(value).strip() if isinstance(value, str) else value
            cleaned_rows.append(cleaned_row)
        
        return headers, cleaned_rows
        
    except Exception as e:
        raise ValueError(f"读取Sheet数据失败: {str(e)}")
