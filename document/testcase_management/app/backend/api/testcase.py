from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_
from typing import Optional
from database import get_db
from models import TestCase, User, TestCaseZmindLink, TestCaseProject, Module, TestCaseHistory, Project, TeamProject, TestPlanTestCase, TestExecution, TestCaseAttachment, ReviewPlanTestCase, TestExecutionProgress, TestExecutionAttachment
from schemas import TestCaseCreate, TestCaseCreateBatch
from auth import get_current_user, has_permission, is_super_admin
from utils.excel import generate_excel
from utils.logger import log_operation, LogAction, LogModule
from utils.notification_helper import trigger_testcase_notification
from utils.constants import CLOSED_STATUSES
from pydantic import BaseModel
import io
import re
import logging
from datetime import datetime

router = APIRouter()
logger = logging.getLogger(__name__)


def generate_case_number(db: Session, project_id: int, module_path: str = None, seq_cache: dict = None) -> str:
    """
    自动生成用例编号
    格式: {用例库Tag}+{模块Tag}+{4位数字序号}
    若模块未设置Tag，则格式为: {用例库Tag}+{4位数字序号}
    规则：
    - 有Tag的模块：按模块的 sort_order 顺序递增，每个子模块的编号连续
    - 没有Tag的模块：按用例库维度递增
    - seq_cache: 可选的序号缓存字典，用于批量导入时避免重复编号
                   格式: {project_id: {prefix: max_seq}}
    """
    new_seq = 1
    prefix = ""
    
    # 获取用例库的Tag
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=400, detail="用例库不存在")
    
    project_tag = project.tag
    if not project_tag:
        project_tag = re.sub(r'[^a-zA-Z0-9]', '', project.name)
        if not project_tag:
            project_tag = "PRJ"
    
    # 获取模块的Tag
    module_tag = None
    current_module_name = None
    
    if module_path:
        module_path = re.sub(r'\s+', ' ', module_path.strip())
        module_parts = module_path.split('/')
        main_module_name = module_parts[0] if module_parts else None
        
        if main_module_name:
            main_module_name = re.sub(r'\s+', ' ', main_module_name.strip())
            main_modules = db.query(Module).filter(
                Module.project_id == project_id,
                Module.name == main_module_name,
                Module.parent_id.is_(None)
            ).all()
            
            main_module = None
            for m in main_modules:
                if m.tag:
                    main_module = m
                    break
            if not main_module and main_modules:
                main_module = main_modules[0]
            
            if main_module:
                module_tag = main_module.tag
                current_module_name = main_module_name
    
    # 确定基础前缀并查询最大序号
    if module_tag and current_module_name:
        prefix = f"{project_tag}_{module_tag}_"
        
        main_module = db.query(Module).filter(
            Module.project_id == project_id,
            Module.name == current_module_name,
            Module.parent_id.is_(None)
        ).first()
        
        if main_module:
            sub_modules = db.query(Module).filter(
                Module.parent_id == main_module.id
            ).order_by(Module.sort_order).all()
            
            sub_module_name = None
            if '/' in module_path:
                # 提取子模块名并清理空格
                sub_path_full = module_path.split('/')[1] if len(module_path.split('/')) > 1 else None
                if sub_path_full:
                    sub_module_name = re.sub(r'\s+', ' ', sub_path_full.strip())
            
            start_seq = 1
            for sm in sub_modules:
                if sub_module_name and sm.name == sub_module_name:
                    break
                sub_path = f"{current_module_name}/{sm.name}"
                # 统计该子模块及其所有深层子模块的用例
                count = db.query(func.count(TestCase.id)).filter(
                    TestCase.primary_project_id == project_id,
                    or_(
                        TestCase.module == sub_path,
                        TestCase.module.like(f"{sub_path}/%")
                    )
                ).scalar() or 0
                start_seq += count
            
            # 先从缓存获取序号（用于批量导入时避免重复）
            cache_key = f"{project_id}_{prefix}"
            cached_max = 0
            if seq_cache and project_id in seq_cache and prefix in seq_cache[project_id]:
                cached_max = seq_cache[project_id][prefix]
            
            # 判断是主模块还是子模块/中间模块
            # module_path == current_module_name 表示主模块（如"UI交互"）
            # module_path == current_module_name/子模块名 表示直接子模块（如"UI交互/Setting"）
            # module_path 包含更多层级 表示深层子模块（如"UI交互/Setting/Network & Connection"）
            is_main_module = (module_path == current_module_name)
            is_direct_submodule = False
            if not is_main_module and '/' in module_path:
                parts = module_path.split('/')
                if len(parts) == 2:
                    is_direct_submodule = True
            
            # 查询最大序号，确保新序号唯一
            if is_main_module:
                # 主模块：查询该前缀下的全局最大序号
                max_seq = db.query(func.max(TestCase.case_number)).filter(
                    TestCase.primary_project_id == project_id,
                    TestCase.case_number.like(f"{prefix}%")
                ).scalar()
            elif is_direct_submodule:
                # 直接子模块：查询该前缀下的全局最大序号
                max_seq = db.query(func.max(TestCase.case_number)).filter(
                    TestCase.primary_project_id == project_id,
                    TestCase.case_number.like(f"{prefix}%")
                ).scalar()
            else:
                # 深层子模块：查询当前模块及其所有子模块的最大序号
                max_seq = db.query(func.max(TestCase.case_number)).filter(
                    TestCase.primary_project_id == project_id,
                    or_(
                        TestCase.module == module_path,
                        TestCase.module.like(f"{module_path}/%")
                    )
                ).scalar()
            
            if max_seq:
                match = re.search(r'(\d+)$', max_seq)
                if match:
                    current_max = int(match.group(1))
                    new_seq = max(current_max + 1, cached_max + 1, start_seq)
                else:
                    new_seq = max(start_seq, cached_max + 1)
            else:
                new_seq = max(start_seq, cached_max + 1)
            
            # 更新缓存
            if seq_cache is not None:
                if project_id not in seq_cache:
                    seq_cache[project_id] = {}
                seq_cache[project_id][prefix] = new_seq
    else:
        prefix = f"{project_tag}_"
        
        # 先从缓存获取序号（用于批量导入时避免重复）
        cache_key = f"{project_id}_{prefix}"
        cached_max = 0
        if seq_cache and project_id in seq_cache and prefix in seq_cache[project_id]:
            cached_max = seq_cache[project_id][prefix]
        
        all_cases = db.query(TestCase.case_number).filter(
            TestCase.primary_project_id == project_id,
            TestCase.case_number.like(f"{prefix}%")
        ).all()
        
        max_seq_num = 0
        tag_str = str(project_tag)
        for cn in all_cases:
            cn_str = cn[0] if cn else ''
            match = re.match(r'^' + tag_str + r'_(\d+)$', cn_str)
            if match:
                seq_num = int(match.group(1))
                if seq_num > max_seq_num:
                    max_seq_num = seq_num
        
        current_max = max_seq_num
        new_seq = max(current_max + 1, cached_max + 1) if current_max > 0 else max(1, cached_max + 1)
        
        # 更新缓存
        if seq_cache is not None:
            if project_id not in seq_cache:
                seq_cache[project_id] = {}
            seq_cache[project_id][prefix] = new_seq
    
    if not new_seq:
        new_seq = 1
    
    case_number = f"{prefix}{str(new_seq).zfill(4)}"
    
    return case_number


def build_module_sort_key(db, project_ids):
    """构建模块路径排序的SQLAlchemy case()表达式（委托给共享工具）"""
    from utils.module_sort import build_module_sort_key_expr
    return build_module_sort_key_expr(db, project_ids)


def apply_testcase_sort(query, project_id, db=None):
    """应用测试用例的标准排序规则 - 基于模块sort_order排序"""
    from sqlalchemy import text, cast, Integer
    
    if db and project_id:
        pid_list = [project_id] if isinstance(project_id, int) else project_id
        module_sort_expr = build_module_sort_key(db, pid_list)
        case_num_sort = cast(text("SUBSTRING(test_cases.case_number FROM '([0-9]+)$')"), Integer).asc()
        query = query.order_by(
            module_sort_expr,
            TestCase.sort_order.asc(),
            case_num_sort,
            TestCase.id.asc()
        )
    else:
        case_num_sort = cast(text("SUBSTRING(test_cases.case_number FROM '([0-9]+)$')"), Integer).asc()
        query = query.order_by(
            TestCase.module.asc().nullslast(),
            TestCase.sort_order.asc(),
            case_num_sort,
            TestCase.id.asc()
        )
    
    return query


def generate_module_tag(db: Session, project_id: int, module_name: str):
    """自动生成模块Tag：4位大写字母
    
    逻辑：
    - 按空格或斜杠分割多个部分
    - 每个部分取首字母（英文单词首字母或中文拼音首字母）
    - 组合成4位Tag
    """
    # 提取模块名称中的字母，转换为大写
    letters = re.findall(r'[a-zA-Z]', module_name)
    if letters:
        # 取前4个字母并大写
        tag = ''.join(letters[:4]).upper()
    else:
        # 按空格或斜杠分割处理多部分
        parts = re.split(r'[/\\]', module_name)
        if len(parts) == 1:
            parts = module_name.split()
        
        # 拼音首字母映射
        pinyin_map = {
            '用': 'Y', '户': 'H', '管': 'G', '理': 'L', '系': 'X', '统': 'T',
            '测': 'C', '试': 'S', '平': 'P', '台': 'T', '订': 'D', '单': 'O',
            '商': 'S', '城': 'C', '消': 'X', '费': 'F', '支': 'Z', '付': 'F',
            '账': 'Z', '中': 'Z', '心': 'X', '会': 'H', '员': 'Y', '卡': 'K',
            '优': 'Y', '惠': 'H', '券': 'Q', '活': 'H', '动': 'D', '数': 'S',
            '据': 'J', '分': 'F', '析': 'X', '报': 'B', '告': 'G', '审': 'S',
            '批': 'P', '通': 'T', '知': 'Z', '维': 'W', '护': 'H', '日': 'R',
            '志': 'Z', '配': 'P', '置': 'Z', '权': 'Q', '限': 'X', '角': 'J',
            '登': 'D', '录': 'L', '注': 'Z', '册': 'C', '注': 'Z', '销': 'X',
            '查': 'C', '看': 'K', '编': 'B', '辑': 'J', '删': 'S', '改': 'G',
            '添': 'T', '加': 'J', '列': 'L', '表': 'B', '详': 'X', '细': 'X',
            '设': 'S', '置': 'Z', '功': 'G', '能': 'N', '权': 'Q', '限': 'X',
            '菜': 'C', '单': 'D', '首': 'S', '页': 'Y', '关': 'G', '于': 'Y',
            '帮': 'B', '助': 'Z', '确': 'Q', '认': 'R', '取': 'Q', '消': 'X',
            '弹': 'T', '窗': 'C', '提': 'T', '示': 'S', '错': 'C', '误': 'W',
            '正': 'Z', '常': 'C', '失': 'S', '败': 'B', '请': 'Q', '稍': 'S',
            '后': 'H', '重': 'Z', '试': 'S', '接': 'J', '入': 'R', '接': 'J',
            '口': 'K', '按': 'A', '钮': 'N', '返': 'F', '回': 'H', '上': 'S',
            '下': 'X', '移': 'Y', '动': 'D', '页': 'Y', '跳': 'T', '转': 'Z',
            '刷': 'S', '新': 'X', '载': 'Z', '入': 'R', '图': 'T', '片': 'P',
            '文': 'W', '件': 'J', '视': 'S', '频': 'P', '播': 'B', '放': 'F',
            '音': 'Y', '乐': 'L', '播': 'B', '放': 'F', '控': 'K', '制': 'Z',
            '播': 'B', '放': 'F', '器': 'Q', '声': 'S', '音': 'Y', '调': 'T',
            '节': 'J', '静': 'J', '音': 'Y', '铃': 'L', '震': 'Z', '动': 'D',
            '提': 'T', '醒': 'X', '闹': 'N', '钟': 'Z', '日': 'R', '历': 'L',
            '记': 'J', '事': 'S', '提': 'T', '示': 'S', '分': 'F', '类': 'L',
            '标': 'B', '签': 'Q', '搜': 'S', '索': 'S', '索': 'S', '引': 'Y',
            '排': 'P', '序': 'X', '筛': 'S', '选': 'X', '过': 'G', '滤': 'L',
            '清': 'Q', '空': 'K', '重': 'Z', '置': 'Z', '确': 'Q', '定': 'D',
            '取': 'Q', '消': 'X', '保': 'B', '存': 'C', '修': 'X', '改': 'G',
            '查': 'C', '看': 'K', '删': 'S', '除': 'C', '复': 'F', '制': 'Z',
            '粘': 'N', '贴': 'T', '剪': 'J', '切': 'Q', '撤': 'C', '销': 'X',
            '重': 'Z', '做': 'Z', '历': 'L', '史': 'S', '记': 'J', '录': 'L',
            '前': 'Q', '进': 'J', '后': 'H', '退': 'T', '返': 'F', '回': 'H',
            '开': 'K', '关': 'G', '启': 'Q', '用': 'Y', '禁': 'J', '止': 'Z',
            '显': 'X', '示': 'S', '隐': 'Y', '藏': 'C', '展': 'Z', '示': 'S',
            '收': 'S', '缩': 'S', '放': 'F', '大': 'D', '小': 'X', '缩': 'S',
            '放': 'F', '原': 'Y', '始': 'S', '缩': 'S', '略': 'L', '图': 'T',
            '预': 'Y', '览': 'L', '查': 'C', '看': 'K', '说': 'S', '明': 'M',
            '告': 'G', '诉': 'S', '请': 'Q', '查': 'C', '看': 'K', '完': 'W',
            '整': 'Z', '版': 'B', '本': 'B', '更': 'G', '新': 'X', '升': 'S',
            '级': 'J', '版': 'B', '本': 'B', '序': 'X', '号': 'H', '版': 'B',
            '版': 'B', '发': 'F', '布': 'B', '公': 'G', '告': 'G', '测': 'C',
            '试': 'S', '环': 'H', '境': 'J', '配': 'P', '置': 'Z', '开': 'K',
            '发': 'F', '生': 'S', '产': 'C', '正': 'Z', '式': 'S', '环': 'H',
            '境': 'J', '本': 'B', '地': 'D', '测': 'C', '试': 'S', '模': 'M',
            '拟': 'N', '调': 'T', '试': 'S', '线': 'X', '上': 'S', '生': 'S',
            '产': 'C', '开': 'K', '发': 'F', '测': 'C', '试': 'S', '预': 'Y',
            '发': 'F', '布': 'B', '测': 'C', '试': 'S', '发': 'F', '布': 'B',
            '灰': 'H', '度': 'D', '发': 'F', '布': 'B', '蓝': 'L', '绿': 'LV',
            '发': 'F', '布': 'B', '版': 'B', '发': 'F', '布': 'B', '首': 'S',
            '发': 'F', '布': 'B', '灰': 'H', '度': 'D', '版': 'B', '全': 'Q',
            '量': 'L', '发': 'F', '布': 'B', '版': 'B', '接': 'J', '口': 'K',
            '应': 'Y', '用': 'Y', '编': 'B', '程': 'C', '接': 'J', '口': 'K',
            '接': 'J', '口': 'K', '文': 'W', '档': 'D', '说': 'S', '明': 'M',
            '接': 'J', '口': 'K', '文': 'W', '档': 'D',  '接': 'J', '口': 'K',
            '请': 'Q', '求': 'Q', '响': 'X', '应': 'Y', '请': 'Q', '求': 'Q',
            '方': 'F', '式': 'S', '请': 'Q', '求': 'Q', '格': 'G', '式': 'S',
            '请': 'Q', '求': 'Q', '路': 'L', '径': 'J', '请': 'Q', '求': 'Q',
            '方': 'F', '法': 'F', '接': 'J', '口': 'K', '开': 'K', '放': 'F',
            '认': 'R', '证': 'Z', '授': 'S', '权': 'Q', '鉴': 'J', '权': 'Q',
            '身': 'S', '认': 'R', '授': 'S', '权': 'Q', '访': 'F', '问': 'W',
            '控': 'K', '制': 'Z', '访': 'F', '问': 'W', '记': 'J', '录': 'L',
            '安': 'A', '全': 'S', '策': 'C', '略': 'L', '防': 'F', '火': 'H',
            '墙': 'Q', '访': 'F', '问': 'W', '控': 'K', '制': 'Z', '数': 'S',
            '据': 'J', '加': 'J', '密': 'M', '密': 'M', '钥': 'Y', '密': 'M',
            '钥': 'Y', '证': 'Z', '书': 'S', '证': 'Z', '书': 'S', '证': 'Z',
            '书': 'S', '证': 'Z', '书': 'S', '证': 'Z', '书': 'S'
        }
        
        # 从每个部分提取拼音首字母（每个部分取多个字符）
        tag = ''
        for part in parts:
            part = part.strip()
            if not part:
                continue
            # 从该部分提取字符直到凑够4位
            for char in part:
                if char in pinyin_map:
                    tag += pinyin_map[char]
                elif char.isalpha():
                    tag += char.upper()
                if len(tag) >= 4:
                    break
            if len(tag) >= 4:
                break
        
        # 如果不足4位，用 M 补齐
        if len(tag) < 4:
            tag = tag.ljust(4, 'M')
    
    # 确保Tag唯一
    base_tag = tag
    counter = 0
    while True:
        existing = db.query(Module).filter(
            Module.project_id == project_id,
            Module.tag == tag,
            Module.parent_id.is_(None)
        ).first()
        if not existing:
            break
        counter += 1
        if counter > 100:
            # 随机生成
            import random
            import string
            tag = ''.join(random.choices(string.ascii_uppercase, k=4))
            break
        tag = f"{base_tag[:2]}{counter:02d}"
    
    return tag


def ensure_module_exists(db: Session, project_id: int, module_name: str, sub_module_name: str = None, user_id: int = 1, auto_create_tag: bool = True):
    """确保模块存在,如果不存在则自动创建。module_name现在存储完整路径如 '模块/子模块/子子模块'
    auto_create_tag: 是否自动为模块生成Tag
    """
    if not module_name:
        return
    
    # 去掉前导斜杠（防止路径格式异常）
    module_name = module_name.lstrip('/')
    if not module_name:
        return
    
    # 解析路径各级
    parts = [p.strip() for p in module_name.split('/') if p.strip()]
    if not parts:
        return
    
    parent_id = None
    for idx, part in enumerate(parts):
        existing = db.query(Module).filter(
            Module.project_id == project_id,
            Module.name == part,
            Module.parent_id == parent_id if parent_id else Module.parent_id.is_(None)
        ).first()
        
        if not existing:
            max_sort = db.query(func.max(Module.sort_order)).filter(
                Module.project_id == project_id,
                Module.parent_id == parent_id if parent_id else Module.parent_id.is_(None)
            ).scalar() or 0
            
            # 自动生成Tag（只为主模块生成，即第一级）
            module_tag = None
            if auto_create_tag and idx == 0:  # 只为主模块生成Tag
                module_tag = generate_module_tag(db, project_id, part)
            
            existing = Module(
                project_id=project_id,
                name=part,
                tag=module_tag,
                parent_id=parent_id,
                sort_order=max_sort + 10,
                created_by=user_id
            )
            db.add(existing)
            db.flush()
        else:
            # 如果模块已存在但Tag为空，且需要自动生成Tag，则补上Tag
            if auto_create_tag and idx == 0 and not existing.tag:
                new_tag = generate_module_tag(db, project_id, part)
                if new_tag:
                    existing.tag = new_tag
                    db.flush()
        
        parent_id = existing.id

class ZmindLinkCreate(BaseModel):
    test_case_id: int
    zmind_issue_id: str
    zmind_issue_subject: str
    zmind_issue_status: Optional[str] = None
    zmind_issue_severity: Optional[str] = None
    test_plan_id: Optional[int] = None  # 关联的测试计划ID

class ZmindLinkUpdate(BaseModel):
    zmind_issue_subject: Optional[str] = None
    zmind_issue_status: Optional[str] = None
    zmind_issue_severity: Optional[str] = None


REGEX_PREFIX = "regex:"

# 匹配 (?=.*<content>) 形式的正向前瞻，content 可包含嵌套括号（简单场景）
_LOOKAHEAD_RE = re.compile(r'\(\?=\.\*(.+?)\)(?=\(\?=|$)')

# 检测仍不支持的 PCRE 语法（排除了已处理的 lookahead 和 (?:...)）
_STILL_UNSUPPORTED_RE = re.compile(
    r'(?:'
    r'\(\?[!<]'            # 负前瞻 (?!  / 后瞻 (?<
    r'|\(\?P[<>]'          # 命名分组
    r'|\(\?#'              # 注释
    r'|\\[dDwWsS]'         # Perl 字符类
    r')'
)


def _convert_non_capturing(pattern: str) -> str:
    """将 (?:...) 非捕获组转换为普通捕获组 (...)，PostgreSQL POSIX ERE 不支持 (?:)"""
    return re.sub(r'\(\?:', '(', pattern)


def _strip_pcre_flags(pattern: str) -> str:
    """
    去除 PCRE inline flag (?i) (?s) (?m) (?x) 及其组合。
    PostgreSQL ~* 本身就是大小写不敏感，(?i) 冗余且不合法；其他 flag 同样不支持。
    """
    return re.sub(r'\(\?[ismx]+\)', '', pattern)


def _try_convert_lookaheads(pattern: str):
    """
    尝试将 (?=.*X)(?=.*Y)... 形式的多前瞻模式转换为多个独立的 POSIX ERE 模式列表。
    每个元素是一个可独立用于 ~* 的模式（AND 语义，所有模式都要匹配）。

    支持的形式：
      (?=.*(?:A|B))(?=.*(?:C|D))   -> ['(A|B)', '(C|D)']
      (?=.*卫星)(?=.*信号)          -> ['卫星', '信号']

    如果无法完整解析，返回 None 表示转换失败。
    """
    # 必须是纯 lookahead 组合，不能有 lookahead 以外的内容
    stripped = pattern.strip()
    # 用正则提取所有 (?=.*...) 块
    lookaheads = re.findall(r'\(\?=\.\*((?:[^()]*|\((?:[^()]*|\([^()]*\))*\))*)\)', stripped)
    if not lookaheads:
        return None

    # 重新拼合已提取的内容，看是否覆盖了整个 pattern（确保没有多余内容）
    reconstructed = ''.join(f'(?=.*{la})' for la in lookaheads)
    if reconstructed != stripped:
        return None

    # 把每个 lookahead 内容转换为 POSIX ERE（去掉 (?:) 的 ?:）
    converted = [_convert_non_capturing(la) for la in lookaheads]
    return converted


def _check_remaining_unsupported(pattern: str):
    """
    检测转换后仍不支持的语法，返回错误描述或 None。
    """
    m = _STILL_UNSUPPORTED_RE.search(pattern)
    if not m:
        return None
    token = m.group(0)
    tips = {
        "(?!": "负前瞻 (?!...)，PostgreSQL 不支持",
        "(?<": "后瞻断言 (?<=...) / (?<!...)，PostgreSQL 不支持",
        "(?P": "命名分组 (?P<name>...)，PostgreSQL 不支持",
        "(?#": "注释语法 (?#...)，PostgreSQL 不支持",
        "\\d": "\\d 不被支持，请改用 [0-9]",
        "\\D": "\\D 不被支持，请改用 [^0-9]",
        "\\w": "\\w 不被支持，请改用 [a-zA-Z0-9_]",
        "\\W": "\\W 不被支持，请改用 [^a-zA-Z0-9_]",
        "\\s": "\\s 不被支持，请改用 [ \\t\\n\\r]",
        "\\S": "\\S 不被支持，请改用 [^ \\t\\n\\r]",
    }
    for k, v in tips.items():
        if token.startswith(k):
            return f"正则语法不支持：{v}"
    return f"正则包含 PostgreSQL 不支持的语法：{token}"


def _build_regex_filter(TestCaseModel, pattern: str):
    """对单个 POSIX ERE pattern 构建 OR 跨字段过滤条件"""
    return or_(
        TestCaseModel.case_number.op("~*")(pattern),
        TestCaseModel.name.op("~*")(pattern),
        TestCaseModel.module.op("~*")(pattern),
        TestCaseModel.steps.op("~*")(pattern),
        TestCaseModel.expected_result.op("~*")(pattern),
        TestCaseModel.precondition.op("~*")(pattern),
        TestCaseModel.remarks.op("~*")(pattern),
    )


def _apply_keyword_filter(query, keyword, db, ZmindLinkModel, TestCaseModel):
    """
    对 query 应用 keyword 过滤。
    - 普通关键词：ILIKE %keyword%
    - 正则关键词（前缀 "regex:"）：PostgreSQL ~* 运算符
      - 自动将 (?=.*A)(?=.*B) 多前瞻语法转换为多个 AND 条件
      - 自动将 (?:...) 非捕获组转换为普通组
    - PR 号格式（#开头或纯数字）：额外搜索 zmind_issue_id
    返回 (query, error_msg)，error_msg 非空时表示正则无效。
    """
    from sqlalchemy import text

    if not keyword:
        return query, None

    is_regex = keyword.startswith(REGEX_PREFIX)

    if is_regex:
        pattern = keyword[len(REGEX_PREFIX):]
        if not pattern:
            return query, None

        # --- 步骤 1：尝试将 (?=.*X)(?=.*Y) 多前瞻转为多个 AND 模式 ---
        converted_patterns = _try_convert_lookaheads(pattern)

        if converted_patterns is not None:
            # 成功转换：每个子模式去掉 inline flag、验证后过滤
            for i, sub in enumerate(converted_patterns):
                converted_patterns[i] = _strip_pcre_flags(sub)
            for sub in converted_patterns:
                err = _check_remaining_unsupported(sub)
                if err:
                    return query, err
            # 设置查询超时
            try:
                db.execute(text("SET LOCAL statement_timeout = '5000'"))
            except Exception:
                pass
            # AND 语义：每个子模式都必须在某个字段中匹配
            for sub in converted_patterns:
                query = query.filter(_build_regex_filter(TestCaseModel, sub))
            return query, None

        # --- 步骤 2：普通正则，转换 (?:...) 和 inline flag 后直接使用 ---
        pattern = _convert_non_capturing(pattern)
        pattern = _strip_pcre_flags(pattern)

        # 兜底：去掉 flag 后如果末尾有多余的 .* 且仍含 lookahead，
        # 尝试去掉末尾 .* 再做一次 lookahead 转换
        if '(?=' in pattern:
            # 去掉末尾可能多余的 .*
            cleaned = re.sub(r'\.\*\s*$', '', pattern).strip()
            retry = _try_convert_lookaheads(cleaned)
            if retry is not None:
                retry = [_strip_pcre_flags(s) for s in retry]
                err = _check_remaining_unsupported(''.join(retry))
                if not err:
                    try:
                        db.execute(text("SET LOCAL statement_timeout = '5000'"))
                    except Exception:
                        pass
                    for sub in retry:
                        query = query.filter(_build_regex_filter(TestCaseModel, sub))
                    return query, None

        # 检测仍不支持的语法
        err = _check_remaining_unsupported(pattern)
        if err:
            return query, err

        # 基本语法验证
        try:
            re.compile(pattern)
        except re.error as e:
            return query, f"正则表达式无效：{e}"

        # 设置查询超时
        try:
            db.execute(text("SET LOCAL statement_timeout = '5000'"))
        except Exception:
            pass

        query = query.filter(_build_regex_filter(TestCaseModel, pattern))
        return query, None

    # 普通关键词
    search_pattern = f"%{keyword}%"
    is_pr_format = keyword.startswith('#') or keyword.isdigit()

    if is_pr_format:
        pr_search = keyword.lstrip('#')
        pr_pattern = f"%{pr_search}%" if pr_search else search_pattern
        linked_query = db.query(ZmindLinkModel.test_case_id).filter(
            ZmindLinkModel.zmind_issue_id.ilike(pr_pattern)
        ).distinct().subquery()
        linked_case_ids = db.query(linked_query).all()
        case_ids_with_link = [c[0] for c in linked_case_ids] if linked_case_ids else [-1]
        query = query.filter(
            or_(
                TestCaseModel.case_number.ilike(search_pattern),
                TestCaseModel.name.ilike(search_pattern),
                TestCaseModel.module.ilike(search_pattern),
                TestCaseModel.steps.ilike(search_pattern),
                TestCaseModel.expected_result.ilike(search_pattern),
                TestCaseModel.precondition.ilike(search_pattern),
                TestCaseModel.remarks.ilike(search_pattern),
                TestCaseModel.id.in_(case_ids_with_link)
            )
        )
    else:
        query = query.filter(
            (TestCaseModel.case_number.ilike(search_pattern)) |
            (TestCaseModel.name.ilike(search_pattern)) |
            (TestCaseModel.module.ilike(search_pattern)) |
            (TestCaseModel.steps.ilike(search_pattern)) |
            (TestCaseModel.expected_result.ilike(search_pattern)) |
            (TestCaseModel.precondition.ilike(search_pattern)) |
            (TestCaseModel.remarks.ilike(search_pattern))
        )
    return query, None


@router.get("/all-ids")
def get_all_testcase_ids(
    project_id: Optional[int] = None,
    project_ids: Optional[str] = None,
    module: Optional[str] = None,
    case_type: Optional[str] = None,
    status: Optional[str] = None,
    status_in: Optional[str] = None,
    level_in: Optional[str] = None,
    keyword: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = None,
    automation: Optional[str] = None,
    has_feedback: Optional[bool] = None,
    archive_source: Optional[str] = None,
    creator_name: Optional[str] = None,
    updater_name: Optional[str] = None,
    tag: Optional[str] = None,
    exclude_in_review: Optional[bool] = None,  # 排除已在评审计划中的用例
    exclude_testplan_id: Optional[int] = None,  # 排除指定测试计划中的用例
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取当前筛选条件下所有用例的ID列表（用于全选功能，不受分页限制）"""
    from utils.project_helper import get_project_ids_with_children
    from utils.data_permission import apply_testcase_data_permission, is_super_admin

    # 查询全选所需的最小字段集：id + 校验/操作所需字段
    query = db.query(TestCase.id, TestCase.status, TestCase.case_number, TestCase.primary_project_id)

    # 应用数据权限过滤
    query = apply_testcase_data_permission(query, current_user, db)

    # 处理多个项目ID
    if project_ids:
        pid_list = [int(pid.strip()) for pid in project_ids.split(',') if pid.strip().isdigit()]
        if pid_list:
            all_project_ids = set()
            for pid in pid_list:
                child_ids = get_project_ids_with_children(db, pid)
                all_project_ids.update(child_ids)
            if all_project_ids:
                query = query.filter(TestCase.primary_project_id.in_(all_project_ids))
    elif project_id:
        project_ids_list = get_project_ids_with_children(db, project_id)
        if project_ids_list:
            query = query.filter(TestCase.primary_project_id.in_(project_ids_list))

    # 按用例类型筛选
    if case_type:
        query = query.filter(TestCase.case_type == case_type)

    # 按状态筛选
    if status:
        query = query.filter(TestCase.status == status)
    elif status_in:
        status_list = [s.strip() for s in status_in.split(',') if s.strip()]
        if status_list:
            query = query.filter(TestCase.status.in_(status_list))

    # 按等级筛选
    if level_in:
        level_list = [l.strip() for l in level_in.split(',') if l.strip()]
        if level_list:
            query = query.filter(TestCase.level.in_(level_list))

    # 按自动化筛选
    if automation is not None:
        if automation == 'D':
            query = query.filter(TestCase.automation == 'D')
        elif automation == 'N':
            query = query.filter(or_(TestCase.automation == 'N', TestCase.automation.like('N-%')))
        elif automation == '' or automation == 'null':
            query = query.filter(or_(TestCase.automation == '', TestCase.automation.is_(None)))
        else:
            query = query.filter(TestCase.automation == automation)

    # 按用户反馈筛选
    if has_feedback is not None:
        if has_feedback:
            query = query.filter(TestCase.feedback.isnot(None), TestCase.feedback != '')
        else:
            query = query.filter(or_(TestCase.feedback.is_(None), TestCase.feedback == ''))

    # 按归档来源筛选
    if archive_source:
        query = query.filter(TestCase.archive_source == archive_source)

    # 按模块筛选（与 list_testcases 保持一致，但避免额外的 COUNT 查询）
    if module:
        actual_module = module
        matched_project_id = None

        pid_prefix_match = re.match(r'^p(\d+)/(.+)$', module)
        if pid_prefix_match:
            matched_project_id = int(pid_prefix_match.group(1))
            actual_module = pid_prefix_match.group(2)
        else:
            if project_ids:
                try:
                    pid_list = [int(pid.strip()) for pid in project_ids.split(',') if pid.strip().isdigit()]
                    if pid_list:
                        projects = db.query(Project).filter(Project.id.in_(pid_list)).all()
                        project_name_to_id = {p.name: p.id for p in projects}
                        for proj_name, proj_id in project_name_to_id.items():
                            if module.startswith(proj_name + '/'):
                                actual_module = module[len(proj_name) + 1:]
                                matched_project_id = proj_id
                                break
                except:
                    pass
            elif project_id:
                project = db.query(Project).filter(Project.id == project_id).first()
                if project and module.startswith(project.name + '/'):
                    actual_module = module[len(project.name) + 1:]
                    matched_project_id = project_id

        if matched_project_id:
            child_ids = get_project_ids_with_children(db, matched_project_id)
            if child_ids:
                query = query.filter(TestCase.primary_project_id.in_(child_ids))

        main_module = actual_module.split('/')[0] if '/' in actual_module else actual_module
        # 优先精确匹配，回退到主模块模糊匹配；用 OR 合并为单次查询，避免额外 COUNT
        query = query.filter(
            or_(
                TestCase.module == actual_module,
                TestCase.module.like(actual_module + '/%'),
                TestCase.module.like(main_module + '/%')
            )
        )

    # 按关键词搜索
    query, regex_error = _apply_keyword_filter(query, keyword, db, TestCaseZmindLink, TestCase)
    if regex_error:
        raise HTTPException(status_code=400, detail=regex_error)

    # 按创建人姓名过滤（通过 User 表 subquery）
    if creator_name:
        creator_subq = db.query(User.id).filter(User.username == creator_name).subquery()
        query = query.filter(TestCase.created_by.in_(creator_subq))

    # 按更新人姓名过滤（通过 User 表 subquery）
    if updater_name:
        updater_subq = db.query(User.id).filter(User.username == updater_name).subquery()
        query = query.filter(TestCase.updated_by.in_(updater_subq))

    # 按标签过滤（tags 字段是 JSON 字符串，用 LIKE 匹配）
    if tag:
        query = query.filter(TestCase.tags.ilike(f'%{tag}%'))

    # 排除已在评审计划中的用例
    if exclude_in_review:
        from models import ReviewPlanTestCase, ReviewPlan
        in_review_subquery = db.query(ReviewPlanTestCase.testcase_id).join(
            ReviewPlan, ReviewPlan.id == ReviewPlanTestCase.review_plan_id
        ).filter(
            ReviewPlan.status != 'COMPLETED',
            ReviewPlanTestCase.testcase_id.isnot(None)
        ).distinct()
        query = query.filter(~TestCase.id.in_(in_review_subquery))

    # 排除指定测试计划中的用例
    if exclude_testplan_id:
        from models import TestPlanTestCase
        in_testplan_subquery = db.query(TestPlanTestCase.test_case_id).filter(
            TestPlanTestCase.test_plan_id == exclude_testplan_id,
            TestPlanTestCase.test_case_id.isnot(None)
        ).distinct()
        query = query.filter(~TestCase.id.in_(in_testplan_subquery))

    try:
        rows = query.all()
    except Exception as e:
        err_str = str(e)
        if "canceling statement due to statement timeout" in err_str or "statement timeout" in err_str:
            db.rollback()
            raise HTTPException(status_code=408, detail="正则搜索超时，请简化正则表达式后重试")
        raise
    return {
        "code": 200,
        "data": {
            "records": [
                {
                    "id": r[0],
                    "status": r[1],
                    "case_number": r[2],
                    "primary_project_id": r[3],
                    "_isIdOnly": True  # 标记为轻量对象，区别于完整行数据
                }
                for r in rows
            ],
            "total": len(rows)
        }
    }


@router.get("")
def list_testcases(
    page: int = 1,
    size: int = 10,
    project_id: Optional[int] = None,
    project_ids: Optional[str] = None,  # 新增：支持多个项目ID，逗号分隔
    module: Optional[str] = None,
    case_type: Optional[str] = None,
    status: Optional[str] = None,
    status_in: Optional[str] = None,  # 新增：支持多个状态筛选，逗号分隔
    level_in: Optional[str] = None,  # 支持多个等级筛选，逗号分隔（如 L1,L2）
    keyword: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = None,
    exclude_in_review: Optional[bool] = None,  # 新增：排除已在评审计划中的用例
    exclude_testplan_id: Optional[int] = None,  # 新增：排除指定测试计划中的用例
    test_plan_id: Optional[int] = None,  # 新增：只显示指定测试计划中的用例
    automation: Optional[str] = None,  # 自动化筛选（Y/N）
    has_feedback: Optional[bool] = None,  # 新增：筛选有用户反馈的用例
    archive_source: Optional[str] = None,  # 归档来源筛选
    req: Request = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from utils.project_helper import get_project_ids_with_children
    from utils.data_permission import apply_testcase_data_permission, is_super_admin

    # 记录尝试请求大量数据的行为（在限制之前检测）
    if size >= 101 and req:
        client_ip = req.headers.get("X-Forwarded-For", "").split(",")[0].strip() or req.headers.get("X-Real-IP") or (req.client.host if req.client else "unknown")
        logger.warning(f"[数据爬取风险] 用户 {current_user.username}({client_ip}) 尝试请求大量数据: size={size}")
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.TESTCASES,
            action=LogAction.VIEW,
            description=f"[安全告警] 用户 {current_user.username}({client_ip}) 尝试请求大量测试用例数据: size={size}，已限制为100",
            request=req,
            response_status=200
        )

    # 限制分页大小，防止一次获取大量数据
    MAX_PAGE_SIZE = 100
    if size > MAX_PAGE_SIZE:
        size = MAX_PAGE_SIZE
    if page < 1:
        page = 1

    query = db.query(TestCase)
    
    # 应用数据权限过滤
    query = apply_testcase_data_permission(query, current_user, db)
    
    # 如果指定了test_plan_id，只返回该测试计划关联的用例
    if test_plan_id:
        from models import TestPlanTestCase
        # 子查询：获取指定测试计划中的用例ID
        in_testplan_subquery = db.query(TestPlanTestCase.test_case_id).filter(
            TestPlanTestCase.test_plan_id == test_plan_id
        ).distinct()
        query = query.filter(TestCase.id.in_(in_testplan_subquery))
    
    # 处理多个项目ID（优先使用project_ids）
    if project_ids and not test_plan_id:
        # 解析逗号分隔的项目ID列表
        pid_list = [int(pid.strip()) for pid in project_ids.split(',') if pid.strip().isdigit()]
        if pid_list:
            # 获取所有项目ID（包括子项目）
            all_project_ids = set()
            for pid in pid_list:
                child_ids = get_project_ids_with_children(db, pid)
                all_project_ids.update(child_ids)
            
            if all_project_ids:
                query = query.filter(TestCase.primary_project_id.in_(all_project_ids))
    elif project_id and not test_plan_id:  # 只有在没有指定test_plan_id时才按项目过滤
        # 获取项目ID列表（包括子项目）
        project_ids_list = get_project_ids_with_children(db, project_id)
        
        if project_ids_list:
            # 查询这些项目的所有测试用例
            query = query.filter(TestCase.primary_project_id.in_(project_ids_list))
    
    # 按用例类型筛选
    if case_type:
        query = query.filter(TestCase.case_type == case_type)
    
    # 按状态筛选
    if status:
        query = query.filter(TestCase.status == status)
    elif status_in:
        # 支持多个状态筛选，逗号分隔
        status_list = [s.strip() for s in status_in.split(',') if s.strip()]
        if status_list:
            query = query.filter(TestCase.status.in_(status_list))
    
    # 按等级筛选
    if level_in:
        level_list = [l.strip() for l in level_in.split(',') if l.strip()]
        if level_list:
            query = query.filter(TestCase.level.in_(level_list))
    
    # 按自动化筛选
    if automation is not None:
        if automation == 'D':
            query = query.filter(TestCase.automation == 'D')  # 自动化已完成
        elif automation == 'N':
            query = query.filter(or_(TestCase.automation == 'N', TestCase.automation.like('N-%')))
        elif automation == '' or automation == 'null':
            query = query.filter(or_(TestCase.automation == '', TestCase.automation.is_(None)))
        else:
            query = query.filter(TestCase.automation == automation)
    else:
        pass
    
    # 按用户反馈筛选
    if has_feedback is not None:
        if has_feedback:
            query = query.filter(TestCase.feedback.isnot(None), TestCase.feedback != '')
        else:
            query = query.filter(or_(TestCase.feedback.is_(None), TestCase.feedback == ''))
    
    # 按归档来源筛选
    if archive_source:
        query = query.filter(TestCase.archive_source == archive_source)
    
    # 排除已在评审计划中的用例
    if exclude_in_review:
        from models import ReviewPlanTestCase, ReviewPlan
        # 子查询：获取所有已在非已完成评审计划中的用例ID
        # 过滤 NULL：避免 NOT IN (NULL, ...) 永远返回 false 的三值逻辑陷阱
        in_review_subquery = db.query(ReviewPlanTestCase.testcase_id).join(
            ReviewPlan, ReviewPlan.id == ReviewPlanTestCase.review_plan_id
        ).filter(
            ReviewPlan.status != 'COMPLETED',
            ReviewPlanTestCase.testcase_id.isnot(None)
        ).distinct()
        query = query.filter(~TestCase.id.in_(in_review_subquery))
    
    # 排除指定测试计划中的用例
    if exclude_testplan_id:
        from models import TestPlanTestCase
        # 子查询：获取指定测试计划中的用例ID
        # 关键：必须过滤 NULL，否则 NOT IN (NULL, ...) 在 SQL 三值逻辑下永远返回 false
        # （历史数据中 test_plan_test_cases 可能有僵尸记录，关联用例已被删除）
        in_testplan_subquery = db.query(TestPlanTestCase.test_case_id).filter(
            TestPlanTestCase.test_plan_id == exclude_testplan_id,
            TestPlanTestCase.test_case_id.isnot(None)
        ).distinct()
        query = query.filter(~TestCase.id.in_(in_testplan_subquery))
    
    # 按模块筛选（module字段存储完整路径，支持前缀匹配以包含子模块）
    # 多项目模式下 path 格式为 "p{pid}/模块路径"，需要解析出 pid 和实际模块路径
    # 用例表中 module 字段存储的是 "模块名" 或 "模块名/子模块名" 格式
    if module:
        actual_module = module
        matched_project_id = None
        
        # 处理多项目模式的 "p{pid}/模块路径" 格式
        pid_prefix_match = re.match(r'^p(\d+)/(.+)$', module)
        if pid_prefix_match:
            matched_project_id = int(pid_prefix_match.group(1))
            actual_module = pid_prefix_match.group(2)
        else:
            # 兼容旧格式：尝试匹配项目名前缀
            if project_ids:
                try:
                    pid_list = [int(pid.strip()) for pid in project_ids.split(',') if pid.strip().isdigit()]
                    if pid_list:
                        projects = db.query(Project).filter(Project.id.in_(pid_list)).all()
                        project_name_to_id = {p.name: p.id for p in projects}
                        for proj_name, proj_id in project_name_to_id.items():
                            if module.startswith(proj_name + '/'):
                                actual_module = module[len(proj_name) + 1:]
                                matched_project_id = proj_id
                                break
                except:
                    pass
            elif project_id:
                project = db.query(Project).filter(Project.id == project_id).first()
                if project and module.startswith(project.name + '/'):
                    actual_module = module[len(project.name) + 1:]
                    matched_project_id = project_id
        
        # 如果匹配到了具体用例库，需要同时用project_id过滤
        if matched_project_id:
            child_ids = get_project_ids_with_children(db, matched_project_id)
            if child_ids:
                query = query.filter(TestCase.primary_project_id.in_(child_ids))
        
        # 提取主模块名（第一个 '/' 之前的部分）
        main_module = actual_module.split('/')[0] if '/' in actual_module else actual_module
        
        # 先尝试精确匹配和完整前缀匹配
        exact_match = or_(
            TestCase.module == actual_module,
            TestCase.module.like(actual_module + '/%')
        )
        
        # 备用匹配：以主模块名开头（处理简繁体差异）
        fuzzy_match = TestCase.module.like(main_module + '/%')
        
        # 先用精确匹配，如果无结果再尝试模糊匹配
        exact_results = query.filter(exact_match).count()
        if exact_results > 0:
            query = query.filter(exact_match)
        else:
            query = query.filter(fuzzy_match)
    
    # 按关键词搜索（搜索用例编号、标题、模块、步骤、预期结果、前置条件、备注、PR号）
    query, regex_error = _apply_keyword_filter(query, keyword, db, TestCaseZmindLink, TestCase)
    if regex_error:
        raise HTTPException(status_code=400, detail=regex_error)

    try:
        total = query.count()
    except Exception as e:
        err_str = str(e)
        if "canceling statement due to statement timeout" in err_str or "statement timeout" in err_str:
            db.rollback()
            raise HTTPException(status_code=408, detail="正则搜索超时，请简化正则表达式后重试")
        raise
    
    # 排序
    from sqlalchemy import case, desc, asc, text
    
    # 如果有自定义排序，使用自定义排序
    if sort_by and sort_order:
        if sort_by == 'zmind_link_count':
            # 按PR数量排序，需要子查询
            link_count_subquery = db.query(
                TestCaseZmindLink.test_case_id,
                func.count(TestCaseZmindLink.id).label('link_count')
            ).group_by(TestCaseZmindLink.test_case_id).subquery()
            
            query = query.outerjoin(
                link_count_subquery,
                TestCase.id == link_count_subquery.c.test_case_id
            )
            
            if sort_order == 'descending':
                query = query.order_by(desc(func.coalesce(link_count_subquery.c.link_count, 0)), TestCase.id.asc())
            else:
                query = query.order_by(func.coalesce(link_count_subquery.c.link_count, 0).asc(), TestCase.id.asc())
        else:
            # 其他字段的排序
            sort_column = getattr(TestCase, sort_by, None)
            if sort_column is not None:
                if sort_by == 'case_number':
                    # 用例编号使用自然排序（提取数字部分）
                    natural_sort = text("""
                        LPAD(COALESCE(SUBSTRING(test_cases.case_number FROM '([0-9]+)$'), '0'), 20, '0')
                    """)
                    if sort_order == 'descending':
                        query = query.order_by(desc(natural_sort), TestCase.id.asc())
                    else:
                        query = query.order_by(asc(natural_sort), TestCase.id.asc())
                else:
                    if sort_order == 'descending':
                        query = query.order_by(desc(sort_column), TestCase.id.asc())
                    else:
                        query = query.order_by(sort_column.asc(), TestCase.id.asc())
    else:
        # 默认排序：使用统一的排序函数
        sort_pid_list = []
        if project_ids:
            sort_pid_list = [int(pid.strip()) for pid in project_ids.split(',') if pid.strip().isdigit()]
        elif project_id:
            sort_pid_list = [project_id]
        
        query = apply_testcase_sort(query, sort_pid_list if sort_pid_list else None, db=db if sort_pid_list else None)
    
    try:
        testcases = query.offset((page - 1) * size).limit(size).all()
    except Exception as e:
        err_str = str(e)
        if "canceling statement due to statement timeout" in err_str or "statement timeout" in err_str:
            db.rollback()
            raise HTTPException(status_code=408, detail="正则搜索超时，请简化正则表达式后重试")
        raise
    
    # 性能优化：批量获取所有需要的数据，避免N+1查询
    testcase_ids = [tc.id for tc in testcases]
    creator_ids = list(set([tc.created_by for tc in testcases if tc.created_by]))
    updater_ids = list(set([tc.updated_by for tc in testcases if tc.updated_by]))
    
    # 批量查询所有创建人和更新人
    all_user_ids = list(set(creator_ids + updater_ids))
    users_dict = {}
    if all_user_ids:
        users = db.query(User).filter(User.id.in_(all_user_ids)).all()
        users_dict = {user.id: user.username for user in users}
    
    # 批量查询PR总数
    link_counts = dict(
        db.query(
            TestCaseZmindLink.test_case_id,
            func.count(TestCaseZmindLink.id)
        )
        .filter(TestCaseZmindLink.test_case_id.in_(testcase_ids))
        .group_by(TestCaseZmindLink.test_case_id)
        .all()
    ) if testcase_ids else {}
    
    # 批量查询打开的PR数量
    open_counts = dict(
        db.query(
            TestCaseZmindLink.test_case_id,
            func.count(TestCaseZmindLink.id)
        )
        .filter(
            TestCaseZmindLink.test_case_id.in_(testcase_ids),
            ~TestCaseZmindLink.zmind_issue_status.in_(list(CLOSED_STATUSES))
        )
        .group_by(TestCaseZmindLink.test_case_id)
        .all()
    ) if testcase_ids else {}
    
    # 批量查询关闭的PR数量
    close_counts = dict(
        db.query(
            TestCaseZmindLink.test_case_id,
            func.count(TestCaseZmindLink.id)
        )
        .filter(
            TestCaseZmindLink.test_case_id.in_(testcase_ids),
            TestCaseZmindLink.zmind_issue_status.in_(list(CLOSED_STATUSES))
        )
        .group_by(TestCaseZmindLink.test_case_id)
        .all()
    ) if testcase_ids else {}
    
    result_testcases = []
    for tc in testcases:
        # 从预加载的字典中获取用户名
        creator_name = users_dict.get(tc.created_by, "-")
        updater_name = users_dict.get(tc.updated_by, creator_name if tc.updated_by == tc.created_by else "-")
        
        tc_dict = {
            "id": tc.id,
            "primary_project_id": tc.primary_project_id,
            "project_id": tc.primary_project_id,  # 兼容前端
            "case_number": tc.case_number,
            "module": tc.module,
            "name": tc.name,
            "precondition": tc.precondition,
            "steps": tc.steps,
            "expected_result": tc.expected_result,
            "level": tc.level,
            "remarks": tc.remarks,
            "automation": tc.automation,
            "status": tc.status,  # 状态
            "case_type": tc.case_type,
            "priority": tc.priority,  # 保留兼容
            "category": tc.category,  # 保留兼容
            "tags": tc.tags,
            "archive_source": tc.archive_source,
            "share_scope": tc.share_scope,
            "source": tc.source,
            "zmind_id": tc.zmind_id,
            "created_at": tc.created_at,
            "updated_at": tc.updated_at,
            "created_by": tc.created_by,
            "creator_name": creator_name,  # 创建人名称
            "updated_by": tc.updated_by,
            "updater_name": updater_name,  # 更新人名称
            "feedback": tc.feedback,
            "sort_order": tc.sort_order or 0,
            "zmind_link_count": link_counts.get(tc.id, 0),
            "zmind_link_open_count": open_counts.get(tc.id, 0),
            "zmind_link_close_count": close_counts.get(tc.id, 0)
        }
        
        result_testcases.append(tc_dict)
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "records": result_testcases,
            "total": total,
            "page": page,
            "size": size
        }
    }

@router.post("")
def create_testcase(
    testcase: TestCaseCreate,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from datetime import datetime
    
    try:
        # 创建测试用例，显式设置 source 字段
        testcase_data = testcase.dict()
        
        # 验证 level 字段
        if testcase_data.get('level') not in ['L1', 'L2', 'L3', 'L4']:
            raise HTTPException(status_code=400, detail="用例等级必须是 L1/L2/L3/L4")
        
        # 验证 automation 字段
        automation = testcase_data.get('automation')
        if automation:
            valid_values = ['Y', 'D', 'N', '']
            valid_values.extend([
                'N-HW_PHYSICAL', 'N-VISUAL_JUDGE', 'N-BOOT_PROCESS', 'N-MEDIA_PLAY',
                'N-OTA_UPGRADE', 'N-DATA_CONFIG', 'N-LOG_CHECK', 'N-BACKEND_CONFIG', 'N-DATA_DYNAMIC', 'N-OTHER_SPECIAL'
            ])
            if automation not in valid_values:
                raise HTTPException(status_code=400, detail="自动化字段值无效")
        
        # 用例编号改为自动生成
        if not testcase_data.get('case_number'):
            # 自动生成用例编号
            testcase_data['case_number'] = generate_case_number(
                db=db,
                project_id=testcase_data['primary_project_id'],
                module_path=testcase_data.get('module')
            )
        
        # 自动创建模块(如果不存在)
        ensure_module_exists(
            db=db,
            project_id=testcase_data['primary_project_id'],
            module_name=testcase_data.get('module'),
            user_id=current_user.id
        )
        
        testcase_data['source'] = 'LOCAL'
        testcase_data['created_by'] = current_user.id
        
        # 新建用例默认状态为待评审
        if 'status' not in testcase_data or not testcase_data['status']:
            testcase_data['status'] = 'PENDING'
        
        # sort_order: 新建用例排到同模块末尾
        if 'sort_order' not in testcase_data or not testcase_data['sort_order']:
            max_sort = db.query(func.max(TestCase.sort_order)).filter(
                TestCase.primary_project_id == testcase_data['primary_project_id'],
                TestCase.module == testcase_data.get('module')
            ).scalar() or 0
            testcase_data['sort_order'] = max_sort + 10
        
        db_testcase = TestCase(**testcase_data)
        db.add(db_testcase)
        db.flush()  # 获取ID
        
        # 创建用例-项目关联
        test_case_project = TestCaseProject(
            test_case_id=db_testcase.id,
            project_id=testcase.primary_project_id,
            relation_type='OWNED',
            is_editable=1,
            created_at=datetime.now(),
            created_by=current_user.id
        )
        db.add(test_case_project)
        
        # 记录创建历史
        create_history = TestCaseHistory(
            testcase_id=db_testcase.id,
            field_name='创建用例',
            old_value=None,
            new_value=f"{db_testcase.case_number} - {db_testcase.name}",
            changed_by=current_user.id,
            changed_by_name=current_user.username
        )
        db.add(create_history)
        
        db.commit()
        db.refresh(db_testcase)
        
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.TESTCASES,
            action=LogAction.CREATE,
            description=f"创建测试用例：{db_testcase.name}（编号: {db_testcase.case_number}，ID: {db_testcase.id}）",
            request=req
        )
        
        # 触发通知
        trigger_testcase_notification(
            db=db,
            event_type='created',
            testcase_id=db_testcase.id,
            testcase_name=db_testcase.name,
            operator_name=current_user.username,
            changes=f'创建了新用例：{db_testcase.name}',
            project_id=db_testcase.primary_project_id
        )
        
        return {"code": 200, "message": "success", "data": db_testcase}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"创建测试用例错误: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"创建测试用例失败: {str(e)}")

@router.post("/batch-create")
def batch_create_testcases(
    batch: TestCaseCreateBatch,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    批量创建测试用例（防N+1优化）
    - 预加载所有涉及的项目和模块信息（最多3次查询）
    - 使用 seq_cache 复用序号缓存，避免重复编号查询
    - 使用 SAVEPOINT 实现逐条独立事务，部分失败不影响其他用例
    - 最终一次性 commit 所有成功的记录
    """
    from datetime import datetime
    import logging
    logger = logging.getLogger(__name__)

    total = len(batch.testcases)
    results = []
    errors = []
    seq_cache = {}

    # ====== 预加载阶段：最多2次查询 ======
    project_ids = set()
    need_modules = set()
    for tc in batch.testcases:
        project_ids.add(tc.primary_project_id)
        if tc.module:
            need_modules.add((tc.primary_project_id, tc.module))

    # 查询1：一次查询所有项目
    projects = {p.id: p for p in db.query(Project).filter(Project.id.in_(project_ids)).all()}

    # 查询2：一次查询各模块当前最大 sort_order（每 (project, module) 一条查询，不会随记录数增长）
    module_max_sort = {}
    for pid, mod in need_modules:
        max_sort = db.query(func.max(TestCase.sort_order)).filter(
            TestCase.primary_project_id == pid, TestCase.module == mod
        ).scalar() or 0
        module_max_sort[(pid, mod)] = max_sort

    valid_automation = {'Y', 'D', 'N', '',
        'N-HW_PHYSICAL', 'N-VISUAL_JUDGE', 'N-BOOT_PROCESS', 'N-MEDIA_PLAY',
        'N-OTA_UPGRADE', 'N-DATA_CONFIG', 'N-LOG_CHECK', 'N-BACKEND_CONFIG',
        'N-DATA_DYNAMIC', 'N-OTHER_SPECIAL'}

    for idx, tc in enumerate(batch.testcases):
        try:
            testcase_data = tc.dict()

            # 预验证（无需查库）
            level = testcase_data.get('level', 'L3')
            if level not in ('L1', 'L2', 'L3', 'L4'):
                raise ValueError(f"用例等级无效: {level}")

            automation = testcase_data.get('automation')
            if automation and automation not in valid_automation:
                raise ValueError(f"自动化字段值无效: {automation}")

            pid = testcase_data['primary_project_id']
            if pid not in projects:
                raise ValueError(f"用例库ID {pid} 不存在")

            # SAVEPOINT：隔离单条写入
            with db.begin_nested():
                testcase_data['case_number'] = generate_case_number(
                    db=db, project_id=pid,
                    module_path=testcase_data.get('module'),
                    seq_cache=seq_cache
                )

                ensure_module_exists(
                    db=db, project_id=pid,
                    module_name=testcase_data.get('module'),
                    user_id=current_user.id
                )

                testcase_data['source'] = 'LOCAL'
                testcase_data['created_by'] = current_user.id
                testcase_data['status'] = testcase_data.get('status') or 'PENDING'

                sort_key = (pid, testcase_data.get('module', ''))
                module_max_sort[sort_key] = module_max_sort.get(sort_key, 0) + 10
                testcase_data['sort_order'] = module_max_sort[sort_key]

                db_testcase = TestCase(**testcase_data)
                db.add(db_testcase)
                db.flush()

                tc_id = db_testcase.id
                tc_number = db_testcase.case_number

                # 项目关联 & 历史（可批量，但简单起见逐条 add）
                db.add(TestCaseProject(
                    test_case_id=tc_id, project_id=pid,
                    relation_type='OWNED', is_editable=1,
                    created_at=datetime.now(), created_by=current_user.id
                ))
                db.add(TestCaseHistory(
                    testcase_id=tc_id,
                    field_name='创建用例', old_value=None,
                    new_value=f"{tc_number} - {tc.name}",
                    changed_by=current_user.id,
                    changed_by_name=current_user.username
                ))

            # SAVEPOINT 成功释放，记录结果
            results.append({
                "id": tc_id, "case_number": tc_number, "name": tc.name
            })

        except Exception as e:
            logger.warning(f"批量创建第{idx+1}条失败: {str(e)}")
            errors.append({"row": idx + 1, "message": str(e)})

    # 一次性提交所有已释放的 SAVEPOINT
    try:
        db.commit()

        if results:
            log_operation(
                db=db, user_id=current_user.id,
                username=current_user.username,
                module=LogModule.TESTCASES, action=LogAction.CREATE,
                description=f"批量创建测试用例：共{total}条，成功{len(results)}条，失败{len(errors)}条",
                request=req
            )

            first = results[0]
            trigger_testcase_notification(
                db=db, event_type='created',
                testcase_id=first['id'],
                testcase_name=first['name'],
                operator_name=current_user.username,
                changes=f'批量创建了{len(results)}个用例',
                project_id=batch.testcases[0].primary_project_id
            )

    except Exception as e:
        db.rollback()
        logger.error(f"批量创建提交失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"批量创建提交失败: {str(e)}")

    return {
        "code": 200,
        "message": "success",
        "data": {
            "success_count": len(results),
            "fail_count": len(errors),
            "results": results,
            "errors": errors
        }
    }

@router.get("/filter-options")
def get_filter_options(
    project_id: Optional[int] = None,
    project_ids: Optional[str] = None,
    module: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取筛选选项（归档来源、标签等唯一值列表）"""
    try:
        from utils.project_helper import get_project_ids_with_children
        
        # 解析项目ID列表
        pid_list = []
        if project_ids:
            pid_list = [int(pid.strip()) for pid in project_ids.split(',') if pid.strip().isdigit()]
        elif project_id:
            pid_list = [project_id]
        
        if not pid_list:
            return {
                "code": 200,
                "message": "success",
                "data": {
                    "archive_sources": [],
                    "tags": []
                }
            }
        
        # 获取所有项目ID（包括子项目）
        all_project_ids = set()
        for pid in pid_list:
            child_ids = get_project_ids_with_children(db, pid)
            all_project_ids.update(child_ids)
        
        # 基础查询条件
        base_filters = [TestCase.primary_project_id.in_(list(all_project_ids))]
        
        # 模块筛选
        if module:
            base_filters.append(
                or_(
                    TestCase.module == module,
                    TestCase.module.like(module + '/%')
                )
            )
        
        # 查询所有唯一的归档来源（排除空值）
        archive_sources = db.query(func.distinct(TestCase.archive_source)).filter(
            *base_filters,
            TestCase.archive_source.isnot(None),
            TestCase.archive_source != ''
        ).all()
        archive_source_list = sorted([s[0] for s in archive_sources if s[0]])
        
        # 查询所有唯一的标签（排除空值）
        tags_query = db.query(TestCase.tags).filter(
            *base_filters,
            TestCase.tags.isnot(None),
            TestCase.tags != '',
            TestCase.tags != '[]'
        ).all()
        
        # 解析JSON格式的标签
        tag_set = set()
        for (tags_text,) in tags_query:
            try:
                import json
                parsed = json.loads(tags_text)
                if isinstance(parsed, list):
                    for tag in parsed:
                        if tag and isinstance(tag, str):
                            tag_set.add(tag)
            except:
                # 兼容逗号分隔格式
                for tag in tags_text.split(','):
                    tag = tag.strip()
                    if tag:
                        tag_set.add(tag)
        
        return {
            "code": 200,
            "message": "success",
            "data": {
                "archive_sources": archive_source_list,
                "tags": sorted(list(tag_set))
            }
        }
    except Exception as e:
        print(f"获取筛选选项失败: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取筛选选项失败: {str(e)}")

@router.get("/statistics/zmind-links")
def get_zmind_link_statistics(
    project_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(TestCase)
    if project_id:
        query = query.filter(TestCase.primary_project_id == project_id)
    
    total_testcases = query.count()
    linked_testcases = db.query(func.count(func.distinct(TestCaseZmindLink.test_case_id))).filter(
        TestCaseZmindLink.test_case_id.in_(query.with_entities(TestCase.id))
    ).scalar()
    total_links = db.query(func.count(TestCaseZmindLink.id)).filter(
        TestCaseZmindLink.test_case_id.in_(query.with_entities(TestCase.id))
    ).scalar()
    
    return {
        "code": 200,
        "message": "success",
        "data": {
            "total_testcases": total_testcases,
            "linked_testcases": linked_testcases,
            "total_links": total_links,
            "unlinked_testcases": total_testcases - linked_testcases
        }
    }

@router.get("/template/download")
def download_template(
    team_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """下载Excel导入模板（包含用例库和模块信息）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.datavalidation import DataValidation
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例模板"
    
    # 新表头：移除用例编号列（由系统自动生成），移除子模块列，添加所属用例库列，模块改为层级格式
    headers = [
        "所属用例库*", "用例类型", "模块*", "用例标题*", "前置条件", 
        "操作步骤*", "预期结果*", "用例等级*", "备注", "自动化"
    ]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="8B9AEE", end_color="8B9AEE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加示例数据（用例编号由系统自动生成）
    ws.append([
        "示例用例库",
        "功能测试",
        "用户管理/登录模块",  # 新格式：模块/子模块
        "验证用户登录功能",
        "用户已注册",
        "1. 打开登录页面\n2. 输入用户名和密码\n3. 点击登录按钮",
        "1. 页面跳转到首页\n2. 显示用户信息\n3. 登录成功提示",
        "L1",
        "这是一个示例用例",
        "Y"
    ])
    
    # 设置列宽（用例编号列已移除）
    ws.column_dimensions['A'].width = 20  # 所属用例库
    ws.column_dimensions['B'].width = 12  # 用例类型
    ws.column_dimensions['C'].width = 30  # 模块（层级格式需要更宽）
    ws.column_dimensions['D'].width = 30  # 用例标题
    ws.column_dimensions['E'].width = 30  # 前置条件
    ws.column_dimensions['F'].width = 40  # 操作步骤
    ws.column_dimensions['G'].width = 40  # 预期结果
    ws.column_dimensions['H'].width = 12  # 用例等级
    ws.column_dimensions['I'].width = 30  # 备注
    ws.column_dimensions['J'].width = 10  # 自动化
    
    # 添加说明sheet（用例编号改为系统自动生成）
    ws_info = wb.create_sheet("填写说明")
    ws_info.append(["字段名", "是否必填", "说明"])
    ws_info.append(["所属用例库", "是", "用例所属的用例库名称，必须是系统中已存在的用例库"])
    ws_info.append(["用例编号", "否", "由系统自动生成，无需填写"])
    ws_info.append(["用例类型", "否", "可选值：功能测试、性能测试、安全测试、接口测试、安装部署、配置相关、其他。默认：功能测试"])
    ws_info.append(["模块", "是", "测试用例所属模块，格式：模块名 或 模块名/子模块名，必须是系统中已存在的模块"])
    ws_info.append(["用例标题", "是", "测试用例的名称"])
    ws_info.append(["前置条件", "否", "执行测试前需要满足的条件"])
    ws_info.append(["操作步骤", "是", "测试步骤，每行一个步骤"])
    ws_info.append(["预期结果", "是", "每个步骤的预期结果，每行一个结果"])
    ws_info.append(["用例等级", "是", "可选值：L1（最高）、L2（高）、L3（中）、L4（低）。默认：L3"])
    ws_info.append(["备注", "否", "其他说明信息"])
    ws_info.append(["自动化", "否", "可选值：Y（是）、N（否）"])
    
    ws_info.column_dimensions['A'].width = 15
    ws_info.column_dimensions['B'].width = 12
    ws_info.column_dimensions['C'].width = 80
    
    # 设置说明sheet表头样式
    for cell in ws_info[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 获取用户有权限的用例库和模块信息
    from models import Team, UserTeam, TeamProject
    
    # 获取用户可访问的项目组
    if team_id:
        # 指定了项目组
        teams = db.query(Team).filter(Team.id == team_id).all()
    else:
        # 获取用户所属的所有项目组
        user_team_ids = db.query(UserTeam.team_id).filter(UserTeam.user_id == current_user.id).all()
        user_team_ids = [t[0] for t in user_team_ids]

        # 超级管理员可以看到所有项目组
        if is_super_admin(current_user):
            teams = db.query(Team).filter(Team.status == 1).all()
        elif user_team_ids:
            teams = db.query(Team).filter(Team.id.in_(user_team_ids), Team.status == 1).all()
        else:
            teams = []
    
    # 创建用例库和模块参考sheet
    ws_ref = wb.create_sheet("用例库和模块参考")
    ws_ref.append(["用例库名称", "模块路径", "说明"])
    
    # 设置参考sheet表头样式
    for cell in ws_ref[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 收集所有用例库和模块信息
    project_names = []
    for team in teams:
        # 通过TeamProject关联表获取项目组下的所有用例库
        team_project_ids = db.query(TeamProject.project_id).filter(TeamProject.team_id == team.id).all()
        team_project_ids = [tp[0] for tp in team_project_ids]
        
        if not team_project_ids:
            continue
            
        projects = db.query(Project).filter(Project.id.in_(team_project_ids)).all()
        
        for project in projects:
            project_names.append(project.name)
            
            # 获取该用例库下的所有模块
            modules = db.query(Module).filter(
                Module.project_id == project.id,
                Module.parent_id.is_(None)  # 只获取主模块
            ).order_by(Module.sort_order).all()
            
            if not modules:
                # 如果没有模块，添加一行说明
                ws_ref.append([project.name, "（暂无模块）", "请先在系统中创建模块"])
            else:
                for module in modules:
                    # 添加主模块
                    ws_ref.append([project.name, module.name, "主模块"])
                    
                    # 获取子模块
                    sub_modules = db.query(Module).filter(
                        Module.parent_id == module.id
                    ).order_by(Module.sort_order).all()
                    
                    for sub_module in sub_modules:
                        # 添加子模块（层级格式）
                        ws_ref.append([project.name, f"{module.name}/{sub_module.name}", "子模块"])
    
    ws_ref.column_dimensions['A'].width = 25
    ws_ref.column_dimensions['B'].width = 40
    ws_ref.column_dimensions['C'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=testcase_template.xlsx"}
    )

@router.get("/statistics")
def get_testcase_statistics(
    project_id: Optional[int] = None,
    project_ids: Optional[str] = None,  # 新增：支持多个项目ID，逗号分隔
    module: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取测试用例统计信息
    
    统计内容:
    1. 总用例数
    2. L1-L4各等级数量
    3. 自动化状态数量(已自动化Y、无法自动化N、未自动化空)
    4. 待评审数量(status=PENDING)
    5. 评审未通过数量(status=REJECTED)
    """
    try:
        from utils.project_helper import get_project_ids_with_children
        
        # 解析项目ID列表
        pid_list = []
        if project_ids:
            pid_list = [int(pid.strip()) for pid in project_ids.split(',') if pid.strip().isdigit()]
        elif project_id:
            pid_list = [project_id]
        
        if not pid_list:
            return {
                "code": 200,
                "message": "获取统计信息成功",
                "data": {
                    "total": 0,
                    "level": {'L1': 0, 'L2': 0, 'L3': 0, 'L4': 0},
                    "automation": {'automated': 0, 'cannot_automate': 0, 'not_automated': 0},
                    "review": {'pending': 0, 'rejected': 0}
                }
            }
        
        # 基础查询条件
        base_filters = [TestCase.primary_project_id.in_(pid_list)]
        
        # 模块筛选（路径前缀匹配，包含子模块）
        # 多项目模式下 path 格式为 "p{pid}/模块路径"
        if module:
            actual_module = module
            matched_project_id = None
            # 处理多项目模式的 "p{pid}/模块路径" 格式
            pid_prefix_match = re.match(r'^p(\d+)/(.+)$', module)
            if pid_prefix_match:
                matched_project_id = int(pid_prefix_match.group(1))
                actual_module = pid_prefix_match.group(2)
            else:
                # 兼容旧格式：尝试匹配项目名前缀
                if pid_list:
                    try:
                        projects = db.query(Project).filter(Project.id.in_(pid_list)).all()
                        project_name_to_id = {p.name: p.id for p in projects}
                        for proj_name, proj_id in project_name_to_id.items():
                            if module.startswith(proj_name + '/'):
                                actual_module = module[len(proj_name) + 1:]
                                matched_project_id = proj_id
                                break
                    except:
                        pass
            
            # 如果匹配到了具体用例库，需要同时用project_id过滤
            if matched_project_id:
                child_ids = get_project_ids_with_children(db, matched_project_id)
                if child_ids:
                    base_filters.append(TestCase.primary_project_id.in_(child_ids))
            
            # 提取主模块名
            main_module = actual_module.split('/')[0] if '/' in actual_module else actual_module
            
            # 精确匹配
            exact_match = or_(
                TestCase.module == actual_module,
                TestCase.module.like(actual_module + '/%')
            )
            # 模糊匹配备用
            fuzzy_match = TestCase.module.like(main_module + '/%')
            
            # 检查是否有精确匹配结果
            exact_count = db.query(TestCase.id).filter(*base_filters).filter(exact_match).count()
            if exact_count > 0:
                base_filters.append(exact_match)
            else:
                base_filters.append(fuzzy_match)
        
        # 总用例数
        total_count = db.query(TestCase).filter(*base_filters).count()
        
        # 等级统计
        level_stats = {
            'L1': db.query(TestCase).filter(*base_filters).filter(TestCase.level == 'L1').count(),
            'L2': db.query(TestCase).filter(*base_filters).filter(TestCase.level == 'L2').count(),
            'L3': db.query(TestCase).filter(*base_filters).filter(TestCase.level == 'L3').count(),
            'L4': db.query(TestCase).filter(*base_filters).filter(TestCase.level == 'L4').count(),
        }
        
        # 自动化状态统计
        automation_stats = {
            'automated': db.query(TestCase).filter(*base_filters).filter(TestCase.automation == 'Y').count(),  # 可自动化
            'completed': db.query(TestCase).filter(*base_filters).filter(TestCase.automation == 'D').count(),  # 已完成自动化
            'cannot_automate': db.query(TestCase).filter(*base_filters).filter(TestCase.automation.like('N-%')).count(),  # 无法自动化（含细分类别）
            'not_automated': db.query(TestCase).filter(*base_filters).filter(or_(TestCase.automation == '', TestCase.automation.is_(None))).count(),  # 未自动化
        }
        
        # 评审状态统计
        review_stats = {
            'pending': db.query(TestCase).filter(*base_filters).filter(TestCase.status == 'PENDING').count(),  # 待评审
            'rejected': db.query(TestCase).filter(*base_filters).filter(TestCase.status == 'REJECTED').count(),  # 评审未通过
        }
        
        return {
            "code": 200,
            "message": "获取统计信息成功",
            "data": {
                "total": total_count,
                "level": level_stats,
                "automation": automation_stats,
                "review": review_stats
            }
        }
    except Exception as e:
        print(f"获取统计信息失败: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取统计信息失败: {str(e)}")

@router.post("/refresh-all-pr-status")
def refresh_all_pr_status(
    project_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """批量刷新项目下所有PR的状态"""
    try:
        import requests
        from config import ZMIND_API_URL, ZMIND_API_KEY
        
        # 使用配置文件中的全局API Key
        if not ZMIND_API_KEY:
            raise HTTPException(status_code=400, detail="系统未配置Zmind API Key")
        
        headers = {
            'X-Redmine-API-Key': ZMIND_API_KEY,
            'Content-Type': 'application/json'
        }
        
        # 获取该项目下所有有PR关联的测试用例
        testcases_with_links = db.query(TestCase).join(
            TestCaseZmindLink, TestCase.id == TestCaseZmindLink.test_case_id
        ).filter(TestCase.primary_project_id == project_id).distinct().all()
        
        updated_count = 0
        failed_count = 0
        
        for testcase in testcases_with_links:
            # 获取该测试用例的所有PR关联
            links = db.query(TestCaseZmindLink).filter(
                TestCaseZmindLink.test_case_id == testcase.id
            ).all()
            
            for link in links:
                try:
                    # 从Zmind API获取最新状态
                    response = requests.get(
                        f"{ZMIND_API_URL}/issues/{link.zmind_issue_id}.json",
                        headers=headers,
                        timeout=10
                    )
                    
                    if response.status_code == 200:
                        data = response.json()
                        issue = data.get('issue', {})
                        
                        # 更新PR信息
                        link.zmind_issue_subject = issue.get('subject', link.zmind_issue_subject)
                        link.zmind_issue_status = issue.get('status', {}).get('name', link.zmind_issue_status)
                        
                        # 获取Severity自定义字段值
                        custom_fields = issue.get('custom_fields', [])
                        for field in custom_fields:
                            if field.get('name') == 'Severity':
                                severity_value = field.get('value')
                                if severity_value:
                                    link.zmind_issue_severity = severity_value
                                break
                        
                        updated_count += 1
                except Exception as e:
                    print(f"刷新PR {link.zmind_issue_id} 状态失败: {str(e)}")
                    failed_count += 1
                    continue
        
        db.commit()
        
        return {
            "code": 200,
            "message": f"刷新完成，成功: {updated_count}，失败: {failed_count}",
            "data": {
                "updated": updated_count,
                "failed": failed_count
            }
        }
    except Exception as e:
        db.rollback()
        print(f"批量刷新PR状态失败: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"刷新失败: {str(e)}")


@router.post("/trigger-pr-sync")
def trigger_pr_sync(
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """手动触发全量PR同步任务（后台异步执行）"""
    from services.zmind_sync_service import ZmindSyncService
    from datetime import datetime

    def _do_sync():
        db_session = None
        try:
            from database import engine
            from sqlalchemy.orm import sessionmaker
            from models import SystemLog
            SessionLocal = sessionmaker(bind=engine)
            db_session = SessionLocal()
            
            service = ZmindSyncService(db_session)
            result = service.sync_all_pr_status()
            logger.info(f"全量PR同步任务完成: {result}")
            
            # 记录到系统日志
            log = SystemLog(
                user_id=current_user.id,
                username=current_user.username,
                module="定时任务",
                action="complete",
                description=f"手动触发Zmind PR同步任务完成: {result}",
                created_at=datetime.now()
            )
            db_session.add(log)
            db_session.commit()
        except Exception as e:
            logger.error(f"全量PR同步任务异常: {e}")
            # 记录错误到系统日志
            try:
                from database import engine
                from sqlalchemy.orm import sessionmaker
                from models import SystemLog
                SessionLocal = sessionmaker(bind=engine)
                db_session = SessionLocal()
                log = SystemLog(
                    user_id=current_user.id,
                    username=current_user.username,
                    module="定时任务",
                    action="error",
                    description=f"手动触发Zmind PR同步任务失败: {str(e)}",
                    created_at=datetime.now()
                )
                db_session.add(log)
                db_session.commit()
                db_session.close()
            except:
                pass
        finally:
            if db_session:
                db_session.close()

    import threading
    t = threading.Thread(target=_do_sync, daemon=True)
    t.start()

    return {
        "code": 200,
        "message": "全量PR同步任务已启动，请在系统日志中查看执行结果",
        "data": {"task_started": True}
    }


@router.get("/pr-sync-status")
def get_sync_status(
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取各项目的PR同步状态"""
    from models import ZmindSyncStatus, Project

    try:
        # 获取所有项目的同步状态
        sync_statuses = db.query(ZmindSyncStatus).all()
        projects = db.query(Project).all()
        project_map = {p.id: p.name for p in projects}

        result = []
        for status in sync_statuses:
            result.append({
                "project_id": status.project_id,
                "project_name": project_map.get(status.project_id, f"项目{status.project_id}"),
                "status": status.status,
                "last_synced_at": status.last_synced_at.isoformat() if status.last_synced_at else None,
                "last_heartbeat": status.last_heartbeat.isoformat() if status.last_heartbeat else None,
                "error_message": status.error_message
            })

        return {"code": 200, "data": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取同步状态失败: {str(e)}")


@router.post("/{testcase_id}/refresh-pr-status")
def refresh_testcase_pr_status(
    testcase_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """刷新单个测试用例下所有PR的状态"""
    try:
        import requests as http_requests
        from config import ZMIND_API_URL, ZMIND_API_KEY
        
        if not ZMIND_API_KEY:
            raise HTTPException(status_code=400, detail="系统未配置Zmind API Key")
        
        headers = {
            'X-Redmine-API-Key': ZMIND_API_KEY,
            'Content-Type': 'application/json'
        }
        
        links = db.query(TestCaseZmindLink).filter(
            TestCaseZmindLink.test_case_id == testcase_id
        ).all()
        
        if not links:
            return {"code": 200, "message": "无关联PR", "data": []}
        
        updated_links = []
        for link in links:
            try:
                response = http_requests.get(
                    f"{ZMIND_API_URL}/issues/{link.zmind_issue_id}.json",
                    headers=headers,
                    timeout=10
                )
                
                if response.status_code == 200:
                    data = response.json()
                    issue = data.get('issue', {})
                    
                    link.zmind_issue_subject = issue.get('subject', link.zmind_issue_subject)
                    link.zmind_issue_status = issue.get('status', {}).get('name', link.zmind_issue_status)
                    
                    custom_fields = issue.get('custom_fields', [])
                    for field in custom_fields:
                        if field.get('name') == 'Severity':
                            severity_value = field.get('value')
                            if severity_value:
                                link.zmind_issue_severity = severity_value
                            break
            except Exception as e:
                print(f"刷新PR {link.zmind_issue_id} 状态失败: {str(e)}")
                continue
            
            updated_links.append(link)
        
        db.commit()
        
        # 重新查询返回最新数据
        links = db.query(TestCaseZmindLink).filter(
            TestCaseZmindLink.test_case_id == testcase_id
        ).all()
        
        return {"code": 200, "message": "success", "data": links}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"刷新失败: {str(e)}")

@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    overwrite: bool = Form(False),
    auto_create_module: bool = Form(False),
    team_id: Optional[int] = Form(None),
    req: Request = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导入Excel/CSV测试用例（异步模式）
    立即返回task_id，后台线程处理，前端通过 /import/{task_id}/progress 轮询进度
    auto_create_module: 是否自动创建不存在的模块（并自动生成4位Tag）
    """
    from services.import_task_service import import_task_manager
    import threading
    
    contents = await file.read()
    filename = file.filename or ''
    
    # 记录文件信息用于调试
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"接收到上传文件: filename={filename}, size={len(contents) if contents else 0}, content_type={file.content_type}")
    if contents:
        logger.info(f"文件头内容: {contents[:20]}")
    is_csv = filename.lower().endswith('.csv')
    
    # 创建任务
    task_id = import_task_manager.create_task()
    
    # 保存当前用户信息（线程中不能用request-scoped的db和user）
    user_id = current_user.id
    username = current_user.username
    
    # 启动后台线程处理导入
    t = threading.Thread(
        target=_do_import_in_background,
        args=(task_id, contents, filename, is_csv, overwrite, auto_create_module,
              team_id, user_id, username),
        daemon=True
    )
    t.start()
    
    return {"code": 200, "data": {"task_id": task_id}, "message": "导入任务已创建"}


@router.get("/import/{task_id}/progress")
def get_import_progress(task_id: str, current_user: User = Depends(get_current_user)):
    """查询导入任务进度"""
    from services.import_task_service import import_task_manager
    
    task = import_task_manager.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在或已过期")
    
    return {
        "code": 200,
        "data": {
            "task_id": task.task_id,
            "status": task.status,
            "message": task.message,
            "progress": task.progress,
            "current": task.current,
            "total": task.total,
            "created": task.created,
            "updated": task.updated,
            "skipped": task.skipped,
            "errors": task.errors,
            "result": task.result,
        }
    }


def _do_import_in_background(
    task_id: str, contents: bytes, filename: str, is_csv: bool,
    overwrite: bool, auto_create_module: bool, team_id: Optional[int], user_id: int, username: str
):
    """后台线程执行导入逻辑，使用独立的数据库会话"""
    from services.import_task_service import import_task_manager
    from database import SessionLocal
    from datetime import datetime
    from utils.excel import parse_excel_v2, parse_csv
    import logging
    logger = logging.getLogger(__name__)
    
    db = SessionLocal()
    try:
        logger.info(f"导入任务 {task_id} 开始，文件大小: {len(contents) if contents else 0} bytes")
        _execute_import(
            db, task_id, contents, filename, is_csv, auto_create_module,
            team_id, user_id, username, import_task_manager, logger
        )
    except Exception as e:
        db.rollback()
        logger.error(f"导入任务 {task_id} 异常: {str(e)}")
        import traceback
        traceback.print_exc()
        error_msg = str(e)
        if error_msg.startswith("无法读取Excel文件"):
            import_task_manager.update_task(task_id, status="error", message=error_msg)
        else:
            import_task_manager.update_task(task_id, status="error", message=f"导入失败: {error_msg}")
    finally:
        db.close()


def _execute_import(
    db, task_id, contents, filename, is_csv, auto_create_module,
    team_id, user_id, username, task_mgr, logger
):
    """实际的导入执行逻辑（在后台线程中运行）"""
    from datetime import datetime
    from utils.excel import parse_excel_v2, parse_csv

    file_type = 'CSV' if is_csv else 'Excel'

    logger.info(f"=== 开始导入{file_type} (task={task_id}) ===")
    logger.info(f"用户: {username}")
    logger.info(f"文件名: {filename}, 文件大小: {len(contents)} bytes")

    # 解析阶段
    task_mgr.update_task(task_id, status="parsing", message=f"正在解析{file_type}文件...", progress=0)

    if is_csv:
        testcases = parse_csv(contents, user_id, db)
    else:
        testcases = parse_excel_v2(contents, user_id, db, task_id, task_mgr)
    total = len(testcases)
    logger.info(f"解析成功，共 {total} 个用例")

    task_mgr.update_task(task_id, status="parsing", message=f"解析完成，共 {total} 个用例", progress=100, total=total)

    # 验证用例库
    projects = db.query(Project).all()
    project_name_to_id = {p.name: p.id for p in projects}

    invalid_projects = set()
    for tc in testcases:
        project_name = tc.get('project_name')
        if project_name and project_name not in project_name_to_id:
            invalid_projects.add(project_name)

    if invalid_projects:
        invalid_list = ', '.join(invalid_projects)
        msg = f'以下用例库不存在: {invalid_list}。请先在系统中创建这些用例库并设置Tag。'
        task_mgr.update_task(task_id, status="error", message=msg)
        return

    # 清理模块名的辅助函数
    def clean_module_path(module_path):
        """清理模块路径，修复常见格式问题"""
        if not module_path:
            return module_path
        # 移除前导/尾随空格
        module_path = module_path.strip()
        # 将多个连续空格替换为单个空格
        module_path = re.sub(r'\s+', ' ', module_path)
        # 修复 "模块名/ 子模块" 格式（斜杠后有多余空格）
        module_path = re.sub(r'/\s+', '/', module_path)
        # 修复 "模块名 /子模块" 格式（斜杠前有多余空格）
        module_path = re.sub(r'\s+/', '/', module_path)
        return module_path

    # 构建模块路径缓存和排序映射
    task_mgr.update_task(task_id, status="parsing", message="正在验证模块信息...", progress=100)

    verify_project_ids = set()
    for tc in testcases:
        pname = tc.get('project_name')
        pid = project_name_to_id.get(pname)
        if pid:
            verify_project_ids.add(pid)

    module_path_cache = {}
    module_sort_map = {}  # 用于排序的模块sort_order映射
    
    for vpid in verify_project_ids:
        all_modules = db.query(Module).filter(Module.project_id == vpid).all()
        mod_map = {m.id: m for m in all_modules}
        paths = set()
        for m in all_modules:
            # 清理模块名称
            name = clean_module_path(m.name) if m.name else ''
            parts = [name] if name else []
            current = m
            while current.parent_id and current.parent_id in mod_map:
                current = mod_map[current.parent_id]
                parent_name = clean_module_path(current.name) if current.name else ''
                if parent_name:
                    parts.insert(0, parent_name)
            if parts:
                full_path = '/'.join(parts)
                paths.add(full_path)
                # 构建排序key: 主模块sort_order.子模块sort_order.完整路径
                main_sort = m.sort_order if m.parent_id is None else 0
                if m.parent_id and m.parent_id in mod_map:
                    parent = mod_map[m.parent_id]
                    main_sort = parent.sort_order
                sub_sort = m.sort_order if m.parent_id is not None else 0
                sort_key = (main_sort, sub_sort, full_path)
                module_sort_map[(vpid, full_path)] = sort_key
        module_path_cache[vpid] = paths

    # ========== 预验证阶段：检查所有用例数据，如果有错误则不导入任何数据 ==========
    task_mgr.update_task(task_id, status="validating", message="正在验证用例数据...", progress=100)
    
    validation_errors = []
    for idx, tc in enumerate(testcases, 1):
        project_name = tc.get('project_name')
        current_project_id = project_name_to_id.get(project_name)
        
        if not current_project_id:
            validation_errors.append(f"第{idx}行: 用例库 '{project_name}' 不存在")
            continue
        
        module_path = tc.get('module_path', '未分类')
        if not module_path:
            module_path = '未分类'
        module_path = clean_module_path(module_path)
        
        valid_paths = module_path_cache.get(current_project_id, set())
        if module_path not in valid_paths:
            if auto_create_module:
                continue
            else:
                validation_errors.append(f"第{idx}行: 模块 '{module_path}' 在用例库 '{project_name}' 中不存在")
    
    if validation_errors:
        error_msg = "验证失败，发现以下问题：\n" + "\n".join(validation_errors[:20])
        if len(validation_errors) > 20:
            error_msg += f"\n...还有 {len(validation_errors) - 20} 个错误"
        task_mgr.update_task(task_id, status="error", message=error_msg)
        return

    # 按模块排序顺序对用例进行排序
    # 排序规则：按Excel中从前到后的出现顺序
    # 1. 主模块：按Excel中首次出现的顺序
    # 2. 子模块：在对应主模块下按首次出现的顺序
    # 3. 已存在于数据库的模块：排在Excel新模块之后，保持原有排序
    excel_main_module_order = {}  # (pid, main_module_name) -> 首次出现的行号
    excel_sub_module_order = {}  # (pid, full_module_path) -> 首次出现的行号
    
    for idx, tc in enumerate(testcases):
        pname = tc.get('project_name')
        pid = project_name_to_id.get(pname)
        if not pid:
            continue
        mp = clean_module_path(tc.get('module_path', '未分类')) if tc.get('module_path') else '未分类'
        main_mod = mp.split('/')[0] if mp else '未分类'
        
        # 记录主模块首次出现顺序
        main_key = (pid, main_mod)
        if main_key not in excel_main_module_order:
            excel_main_module_order[main_key] = idx
        
        # 记录子模块首次出现顺序
        sub_key = (pid, mp)
        if sub_key not in excel_sub_module_order:
            excel_sub_module_order[sub_key] = idx

    def get_case_sort_key(tc):
        project_name = tc.get('project_name')
        pid = project_name_to_id.get(project_name)
        module_path = tc.get('module_path', '未分类')
        module_path = clean_module_path(module_path) if module_path else '未分类'
        main_mod = module_path.split('/')[0] if module_path else '未分类'
        sub_mod = module_path
        
        # 无论模块是否存在于数据库，都优先使用Excel中的顺序
        # Excel中的顺序：主模块按首次出现顺序，子模块在对应主模块下按首次出现顺序
        main_order = excel_main_module_order.get((pid, main_mod), 999999)
        sub_order = excel_sub_module_order.get((pid, sub_mod), 0)
        
        # 尝试获取数据库中的sort_order作为辅助排序（存在时使用，不存在则用0）
        db_sort_key = module_sort_map.get((pid, module_path), (0, 0, ''))
        db_main_sort = db_sort_key[0] if db_sort_key else 0
        db_sub_sort = db_sort_key[1] if db_sort_key else 0
        
        # 排序规则：优先按Excel中的顺序，相同模块内再按数据库sort_order
        # 使用 (Excel主模块顺序, Excel子模块顺序, 数据库主模块sort_order, 数据库子模块sort_order, 模块路径)
        return (main_order, sub_order, db_main_sort, db_sub_sort, module_path)
    
    testcases = sorted(testcases, key=get_case_sort_key)

    # ========== 性能优化：预加载所有数据到内存 ==========
    task_mgr.update_task(task_id, status="preloading", message="正在准备导入数据...", progress=100)
    
    # 预加载所有项目信息（包括tag）
    all_projects = db.query(Project).all()
    project_info_cache = {p.id: {'tag': p.tag or re.sub(r'[^a-zA-Z0-9]', '', p.name) or 'PRJ', 'name': p.name} for p in all_projects}
    
    # 预加载所有模块信息（按项目分组）
    all_modules = db.query(Module).all()
    modules_by_project = {}
    for m in all_modules:
        if m.project_id not in modules_by_project:
            modules_by_project[m.project_id] = {}
        modules_by_project[m.project_id][m.id] = m
    
    # 预构建模块路径映射（id -> path）和反向映射（path -> id）
    module_path_to_id = {}
    module_id_to_path = {}
    for pid, mods in modules_by_project.items():
        mod_map = {m.id: m for m in mods.values()}
        for m in mods.values():
            parts = []
            current = m
            while current and current.id in mod_map:
                if current.name:
                    parts.insert(0, current.name)
                current = mod_map.get(current.parent_id) if current.parent_id else None
            if parts:
                full_path = '/'.join(parts)
                module_path_to_id[(pid, full_path)] = m.id
                module_id_to_path[m.id] = full_path
    
    # 预加载每个项目的最大序号（一次性查询，避免N+1）
    max_seq_by_prefix = {}
    for pid in verify_project_ids:
        # 查询所有以项目tag开头的用例编号
        proj_info = project_info_cache.get(pid, {})
        proj_tag = proj_info.get('tag', 'PRJ')
        
        # 查询该项目下所有模块的tag
        project_mods = modules_by_project.get(pid, {})
        module_tags = {}
        for m in project_mods.values():
            if m.tag:
                full_path = module_id_to_path.get(m.id, m.name)
                main_name = full_path.split('/')[0] if full_path else m.name
                module_tags[main_name] = m.tag
        
        # 生成可能的前缀
        prefixes = [f"{proj_tag}_"]
        for mod_name, mod_tag in module_tags.items():
            prefixes.append(f"{proj_tag}_{mod_tag}_")
        
        # 批量查询最大序号
        for prefix in prefixes:
            max_result = db.query(func.max(TestCase.case_number)).filter(
                TestCase.primary_project_id == pid,
                TestCase.case_number.like(f"{prefix}%")
            ).scalar()
            if max_result:
                match = re.search(r'(\d+)$', max_result)
                if match:
                    max_seq_by_prefix[(pid, prefix)] = int(match.group(1))
    
    # 预统计每个子模块的用例数（包含所有子模块及其深层子模块）
    sub_module_case_count = {}
    for pid, mods in modules_by_project.items():
        for mod in mods.values():
            if mod.parent_id:
                full_path = module_id_to_path.get(mod.id, mod.name)
                # 统计该子模块及其所有深层子模块的用例数（与原逻辑一致）
                count = db.query(func.count(TestCase.id)).filter(
                    TestCase.primary_project_id == pid,
                    or_(
                        TestCase.module == full_path,
                        TestCase.module.like(f"{full_path}/%")
                    )
                ).scalar() or 0
                sub_module_case_count[(pid, full_path)] = count
    
    # 预构建主模块tag缓存 (project_id, main_module_name) -> tag
    sub_module_tags_cache = {}
    for pid, mods in modules_by_project.items():
        for m in mods.values():
            if m.tag and m.parent_id is None:
                full_path = module_id_to_path.get(m.id, m.name)
                main_name = full_path.split('/')[0] if full_path else m.name
                sub_module_tags_cache[(pid, main_name)] = m.tag
    
    # 预加载每个模块的最大sort_order（用于新用例排序）
    max_sort_order_by_module = {}
    for pid, mods in modules_by_project.items():
        for mod in mods.values():
            full_path = module_id_to_path.get(mod.id, mod.name)
            if full_path:
                max_sort_order_by_module[(pid, full_path)] = mod.sort_order
    
    # 序号递增缓存
    seq_increment = {}
    
    # sort_order递增缓存
    sort_order_increment = {}
    
    # ========== 预加载完成 ==========
    logger.info(f"数据预加载完成：{len(project_info_cache)} 个项目，{len(module_path_to_id)} 个模块路径")

    # 导入阶段：直接创建用例，不判断是否已存在
    created_count = 0
    skipped_count = 0
    errors = []
    skipped_cases = []
    batch_size = max(1, total // 50)
    
    # 批量创建用例，每100条提交一次
    for idx, testcase_data in enumerate(testcases, 1):
        try:
            project_name = testcase_data.get('project_name')
            current_project_id = project_name_to_id.get(project_name)

            if not current_project_id:
                errors.append(f"用例库 '{project_name}' 不存在")
                skipped_cases.append(testcase_data.get('name') or f'用例{idx}')
                skipped_count += 1
                continue

            module_path = testcase_data.get('module_path', '未分类')
            if not module_path:
                module_path = '未分类'
            # 使用清理函数处理模块路径
            module_path = clean_module_path(module_path)
            valid_paths = module_path_cache.get(current_project_id, set())
            if module_path not in valid_paths:
                if auto_create_module:
                    ensure_module_exists(db=db, project_id=current_project_id, module_name=module_path, user_id=user_id, auto_create_tag=True)
                    if current_project_id not in module_path_cache:
                        module_path_cache[current_project_id] = set()
                    module_path_cache[current_project_id].add(module_path)
                    # 刷新模块信息缓存，确保新创建的模块及其Tag被包含
                    if current_project_id not in modules_by_project:
                        modules_by_project[current_project_id] = {}
                    # 重新查询该模块的tag并更新缓存
                    main_mod_name = module_path.split('/')[0] if module_path else None
                    if main_mod_name:
                        new_mod = db.query(Module).filter(
                            Module.project_id == current_project_id,
                            Module.name == main_mod_name,
                            Module.parent_id.is_(None)
                        ).first()
                        if new_mod and new_mod.tag:
                            sub_module_tags_cache[(current_project_id, main_mod_name)] = new_mod.tag
                else:
                    errors.append(f"模块 '{module_path}' 在用例库 '{project_name}' 中不存在")
                    skipped_cases.append(testcase_data.get('name') or f'用例{idx}')
                    skipped_count += 1
                    continue
            
            # 检查是否需要自动创建主模块名/主模块名的子模块
            # 条件：模块是主模块，且该主模块下有其他子模块
            module_parts = module_path.split('/')
            main_mod_name = module_parts[0] if module_parts else None
            if main_mod_name and len(module_parts) == 1:  # 当前是主模块
                # 去掉主模块名中的（Tag）后缀
                clean_main_mod_name = re.sub(r'\s*（[^）]+）\s*$', '', main_mod_name.strip())
                if not clean_main_mod_name:
                    clean_main_mod_name = main_mod_name
                
                # 检查该主模块下是否有其他子模块
                main_mod = db.query(Module).filter(
                    Module.project_id == current_project_id,
                    Module.name == main_mod_name,
                    Module.parent_id.is_(None)
                ).first()
                if main_mod:
                    # 检查是否有子模块
                    has_sub_modules = db.query(Module.id).filter(
                        Module.parent_id == main_mod.id
                    ).first() is not None
                    if has_sub_modules:
                        # 自动创建 主模块名/主模块名 的子模块
                        sub_module_name = f"{clean_main_mod_name}/{clean_main_mod_name}"
                        # 检查是否已存在
                        existing_sub = db.query(Module).filter(
                            Module.project_id == current_project_id,
                            Module.name == clean_main_mod_name,
                            Module.parent_id == main_mod.id
                        ).first()
                        if not existing_sub:
                            # 创建子模块
                            max_sort = db.query(func.max(Module.sort_order)).filter(
                                Module.project_id == current_project_id,
                                Module.parent_id == main_mod.id
                            ).scalar() or 0
                            new_sub_module = Module(
                                project_id=current_project_id,
                                name=clean_main_mod_name,
                                tag=main_mod.tag,  # 继承主模块的tag
                                parent_id=main_mod.id,
                                sort_order=max_sort + 10,
                                created_by=user_id
                            )
                            db.add(new_sub_module)
                            db.flush()
                        else:
                            sub_module_name = f"{clean_main_mod_name}/{clean_main_mod_name}"
                        # 更新module_path
                        module_path = sub_module_name

            # 新建用例：自动生成编号（使用预加载的缓存数据）
            proj_info = project_info_cache.get(current_project_id, {})
            proj_tag = proj_info.get('tag', 'PRJ')
            
            # 获取模块tag（仅当模块真正设置了Tag时才使用，否则只用用例库Tag）
            mod_tag = None
            module_parts = module_path.split('/')
            main_mod = module_parts[0] if module_parts else None
            if main_mod:
                mod_tag = sub_module_tags_cache.get((current_project_id, main_mod))
            
            # 构建前缀：规则是
            # - 有主模块Tag：{用例库Tag}_{主模块Tag}_
            # - 无主模块Tag：{用例库Tag}_
            if mod_tag:
                prefix = f"{proj_tag}_{mod_tag}_"
            else:
                prefix = f"{proj_tag}_"
            
            # 获取当前序号（跨子模块连续编号）
            current_key = (current_project_id, prefix)
            current_seq = seq_increment.get(current_key, 0)
            
            # 预加载的最大序号 + 缓存序号
            cached_max = max_seq_by_prefix.get(current_key, 0)
            new_seq = max(cached_max + current_seq + 1, 1)
            
            # 构建编号
            case_number = f"{prefix}{new_seq:04d}"
            
            # 更新缓存
            seq_increment[current_key] = current_seq + 1
            
            # 计算sort_order：新用例添加到模块最后
            sort_key = (current_project_id, module_path)
            base_sort = max_sort_order_by_module.get(sort_key, 0)
            current_sort_offset = sort_order_increment.get(sort_key, 0)
            sort_order = base_sort + current_sort_offset + 1
            sort_order_increment[sort_key] = current_sort_offset + 1
            
            db_data = {
                "primary_project_id": current_project_id,
                "case_number": case_number,
                "case_type": testcase_data.get('case_type', 'COMMON'),
                "module": module_path,
                "name": testcase_data.get('name'),
                "precondition": testcase_data.get('precondition'),
                "steps": testcase_data.get('steps'),
                "expected_result": testcase_data.get('expected_result'),
                "level": testcase_data.get('level', 'L3'),
                "remarks": testcase_data.get('remarks'),
                "automation": testcase_data.get('automation'),
                "status": "PENDING",
                "source": "LOCAL",
                "created_by": user_id,
                "sort_order": sort_order,
            }

            db_testcase = TestCase(**db_data)
            db.add(db_testcase)
            db.flush()  # 获取 testcase id

            db.add(TestCaseProject(
                test_case_id=db_testcase.id,
                project_id=current_project_id,
                relation_type='OWNED',
                is_editable=1,
                created_at=datetime.now(),
                created_by=user_id
            ))

            db.add(TestCaseHistory(
                testcase_id=db_testcase.id,
                field_name='创建用例',
                old_value=None,
                new_value=f"{db_testcase.case_number} - {db_testcase.name}",
                changed_by=user_id,
                changed_by_name=username
            ))
            created_count += 1

            # 每100条提交一次
            if idx % 100 == 0:
                db.commit()

        except Exception as e:
            error_msg = f"处理用例 {testcase_data.get('case_number', 'unknown')} 失败: {str(e)}"
            logger.error(error_msg)
            errors.append(error_msg)

        if idx % batch_size == 0 or idx == total:
            progress = round(idx / total * 100)
            task_mgr.update_task(
                task_id, status="importing",
                message=f"正在导入 {idx}/{total}...",
                progress=progress, current=idx, total=total,
                created=created_count, skipped=skipped_count
            )

    # 提交阶段
    task_mgr.update_task(task_id, status="committing", message="正在保存到数据库...", progress=100)

    db.commit()
    logger.info(f"导入完成 (task={task_id}): 新建 {created_count}, 跳过 {skipped_count}")

    log_operation(
        db=db,
        user_id=user_id,
        username=username,
        module=LogModule.TESTCASES,
        action=LogAction.IMPORT,
        description=f"从{file_type}导入测试用例：新建 {created_count} 个，跳过 {skipped_count} 个",
        request=None
    )

    result = {
        'created': created_count, 
        'skipped': skipped_count, 
        'errors': errors,
        'skipped_cases': skipped_cases
    }
    task_mgr.update_task(
        task_id, status="done", message="导入完成",
        progress=100, created=created_count, skipped=skipped_count, errors=errors, result=result
    )

@router.get("/export")
def export_excel(
    project_id: int,
    module: Optional[str] = None,
    req: Request = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 权限检查
    if not is_super_admin(current_user) and not has_permission(current_user, db, 'testcases.export'):
        raise HTTPException(status_code=403, detail="您没有导出测试用例的权限")

    # 与列表接口保持一致：展开子项目 ID，确保导出范围与页面显示一致
    from utils.project_helper import get_project_ids_with_children
    project_ids_list = get_project_ids_with_children(db, project_id)

    # 构建查询
    if project_ids_list:
        query = db.query(TestCase).filter(TestCase.primary_project_id.in_(project_ids_list))
    else:
        query = db.query(TestCase).filter(TestCase.primary_project_id == project_id)

    # 添加模块筛选（路径前缀匹配）
    if module:
        query = query.filter(
            or_(
                TestCase.module == module,
                TestCase.module.like(module + '/%')
            )
        )
    
    # 应用标准排序
    query = apply_testcase_sort(query, project_id, db)

    # 先统计总数，避免超大数据集一次性加载到内存
    # 超过 5000 条时分批加载（yield_per），防止 OOM；generate_excel 仍接收完整列表
    total_count = query.count()
    BATCH_THRESHOLD = 5000
    if total_count > BATCH_THRESHOLD:
        # 分批加载，避免一次性建立所有 ORM 对象
        testcases = list(query.yield_per(500))
    else:
        testcases = query.all()
    
    # 当用例库为空时，导出文件需要在表头下方填入用例库名称
    # 这样用户可以基于导出的模板直接填写用例（避免手填名称错误）
    project = db.query(Project).filter(Project.id == project_id).first()
    project_name = project.name if project else None
    
    excel_data = generate_excel(testcases, db=db, default_project_name=project_name)
    
    # 构建文件名
    if module:
        filename = f"testcase_{module.replace('/', '_')}.xlsx"
    else:
        filename = "testcase_all.xlsx"
    
    # 构建导出描述
    export_desc = f"导出 {len(testcases)} 个测试用例到Excel"
    if module:
        export_desc += f" (模块: {module})"
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTCASES,
        action=LogAction.EXPORT,
        description=export_desc,
        request=req
    )
    
    # URL编码文件名以支持中文
    from urllib.parse import quote
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        io.BytesIO(excel_data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

@router.get("/{testcase_id}")
def get_testcase(
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    return {"code": 200, "message": "success", "data": testcase}

@router.put("/{testcase_id}")
def update_testcase(
    testcase_id: int,
    testcase: TestCaseCreate,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    db_testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not db_testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    # 验证 level 字段
    testcase_data = testcase.dict()
    if testcase_data.get('level') not in ['L1', 'L2', 'L3', 'L4']:
        raise HTTPException(status_code=400, detail="用例等级必须是 L1/L2/L3/L4")
    
    # 验证 automation 字段
    automation = testcase_data.get('automation')
    if automation:
        valid_values = ['Y', 'D', 'N', '']
        valid_values.extend([
            'N-HW_PHYSICAL', 'N-VISUAL_JUDGE', 'N-BOOT_PROCESS', 'N-MEDIA_PLAY',
            'N-OTA_UPGRADE', 'N-DATA_CONFIG', 'N-LOG_CHECK', 'N-BACKEND_CONFIG', 'N-DATA_DYNAMIC', 'N-OTHER_SPECIAL'
        ])
        if automation not in valid_values:
            raise HTTPException(status_code=400, detail="自动化字段值无效")
    
    # 自动创建模块(如果不存在)
    ensure_module_exists(
        db=db,
        project_id=testcase_data['primary_project_id'],
        module_name=testcase_data.get('module'),
        user_id=current_user.id
    )
    
    # 检测用例库或模块是否变更，变更则重新生成编号、迁移关联、更新排序
    old_project_id = db_testcase.primary_project_id
    new_project_id = testcase_data.get('primary_project_id', old_project_id)
    old_module = db_testcase.module or ''
    new_module = testcase_data.get('module') or ''
    project_or_module_changed = (old_project_id != new_project_id) or (old_module.strip() != new_module.strip())

    if project_or_module_changed:
        new_case_number = generate_case_number(
            db=db,
            project_id=new_project_id,
            module_path=new_module
        )
        testcase_data['case_number'] = new_case_number
        
        # 移入新模块时，sort_order 追加到末尾
        max_sort = db.query(func.max(TestCase.sort_order)).filter(
            TestCase.primary_project_id == new_project_id,
            TestCase.module == new_module
        ).scalar() or 0
        testcase_data['sort_order'] = max_sort + 10
        
        # 如果用例库变更，迁移 TestCaseProject 关联
        if old_project_id != new_project_id:
            old_link = db.query(TestCaseProject).filter(
                TestCaseProject.test_case_id == testcase_id,
                TestCaseProject.project_id == old_project_id,
                TestCaseProject.relation_type == 'OWNED'
            ).first()
            if old_link:
                db.delete(old_link)
            existing_new = db.query(TestCaseProject).filter(
                TestCaseProject.test_case_id == testcase_id,
                TestCaseProject.project_id == new_project_id
            ).first()
            if not existing_new:
                new_link = TestCaseProject(
                    test_case_id=testcase_id,
                    project_id=new_project_id,
                    relation_type='OWNED',
                    is_editable=1,
                    created_at=datetime.now(),
                    created_by=current_user.id
                )
                db.add(new_link)
    
    old_name = db_testcase.name
    changes = []  # 记录变更内容（用于日志）
    
    def _parse_steps_text(raw):
        """将步骤/预期结果的JSON格式解析为可读文本"""
        if not raw:
            return ''
        try:
            import json as _json
            data = _json.loads(raw)
            if isinstance(data, list):
                parts = []
                for i, item in enumerate(data, 1):
                    if isinstance(item, dict):
                        text = item.get('step') or item.get('result') or item.get('text') or str(item)
                    else:
                        text = str(item)
                    parts.append(f"{i}. {text}")
                return '\n'.join(parts)
        except:
            pass
        return raw
    
    # 定义关键字段：修改这些字段将触发状态重置
    CRITICAL_FIELDS = [
        'case_number', 'module', 'name', 'precondition',
        'steps', 'expected_result', 'level', 'automation', 'case_type'
    ]
    
    # 辅助函数：比较JSON字段是否真正改变（用于状态判断）
    def json_fields_equal_for_status(old_val, new_val):
        """比较两个JSON字段是否相等（忽略格式差异）"""
        if old_val == new_val:
            return True
        try:
            import json
            old_json = json.loads(old_val) if isinstance(old_val, str) else old_val
            new_json = json.loads(new_val) if isinstance(new_val, str) else new_val
            return old_json == new_json
        except:
            return old_val == new_val
    
    # 辅助函数：深度比较 steps 字段（忽略序号差异）
    def steps_equal_ignore_prefix(old_val, new_val):
        """比较 steps 字段，忽略序号差异"""
        try:
            import json
            old_json = json.loads(old_val) if isinstance(old_val, str) else old_val
            new_json = json.loads(new_val) if isinstance(new_val, str) else new_val
            
            if isinstance(old_json, list) and isinstance(new_json, list) and len(old_json) == len(new_json):
                def remove_prefix(text):
                    if not text:
                        return ''
                    text = str(text).strip()
                    text = re.sub(r'^(\d+)([.。、])\s*', '', text)
                    return text.strip()
                
                for old_item, new_item in zip(old_json, new_json):
                    old_step = remove_prefix(old_item.get('step', ''))
                    new_step = remove_prefix(new_item.get('step', ''))
                    old_expected = remove_prefix(old_item.get('expected', ''))
                    new_expected = remove_prefix(new_item.get('expected', ''))
                    
                    if old_step != new_step or old_expected != new_expected:
                        return False
                return True
        except:
            pass
        return False
    
    # 辅助函数：规范化文本比较
    def text_equal_normalized(old_val, new_val):
        """规范化后比较文本（移除所有空白字符）"""
        try:
            def normalize(text):
                if not text:
                    return ''
                return ''.join(str(text).split())
            return normalize(old_val) == normalize(new_val)
        except:
            return False
    
    # 检查是否修改了关键字段（使用智能比较）
    critical_field_changed = False
    old_status = db_testcase.status
    if db_testcase.status in ['REVIEWED', 'REJECTED']:
        for field in CRITICAL_FIELDS:
            old_value = getattr(db_testcase, field)
            new_value = testcase_data.get(field)
            
            # 如果字段不在更新数据中，跳过
            if field not in testcase_data:
                continue
            
            # 对于 steps 和 expected_result 字段，使用智能比较
            if field == 'steps':
                # 先尝试 JSON 比较
                if json_fields_equal_for_status(old_value, new_value):
                    continue
                # 再尝试深度比较（忽略序号）
                if steps_equal_ignore_prefix(old_value, new_value):
                    continue
                # 如果都不相同，说明真的修改了
                critical_field_changed = True
                break
            elif field == 'expected_result':
                # 先尝试 JSON 比较
                if json_fields_equal_for_status(old_value, new_value):
                    continue
                # 再尝试规范化比较
                if text_equal_normalized(old_value, new_value):
                    continue
                # 如果都不相同，说明真的修改了
                critical_field_changed = True
                break
            else:
                # 其他字段使用简单字符串比较
                old_str = str(old_value) if old_value is not None else ''
                new_str = str(new_value) if new_value is not None else ''
                if old_str != new_str:
                    critical_field_changed = True
                    break
    
    # 如果是已评审或评审未通过状态且关键字段被修改，重置为待评审
    status_auto_changed = False
    if critical_field_changed:
        testcase_data['status'] = 'PENDING'
        status_auto_changed = True  # 标记状态是自动变更的
    
    # 辅助函数：格式化步骤和结果（将JSON数组转换为可读文本）
    def format_steps_or_results(value, field_type='steps'):
        """将步骤或结果的JSON格式转换为纯文本"""
        if not value:
            return ""
        try:
            import json
            data = json.loads(value) if isinstance(value, str) else value
            if isinstance(data, list):
                lines = []
                for i, item in enumerate(data, 1):
                    if isinstance(item, dict):
                        if field_type == 'steps':
                            # 操作步骤只显示step字段
                            step = item.get('step', '')
                            if step:
                                lines.append(f"{i}. {step}")
                        else:
                            # 预期结果只显示expected字段
                            expected = item.get('expected', '')
                            if expected:
                                lines.append(f"{i}. {expected}")
                    else:
                        lines.append(f"{i}. {item}")
                return "\n".join(lines)
            return str(data)
        except:
            return str(value)
    
    # 辅助函数：格式化用例类型
    def format_case_type(value):
        """将用例类型转换为中文"""
        case_type_map = {
            'COMMON': '通用',
            'FUNCTIONAL': '功能测试',
            'PERFORMANCE': '性能测试',
            'SECURITY': '安全测试',
            'INTERFACE': '接口测试',
            'INSTALLATION': '安装部署',
            'CONFIGURATION': '配置相关',
            'OTHER': '其他'
        }
        return case_type_map.get(value, value) if value else ''
    
    # 字段映射（中文名称，用于历史记录）
    field_names = {
        'case_number': '用例编号',
        'module': '所属模块',
        'name': '用例标题',
        'precondition': '前置条件',
        'steps': '操作步骤',
        'expected_result': '预期结果',
        'level': '用例等级',
        'automation': '自动化',
        'status': '状态',
        'case_type': '用例类型',
        'tags': '标签',
        'archive_source': '归档来源',
        'remarks': '备注'
    }
    
    # 比较并记录变更到历史表
    # 记录哪些关键字段被修改了（用于状态自动变更时的说明）
    changed_critical_fields = []
    
    # 辅助函数：比较JSON字段是否真正改变
    def json_fields_equal(old_val, new_val):
        """比较两个JSON字段是否相等（忽略格式差异）"""
        if old_val == new_val:
            return True
        try:
            import json
            old_json = json.loads(old_val) if isinstance(old_val, str) else old_val
            new_json = json.loads(new_val) if isinstance(new_val, str) else new_val
            return old_json == new_json
        except:
            return old_val == new_val
    
    for field, chinese_name in field_names.items():
        # 只处理实际在更新数据中的字段
        if field not in testcase_data:
            continue
            
        # 如果状态是自动变更的，跳过状态字段的历史记录（稍后会合并到修改字段的记录中）
        if field == 'status' and status_auto_changed:
            continue
            
        old_value = getattr(db_testcase, field)
        new_value = testcase_data.get(field)
        
        # 对于JSON字段（steps）和从JSON生成的字段（expected_result），使用智能比较
        if field in ['steps', 'expected_result']:
            # 首先尝试 JSON 比较
            if json_fields_equal(old_value, new_value):
                continue  # 值实际上没有改变，跳过
            
            # 对于 steps 字段，进行深度 JSON 比较（忽略序号差异）
            if field == 'steps':
                try:
                    import json
                    old_json = json.loads(old_value) if isinstance(old_value, str) else old_value
                    new_json = json.loads(new_value) if isinstance(new_value, str) else new_value
                    
                    if isinstance(old_json, list) and isinstance(new_json, list) and len(old_json) == len(new_json):
                        # 比较每个步骤，忽略序号
                        def remove_prefix(text):
                            """移除文本开头的序号和空格"""
                            if not text:
                                return ''
                            text = str(text).strip()
                            # 移除序号（1. 1。 1、等）
                            text = re.sub(r'^(\d+)([.。、])\s*', '', text)
                            return text.strip()
                        
                        all_same = True
                        for i, (old_item, new_item) in enumerate(zip(old_json, new_json)):
                            old_step = remove_prefix(old_item.get('step', ''))
                            new_step = remove_prefix(new_item.get('step', ''))
                            old_expected = remove_prefix(old_item.get('expected', ''))
                            new_expected = remove_prefix(new_item.get('expected', ''))
                            
                            if old_step != new_step or old_expected != new_expected:
                                all_same = False
                                break
                        
                        if all_same:
                            continue  # 内容实际相同，只是序号格式不同
                except Exception as e:
                    pass
            
            # 对于文本字段，也尝试规范化后比较（去除多余空格、换行和标点符号后的空格）
            try:
                # 更彻底的规范化：移除所有空白字符
                def normalize_text(text):
                    if not text:
                        return ''
                    # 移除所有空白字符
                    text = ''.join(str(text).split())
                    return text
                
                old_normalized = normalize_text(old_value)
                new_normalized = normalize_text(new_value)
                
                if old_normalized == new_normalized:
                    continue  # 内容实际相同，只是格式不同
            except Exception as e:
                # 如果规范化比较失败，继续使用字符串比较
                pass
        
        # 处理None值
        old_value_str = str(old_value) if old_value is not None else ''
        new_value_str = str(new_value) if new_value is not None else ''
        
        # 如果值完全相同，跳过
        if old_value_str == new_value_str:
            continue
        
        # 对不同字段进行格式化
        if field == 'steps':
            old_value_display = format_steps_or_results(old_value, 'steps')
            new_value_display = format_steps_or_results(new_value, 'steps')
        elif field == 'expected_result':
            old_value_display = format_steps_or_results(old_value, 'expected')
            new_value_display = format_steps_or_results(new_value, 'expected')
        elif field == 'case_type':
            old_value_display = format_case_type(old_value)
            new_value_display = format_case_type(new_value)
        else:
            old_value_display = old_value_str
            new_value_display = new_value_str
        
        # 如果这是关键字段且状态会自动变更，记录下来
        if status_auto_changed and field in CRITICAL_FIELDS:
            changed_critical_fields.append(chinese_name)
        
        # 构建字段名称（如果状态自动变更，在第一个修改的关键字段上添加状态变更说明）
        field_name_display = chinese_name
        if status_auto_changed and field in CRITICAL_FIELDS and len(changed_critical_fields) == 1:
            status_map = {
                'PENDING': '待评审',
                'REVIEWED': '已评审',
                'REJECTED': '评审未通过',
                'DEPRECATED': '已废弃'
            }
            old_status_display = status_map.get(old_status, old_status)
            new_status_display = status_map.get('PENDING', '待评审')
            field_name_display = f"{chinese_name}，状态从{old_status_display}变更为{new_status_display}"
        
        history = TestCaseHistory(
            testcase_id=testcase_id,
            field_name=field_name_display,
            old_value=old_value_display,
            new_value=new_value_display,
            changed_by=current_user.id,
            changed_by_name=current_user.username
        )
        db.add(history)
    
    # 记录简要变更（用于日志）
    if testcase.case_number != db_testcase.case_number:
        changes.append(f"用例编号: {db_testcase.case_number} → {testcase.case_number}")
    if testcase.name != db_testcase.name:
        changes.append(f"用例名称: {db_testcase.name} → {testcase.name}")
    if testcase.module != db_testcase.module:
        changes.append(f"模块: {db_testcase.module} → {testcase.module}")
    if testcase.level != db_testcase.level:
        changes.append(f"等级: {db_testcase.level} → {testcase.level}")
    if (testcase.precondition or '') != (db_testcase.precondition or ''):
        changes.append(f"前置条件已更新")
    if testcase.steps != db_testcase.steps:
        # JSON 规范化比较，避免格式差异导致误判
        try:
            import json as _json2
            old_parsed = _json2.loads(db_testcase.steps) if db_testcase.steps else None
            new_parsed = _json2.loads(testcase.steps) if testcase.steps else None
            steps_really_changed = old_parsed != new_parsed
        except:
            steps_really_changed = True
        if steps_really_changed:
            changes.append(f"测试步骤已更新")
    if testcase.expected_result != db_testcase.expected_result:
        try:
            import json as _json3
            old_parsed = _json3.loads(db_testcase.expected_result) if db_testcase.expected_result else None
            new_parsed = _json3.loads(testcase.expected_result) if testcase.expected_result else None
            result_really_changed = old_parsed != new_parsed
        except:
            result_really_changed = True
        if result_really_changed:
            changes.append(f"预期结果已更新")
    
    # 允许修改用例编号
    for key, value in testcase_data.items():
        setattr(db_testcase, key, value)
    db_testcase.updated_by = current_user.id
    
    db.commit()
    
    # 如果用例状态变为待评审，重置未完成评审计划中的评审状态
    if status_auto_changed and testcase_data.get('status') == 'PENDING':
        from models import ReviewPlan, ReviewPlanTestCase

        # 查找包含该用例的所有未完成评审计划
        pending_plans = db.query(ReviewPlanTestCase).join(
            ReviewPlan, ReviewPlanTestCase.review_plan_id == ReviewPlan.id
        ).filter(
            ReviewPlanTestCase.testcase_id == testcase_id,
            ReviewPlan.status.in_(['PENDING', 'IN_PROGRESS'])  # 只影响未完成的计划
        ).all()

        # 重置这些评审计划中该用例的评审状态
        for plan_tc in pending_plans:
            plan_tc.review_status = 'PENDING'
            plan_tc.review_result = None
            plan_tc.reviewer_id = None
            plan_tc.review_comment = None
            plan_tc.reviewed_at = None
            # 关键字段被改后，对应的"评审草稿"也清空，避免基于旧内容做评审
            plan_tc.pending_review_result = None
            plan_tc.pending_review_comment = None
            plan_tc.pending_reviewer_id = None
            plan_tc.pending_reviewed_at = None

        if pending_plans:
            db.commit()
    
    # 只有在有实际变更时才记录日志
    if changes:
        change_detail = "；".join(changes)
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.TESTCASES,
            action=LogAction.UPDATE,
            description=f"更新测试用例：{old_name}（编号: {db_testcase.case_number}，ID: {testcase_id}，{change_detail}）",
            request=req
        )
        
        # 触发通知
        trigger_testcase_notification(
            db=db,
            event_type='updated',
            testcase_id=testcase_id,
            testcase_name=db_testcase.name,
            operator_name=current_user.username,
            changes=change_detail,
            project_id=db_testcase.primary_project_id
        )
    
    return {"code": 200, "message": "success", "data": None}

@router.get("/{testcase_id}/history")
def get_testcase_history(
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取测试用例的历史记录"""
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    history = db.query(TestCaseHistory).filter(
        TestCaseHistory.testcase_id == testcase_id
    ).order_by(TestCaseHistory.changed_at.desc()).all()
    
    return {
        "code": 200,
        "message": "success",
        "data": [
            {
                "id": h.id,
                "field_name": h.field_name,
                "old_value": h.old_value,
                "new_value": h.new_value,
                "changed_by": h.changed_by,
                "changed_by_name": h.changed_by_name,
                "changed_at": h.changed_at.strftime('%Y-%m-%d %H:%M:%S') if h.changed_at else None
            }
            for h in history
        ]
    }


@router.post("/{testcase_id}/feedback")
def submit_testcase_feedback(
    testcase_id: int,
    body: dict,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """提交用例问题反馈"""
    from datetime import datetime as dt

    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")

    content = (body.get("content") or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="反馈内容不能为空")
    if len(content) > 2000:
        raise HTTPException(status_code=400, detail="反馈内容不能超过2000字")

    submitter = current_user.username
    timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S')

    # 新反馈覆盖旧反馈（对话框只显示最新），历史记录永远保留
    new_entry = f"[{timestamp}] {submitter}：{content}"
    testcase.feedback = new_entry

    # 状态从已评审变为待评审
    old_status = testcase.status
    if testcase.status == 'REVIEWED':
        testcase.status = 'PENDING'

    # 记录历史
    db.add(TestCaseHistory(
        testcase_id=testcase_id,
        field_name='问题反馈',
        old_value=None,
        new_value=content,
        changed_by=current_user.id,
        changed_by_name=submitter
    ))

    if old_status == 'REVIEWED' and testcase.status == 'PENDING':
        db.add(TestCaseHistory(
            testcase_id=testcase_id,
            field_name='状态（因问题反馈自动变更）',
            old_value='已评审',
            new_value='待评审',
            changed_by=current_user.id,
            changed_by_name=submitter
        ))

    db.commit()

    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTCASES,
        action=LogAction.UPDATE,
description=f"提交用例问题反馈：{testcase.name}（编号: {testcase.case_number}，ID: {testcase_id}）",
            request=req
        )

    return {"code": 200, "message": "反馈提交成功", "data": None}


@router.delete("/{testcase_id}/feedback")
def clear_testcase_feedback(
    testcase_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """清除用例反馈信息（移除提示）"""
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")

    testcase.feedback = None
    db.commit()

    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTCASES,
        action=LogAction.UPDATE,
        description=f"移除用例问题反馈提示：{testcase.name}（编号: {testcase.case_number}，ID: {testcase_id}）",
        request=req
    )

    return {"code": 200, "message": "反馈已移除", "data": None}


@router.delete("/{testcase_id}")
def delete_testcase(
    testcase_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    db_testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not db_testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    testcase_name = db_testcase.name
    testcase_project_id = db_testcase.primary_project_id
    case_number = db_testcase.case_number
    
    try:
        # 先删除相关的关联记录
        # 1. 删除Zmind关联
        db.query(TestCaseZmindLink).filter(
            TestCaseZmindLink.test_case_id == testcase_id
        ).delete(synchronize_session=False)
        
        # 2. 删除项目关联
        db.query(TestCaseProject).filter(
            TestCaseProject.test_case_id == testcase_id
        ).delete(synchronize_session=False)
        
        # 3. 删除历史记录
        db.query(TestCaseHistory).filter(
            TestCaseHistory.testcase_id == testcase_id
        ).delete(synchronize_session=False)
        
        # 4. 删除附件
        db.query(TestCaseAttachment).filter(
            TestCaseAttachment.test_case_id == testcase_id
        ).delete(synchronize_session=False)
        
        # 5. 删除测试执行附件（先于执行记录）
        execution_ids = [e.id for e in db.query(TestExecution.id).filter(
            TestExecution.test_case_id == testcase_id
        ).all()]
        if execution_ids:
            db.query(TestExecutionAttachment).filter(
                TestExecutionAttachment.execution_id.in_(execution_ids)
            ).delete(synchronize_session=False)

        # 6. 只断开测试执行记录的关联，保留执行记录（快照字段已保存用例信息）
        # 注意：不再删除 test_executions，记录中的快照字段已保存完整用例信息
        db.query(TestExecution).filter(
            TestExecution.test_case_id == testcase_id
        ).update({TestExecution.test_case_id: None}, synchronize_session=False)

        # 7. 保留测试计划关联（只断开 test_case_id 引用）
        db.query(TestPlanTestCase).filter(
            TestPlanTestCase.test_case_id == testcase_id
        ).update({TestPlanTestCase.test_case_id: None}, synchronize_session=False)

        # 8. 删除评审计划关联
        db.query(ReviewPlanTestCase).filter(
            ReviewPlanTestCase.testcase_id == testcase_id
        ).delete(synchronize_session=False)
        
        # 8.5 删除测试套件关联
        from models import TestSuiteTestCase
        db.query(TestSuiteTestCase).filter(
            TestSuiteTestCase.test_case_id == testcase_id
        ).delete(synchronize_session=False)
        
        # 9. 清除执行进度中的引用
        db.query(TestExecutionProgress).filter(
            TestExecutionProgress.current_testcase_id == testcase_id
        ).update({TestExecutionProgress.current_testcase_id: None}, synchronize_session=False)
        
        # 10. 删除测试用例本身
        db.delete(db_testcase)
        
        # 4. 提交所有删除操作
        db.commit()
        
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.TESTCASES,
            action=LogAction.DELETE,
            description=f"删除测试用例：{testcase_name}（编号: {case_number}，ID: {testcase_id}）",
            request=req
        )
        
        # 触发通知
        trigger_testcase_notification(
            db=db,
            event_type='deleted',
            testcase_id=testcase_id,
            testcase_name=testcase_name,
            operator_name=current_user.username,
            project_id=testcase_project_id
        )
        
        return {"code": 200, "message": "success", "data": None}
    except Exception as e:
        db.rollback()
        print(f"删除测试用例错误: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"删除测试用例失败: {str(e)}")


class BatchDeleteRequest(BaseModel):
    ids: list[int]


@router.post("/batch-delete")
def batch_delete_testcases(
    request_data: BatchDeleteRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """批量删除测试用例（异步模式）"""
    from services.delete_task_service import delete_task_manager
    import threading
    
    ids = request_data.ids
    if not ids:
        raise HTTPException(status_code=400, detail="请选择要删除的用例")
    
    if len(ids) > 9999:
        raise HTTPException(status_code=400, detail="单次最多删除9999条用例")
    
    # 创建删除任务
    task_id = delete_task_manager.create_task(total=len(ids))
    
    # 保存用户信息供后台线程使用
    user_id = current_user.id
    username = current_user.username
    
    # 启动后台线程处理删除
    t = threading.Thread(
        target=_do_delete_in_background,
        args=(task_id, ids, user_id, username),
        daemon=True
    )
    t.start()
    
    return {"code": 200, "data": {"task_id": task_id}, "message": "删除任务已创建"}


def _do_delete_in_background(task_id: str, ids: list, user_id: int, username: str):
    """后台线程执行删除逻辑"""
    from services.delete_task_service import delete_task_manager
    from database import SessionLocal
    from utils.logger import log_operation, LogModule, LogAction
    import logging
    logger = logging.getLogger(__name__)
    
    db = SessionLocal()
    try:
        _execute_batch_delete(
            db, task_id, ids, user_id, username, delete_task_manager, logger
        )
    except Exception as e:
        logger.error(f"删除任务 {task_id} 异常: {str(e)}")
        import traceback
        traceback.print_exc()
        delete_task_manager.update_task(task_id, status="error", message=f"删除失败: {str(e)}")
    finally:
        db.close()


def _execute_batch_delete(db, task_id: str, ids: list, user_id: int, username: str, task_mgr, logger):
    """实际的批量删除执行逻辑（批量模式）"""
    success_ids = []
    failed_items = []
    total = len(ids)
    
    logger.info(f"=== 开始批量删除 (task={task_id}, 总数={total}) ===")
    
    task_mgr.update_task(
        task_id,
        status="deleting",
        message=f"正在准备批量删除...",
        progress=0,
        current=0,
        total=total
    )
    
    try:
        # 1. 批量删除Zmind关联
        deleted_zmind = db.query(TestCaseZmindLink).filter(
            TestCaseZmindLink.test_case_id.in_(ids)
        ).delete(synchronize_session=False)
        logger.info(f"删除Zmind关联: {deleted_zmind}条")
        
        # 2. 批量删除项目关联
        deleted_projects = db.query(TestCaseProject).filter(
            TestCaseProject.test_case_id.in_(ids)
        ).delete(synchronize_session=False)
        logger.info(f"删除项目关联: {deleted_projects}条")
        
        # 3. 批量删除历史记录
        deleted_history = db.query(TestCaseHistory).filter(
            TestCaseHistory.testcase_id.in_(ids)
        ).delete(synchronize_session=False)
        logger.info(f"删除历史记录: {deleted_history}条")
        
        # 4. 批量删除附件
        deleted_attachments = db.query(TestCaseAttachment).filter(
            TestCaseAttachment.test_case_id.in_(ids)
        ).delete(synchronize_session=False)
        logger.info(f"删除附件: {deleted_attachments}条")
        
        # 5. 批量删除测试执行附件（先获取执行ID）
        execution_ids = [e[0] for e in db.query(TestExecution.id).filter(
            TestExecution.test_case_id.in_(ids)
        ).distinct().all()]
        if execution_ids:
            deleted_exec_attachments = db.query(TestExecutionAttachment).filter(
                TestExecutionAttachment.execution_id.in_(execution_ids)
            ).delete(synchronize_session=False)
            logger.info(f"删除测试执行附件: {deleted_exec_attachments}条")

        # 6. 只断开测试执行记录的关联，保留执行记录（快照字段已保存用例信息）
        # 注意：不再删除 test_executions，记录中的快照字段已保存完整用例信息
        db.query(TestExecution).filter(
            TestExecution.test_case_id.in_(ids)
        ).update({TestExecution.test_case_id: None}, synchronize_session=False)
        logger.info(f"断开执行记录关联: {len(ids)}条")

        # 7. 保留测试计划关联（只断开 test_case_id 引用）
        updated_testplan = db.query(TestPlanTestCase).filter(
            TestPlanTestCase.test_case_id.in_(ids)
        ).update({TestPlanTestCase.test_case_id: None}, synchronize_session=False)
        logger.info(f"断开测试计划关联: {updated_testplan}条")

        # 8. 批量删除评审计划关联
        deleted_review = db.query(ReviewPlanTestCase).filter(
            ReviewPlanTestCase.testcase_id.in_(ids)
        ).delete(synchronize_session=False)
        logger.info(f"删除评审计划关联: {deleted_review}条")
        
        # 8.5 批量删除测试套件关联
        from models import TestSuiteTestCase
        deleted_suite = db.query(TestSuiteTestCase).filter(
            TestSuiteTestCase.test_case_id.in_(ids)
        ).delete(synchronize_session=False)
        logger.info(f"删除测试套件关联: {deleted_suite}条")
        
        # 9. 清除执行进度中的引用
        db.query(TestExecutionProgress).filter(
            TestExecutionProgress.current_testcase_id.in_(ids)
        ).update({TestExecutionProgress.current_testcase_id: None}, synchronize_session=False)
        
        # 10. 批量删除测试用例本身
        deleted_cases = db.query(TestCase).filter(
            TestCase.id.in_(ids)
        ).delete(synchronize_session=False)
        logger.info(f"删除测试用例: {deleted_cases}条")
        
        # 一次性提交所有删除
        db.commit()
        success_ids = ids
        logger.info(f"批量删除提交成功，共 {len(success_ids)} 条")
        
    except Exception as e:
        db.rollback()
        logger.error(f"批量删除失败: {str(e)}")
        failed_items = [{"id": tid, "reason": str(e)} for tid in ids]
        task_mgr.update_task(task_id, status="error", message=f"删除失败: {str(e)}")
    
    # 记录日志
    try:
        log_operation(
            db=db,
            user_id=user_id,
            username=username,
            module=LogModule.TESTCASES,
            action=LogAction.DELETE,
            description=f"批量删除测试用例：成功{len(success_ids)}条，失败{len(failed_items)}条"
        )
    except Exception as e:
        logger.error(f"记录日志失败: {str(e)}")
    
    # 更新任务状态
    task_mgr.update_task(
        task_id,
        status="done",
        message=f"删除完成：成功{len(success_ids)}条，失败{len(failed_items)}条",
        progress=100,
        current=total,
        total=total,
        success_count=len(success_ids),
        failed_count=len(failed_items),
        success_ids=success_ids,
        failed_items=failed_items
    )
    
    logger.info(f"=== 批量删除完成 (task={task_id}, 成功={len(success_ids)}, 失败={len(failed_items)}) ===")


@router.get("/batch-delete/{task_id}/progress")
def get_delete_progress(task_id: str, current_user: User = Depends(get_current_user)):
    """查询删除任务进度"""
    from services.delete_task_service import delete_task_manager
    
    task = delete_task_manager.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    return {
        "code": 200,
        "data": {
            "task_id": task.task_id,
            "status": task.status,
            "message": task.message,
            "progress": task.progress,
            "current": task.current,
            "total": task.total,
            "success_count": task.success_count,
            "failed_count": task.failed_count,
            "success_ids": task.success_ids if task.status == "done" else [],
            "failed_items": task.failed_items if task.status == "done" else []
        }
    }


@router.post("/{testcase_id}/zmind-links")
def link_zmind_issue(
    testcase_id: int,
    link: ZmindLinkCreate,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    existing = db.query(TestCaseZmindLink).filter(
        TestCaseZmindLink.test_case_id == testcase_id,
        TestCaseZmindLink.zmind_issue_id == link.zmind_issue_id
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="该PR已经关联到此测试用例")
    
    db_link = TestCaseZmindLink(
        test_case_id=testcase_id,
        zmind_issue_id=link.zmind_issue_id,
        zmind_issue_subject=link.zmind_issue_subject,
        zmind_issue_status=link.zmind_issue_status,
        zmind_issue_severity=link.zmind_issue_severity,
        test_plan_id=link.test_plan_id,  # 保存测试计划ID
        created_by=current_user.id,
        created_by_name=current_user.username  # 添加创建者名称
    )
    db.add(db_link)
    db.commit()
    db.refresh(db_link)
    
    # 记录历史
    history = TestCaseHistory(
        testcase_id=testcase_id,
        field_name="关联PR",
        old_value="",
        new_value=f"#{link.zmind_issue_id} - {link.zmind_issue_subject}",
        changed_by=current_user.id,
        changed_by_name=current_user.username
    )
    db.add(history)
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTCASES,
        action=LogAction.CREATE,
        description=f"关联PR：测试用例 {testcase.name}（编号: {testcase.case_number}，ID: {testcase_id}）关联 Zmind Issue #{link.zmind_issue_id}（链接ID: {db_link.id}）",
        request=req
    )
    
    return {"code": 200, "message": "关联成功", "data": db_link}

@router.get("/{testcase_id}/zmind-links")
def get_zmind_links(
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    links = db.query(TestCaseZmindLink).filter(
        TestCaseZmindLink.test_case_id == testcase_id
    ).all()
    
    return {"code": 200, "message": "success", "data": links}

@router.delete("/{testcase_id}/zmind-links/{link_id}")
def unlink_zmind_issue(
    testcase_id: int,
    link_id: int,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    link = db.query(TestCaseZmindLink).filter(
        TestCaseZmindLink.id == link_id,
        TestCaseZmindLink.test_case_id == testcase_id
    ).first()
    
    if not link:
        raise HTTPException(status_code=404, detail="关联不存在")
    
    # 获取测试用例信息
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    testcase_name = testcase.name if testcase else f"ID:{testcase_id}"
    case_number = testcase.case_number if testcase else ""
    
    issue_id = link.zmind_issue_id
    issue_subject = link.zmind_issue_subject
    
    # 记录历史
    history = TestCaseHistory(
        testcase_id=testcase_id,
        field_name="取消关联PR",
        old_value=f"#{issue_id} - {issue_subject}",
        new_value="",
        changed_by=current_user.id,
        changed_by_name=current_user.username
    )
    db.add(history)
    
    db.delete(link)
    db.commit()
    
    log_operation(
        db=db,
        user_id=current_user.id,
        username=current_user.username,
        module=LogModule.TESTCASES,
        action=LogAction.DELETE,
        description=f"取消关联PR：测试用例 {testcase_name}（编号: {case_number}，ID: {testcase_id}）取消关联 Zmind Issue #{issue_id}（链接ID: {link_id}）",
        request=req
    )
    
    return {"code": 200, "message": "取消关联成功", "data": None}

@router.put("/{testcase_id}/zmind-links/{link_id}")
def update_zmind_link(
    testcase_id: int,
    link_id: int,
    link: ZmindLinkUpdate,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新PR关联信息（用于刷新PR状态）"""
    db_link = db.query(TestCaseZmindLink).filter(
        TestCaseZmindLink.id == link_id,
        TestCaseZmindLink.test_case_id == testcase_id
    ).first()
    
    if not db_link:
        raise HTTPException(status_code=404, detail="关联不存在")
    
    # 只更新传递的字段（使用 exclude_unset 避免覆盖未传递的字段）
    update_data = link.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_link, field, value)
    
    db.commit()
    db.refresh(db_link)
    
    return {"code": 200, "message": "更新成功", "data": db_link}


@router.get("/{testcase_id}/execution-history")
def get_testcase_execution_history(
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取测试用例在所有测试计划中的执行历史"""
    from models import TestExecution, TestPlan, User as UserModel
    
    # 验证测试用例是否存在
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    # 查询该用例在所有测试计划中的执行记录
    executions = db.query(TestExecution, TestPlan, UserModel).join(
        TestPlan, TestExecution.test_plan_id == TestPlan.id
    ).outerjoin(
        UserModel, TestExecution.executor_id == UserModel.id
    ).filter(
        TestExecution.test_case_id == testcase_id
    ).order_by(
        TestExecution.executed_at.desc()
    ).all()
    
    result = []
    for execution, testplan, executor in executions:
        # 解析PR快照
        import json
        pr_links_snapshot = []
        if execution.pr_links_snapshot:
            try:
                pr_links_snapshot = json.loads(execution.pr_links_snapshot)
            except:
                pr_links_snapshot = []
        
        result.append({
            "id": execution.id,
            "testplan_id": testplan.id,
            "testplan_name": testplan.name,
            "result": execution.result,
            "executor_id": execution.executor_id,
            "executor_name": executor.username if executor else None,
            "remarks": execution.remarks,
            "actual_result": execution.actual_result,
            "failure_reason": execution.failure_reason,
            "version_info": execution.version_info,
            "pr_links_snapshot": pr_links_snapshot,
            "executed_at": execution.executed_at.isoformat() if execution.executed_at else None
        })
    
    return {"code": 200, "message": "success", "data": result}


# 批量更新测试用例
class BatchUpdateRequest(BaseModel):
    ids: list[int]
    updates: dict

class BatchMoveRequest(BaseModel):
    ids: list[int]
    target_project_id: int
    target_module: str

@router.post("/batch-move")
def batch_move_testcases(
    request: BatchMoveRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """批量移动测试用例到指定用例库的指定模块（跨用例库移动）"""
    try:
        if not request.ids:
            return {"code": 200, "message": "没有需要移动的用例", "data": None}

        # 验证目标用例库
        target_project = db.query(Project).filter(Project.id == request.target_project_id).first()
        if not target_project:
            raise HTTPException(status_code=404, detail="目标用例库不存在")

        # 确保目标模块存在
        ensure_module_exists(
            db=db,
            project_id=request.target_project_id,
            module_name=request.target_module,
            user_id=current_user.id
        )

        # 计算目标模块的当前最大 sort_order
        max_sort = db.query(func.max(TestCase.sort_order)).filter(
            TestCase.primary_project_id == request.target_project_id,
            TestCase.module == request.target_module
        ).scalar() or 0

        moved_count = 0
        # 使用序号缓存避免重复查询 DB，大幅提升批量性能
        seq_cache = {}
        for idx, testcase_id in enumerate(request.ids):
            db_testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
            if not db_testcase:
                continue

            old_project_id = db_testcase.primary_project_id
            old_case_number = db_testcase.case_number
            old_module = db_testcase.module or ''

            # 利用缓存生成编号，避免每调一次 query 一次 DB
            new_case_number = generate_case_number(
                db=db,
                project_id=request.target_project_id,
                module_path=request.target_module,
                seq_cache=seq_cache
            )

            # 更新用例
            old_name = db_testcase.name
            db_testcase.primary_project_id = request.target_project_id
            db_testcase.module = request.target_module
            db_testcase.case_number = new_case_number
            db_testcase.sort_order = max_sort + (idx + 1) * 10
            db_testcase.updated_by = current_user.id
            # 移动属于关键字段变更，已评审/未通过的用例重置为待评审
            if db_testcase.status in ('REVIEWED', 'REJECTED'):
                db_testcase.status = 'PENDING'

            # 处理 TestCaseProject 关联迁移
            if old_project_id != request.target_project_id:
                old_link = db.query(TestCaseProject).filter(
                    TestCaseProject.test_case_id == testcase_id,
                    TestCaseProject.project_id == old_project_id,
                    TestCaseProject.relation_type == 'OWNED'
                ).first()
                if old_link:
                    db.delete(old_link)

                existing_new = db.query(TestCaseProject).filter(
                    TestCaseProject.test_case_id == testcase_id,
                    TestCaseProject.project_id == request.target_project_id
                ).first()
                if not existing_new:
                    new_link = TestCaseProject(
                        test_case_id=testcase_id,
                        project_id=request.target_project_id,
                        relation_type='OWNED',
                        is_editable=1,
                        created_at=datetime.now(),
                        created_by=current_user.id
                    )
                    db.add(new_link)

            # 记录历史
            changes_log = []
            if old_project_id != request.target_project_id:
                old_project_name = db.query(Project.name).filter(Project.id == old_project_id).scalar() or str(old_project_id)
                changes_log.append(f"用例库: {old_project_name} → {target_project.name}")
            if old_module != request.target_module:
                changes_log.append(f"模块: {old_module} → {request.target_module}")
            changes_log.append(f"用例编号: {old_case_number} → {new_case_number}")

            for change in changes_log:
                history = TestCaseHistory(
                    testcase_id=testcase_id,
                    field_name='移动用例',
                    old_value='',
                    new_value=change,
                    changed_by=current_user.id,
                    changed_by_name=current_user.username
                )
                db.add(history)

            moved_count += 1

        db.commit()

        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.TESTCASES,
            action=LogAction.UPDATE,
            description=f"批量移动 {moved_count} 个测试用例到用例库「{target_project.name}」的模块「{request.target_module}」",
            request=req
        )

        return {"code": 200, "message": f"成功移动 {moved_count} 个用例", "data": None}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"批量移动失败: {str(e)}")

@router.post("/batch-update")
def batch_update_testcases(
    request: BatchUpdateRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """批量更新测试用例（优化：使用批量SQL操作）"""
    try:
        if not request.ids:
            return {"code": 200, "message": "没有需要更新的用例", "data": None}
        
        # 定义关键字段（修改后需要重置状态为待评审）
        CRITICAL_FIELDS = [
            'case_number', 'module', 'name', 'precondition',
            'steps', 'expected_result', 'level', 'automation', 'case_type'
        ]
        
        # 检查是否修改了关键字段
        critical_field_updated = any(field in request.updates for field in CRITICAL_FIELDS)
        
        # 过滤出有效的更新字段
        valid_updates = {}
        for key, value in request.updates.items():
            if hasattr(TestCase, key):
                valid_updates[key] = value
        
        if not valid_updates:
            return {"code": 200, "message": "没有有效的更新字段", "data": None}
        
        # 分批处理，每批1000条，避免SQL参数过多
        batch_size = 1000
        updated_count = 0
        
        for i in range(0, len(request.ids), batch_size):
            batch_ids = request.ids[i:i + batch_size]
            
            # 如果修改了关键字段，需要同时重置已评审/评审未通过的用例状态
            if critical_field_updated:
                # 先把需要重置状态的用例一起处理
                update_values = {getattr(TestCase, k): v for k, v in valid_updates.items()}
                
                # 对已评审/评审未通过的用例，额外重置状态为待评审
                db.query(TestCase).filter(
                    TestCase.id.in_(batch_ids),
                    TestCase.status.in_(['REVIEWED', 'REJECTED'])
                ).update(
                    {**{getattr(TestCase, k): v for k, v in valid_updates.items()},
                     TestCase.status: 'PENDING',
                     TestCase.updated_by: current_user.id},
                    synchronize_session=False
                )
                
                # 对其他状态的用例，只更新指定字段
                db.query(TestCase).filter(
                    TestCase.id.in_(batch_ids),
                    ~TestCase.status.in_(['REVIEWED', 'REJECTED'])
                ).update(
                    {**{getattr(TestCase, k): v for k, v in valid_updates.items()},
                     TestCase.updated_by: current_user.id},
                    synchronize_session=False
                )
            else:
                # 不涉及关键字段，直接批量更新
                db.query(TestCase).filter(
                    TestCase.id.in_(batch_ids)
                ).update(
                    {**{getattr(TestCase, k): v for k, v in valid_updates.items()},
                     TestCase.updated_by: current_user.id},
                    synchronize_session=False
                )
            
            updated_count += len(batch_ids)
        
        db.commit()
        
        # 记录日志
        update_fields = ", ".join(valid_updates.keys())
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.TESTCASES,
            action=LogAction.UPDATE,
            description=f"批量更新 {updated_count} 个测试用例的字段: {update_fields}",
            request=req
        )
        
        return {"code": 200, "message": f"成功更新 {updated_count} 个用例", "data": None}
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"批量更新失败: {str(e)}")


# 按ID导出测试用例
class ExportByIdsRequest(BaseModel):
    project_id: Optional[int] = None  # 兼容旧调用，不再用于过滤
    ids: list[int]

# 单次导出的 ID 数量上限，超出时拒绝并告警
EXPORT_BY_IDS_LIMIT = 5000

@router.post("/export-by-ids")
def export_by_ids(
    request: ExportByIdsRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """按ID导出选中的测试用例（支持跨用例库）"""
    # 权限检查
    if not is_super_admin(current_user) and not has_permission(current_user, db, 'testcases.export'):
        raise HTTPException(status_code=403, detail="您没有导出测试用例的权限")

    # 大批量导出记录日志（不拒绝，仅留审计记录）
    if len(request.ids) > EXPORT_BY_IDS_LIMIT:
        client_ip = req.headers.get("X-Forwarded-For", "").split(",")[0].strip() \
            or req.headers.get("X-Real-IP") \
            or (req.client.host if req.client else "unknown")
        logger.info(
            f"[大批量导出] 用户 {current_user.username}({client_ip}) "
            f"导出大量用例: ids数量={len(request.ids)}"
        )
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.TESTCASES,
            action=LogAction.EXPORT,
            description=f"[大批量导出] 用户 {current_user.username}({client_ip}) 导出 {len(request.ids)} 个用例（超过 {EXPORT_BY_IDS_LIMIT} 条审计阈值）",
            request=req,
            response_status=200
        )

    try:
        query = db.query(TestCase).filter(
            TestCase.id.in_(request.ids)
        )

        # 按用例标准排序（模块顺序 → sort_order → case_number 自然排序），与列表页保持一致
        if request.ids:
            # 取第一个用例的 primary_project_id 作为排序基准（同一次导出通常属于同一项目）
            first_tc = db.query(TestCase.primary_project_id).filter(
                TestCase.id == request.ids[0]
            ).first()
            pid_list = [first_tc.primary_project_id] if first_tc and first_tc.primary_project_id else None
            query = apply_testcase_sort(query, pid_list, db=db if pid_list else None)
        
        testcases = query.all()
        
        excel_data = generate_excel(testcases, db=db)
        
        # 超阈值时已在前面写过审计日志，正常情况下写常规导出日志
        if len(request.ids) <= EXPORT_BY_IDS_LIMIT:
            log_operation(
                db=db,
                user_id=current_user.id,
                username=current_user.username,
                module=LogModule.TESTCASES,
                action=LogAction.EXPORT,
                description=f"导出选中的 {len(testcases)} 个测试用例到Excel",
                request=req
            )
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=testcases_selected.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


class ReorderRequest(BaseModel):
    """拖拽排序请求"""
    testcase_id: int
    target_position: int  # 目标位置（在目标用例之前或之后）
    target_testcase_id: Optional[int] = None  # 目标用例ID
    insert_before: bool = True  # True=插入到目标前面，False=插入到目标后面

@router.post("/reorder")
def reorder_testcase(
    request: ReorderRequest,
    req: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """拖拽调整用例顺序"""
    try:
        # 获取被拖拽的用例
        testcase = db.query(TestCase).filter(TestCase.id == request.testcase_id).first()
        if not testcase:
            raise HTTPException(status_code=404, detail="测试用例不存在")
        
        # 获取目标用例
        target_testcase = db.query(TestCase).filter(TestCase.id == request.target_testcase_id).first()
        if not target_testcase:
            raise HTTPException(status_code=404, detail="目标用例不存在")
        
        # 检查是否在同一模块路径前缀下
        # 允许在同一父模块下跨子模块拖拽
        
        # 如果跨模块拖拽，更新被拖拽用例的模块路径
        if testcase.module != target_testcase.module:
            testcase.module = target_testcase.module
        
        # 获取同一模块下的所有用例（按当前排序）
        query = db.query(TestCase).filter(
            TestCase.primary_project_id == testcase.primary_project_id,
            TestCase.module == testcase.module
        )
        
        # 按模块sort_order → sort_order → case_number自然排序
        from sqlalchemy import text
        natural_sort = text("""
            LPAD(COALESCE(SUBSTRING(test_cases.case_number FROM '(\d+)$'), '0'), 20, '0')
        """)
        
        module_sort_expr = build_module_sort_key(db, [testcase.primary_project_id])
        query = query.order_by(
            module_sort_expr,
            TestCase.sort_order.asc(),
            natural_sort,
            TestCase.id.asc()
        )
        
        testcases = query.all()
        
        # 找到被拖拽用例的当前位置
        current_index = next((i for i, tc in enumerate(testcases) if tc.id == request.testcase_id), None)
        if current_index is None:
            raise HTTPException(status_code=404, detail="无法找到用例位置")
        
        # 找到目标位置
        target_index = next((i for i, tc in enumerate(testcases) if tc.id == request.target_testcase_id), None)
        if target_index is None:
            raise HTTPException(status_code=404, detail="无法找到目标位置")
        
        # 调整目标位置（如果是插入到后面）
        if not request.insert_before and target_index < len(testcases) - 1:
            target_index += 1
        
        # 移除当前用例
        testcases.pop(current_index)
        
        # 插入到新位置
        testcases.insert(target_index if target_index <= current_index else target_index - 1, testcase)
        
        # 重新分配排序号（使用10的倍数，便于后续插入）
        for i, tc in enumerate(testcases):
            tc.sort_order = (i + 1) * 10
        
        db.commit()
        
        log_operation(
            db=db,
            user_id=current_user.id,
            username=current_user.username,
            module=LogModule.TESTCASES,
            action=LogAction.UPDATE,
            description=f"调整测试用例顺序：{testcase.name}（编号: {testcase.case_number}）",
            request=req
        )
        
        return {"code": 200, "message": "排序成功", "data": None}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"调整用例顺序错误: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"调整顺序失败: {str(e)}")
